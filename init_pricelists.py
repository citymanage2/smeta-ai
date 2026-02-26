"""
Скрипт для инициализации примеров прайс-листов.
Запустите этот скрипт один раз перед использованием сервиса.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

# Директория для прайс-листов
pricelists_dir = Path(__file__).parent / "pricelists"
pricelists_dir.mkdir(exist_ok=True)

# ==================== price_works.xlsx ====================
wb_works = Workbook()
ws_works = wb_works.active
ws_works.title = "Прайс на работы"

# Заголовок
headers_works = ["Наименование", "Ед. изм.", "Подрядчик 1", "Подрядчик 2"]
ws_works.append(headers_works)

# Форматирование заголовка
for cell in ws_works[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")

# Пример данных
sample_works = [
    ["Демонтаж стен из кирпича", "м²", 450, 480],
    ["Кладка стен из кирпича", "м²", 550, 600],
    ["Штукатурка поверхностей", "м²", 320, 350],
    ["Шпатлевание поверхностей", "м²", 280, 300],
    ["Окраска стен", "м²", 200, 220],
    ["Монтаж входной двери", "комп.", 3500, 4000],
    ["Монтаж окна ПВХ", "комп.", 2000, 2200],
    ["Укладка ламината", "м²", 400, 450],
    ["Укладка керамической плитки", "м²", 600, 700],
    ["Установка унитаза", "шт.", 2000, 2500],
]

for row in sample_works:
    ws_works.append(row)

# Ширина колонок
ws_works.column_dimensions['A'].width = 35
ws_works.column_dimensions['B'].width = 12
ws_works.column_dimensions['C'].width = 15
ws_works.column_dimensions['D'].width = 15

# Сохранение
wb_works.save(pricelists_dir / "price_works.xlsx")
print("✓ Создан price_works.xlsx")

# ==================== price_materials.xlsx ====================
wb_materials = Workbook()
ws_materials = wb_materials.active
ws_materials.title = "Прайс на материалы"

# Заголовок
headers_materials = ["Наименование", "Ед. изм.", "Цена, руб."]
ws_materials.append(headers_materials)

# Форматирование заголовка
for cell in ws_materials[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")

# Пример данных
sample_materials = [
    ["Кирпич рядовой красный", "шт.", 14.5],
    ["Цемент М400", "кг", 8.5],
    ["Песок строительный", "м³", 450],
    ["Щебень фракция 20-40 мм", "м³", 650],
    ["Стекловидный блок", "м²", 280],
    ["Монтажная пена", "баллон", 250],
    ["Гипсокартон стандартный", "м²", 180],
    ["Ламинат класс 32", "м²", 350],
    ["Керамическая плитка", "м²", 550],
    ["Краска акриловая 14 кг", "л", 450],
    ["Грунтовка глубокого проникновения", "л", 180],
    ["Дверное полотно ДСП", "шт.", 1500],
    ["Окно ПВХ 1400х1400", "комп.", 8000],
    ["Унитаз керамический", "шт.", 5000],
    ["Раковина для ванной", "шт.", 3000],
]

for row in sample_materials:
    ws_materials.append(row)

# Ширина колонок
ws_materials.column_dimensions['A'].width = 35
ws_materials.column_dimensions['B'].width = 12
ws_materials.column_dimensions['C'].width = 15

# Сохранение
wb_materials.save(pricelists_dir / "price_materials.xlsx")
print("✓ Создан price_materials.xlsx")

print("\n✓ Прайс-листы инициализированы успешно!")
print(f"Расположение: {pricelists_dir}")
