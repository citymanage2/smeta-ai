import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelBuilder:
    """Построитель Excel файлов"""

    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_list_workbook(self, data: List[Dict[str, Any]]) -> bytes:
        """Создать Excel файл с Перечнем работ и материалов"""
        
        wb = Workbook()
        
        # Лист 1: Полный перечень
        ws_all = wb.active
        ws_all.title = "Перечень работ и материалов"
        self._create_list_sheet(ws_all, data)
        
        # Лист 2: Только работы
        works = [item for item in data if item.get('type') == 'Работа']
        ws_works = wb.create_sheet("Перечень работ")
        self._create_works_sheet(ws_works, works)
        
        # Лист 3: Только материалы
        materials = [item for item in data if item.get('type') == 'Материал']
        ws_materials = wb.create_sheet("Перечень материалов")
        self._create_materials_sheet(ws_materials, materials)
        
        # Сохранить в памяти
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def create_estimate_workbook(self, data: List[Dict[str, Any]]) -> bytes:
        """Создать Excel файл со сметой"""
        
        wb = Workbook()
        
        # Лист 1: Полная смета
        ws_all = wb.active
        ws_all.title = "Перечень работ и материалов"
        self._create_estimate_sheet(ws_all, data)
        
        # Лист 2: Только работы
        works = [item for item in data if item.get('type') == 'Работа']
        ws_works = wb.create_sheet("Перечень работ")
        self._create_estimate_works_sheet(ws_works, works)
        
        # Лист 3: Только материалы
        materials = [item for item in data if item.get('type') == 'Материал']
        ws_materials = wb.create_sheet("Перечень материалов")
        self._create_estimate_materials_sheet(ws_materials, materials)
        
        # Сохранить в памяти
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_list_sheet(self, ws, data: List[Dict[str, Any]]):
        """Создать лист с Перечнем"""
        
        # Заголовок
        headers = ["№ п/п", "Работа/Материал", "Наименование", "Ед. изм.", "Кол-во"]
        ws.append(headers)
        
        # Форматирование заголовка
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Данные
        for idx, item in enumerate(data, 1):
            ws.append([
                idx,
                item.get('type', ''),
                item.get('name', ''),
                item.get('unit', ''),
                item.get('quantity', '')
            ])
            
            # Чередующаяся заливка
            if idx % 2 == 0:
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Границы
            for cell in ws[idx + 1]:
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Установить ширину колонок
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        # Закрепить заголовок
        ws.freeze_panes = "A2"

    def _create_works_sheet(self, ws, data: List[Dict[str, Any]]):
        """Создать лист только с работами"""
        
        headers = ["№ п/п", "Наименование", "Ед. изм.", "Кол-во"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for idx, item in enumerate(data, 1):
            ws.append([
                idx,
                item.get('name', ''),
                item.get('unit', ''),
                item.get('quantity', '')
            ])
            
            if idx % 2 == 0:
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            for cell in ws[idx + 1]:
                cell.border = self.thin_border
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.freeze_panes = "A2"

    def _create_materials_sheet(self, ws, data: List[Dict[str, Any]]):
        """Создать лист только с материалами"""
        
        headers = ["№ п/п", "Наименование", "Ед. изм.", "Кол-во"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for idx, item in enumerate(data, 1):
            ws.append([
                idx,
                item.get('name', ''),
                item.get('unit', ''),
                item.get('quantity', '')
            ])
            
            if idx % 2 == 0:
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            for cell in ws[idx + 1]:
                cell.border = self.thin_border
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.freeze_panes = "A2"

    def _create_estimate_sheet(self, ws, data: List[Dict[str, Any]]):
        """Создать лист со сметой"""
        
        headers = [
            "№ п/п", "Работа/Материал", "Наименование", "Ед. изм.", "Кол-во",
            "Цена за ед. (Работа)", "Стоимость (Работа), руб.",
            "Цена за ед. (Материал)", "Стоимость (Материал), руб.",
            "Наименование в прайсе", "Примечание"
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        total_work = 0
        total_material = 0
        
        for idx, item in enumerate(data, 1):
            work_cost = 0
            material_cost = 0
            
            if item.get('type') == 'Работа':
                work_cost = (item.get('quantity', 0) or 0) * (item.get('price_work_per_unit', 0) or 0)
                total_work += work_cost
            else:
                material_cost = (item.get('quantity', 0) or 0) * (item.get('price_material_per_unit', 0) or 0)
                total_material += material_cost
            
            ws.append([
                idx,
                item.get('type', ''),
                item.get('name', ''),
                item.get('unit', ''),
                item.get('quantity', ''),
                item.get('price_work_per_unit', '') if item.get('type') == 'Работа' else '',
                work_cost if item.get('type') == 'Работа' else '',
                item.get('price_material_per_unit', '') if item.get('type') == 'Материал' else '',
                material_cost if item.get('type') == 'Материал' else '',
                item.get('name_in_pricelist', ''),
                item.get('note', '')
            ])
            
            # Выделение строк без цены
            has_price = (item.get('price_work_per_unit') or item.get('price_material_per_unit'))
            if not has_price:
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif idx % 2 == 0:
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            for cell in ws[idx + 1]:
                cell.border = self.thin_border
        
        # Итоговые строки
        last_row = len(data) + 2
        ws[f"A{last_row}"] = "ИТОГО БЕЗ НДС"
        ws[f"F{last_row}"] = total_work
        ws[f"H{last_row}"] = total_material
        
        ws[f"A{last_row + 1}"] = "НДС 22%"
        ws[f"F{last_row + 1}"] = round(total_work * 0.22, 2)
        ws[f"H{last_row + 1}"] = round(total_material * 0.22, 2)
        
        ws[f"A{last_row + 2}"] = "ИТОГО С НДС"
        ws[f"F{last_row + 2}"] = round(total_work * 1.22, 2)
        ws[f"H{last_row + 2}"] = round(total_material * 1.22, 2)
        
        # Форматирование итогов
        for row in [last_row, last_row + 1, last_row + 2]:
            ws[f"A{row}"].font = Font(bold=True)
            for col in ['F', 'H']:
                ws[f"{col}{row}"].font = Font(bold=True)
                ws[f"{col}{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        # Установить ширину колонок
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 25
        ws.column_dimensions['K'].width = 20
        
        ws.freeze_panes = "A2"

    def _create_estimate_works_sheet(self, ws, data: List[Dict[str, Any]]):
        """Создать лист только с работами в смете"""
        
        headers = ["№ п/п", "Наименование", "Ед. изм.", "Кол-во", "Цена за ед.", "Стоимость, руб.", "Наименование в прайсе", "Примечание"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        total = 0
        
        for idx, item in enumerate(data, 1):
            cost = (item.get('quantity', 0) or 0) * (item.get('price_work_per_unit', 0) or 0)
            total += cost
            
            ws.append([
                idx,
                item.get('name', ''),
                item.get('unit', ''),
                item.get('quantity', ''),
                item.get('price_work_per_unit', ''),
                cost,
                item.get('name_in_pricelist', ''),
                item.get('note', '')
            ])
            
            if not item.get('price_work_per_unit'):
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif idx % 2 == 0:
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            for cell in ws[idx + 1]:
                cell.border = self.thin_border
        
        # Итоги
        last_row = len(data) + 2
        ws[f"A{last_row}"] = "ИТОГО БЕЗ НДС"
        ws[f"F{last_row}"] = total
        
        ws[f"A{last_row + 1}"] = "НДС 22%"
        ws[f"F{last_row + 1}"] = round(total * 0.22, 2)
        
        ws[f"A{last_row + 2}"] = "ИТОГО С НДС"
        ws[f"F{last_row + 2}"] = round(total * 1.22, 2)
        
        for row in [last_row, last_row + 1, last_row + 2]:
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"F{row}"].font = Font(bold=True)
            ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 20
        ws.freeze_panes = "A2"

    def create_scan_excel_workbook(self, scan_data: Dict[str, Any]) -> bytes:
        """Создать Excel файл из распознанных данных скана сметы"""

        wb = Workbook()
        ws = wb.active
        ws.title = "Смета"

        self._write_scan_sheet(ws, scan_data)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _write_scan_sheet(self, ws, scan_data: Dict[str, Any]):
        """Заполнить лист данными из распознанного скана"""

        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        section_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        total_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        alt_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        warn_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        white_font = Font(bold=True, color="FFFFFF", size=10)
        bold_font = Font(bold=True)
        section_font = Font(bold=True, color="1F4788")
        number_fmt = '#,##0.00'

        # ── Шапка документа ──────────────────────────────────────────────
        metadata = scan_data.get("metadata", {})

        def add_meta_row(label: str, value: Optional[str]):
            if value:
                ws.append([label, value])
                ws[ws.max_row][1 - 1 + 1].font = bold_font  # column A
                ws.merge_cells(
                    start_row=ws.max_row, start_column=2,
                    end_row=ws.max_row, end_column=7
                )

        if metadata.get("object_name"):
            add_meta_row("Объект:", metadata["object_name"])
        if metadata.get("estimate_number"):
            add_meta_row("Номер сметы:", metadata["estimate_number"])
        if metadata.get("date"):
            add_meta_row("Дата:", metadata["date"])
        if metadata.get("author"):
            add_meta_row("Составитель:", metadata["author"])
        if metadata.get("estimate_type"):
            add_meta_row("Тип сметы:", metadata["estimate_type"])
        if metadata.get("normalization_system"):
            add_meta_row("Нормирование:", metadata["normalization_system"])

        if ws.max_row > 0:
            ws.append([])  # пустая строка-разделитель

        # ── Заголовок таблицы ─────────────────────────────────────────────
        col_headers = [
            "№ п/п", "Шифр / код", "Наименование работ и затрат",
            "Ед. изм.", "Объём / кол-во", "Цена за ед., руб.", "Сумма, руб."
        ]
        header_row = ws.max_row + 1
        ws.append(col_headers)
        for cell in ws[header_row]:
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.thin_border
        ws.row_dimensions[header_row].height = 32

        # ── Позиции по разделам ──────────────────────────────────────────
        sections = scan_data.get("sections", [])
        position_row_start = header_row + 1
        item_counter = 0

        for section in sections:
            section_name = section.get("name")
            items = section.get("items", [])

            # Заголовок раздела
            if section_name:
                ws.append(["", "", section_name, "", "", "", ""])
                sec_row = ws.max_row
                for cell in ws[sec_row]:
                    cell.fill = section_fill
                    cell.font = section_font
                    cell.border = self.thin_border
                ws.merge_cells(
                    start_row=sec_row, start_column=3,
                    end_row=sec_row, end_column=7
                )

            # Позиции раздела
            for item in items:
                item_counter += 1
                unreadable = item.get("unreadable", False)

                quantity = item.get("quantity")
                price = item.get("price_per_unit")
                total = item.get("total")

                ws.append([
                    item.get("number", item_counter),
                    item.get("code") or "",
                    item.get("name", ""),
                    item.get("unit") or "",
                    quantity if quantity is not None else "",
                    price if price is not None else "",
                    total if total is not None else "",
                ])
                row = ws.max_row
                fill = warn_fill if unreadable else (alt_fill if item_counter % 2 == 0 else None)
                for col_idx, cell in enumerate(ws[row], 1):
                    cell.border = self.thin_border
                    cell.alignment = Alignment(
                        horizontal="right" if col_idx >= 5 else "left",
                        vertical="center",
                        wrap_text=(col_idx == 3)
                    )
                    if fill:
                        cell.fill = fill
                    if col_idx in (5, 6, 7) and isinstance(cell.value, (int, float)):
                        cell.number_format = number_fmt

            # Промежуточный итог раздела
            subtotal = section.get("subtotal")
            if subtotal is not None and section_name:
                ws.append(["", "", f"Итого по разделу: {section_name}", "", "", "", subtotal])
                row = ws.max_row
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
                for cell in ws[row]:
                    cell.font = bold_font
                    cell.fill = total_fill
                    cell.border = self.thin_border
                ws[f"G{row}"].number_format = number_fmt

        # ── Итоговый блок ─────────────────────────────────────────────────
        totals = scan_data.get("totals", {})
        ws.append([])

        def add_total_row(label: str, value):
            if value is not None:
                ws.append(["", "", label, "", "", "", value])
                row = ws.max_row
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
                for cell in ws[row]:
                    cell.font = bold_font
                    cell.fill = total_fill
                    cell.border = self.thin_border
                ws[f"G{row}"].number_format = number_fmt

        add_total_row("Итого без накладных:", totals.get("subtotal"))
        add_total_row("Накладные расходы:", totals.get("overhead"))
        add_total_row("Сметная прибыль:", totals.get("profit"))
        add_total_row("НДС:", totals.get("vat"))
        if totals.get("grand_total") is not None:
            ws.append(["", "", "ИТОГО ПО СМЕТЕ:", "", "", "", totals["grand_total"]])
            row = ws.max_row
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            for cell in ws[row]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.border = self.thin_border
            ws[f"G{row}"].number_format = number_fmt

        # ── Ширина колонок ────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 16

        ws.freeze_panes = f"A{header_row + 1}"

    def _create_estimate_materials_sheet(self, ws, data: List[Dict[str, Any]]):
        """Создать лист только с материалами в смете"""
        
        headers = ["№ п/п", "Наименование", "Ед. изм.", "Кол-во", "Цена за ед.", "Стоимость, руб.", "Наименование в прайсе", "Примечание"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        total = 0
        
        for idx, item in enumerate(data, 1):
            cost = (item.get('quantity', 0) or 0) * (item.get('price_material_per_unit', 0) or 0)
            total += cost
            
            ws.append([
                idx,
                item.get('name', ''),
                item.get('unit', ''),
                item.get('quantity', ''),
                item.get('price_material_per_unit', ''),
                cost,
                item.get('name_in_pricelist', ''),
                item.get('note', '')
            ])
            
            if not item.get('price_material_per_unit'):
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif idx % 2 == 0:
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            for cell in ws[idx + 1]:
                cell.border = self.thin_border
        
        # Итоги
        last_row = len(data) + 2
        ws[f"A{last_row}"] = "ИТОГО БЕЗ НДС"
        ws[f"F{last_row}"] = total
        
        ws[f"A{last_row + 1}"] = "НДС 22%"
        ws[f"F{last_row + 1}"] = round(total * 0.22, 2)
        
        ws[f"A{last_row + 2}"] = "ИТОГО С НДС"
        ws[f"F{last_row + 2}"] = round(total * 1.22, 2)
        
        for row in [last_row, last_row + 1, last_row + 2]:
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"F{row}"].font = Font(bold=True)
            ws[f"F{row}"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 20
        ws.freeze_panes = "A2"
