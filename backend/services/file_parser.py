import os
import json
import tempfile
from pathlib import Path
from typing import Dict, List, Any
import pdfplumber
import openpyxl
from openpyxl import load_workbook
import pandas as pd
from lxml import etree
import zipfile

class FileParser:
    """Парсер для различных форматов файлов"""

    @staticmethod
    def parse_pdf(file_path: str) -> str:
        """Извлечь текст из PDF"""
        try:
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            raise Exception(f"Ошибка при парсинге PDF: {str(e)}")

    @staticmethod
    def parse_excel(file_path: str) -> Dict[str, Any]:
        """Парсить Excel файл и вернуть структурированные данные"""
        try:
            workbook = load_workbook(file_path)
            result = {}
            
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                data = []
                headers = []
                
                for idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if idx == 0:
                        headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(row)]
                    else:
                        row_data = {headers[i]: val for i, val in enumerate(row)}
                        data.append(row_data)
                
                result[sheet_name] = {
                    "headers": headers,
                    "data": data
                }
            
            return result
        except Exception as e:
            raise Exception(f"Ошибка при парсинге Excel: {str(e)}")

    @staticmethod
    def parse_xml(file_path: str) -> Dict[str, Any]:
        """Парсить XML файл (ГрандСмета)"""
        try:
            tree = etree.parse(file_path)
            root = tree.getroot()
            
            # Конвертировать XML в словарь
            result = FileParser._xml_to_dict(root)
            return result
        except Exception as e:
            raise Exception(f"Ошибка при парсинге XML: {str(e)}")

    @staticmethod
    def parse_gsn(file_path: str) -> Dict[str, Any]:
        """Парсить GSN файл (ГрандСмета, zip архив)"""
        try:
            # GSN - это ZIP архив с XML файлами
            with tempfile.TemporaryDirectory() as tmpdir:
                with zipfile.ZipFile(file_path, 'r') as zip_ref:
                    zip_ref.extractall(tmpdir)
                
                # Найти XML файл с данными (обычно main.xml или content.xml)
                xml_files = list(Path(tmpdir).glob("**/*.xml"))
                if not xml_files:
                    raise Exception("XML файлы не найдены в GSN архиве")
                
                # Парсить первый найденный XML
                result = FileParser.parse_xml(str(xml_files[0]))
                return result
        except Exception as e:
            raise Exception(f"Ошибка при парсинге GSN: {str(e)}")

    @staticmethod
    def _xml_to_dict(element) -> Dict[str, Any]:
        """Рекурсивно конвертировать XML элемент в словарь"""
        result = {}
        
        # Атрибуты
        if element.attrib:
            result['@attributes'] = element.attrib
        
        # Дочерние элементы
        children = {}
        for child in element:
            child_data = FileParser._xml_to_dict(child)
            if child.tag in children:
                if not isinstance(children[child.tag], list):
                    children[child.tag] = [children[child.tag]]
                children[child.tag].append(child_data)
            else:
                children[child.tag] = child_data
        
        if children:
            result.update(children)
        
        # Текстовое содержимое
        if element.text and element.text.strip():
            if children or element.attrib:
                result['#text'] = element.text.strip()
            else:
                return element.text.strip()
        
        return result if result else None

    @staticmethod
    def detect_file_type(file_name: str) -> str:
        """Определить тип файла по расширению"""
        ext = Path(file_name).suffix.lower()
        
        if ext == ".pdf":
            return "pdf"
        elif ext in [".xlsx", ".xls"]:
            return "excel"
        elif ext == ".xml":
            return "xml"
        elif ext == ".gsn":
            return "gsn"
        else:
            return "unknown"

    @staticmethod
    def parse_file(file_path: str) -> Dict[str, Any]:
        """Парсить файл в зависимости от типа"""
        file_type = FileParser.detect_file_type(file_path)
        
        if file_type == "pdf":
            text = FileParser.parse_pdf(file_path)
            return {
                "type": "pdf",
                "content": text,
                "text": text
            }
        elif file_type == "excel":
            data = FileParser.parse_excel(file_path)
            return {
                "type": "excel",
                "content": data,
                "sheets": data
            }
        elif file_type == "xml":
            data = FileParser.parse_xml(file_path)
            return {
                "type": "xml",
                "content": json.dumps(data, ensure_ascii=False, indent=2),
                "data": data
            }
        elif file_type == "gsn":
            data = FileParser.parse_gsn(file_path)
            return {
                "type": "gsn",
                "content": json.dumps(data, ensure_ascii=False, indent=2),
                "data": data
            }
        else:
            raise Exception(f"Неподдерживаемый тип файла: {file_type}")
