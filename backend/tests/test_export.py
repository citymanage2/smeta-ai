import pytest
from httpx import AsyncClient


@pytest.mark.asyncio
async def test_export_xlsx_returns_file(async_client: AsyncClient, user_token: str):
    """Creating a project and exporting it as xlsx returns valid bytes."""
    create_resp = await async_client.post(
        "/projects",
        json={"name": "Экспорт проект", "description": "Тест"},
        headers={"Authorization": user_token},
    )
    assert create_resp.status_code == 200
    project_id = create_resp.json()["id"]

    resp = await async_client.get(
        f"/projects/{project_id}/export",
        params={"format": "xlsx"},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(resp.content) > 0
    assert "attachment" in resp.headers.get("content-disposition", "")


@pytest.mark.asyncio
async def test_export_pdf_returns_file(async_client: AsyncClient, user_token: str):
    """Creating a project and exporting it as pdf returns valid bytes."""
    create_resp = await async_client.post(
        "/projects",
        json={"name": "PDF проект"},
        headers={"Authorization": user_token},
    )
    project_id = create_resp.json()["id"]

    resp = await async_client.get(
        f"/projects/{project_id}/export",
        params={"format": "pdf"},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/pdf")
    assert len(resp.content) > 0


@pytest.mark.asyncio
async def test_export_invalid_format_returns_400(async_client: AsyncClient, user_token: str):
    create_resp = await async_client.post(
        "/projects",
        json={"name": "Проект формат"},
        headers={"Authorization": user_token},
    )
    project_id = create_resp.json()["id"]

    resp = await async_client.get(
        f"/projects/{project_id}/export",
        params={"format": "docx"},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 400


@pytest.mark.asyncio
async def test_export_project_not_found(async_client: AsyncClient, user_token: str):
    resp = await async_client.get(
        "/projects/c1000000-0000-0000-0000-000000000099/export",
        params={"format": "xlsx"},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_export_xlsx_is_valid_workbook(async_client: AsyncClient, user_token: str):
    """Exported xlsx can be opened with openpyxl."""
    import io
    import openpyxl

    create_resp = await async_client.post(
        "/projects",
        json={"name": "Валидный xlsx"},
        headers={"Authorization": user_token},
    )
    project_id = create_resp.json()["id"]

    resp = await async_client.get(
        f"/projects/{project_id}/export",
        params={"format": "xlsx"},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Задачи" in wb.sheetnames
    assert "Исходные файлы" in wb.sheetnames
