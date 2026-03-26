import io
import re
from decimal import Decimal
from datetime import datetime, timezone

import openpyxl
import pytest

from app.utils.xlsx_exporter import generate_project_xlsx


class _FakeProject:
    id = "proj-1"
    name = "Тест проект"
    description = "Описание"


class _FakeTask:
    def __init__(self, tid, task_type, estimation_status, cost, created_at):
        self.id = tid
        self.task_type = task_type
        self.estimation_status = estimation_status
        self.cost = cost
        self.created_at = created_at


class _FakeResult:
    def __init__(self, task_id, file_name, slot):
        self.task_id = task_id
        self.file_name = file_name
        self.slot = slot


_DT = datetime(2026, 3, 26, 12, 0, 0, tzinfo=timezone.utc)


def _make_data():
    project = _FakeProject()
    tasks = [
        _FakeTask("t1", "SMETA_FROM_LIST", "estimated", Decimal("1500000"), _DT),
        _FakeTask("t2", "LIST_FROM_TZ", "not_applicable", None, _DT),
    ]
    result1 = _FakeResult("t1", "smeta.xlsx", "estimate")
    slot_results = {
        "source": [],
        "estimate": [(tasks[0], result1)],
        "optimized": [],
    }
    return project, tasks, slot_results


def test_xlsx_returns_bytes():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_xlsx_has_tasks_sheet():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Задачи" in wb.sheetnames


def test_xlsx_tasks_sheet_headers():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Задачи"]
    assert ws.cell(1, 1).value == "Тип задачи"
    assert ws.cell(1, 2).value == "Статус сметы"
    assert ws.cell(1, 3).value == "Стоимость (₽)"
    assert ws.cell(1, 4).value == "Дата создания"


def test_xlsx_tasks_sheet_rows():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Задачи"]
    # Row 2 = first task
    assert ws.cell(2, 1).value == "Смета из ТЗ"
    assert ws.cell(2, 2).value == "Рассчитано"
    assert ws.cell(2, 3).value == 1500000.0
    # Row 3 = second task
    assert ws.cell(3, 1).value == "Список из ТЗ"
    assert ws.cell(3, 2).value == "—"
    assert ws.cell(3, 3).value is None


def test_xlsx_total_row_is_bold():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Задачи"]
    total_row = len(tasks) + 2
    assert ws.cell(total_row, 1).font.bold is True
    assert ws.cell(total_row, 1).value == "ИТОГО"


def test_xlsx_has_all_slot_sheets():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Исходные файлы" in wb.sheetnames
    assert "Расчёты" in wb.sheetnames
    assert "Оптимизированные" in wb.sheetnames


def test_xlsx_slot_sheet_with_file_has_hyperlink():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Расчёты"]
    # Row 2 has the estimate file
    assert ws.cell(2, 2).value == "smeta.xlsx"
    link_value = ws.cell(2, 3).value
    assert isinstance(link_value, str) and link_value.startswith("=HYPERLINK(")
    assert "http://localhost:8000/tasks/t1/files/estimate/download" in link_value


def test_xlsx_date_column_has_no_time():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Задачи"]
    for row_idx in range(2, len(tasks) + 2):
        date_val = ws.cell(row_idx, 4).value
        assert date_val is not None
        assert ":" not in str(date_val), f"Date value {date_val!r} must not contain time"
        # Must match DD.MM.YYYY pattern
        assert re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", str(date_val)), (
            f"Date value {date_val!r} does not match DD.MM.YYYY"
        )


def test_xlsx_empty_slot_shows_placeholder():
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Исходные файлы"]
    assert ws.cell(2, 1).value == "Файлы отсутствуют"


def test_xlsx_total_row_summary_counts():
    # Fixture has: t1 estimated=1, t2 not_applicable (none of the three) → 0/1/0
    project, tasks, slot_results = _make_data()
    data = generate_project_xlsx(project, tasks, slot_results, "http://localhost:8000")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Задачи"]
    total_row = len(tasks) + 2
    expected = "не рассчитано: 0 / рассчитано: 1 / оптимизировано: 0"
    assert ws.cell(total_row, 2).value == expected
