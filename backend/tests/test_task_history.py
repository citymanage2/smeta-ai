"""Tests for GET /tasks/{task_id}/history and POST /tasks/{task_id}/history/{id}/revert."""
import io
import pytest
import pytest_asyncio
import openpyxl
from datetime import datetime, timezone, timedelta
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.models.task import Task
from app.models.result import TaskResult
from app.models.history import TaskHistory


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"
    headers = [
        "№", "Тип", "Наименование", "Ед. изм.", "Кол-во",
        "Цена работы (за ед.)", "Цена материала (за ед.)",
        "Стоимость работ", "Стоимость материалов",
        "Итого без НДС", "НДС (20%)", "Итого с НДС",
        "Наименование в прайсе", "Источники", "Примечание",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=2, value="Материал")
    ws.cell(row=2, column=3, value="Кирпич М150")
    ws.cell(row=2, column=4, value="шт")
    ws.cell(row=2, column=5, value=1000)
    ws.cell(row=2, column=7, value=10.0)
    ws.cell(row=2, column=10, value=10000.0)
    ws.cell(row=2, column=11, value=2000.0)
    ws.cell(row=2, column=12, value=12000.0)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


TASK_ID = "d1000000-0000-0000-0000-000000000001"
_TASK_ID_HEX = TASK_ID.replace("-", "")

ENTRY_ID_1 = "e1000000-0000-0000-0000-000000000001"
ENTRY_ID_2 = "e2000000-0000-0000-0000-000000000002"

_ENTRY_ID_1_HEX = ENTRY_ID_1.replace("-", "")
_ENTRY_ID_2_HEX = ENTRY_ID_2.replace("-", "")

_NOW = datetime(2026, 3, 26, 10, 0, 0, tzinfo=timezone.utc)
_LATER = _NOW + timedelta(hours=1)


@pytest_asyncio.fixture
async def seed_history_task(db_session: AsyncSession):
    """Seed a task with one history entry."""
    task = Task(
        id=TASK_ID,
        user_role="user",
        task_type="SMETA_FROM_LIST",
        status="completed",
        estimation_status="optimized",
        input_files=[],
        input_file_data=[],
        chat_history=[],
    )
    db_session.add(task)
    await db_session.flush()

    entry = TaskHistory(
        id=ENTRY_ID_1,
        task_id=TASK_ID,
        operation_type="optimization",
        slot="optimized",
        description="Оптимизация: найдено 1 из 1 аналогов",
        previous_value={"estimation_status": "estimated"},
        new_value={
            "file_name": "optimized.xlsx",
            "file_data_b64": "dGVzdA==",
            "estimation_status": "optimized",
        },
        created_at=_NOW,
    )
    db_session.add(entry)
    await db_session.commit()
    yield
    await db_session.execute(
        text("DELETE FROM task_history WHERE task_id = :tid"), {"tid": _TASK_ID_HEX}
    )
    await db_session.execute(
        text("DELETE FROM tasks WHERE id = :tid"), {"tid": _TASK_ID_HEX}
    )
    await db_session.commit()


# ---------------------------------------------------------------------------
# Tests: GET /history
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_get_history_returns_entries(
    async_client: AsyncClient, user_token: str, seed_history_task
):
    """GET /history returns 200 with list of entries."""
    resp = await async_client.get(
        f"/tasks/{TASK_ID}/history",
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert isinstance(data, list)
    assert len(data) == 1
    assert data[0]["operation_type"] == "optimization"
    assert data[0]["description"] == "Оптимизация: найдено 1 из 1 аналогов"
    assert "id" in data[0]
    assert "created_at" in data[0]
    # file_data_b64 must NOT be in the response
    assert "file_data_b64" not in data[0]
    assert "previous_value" not in data[0]


@pytest.mark.asyncio
async def test_get_history_task_not_found(async_client: AsyncClient, user_token: str):
    """GET /history returns 404 for nonexistent task."""
    resp = await async_client.get(
        "/tasks/nonexistent-id/history",
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Fixtures for revert tests
# ---------------------------------------------------------------------------

@pytest_asyncio.fixture
async def seed_two_history_entries(db_session: AsyncSession):
    """Seed a task with two sequential history entries (for cascade test)."""
    task = Task(
        id=TASK_ID,
        user_role="user",
        task_type="SMETA_FROM_LIST",
        status="completed",
        estimation_status="optimized",
        input_files=[],
        input_file_data=[],
        chat_history=[],
    )
    db_session.add(task)
    await db_session.flush()

    entry1 = TaskHistory(
        id=ENTRY_ID_1,
        task_id=TASK_ID,
        operation_type="optimization",
        slot="optimized",
        description="Оптимизация первая",
        previous_value={"estimation_status": "estimated"},
        new_value={"file_name": "opt1.xlsx", "file_data_b64": "dGVzdA==", "estimation_status": "optimized"},
        created_at=_NOW,
    )
    entry2 = TaskHistory(
        id=ENTRY_ID_2,
        task_id=TASK_ID,
        operation_type="optimization",
        slot="optimized",
        description="Оптимизация вторая",
        previous_value={"file_name": "opt1.xlsx", "file_data_b64": "dGVzdA==", "estimation_status": "optimized"},
        new_value={"file_name": "opt2.xlsx", "file_data_b64": "dGVzdDI=", "estimation_status": "optimized"},
        created_at=_LATER,
    )
    db_session.add(entry1)
    db_session.add(entry2)
    await db_session.commit()
    yield
    await db_session.execute(
        text("DELETE FROM task_history WHERE task_id = :tid"), {"tid": _TASK_ID_HEX}
    )
    await db_session.execute(
        text("DELETE FROM tasks WHERE id = :tid"), {"tid": _TASK_ID_HEX}
    )
    await db_session.commit()


# ---------------------------------------------------------------------------
# Tests: POST /history/{entry_id}/revert
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_revert_no_dependents_executes_immediately(
    async_client: AsyncClient, user_token: str, seed_history_task
):
    """Revert with confirm=False and no dependents executes immediately."""
    resp = await async_client.post(
        f"/tasks/{TASK_ID}/history/{ENTRY_ID_1}/revert",
        json={"confirm": False},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data.get("reverted") is True


@pytest.mark.asyncio
async def test_revert_with_dependents_returns_warning(
    async_client: AsyncClient, user_token: str, seed_two_history_entries
):
    """Revert entry1 with confirm=False returns warning when entry2 exists."""
    resp = await async_client.post(
        f"/tasks/{TASK_ID}/history/{ENTRY_ID_1}/revert",
        json={"confirm": False},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data.get("warning") is True
    assert isinstance(data.get("dependent_entries"), list)
    assert len(data["dependent_entries"]) == 1
    assert data["dependent_entries"][0]["id"] == ENTRY_ID_2


@pytest.mark.asyncio
async def test_revert_confirm_true_cascades(
    async_client: AsyncClient, user_token: str, seed_two_history_entries
):
    """confirm=True executes cascade rollback and returns reverted=True."""
    resp = await async_client.post(
        f"/tasks/{TASK_ID}/history/{ENTRY_ID_1}/revert",
        json={"confirm": True},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    assert resp.json().get("reverted") is True

    # Verify via GET /history: original entries are gone, a revert entry exists
    hist_resp = await async_client.get(
        f"/tasks/{TASK_ID}/history",
        headers={"Authorization": user_token},
    )
    assert hist_resp.status_code == 200
    history = hist_resp.json()
    ids = [e["id"] for e in history]
    assert ENTRY_ID_1 not in ids
    assert ENTRY_ID_2 not in ids
    assert any(e["operation_type"] == "revert" for e in history)


@pytest.mark.asyncio
async def test_revert_entry_not_found(async_client: AsyncClient, user_token: str, seed_history_task):
    """Revert returns 404 for nonexistent history entry."""
    resp = await async_client.post(
        f"/tasks/{TASK_ID}/history/nonexistent-entry-id/revert",
        json={"confirm": False},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 404
