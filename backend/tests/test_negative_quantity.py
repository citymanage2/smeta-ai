"""
Позиция с отрицательным объёмом — вычет, а не работа.

В перечне встречаются строки вида «на каждые 10 мм изменения глубины сверления
добавляется −0,61»: они корректируют объём соседней позиции. Раньше такая строка
шла общим путём — по ней искалась цена (в том числе платным web-поиском), а в
редакторе сметы стоимость считалась как qty × цена и выходила отрицательной
(−0,61 × 1393 = −849,73), занижая итог.

Правило: объём < 0 → цену не ищем, стоимость не считаем, строку оставляем как
есть.

План: plans/2026-07-30-otricatelnyy-obyom-bez-rascheta.md
"""
import sys
from unittest.mock import AsyncMock, MagicMock

import pytest

sys.modules.setdefault("fitz", MagicMock())

from app.utils.price_coercion import coerce_qty_signed, is_negative_qty  # noqa: E402
from app.utils.xlsx_exporter import generate_estimate_xlsx  # noqa: E402
from app.services.task_processor import NEGATIVE_QTY_NOTE, TaskProcessor  # noqa: E402


# ---------------------------------------------------------------------------
# Признак
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw", [-0.61, "-0,61", "-1 139", -5, "-5 шт"])
def test_is_negative_qty_true(raw):
    assert is_negative_qty(raw) is True


@pytest.mark.parametrize("raw", [0, 0.61, "3", None, "", "мусор", float("nan")])
def test_is_negative_qty_false(raw):
    """Пусто и мусор — это не «отрицательный объём», путь для них прежний."""
    assert is_negative_qty(raw) is False


def test_coerce_qty_signed_keeps_sign():
    assert coerce_qty_signed("-0,61") == pytest.approx(-0.61)
    assert coerce_qty_signed(3) == pytest.approx(3.0)
    assert coerce_qty_signed("мусор") is None


# ---------------------------------------------------------------------------
# Excel: объём виден, стоимость не считается
# ---------------------------------------------------------------------------

def _item(**over):
    base = {"type": "Работа", "name": "Сверление", "unit": "100 отверстий",
            "quantity": 0.61, "work_price": 1393, "material_price": None}
    base.update(over)
    return base


def test_xlsx_negative_qty_has_no_cost():
    import openpyxl, io  # noqa: E401

    data, total = generate_estimate_xlsx([_item(quantity=-0.61)])
    assert total == 0.0

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    # Колонка 4 «Кол-во» — исходное число со знаком (раньше печаталось пусто,
    # потому что coerce_qty приводил минус к нулю).
    assert ws.cell(row=2, column=4).value == pytest.approx(-0.61)
    assert ws.cell(row=2, column=6).value is None   # стоимость работ
    assert ws.cell(row=2, column=8).value is None   # стоимость материалов


def test_xlsx_negative_qty_does_not_lower_grand_total():
    """Вычет не должен уменьшать итог на qty × цену."""
    _, only_positive = generate_estimate_xlsx([_item()])
    _, with_negative = generate_estimate_xlsx([_item(), _item(quantity=-0.61)])
    assert with_negative == pytest.approx(only_positive)


# ---------------------------------------------------------------------------
# Сводная (редактор версий → сводная xlsx): вычет не уменьшает итог
# ---------------------------------------------------------------------------

def _summary(rows: list[dict]) -> MagicMock:
    s = MagicMock()
    s.sections = [{"card_id": "c1", "card_name": "Раздел", "tax_pct": 0, "rows": rows}]
    s.overrides = {}
    return s


def _row(qty, price_work=1000.0):
    return {"type": "work", "name": "Работа", "unit": "м2",
            "qty": qty, "price_work": price_work, "price_material": None}


def test_summary_xlsx_ignores_negative_qty():
    import io  # noqa: E401

    import openpyxl

    from app.utils.xlsx_summary import generate_summary_xlsx

    only_positive = generate_summary_xlsx(_summary([_row(10)]))
    with_negative = generate_summary_xlsx(_summary([_row(10), _row(-5)]))

    def _grand(data: bytes) -> float:
        ws = openpyxl.load_workbook(io.BytesIO(data))["Сводная"]
        return max(
            c.value for row in ws.iter_rows() for c in row
            if isinstance(c.value, (int, float))
        )

    assert _grand(with_negative) == pytest.approx(_grand(only_positive))

    # Объём вычета в листе раздела остаётся видимым — теряется только стоимость.
    ws = openpyxl.load_workbook(io.BytesIO(with_negative))["Раздел"]
    assert ws.cell(row=3, column=4).value == pytest.approx(-5)
    assert ws.cell(row=3, column=6).value is None


# ---------------------------------------------------------------------------
# Шаг 3: строка собирается без цен и с примечанием
# ---------------------------------------------------------------------------

def _proc() -> TaskProcessor:
    db = MagicMock()
    db.execute = AsyncMock(return_value=MagicMock(scalar_one_or_none=MagicMock(return_value=None)))
    db.commit = AsyncMock()
    p = TaskProcessor("tid-neg", db=db)
    p.update_progress = AsyncMock(return_value=None)
    p.save_result = AsyncMock(return_value=None)
    p._save_progress_data = AsyncMock(return_value=None)
    p._result_filename = MagicMock(return_value="Смета.xlsx")
    return p


async def test_step3_negative_item_has_no_prices():
    p = _proc()
    items = [
        {"type": "Работа", "name": "Сверление", "unit": "100 отв.", "quantity": 0.61},
        {"type": "Работа", "name": "Добавка на глубину", "unit": "100 отв.", "quantity": -0.61},
    ]
    # Даже если цена по вычету откуда-то нашлась (старый чекпоинт, ручная
    # правка) — в смету она не идёт: шаг 3 общий для всех путей.
    matched = {1: {"type": "Работа", "name": "Добавка на глубину", "unit": "100 отв.",
                   "quantity": -0.61, "work_price": 1393, "price_list_name": "Прайс"}}
    claude_results = {0: {"id": 0, "type": "Работа", "work_price": 25410, "sources": "src"}}

    await p._run_estimate_step3(MagicMock(), items, matched, claude_results)

    saved = p._save_progress_data.await_args.args[0]["items"]
    negative = saved[1]
    assert negative["quantity"] == -0.61
    assert negative["work_price"] is None
    assert negative["material_price"] is None
    assert negative["price_list_name"] is None
    assert negative["notes"] == NEGATIVE_QTY_NOTE
    # Обычная позиция считается как раньше.
    assert saved[0]["work_price"] == pytest.approx(25410)


# ---------------------------------------------------------------------------
# Поиск цен: вычет не попадает ни в прайс, ни в кеш, ни в ИИ
# ---------------------------------------------------------------------------

class _FakePriceSvc:
    """Прайс, в котором находится всё, — чтобы unmatched остался пустым."""

    def __init__(self):
        self.asked: list[str] = []

    def _exact_match_work(self, name):
        self.asked.append(name)
        return {"min_price": 100.0, "name": name}

    def _exact_match_material(self, name):
        self.asked.append(name)
        return 50.0

    def is_cache_loaded(self):
        return True


async def test_price_lookup_skips_negative_qty(monkeypatch):
    from app.services import task_processor as tp

    fake = _FakePriceSvc()
    monkeypatch.setattr(tp, "_price_svc", fake)

    p = _proc()
    p._check_cancelled = AsyncMock(return_value=None)
    p._run_estimate_step3 = AsyncMock(return_value=None)

    source_task = MagicMock()
    source_task.status = "completed"
    source_task.progress_data = {"items": [
        {"type": "Работа", "name": "Сверление", "unit": "100 отв.", "quantity": 0.61},
        {"type": "Работа", "name": "Добавка на глубину", "unit": "100 отв.", "quantity": -0.61},
        {"type": "Материал", "name": "Сверло", "unit": "шт", "quantity": -2},
    ]}
    p.db.execute = AsyncMock(
        return_value=MagicMock(scalar_one_or_none=MagicMock(return_value=source_task))
    )

    task = MagicMock()
    task.user_prompt = '{"path": "B", "source_task_id": "src-1", "source_stage": 1}'
    task.progress_data = {}
    task.processing_mode = "fast"

    await p._handle_estimate_from_list(task)

    assert fake.asked == ["Сверление"], "цену искали только для положительного объёма"
    # Шаг 3 вызван, вычеты не попали в matched → цены им не проставлены.
    matched = p._run_estimate_step3.await_args.args[2]
    assert list(matched.keys()) == [0]
