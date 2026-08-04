"""Смета из нескольких листов: точные числа (план 2026-08-04, Фаза 3).

Правило проекта — «экран и скачиваемый файл считают по одной формуле» — при
разбивке на листы становится главным местом ошибки: сумм становится не одна, а
N+1. Здесь она закреплена точными числами: итог каждого листа, «ВСЕГО» и то,
что посчитает формула фронта (`frontend/src/utils/estimateCalc.ts`).
"""
import io

import openpyxl
import pytest

from app.services.estimate_parser import parse_estimate_excel
from app.utils.estimate_rows import items_to_rows, rows_to_items
from app.utils.xlsx_cost_parser import extract_total_cost, parse_list_sheet
from app.utils.xlsx_exporter import SUMMARY_SHEET_TITLE, generate_estimate_xlsx

# Раздел 1: работа 12 × 1000 = 12 000; материал 5 × 200 = 1 000
# Раздел 2: работа 3 × 5000 = 15 000; материал 10 × 350 = 3 500
ITEMS = [
    {"type": "Работа", "name": "Демонтаж стен", "unit": "м2", "quantity": 12,
     "work_price": 1000, "material_price": None, "sheet": "Раздел 1"},
    {"type": "Материал", "name": "Кирпич", "unit": "шт", "quantity": 5,
     "work_price": None, "material_price": 200, "sheet": "Раздел 1"},
    {"type": "Работа", "name": "Кладка", "unit": "м3", "quantity": 3,
     "work_price": 5000, "material_price": None, "sheet": "Раздел 2"},
    {"type": "Материал", "name": "Раствор", "unit": "кг", "quantity": 10,
     "work_price": None, "material_price": 350, "sheet": "Раздел 2"},
]

OVERHEAD_PCT = 3.0
TRANSPORT_PCT = 3.0

# Раздел 1: 12 000 + 360 + 1 000 + 30 = 13 390
SHEET1_TOTAL = 13390.0
# Раздел 2: 15 000 + 450 + 3 500 + 105 = 19 055
SHEET2_TOTAL = 19055.0
GRAND_TOTAL = 32445.0


def _wb(data: bytes):
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def _totals_of(ws) -> dict:
    """Блок итогов листа: подпись в первой колонке → число в десятой."""
    found = {}
    for row in ws.iter_rows(min_col=1, max_col=10):
        label = row[0].value
        if isinstance(label, str) and label.strip().endswith(":"):
            found[label.strip()] = row[9].value
    return found


def _screen_total(items: list, sheet: str = None) -> float:
    """Тот же расчёт, что делает `calcEstimateTotals` на экране."""
    sum_work = 0.0
    sum_mat = 0.0
    for item in items:
        if sheet is not None and item.get("sheet") != sheet:
            continue
        qty = max(float(item.get("quantity") or 0), 0.0)
        sum_work += round(qty * float(item.get("work_price") or 0), 2)
        sum_mat += round(qty * float(item.get("material_price") or 0), 2)
    overhead = sum_work * OVERHEAD_PCT / 100
    transport = sum_mat * TRANSPORT_PCT / 100
    return round(sum_work + overhead + sum_mat + transport, 2)


@pytest.fixture
def built():
    data, grand = generate_estimate_xlsx(
        ITEMS, overhead_pct=OVERHEAD_PCT, transport_pct=TRANSPORT_PCT,
    )
    return data, grand


class TestFileStructure:
    def test_sheet_per_source_sheet_plus_summary(self, built):
        data, _ = built
        assert _wb(data).sheetnames == ["Раздел 1", "Раздел 2", SUMMARY_SHEET_TITLE]

    def test_single_sheet_estimate_keeps_the_old_shape(self):
        items = [dict(it) for it in ITEMS]
        for it in items:
            it.pop("sheet")
        data, grand = generate_estimate_xlsx(
            items, overhead_pct=OVERHEAD_PCT, transport_pct=TRANSPORT_PCT,
        )
        wb = _wb(data)

        assert wb.sheetnames == ["Смета"]
        assert "ИТОГО ПО СМЕТЕ:" in _totals_of(wb["Смета"])
        assert grand == GRAND_TOTAL


class TestExactNumbers:
    def test_each_sheet_has_its_own_total(self, built):
        data, _ = built
        wb = _wb(data)

        assert _totals_of(wb["Раздел 1"]) == {
            "Сумма по работам:": 12000.0,
            "Накладные расходы 3%:": 360.0,
            "Сумма по материалам:": 1000.0,
            "Транспортные расходы 3%:": 30.0,
            "ИТОГО ПО ЛИСТУ «Раздел 1»:": SHEET1_TOTAL,
        }
        assert _totals_of(wb["Раздел 2"])["ИТОГО ПО ЛИСТУ «Раздел 2»:"] == SHEET2_TOTAL

    def test_summary_sheet_lists_every_sheet_and_the_grand_total(self, built):
        data, _ = built
        ws = _wb(data)[SUMMARY_SHEET_TITLE]
        values = {row[0].value: row[9].value for row in ws.iter_rows(min_col=1, max_col=10)}

        assert values["Раздел 1"] == SHEET1_TOTAL
        assert values["Раздел 2"] == SHEET2_TOTAL
        assert values["ВСЕГО ПО СМЕТЕ:"] == GRAND_TOTAL

    def test_returned_total_is_the_grand_total(self, built):
        _, grand = built
        assert grand == GRAND_TOTAL

    def test_task_cost_read_back_from_the_file_matches(self, built):
        # `task.cost` дозаполняется чтением файла — оно обязано дать «ВСЕГО».
        data, _ = built
        assert float(extract_total_cost(data)) == GRAND_TOTAL

    def test_screen_formula_agrees_per_sheet_and_overall(self, built):
        _, grand = built

        assert _screen_total(ITEMS, "Раздел 1") == SHEET1_TOTAL
        assert _screen_total(ITEMS, "Раздел 2") == SHEET2_TOTAL
        assert _screen_total(ITEMS) == grand == GRAND_TOTAL


class TestRoundTrip:
    def test_generated_estimate_reopens_with_the_same_tabs(self, built):
        data, _ = built
        rows = parse_estimate_excel(data)

        assert [r["sheet"] for r in rows] == ["Раздел 1", "Раздел 1", "Раздел 2", "Раздел 2"]
        assert [r["name"] for r in rows] == [
            "Демонтаж стен", "Кирпич", "Кладка", "Раствор",
        ]

    def test_summary_sheet_does_not_become_a_tab(self, built):
        data, _ = built
        assert SUMMARY_SHEET_TITLE not in {r["sheet"] for r in parse_estimate_excel(data)}

    def test_material_is_not_linked_to_a_work_from_another_sheet(self, built):
        data, _ = built
        rows = parse_estimate_excel(data)
        by_id = {r["id"]: r for r in rows}

        for row in rows:
            parent = row.get("work_row_id")
            if parent:
                assert by_id[parent]["sheet"] == row["sheet"]

    def test_rows_survive_the_editor_round_trip_with_the_same_total(self, built):
        data, grand = built
        rows = items_to_rows(ITEMS)
        again, grand_again = generate_estimate_xlsx(
            rows_to_items(rows), overhead_pct=OVERHEAD_PCT, transport_pct=TRANSPORT_PCT,
        )

        assert _wb(again).sheetnames == _wb(data).sheetnames
        assert grand_again == grand == GRAND_TOTAL


class TestStep3KeepsSheets:
    """Сборка сметы пересобирает позиции заново — лист обязан пережить сборку.

    У шага 3 четыре ветки: цена из прайса (позиция едет как есть), цена от ИИ,
    цена не определена и вычет. Три последние строят словарь из перечисленных
    полей, и забытый там лист унёс бы позицию с её раздела в «Прочее».
    """

    @pytest.mark.asyncio
    async def test_every_branch_keeps_the_sheet(self):
        from unittest.mock import AsyncMock, MagicMock

        from app.services.task_processor import TaskProcessor

        db = MagicMock()
        db.execute = AsyncMock(
            return_value=MagicMock(scalar_one_or_none=MagicMock(return_value=None)),
        )
        db.commit = AsyncMock()
        processor = TaskProcessor("tid-sheets", db=db)
        processor.update_progress = AsyncMock(return_value=None)
        processor.save_result = AsyncMock(return_value=None)
        processor._save_progress_data = AsyncMock(return_value=None)
        processor._result_filename = MagicMock(return_value="Смета.xlsx")

        items = [
            # из прайса
            {"type": "Работа", "name": "Демонтаж", "unit": "м2", "quantity": 2,
             "sheet": "Раздел 1"},
            # цену дал ИИ
            {"type": "Работа", "name": "Кладка", "unit": "м3", "quantity": 3,
             "sheet": "Раздел 2"},
            # цена не определена
            {"type": "Материал", "name": "Раствор", "unit": "кг", "quantity": 4,
             "sheet": "Раздел 2"},
            # вычет
            {"type": "Работа", "name": "Добавка", "unit": "м2", "quantity": -1,
             "sheet": "Раздел 3"},
        ]
        matched = {0: {**items[0], "work_price": 1000, "price_list_name": "Прайс"}}
        claude_results = {1: {"id": 1, "type": "Работа", "work_price": 500, "sources": "src"}}

        await processor._run_estimate_step3(MagicMock(), items, matched, claude_results)

        saved = processor._save_progress_data.await_args.args[0]["items"]
        assert [it.get("sheet") for it in saved] == [
            "Раздел 1", "Раздел 2", "Раздел 2", "Раздел 3",
        ]

        # И в файле позиции легли по своим листам, а не в «Прочее».
        data = processor.save_result.await_args.args[2]
        assert _wb(data).sheetnames == [
            "Раздел 1", "Раздел 2", "Раздел 3", SUMMARY_SHEET_TITLE,
        ]


class TestEstimateFromUploadedFile:
    """Path A: смета на основании загруженного Excel."""

    def _uploaded(self, sheets: dict) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for title, rows in sheets.items():
            ws = wb.create_sheet(title)
            for col, head in enumerate(["№", "Тип", "Наименование", "Ед. изм", "Кол-во"], 1):
                ws.cell(row=1, column=col, value=head)
            for index, (kind, name, unit, qty) in enumerate(rows, start=2):
                ws.cell(row=index, column=1, value=index - 1)
                ws.cell(row=index, column=2, value=kind)
                ws.cell(row=index, column=3, value=name)
                ws.cell(row=index, column=4, value=unit)
                ws.cell(row=index, column=5, value=qty)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_all_sheets_of_uploaded_file_become_positions(self):
        data = self._uploaded({
            "Раздел 1": [("Работа", "Демонтаж стен", "м2", 12)],
            "Раздел 2": [("Работа", "Кладка", "м3", 3), ("Материал", "Раствор", "кг", 10)],
        })
        items = parse_list_sheet(data)

        assert [it["name"] for it in items] == ["Демонтаж стен", "Кладка", "Раствор"]
        assert [it["sheet"] for it in items] == ["Раздел 1", "Раздел 2", "Раздел 2"]

    def test_single_sheet_upload_leaves_positions_unmarked(self):
        data = self._uploaded({"Смета": [("Работа", "Демонтаж стен", "м2", 12)]})
        items = parse_list_sheet(data)

        assert [it["name"] for it in items] == ["Демонтаж стен"]
        assert all("sheet" not in it for it in items)

    def test_our_own_old_perechen_is_not_counted_twice(self):
        # В нём рядом с «Перечнем» лежат сводки «Работы» и «Материалы» с теми же
        # позициями — взять их второй раз значит удвоить смету.
        from app.services.excel_service import generate_list

        xlsx = generate_list([
            {"type": "Работа", "name": "Демонтаж стен", "unit": "м2", "quantity": 12},
            {"type": "Материал", "name": "Кирпич", "unit": "шт", "quantity": 5},
        ])
        items = parse_list_sheet(xlsx)

        assert [it["name"] for it in items] == ["Демонтаж стен", "Кирпич"]

    def test_customer_sheet_named_works_is_not_dropped(self):
        # Признак «наш файл» строгий: раздел заказчика, честно названный
        # «Работы», обязан попасть в смету.
        data = self._uploaded({
            "Работы": [("Работа", "Демонтаж стен", "м2", 12)],
            "Материалы": [("Материал", "Кирпич", "шт", 5)],
        })
        items = parse_list_sheet(data)

        assert [it["name"] for it in items] == ["Демонтаж стен", "Кирпич"]

    def test_file_without_any_positions_is_reported(self):
        wb = openpyxl.Workbook()
        wb.active.cell(row=1, column=1, value="ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ")
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="ни одной позиции"):
            parse_list_sheet(buf.getvalue())


class TestOptimizationAcrossSheets:
    """Оптимизация по ABC работает по всем листам (критерий D3)."""

    def _estimate(self) -> bytes:
        data, _ = generate_estimate_xlsx(
            ITEMS, overhead_pct=OVERHEAD_PCT, transport_pct=TRANSPORT_PCT,
        )
        return data

    def test_items_of_every_sheet_are_analysed(self):
        from app.utils.xlsx_optimizer import parse_estimate_xlsx

        items = parse_estimate_xlsx(self._estimate())

        assert {it["sheet"] for it in items} == {"Раздел 1", "Раздел 2"}
        assert [it["name"] for it in items] == [
            "Демонтаж стен", "Кирпич", "Кладка", "Раствор",
        ]

    def test_result_lands_on_its_own_sheet(self):
        from app.utils.xlsx_optimizer import generate_optimized_xlsx, parse_estimate_xlsx

        data = self._estimate()
        items = parse_estimate_xlsx(data)
        # Оптимизируем по одной позиции на каждом листе.
        results = [
            {"sheet": it["sheet"], "row_index": it["row_index"], "name": it["name"],
             "original_price": it["price_incl_vat"], "new_price": 1.0,
             "source": "поставщик", "savings_abs": 1.0, "savings_pct": 10.0}
            for it in items if it["name"] in ("Демонтаж стен", "Кладка")
        ]
        wb = _wb(generate_optimized_xlsx(data, results))

        for sheet, name in (("Раздел 1", "Демонтаж стен"), ("Раздел 2", "Кладка")):
            ws = wb[sheet]
            row = next(
                r for r in ws.iter_rows(min_row=2)
                if r[1].value == name
            )
            # «Цена сниженная» дописана в первую колонку после исходных.
            assert any(cell.value == 1.0 for cell in row), sheet

    def test_row_numbers_of_different_sheets_do_not_collide(self):
        from app.utils.xlsx_optimizer import generate_optimized_xlsx, parse_estimate_xlsx

        data = self._estimate()
        items = parse_estimate_xlsx(data)
        second = next(it for it in items if it["sheet"] == "Раздел 2")
        wb = _wb(generate_optimized_xlsx(data, [{
            "sheet": second["sheet"], "row_index": second["row_index"],
            "name": second["name"], "original_price": second["price_incl_vat"],
            "new_price": 7.0, "source": "поставщик",
            "savings_abs": 1.0, "savings_pct": 10.0,
        }]))

        # Результат по «Разделу 2» не должен появиться в «Разделе 1» на строке
        # с тем же номером.
        assert all(
            cell.value != 7.0
            for row in wb["Раздел 1"].iter_rows() for cell in row
        )
        assert any(
            cell.value == 7.0
            for row in wb["Раздел 2"].iter_rows() for cell in row
        )


class TestStatementExport:
    """Выгрузка-ведомость по документу с вкладками (критерий E10)."""

    COLUMNS = [
        {"key": "name", "label": "Наименование", "numeric": False},
        {"key": "cost", "label": "Стоимость", "numeric": True},
    ]

    def test_sheet_per_tab(self):
        from app.utils.xlsx_statement import generate_statement_xlsx

        rows = [
            {"_id": "1", "__sheet": "Раздел 1", "name": "Демонтаж", "cost": 100},
            {"_id": "2", "__sheet": "Раздел 2", "name": "Кладка", "cost": 200},
        ]
        wb = _wb(generate_statement_xlsx(self.COLUMNS, rows, title="Ведомость"))

        assert wb.sheetnames == ["Раздел 1", "Раздел 2"]

    def test_each_sheet_has_its_own_total(self):
        from app.utils.xlsx_statement import generate_statement_xlsx

        rows = [
            {"_id": "1", "__sheet": "Раздел 1", "name": "Демонтаж", "cost": 100},
            {"_id": "2", "__sheet": "Раздел 1", "name": "Кладка", "cost": 250},
            {"_id": "3", "__sheet": "Раздел 2", "name": "Штукатурка", "cost": 200},
        ]
        wb = _wb(generate_statement_xlsx(self.COLUMNS, rows, title="Ведомость"))

        def total(ws):
            for row in ws.iter_rows():
                if row[0].value == "ИТОГО":
                    return row[1].value
            return None

        assert total(wb["Раздел 1"]) == 350
        assert total(wb["Раздел 2"]) == 200

    def test_document_without_tabs_exports_one_sheet_as_before(self):
        from app.utils.xlsx_statement import generate_statement_xlsx

        rows = [{"_id": "1", "name": "Демонтаж", "cost": 100}]
        wb = _wb(generate_statement_xlsx(
            self.COLUMNS, rows, title="Ведомость", sheet_name="Выгрузка",
        ))

        assert wb.sheetnames == ["Выгрузка"]
