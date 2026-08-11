"""Коэффициент к ценам и проценты доп. расходов на уровне проекта.

Фаза 8 плана `plans/2026-08-02-edinyy-redaktor-tablic.md`.

Два правила, ради которых фаза:

  * **коэффициент обратим.** Он настройка документа, а не разовая операция:
    исходные цены в строках не меняются никогда, поэтому снятие коэффициента
    возвращает ровно прежние числа — и на экране, и в файле, и в `task.cost`;
  * **проценты накладных и транспортных живут в проекте.** Раньше 3% были
    зашиты в генераторе файла, и смета проекта с другими ставками считалась
    неверно в скачанном файле, хотя на экране была права.
"""
import pytest
import pytest_asyncio
from sqlalchemy import select

from app.models.estimate_version import EstimateVersion
from app.models.history import TaskHistory
from app.models.project import Project
from app.models.task import Task
from app.models.user import User
from app.models.workflow_card import WorkflowCard
from app.utils.auth import create_access_token, hash_password
from app.utils.xlsx_exporter import generate_estimate_xlsx


def _auth(user_id: int, role: str, username: str = None) -> dict:
    return {"Authorization": f"Bearer {create_access_token(user_id, role, username)}"}


def _item(row_id: str, kind: str, qty, work_price=None, material_price=None) -> dict:
    return {
        "row_id": row_id, "type": kind, "name": f"Позиция {row_id}", "unit": "м2",
        "quantity": qty, "work_price": work_price, "material_price": material_price,
    }


def _items() -> list:
    # 4 × 1000 = 4000 работы, 10 × 500 = 5000 материалы
    return [
        _item("r1", "Работа", 4, work_price=1000),
        _item("r2", "Материал", 10, material_price=500),
    ]


def _rows() -> list:
    return [
        {"id": "r1", "lineage_id": "r1", "num": 1, "type": "work", "name": "Кладка",
         "unit": "м2", "qty": 4, "price_work": 1000, "price_material": None,
         "cost": 4000, "selected": False},
        {"id": "r2", "lineage_id": "r2", "num": 2, "type": "material", "name": "Кирпич",
         "unit": "шт", "qty": 10, "price_work": None, "price_material": 500,
         "cost": 5000, "selected": False},
    ]


# ---------------------------------------------------------------------------
# Генератор файла сметы
# ---------------------------------------------------------------------------

def test_xlsx_defaults_keep_previous_numbers():
    """Регресс: без процентов и коэффициента итог тот же, что был при 3%/3%."""
    _, total = generate_estimate_xlsx(_items())
    # 4000 + 3% + 5000 + 3% = 9270
    assert total == pytest.approx(9270.0)


def test_xlsx_uses_project_percentages():
    """Проценты приходят снаружи, а не зашиты в генераторе."""
    _, total = generate_estimate_xlsx(_items(), overhead_pct=10, transport_pct=0)
    # 4000 + 400 + 5000 + 0
    assert total == pytest.approx(9400.0)


def test_xlsx_applies_coefficient_to_prices():
    """Коэффициент множит цены, а не только итог: в файле цены уже с ним."""
    import io

    import openpyxl

    data, total = generate_estimate_xlsx(
        _items(), coefficient={"work": 1.05, "material": 1.0, "scope": "all"},
    )
    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    assert ws.cell(row=2, column=5).value == pytest.approx(1050.0)   # цена работ
    assert ws.cell(row=2, column=6).value == pytest.approx(4200.0)   # стоимость работ
    assert ws.cell(row=3, column=7).value == pytest.approx(500.0)    # материал не тронут
    # 4200 + 3% + 5000 + 3% = 9476
    assert total == pytest.approx(9476.0)


def test_xlsx_coefficient_scope_touches_only_selected():
    """Область «отмеченные строки» не трогает остальные."""
    _, total = generate_estimate_xlsx(
        _items(), coefficient={"work": 2.0, "material": 2.0, "scope": ["r2"]},
    )
    # работы 4000 без изменений, материалы 5000 × 2 = 10000
    assert total == pytest.approx(4000 * 1.03 + 10000 * 1.03)


# ---------------------------------------------------------------------------
# Смета: файл, итог задачи и коэффициент
# ---------------------------------------------------------------------------

@pytest_asyncio.fixture
async def est_env(db_session, fake_s3):
    for model in (EstimateVersion, TaskHistory, WorkflowCard, Task, Project, User):
        await db_session.execute(model.__table__.delete())
    await db_session.commit()

    pm1 = User(username="pm1", role="project_manager", full_name="Иванов Иван",
               password_hash=hash_password("p1"))
    pm2 = User(username="pm2", role="project_manager", full_name="Петров Пётр",
               password_hash=hash_password("p2"))
    db_session.add_all([pm1, pm2])
    await db_session.flush()

    project = Project(name="Объект", owner_id=pm1.id)
    db_session.add(project)
    await db_session.flush()

    task = Task(
        owner_id=pm1.id, user_role="project_manager", task_type="ESTIMATE_FROM_LIST",
        status="completed", input_files=[], input_file_data=[], chat_history=[],
        project_id=str(project.id),
    )
    db_session.add(task)
    await db_session.flush()

    version = EstimateVersion(
        task_id=str(task.id), version_number=0, version_label="original",
        version_display_name="Смета", rows=_rows(), file_slot="estimate",
        task_type="ESTIMATE_FROM_LIST",
    )
    db_session.add(version)
    card = WorkflowCard(project_id=str(project.id), name="АР", stage="estimate",
                        estimate_task_id=str(task.id))
    db_session.add(card)
    await db_session.commit()

    env = {
        "pm1": pm1.id, "pm2": pm2.id, "project_id": str(project.id),
        "task_id": str(task.id), "card_id": str(card.id), "version_id": str(version.id),
    }
    yield env

    for model in (EstimateVersion, TaskHistory, WorkflowCard, Task, Project, User):
        await db_session.execute(model.__table__.delete())
    await db_session.commit()


@pytest.mark.asyncio
async def test_task_cost_uses_project_percentages(db_session, est_env):
    """Итог задачи считается по процентам проекта, а не по зашитым 3%."""
    from app.services import estimate_store

    project = await db_session.get(Project, est_env["project_id"])
    project.overhead_pct = 10
    project.transport_pct = 0
    await db_session.commit()

    task = await db_session.get(Task, est_env["task_id"])
    await estimate_store.sync_artifacts(db_session, task, _rows())
    await db_session.commit()

    assert float(task.cost) == pytest.approx(9400.0)


@pytest.mark.asyncio
async def test_version_percentages_win_over_project(db_session, est_env):
    """Если у версии свои проценты — считаем по ним: проект даёт лишь значение по умолчанию."""
    from app.services import estimate_store

    project = await db_session.get(Project, est_env["project_id"])
    project.overhead_pct = 10
    project.transport_pct = 10
    version = await db_session.get(EstimateVersion, est_env["version_id"])
    version.overhead_pct = 0
    version.transport_pct = 0
    version.expenses_overridden = True
    await db_session.commit()

    task = await db_session.get(Task, est_env["task_id"])
    await estimate_store.sync_artifacts(db_session, task, _rows())
    await db_session.commit()

    assert float(task.cost) == pytest.approx(9000.0)


@pytest.mark.asyncio
async def test_coefficient_changes_total_and_is_reversible(
    async_client, db_session, est_env
):
    """Коэффициент меняет итог, а снятие возвращает прежний — исходные цены целы."""
    url = f"/documents/{est_env['card_id']}/estimate/coefficient"
    headers = _auth(est_env["pm1"], "project_manager")

    resp = await async_client.put(
        url, json={"work": 1.05, "material": 1.0, "scope": "all"}, headers=headers)
    assert resp.status_code == 200, resp.text

    db_session.expire_all()
    task = await db_session.get(Task, est_env["task_id"])
    assert float(task.cost) == pytest.approx(9476.0)

    version = await db_session.get(EstimateVersion, est_env["version_id"])
    # Исходные цены не тронуты — иначе снятие коэффициента было бы невозможно.
    assert version.rows[0]["price_work"] == 1000

    resp = await async_client.put(url, json=None, headers=headers)
    assert resp.status_code == 200, resp.text

    db_session.expire_all()
    task = await db_session.get(Task, est_env["task_id"])
    assert float(task.cost) == pytest.approx(9270.0)
    version = await db_session.get(EstimateVersion, est_env["version_id"])
    assert version.coefficient is None


@pytest.mark.asyncio
async def test_coefficient_is_written_to_history(async_client, db_session, est_env):
    """Кто и когда поставил коэффициент — видно в истории документа."""
    await async_client.put(
        f"/documents/{est_env['card_id']}/estimate/coefficient",
        json={"work": 1.05, "material": 1.0, "scope": "all"},
        headers=_auth(est_env["pm1"], "project_manager"))

    res = await db_session.execute(
        select(TaskHistory).where(TaskHistory.document_kind == "estimate"))
    entries = list(res.scalars().all())
    assert len(entries) == 1
    assert entries[0].user_name == "Иванов Иван"
    assert "оэффициент" in entries[0].description


@pytest.mark.asyncio
async def test_coefficient_rejects_nonsense(async_client, est_env):
    """Ноль и минус коэффициентом не бывают — обнулили бы смету молча."""
    resp = await async_client.put(
        f"/documents/{est_env['card_id']}/estimate/coefficient",
        json={"work": 0, "material": 1.0, "scope": "all"},
        headers=_auth(est_env["pm1"], "project_manager"))
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_colleague_applies_coefficient(async_client, est_env):
    """Сметы общие: коэффициент ставит и не владелец."""
    resp = await async_client.put(
        f"/documents/{est_env['card_id']}/estimate/coefficient",
        json={"work": 1.05, "material": 1.0, "scope": "all"},
        headers=_auth(est_env["pm2"], "project_manager"))
    assert resp.status_code == 200, resp.text


@pytest.mark.asyncio
async def test_version_export_matches_editor(async_client, db_session, est_env):
    """Выгруженная версия считается так же, как показывает редактор.

    Раньше она брала проценты только из самой версии (у неоверрайженной это 0%)
    и считала накладные от общего базиса — файл расходился с экраном.
    """
    import io

    import openpyxl

    project = await db_session.get(Project, est_env["project_id"])
    project.overhead_pct = 10
    project.transport_pct = 0
    version = await db_session.get(EstimateVersion, est_env["version_id"])
    version.coefficient = {"work": 2.0, "material": 1.0, "scope": "all"}
    await db_session.commit()

    resp = await async_client.get(
        f"/tasks/{est_env['task_id']}/estimate/versions/{est_env['version_id']}/export",
        headers=_auth(est_env["pm1"], "project_manager"))
    assert resp.status_code == 200, resp.text

    ws = openpyxl.load_workbook(io.BytesIO(resp.content)).active
    # Цена работ — с коэффициентом.
    assert ws.cell(row=2, column=6).value == pytest.approx(2000.0)

    labels = {}
    for row in ws.iter_rows(min_col=3, max_col=8):
        if row[0].value:
            labels[str(row[0].value)] = row[-1].value
    overhead = next(v for k, v in labels.items() if "акладн" in k)
    # Накладные = работы (8000 с коэффициентом) × 10%
    assert overhead == pytest.approx(800.0)


# ---------------------------------------------------------------------------
# Проценты на уровне проекта
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_existing_projects_default_to_3_percent(db_session, est_env):
    """У проектов, созданных раньше, ставки остаются прежними 3% — поведение не меняется."""
    project = await db_session.get(Project, est_env["project_id"])
    assert float(project.overhead_pct) == pytest.approx(3.0)
    assert float(project.transport_pct) == pytest.approx(3.0)


@pytest.mark.asyncio
async def test_project_percentages_saved(async_client, db_session, est_env):
    resp = await async_client.patch(
        f"/projects/{est_env['project_id']}",
        json={"overhead_pct": 7.5, "transport_pct": 1},
        headers=_auth(est_env["pm1"], "project_manager"))
    assert resp.status_code == 200, resp.text
    assert float(resp.json()["overhead_pct"]) == pytest.approx(7.5)

    db_session.expire_all()
    project = await db_session.get(Project, est_env["project_id"])
    assert float(project.overhead_pct) == pytest.approx(7.5)
    assert float(project.transport_pct) == pytest.approx(1.0)


@pytest.mark.asyncio
async def test_project_percentages_recalculate_estimates(
    async_client, db_session, est_env
):
    """Смена ставок проекта пересчитывает сметы: иначе на карточке висел бы старый итог."""
    await async_client.patch(
        f"/projects/{est_env['project_id']}",
        json={"overhead_pct": 10, "transport_pct": 0},
        headers=_auth(est_env["pm1"], "project_manager"))

    db_session.expire_all()
    task = await db_session.get(Task, est_env["task_id"])
    assert float(task.cost) == pytest.approx(9400.0)


@pytest.mark.asyncio
async def test_project_percentages_reject_nonsense(async_client, est_env):
    resp = await async_client.patch(
        f"/projects/{est_env['project_id']}",
        json={"overhead_pct": -5},
        headers=_auth(est_env["pm1"], "project_manager"))
    assert resp.status_code == 422
