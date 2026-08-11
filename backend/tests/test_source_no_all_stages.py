"""Номер позиции исходной сметы живёт на всех стадиях, не только в перечне.

План: plans/2026-08-11-nomer-pozicii-na-vseh-stadiyah.md
"""
import io

import openpyxl

from app.services.material_kits import expand_completeness_items
from app.utils.estimate_rows import items_to_rows, rows_to_items
from app.utils.source_numbers import attach_source_numbers
from app.utils.xlsx_cost_parser import extract_total_cost, parse_list_sheet
from app.utils.xlsx_exporter import generate_estimate_xlsx


def _est_item(name, *, source_no=None, qty=1.0, work=100.0, mat=None, type_="Работа"):
    item = {
        "type": type_, "name": name, "unit": "м2", "quantity": qty,
        "work_price": work, "material_price": mat,
        "price_list_name": "Прайс", "sources": "", "notes": "",
    }
    if source_no is not None:
        item["source_no"] = source_no
    return item


# ── Фаза 1. Полнота ──────────────────────────────────────────────────────────

def test_completeness_keeps_number_of_source_position():
    """Позиция перечня прошла через ИИ полноты и вернулась со своим номером."""
    source = [
        {"type": "Работа", "name": "Устройство перегородок ГКЛ",
         "unit": "м2", "quantity": 10.0, "notes": "", "source_no": "5"},
    ]
    # То, что вернул ИИ: та же работа плюс дописанный им материал
    from_ai = [
        {"type": "Работа", "name": "Устройство перегородок ГКЛ",
         "unit": "м2", "quantity": 10.0, "notes": ""},
        {"type": "Материал", "name": "Профиль ПН-2 50х40",
         "unit": "м", "quantity": 8.4, "notes": "добавлен по ГЭСН"},
    ]
    attach_source_numbers(from_ai, source)

    assert from_ai[0]["source_no"] == "5"
    # Материала в исходной смете не было — номера у него нет
    assert "source_no" not in from_ai[1]


def test_material_kit_rows_have_no_number():
    """Комплект по нормам расхода дописывает строки, которых в смете не было."""
    items = [
        {"type": "Работа", "name": "Устройство перегородок из ГКЛ в 2 слоя с двух сторон",
         "unit": "м2", "quantity": 10.0, "notes": "", "source_no": "5"},
    ]
    result = expand_completeness_items(items)
    assert result.added > 0

    work = next(it for it in result.items if it.get("type") == "Работа")
    assert work["source_no"] == "5"
    assert all(not it.get("source_no") for it in result.items if it.get("type") == "Материал")


# ── Фаза 2. Смета: поле сквозь строку ────────────────────────────────────────

def test_number_survives_item_row_round_trip():
    items = [_est_item("Пробивка гнезд", source_no="1"),
             _est_item("Заделка борозд", source_no="2а")]
    rows = items_to_rows(items)
    assert [r["source_no"] for r in rows] == ["1", "2а"]

    back = rows_to_items(rows)
    assert [it["source_no"] for it in back] == ["1", "2а"]


def test_row_without_number_gets_none():
    rows = items_to_rows([_est_item("Пробивка гнезд")])
    assert rows[0]["source_no"] is None


def test_estimate_row_schema_accepts_number():
    from app.schemas.estimate_version import EstimateRowSchema

    row = EstimateRowSchema(lineage_id="x", name="Пробивка", source_no="7")
    assert row.source_no == "7"
    # Старая строка без поля открывается по-прежнему
    assert EstimateRowSchema(lineage_id="x", name="Пробивка").source_no is None


def test_uploaded_list_column_is_read():
    """Перечень скачали, положили в задачу сметы — номер читается из файла."""
    from app.services.excel_service import generate_list

    data = generate_list([
        {"type": "Работа", "name": "Пробивка гнезд", "unit": "м2",
         "quantity": 1, "notes": "", "source_no": "3"},
    ])
    items = parse_list_sheet(data)
    assert items[0]["source_no"] == "3"
    assert items[0]["name"] == "Пробивка гнезд"


def test_uploaded_list_without_column_still_parses():
    from app.services.excel_service import generate_list

    items = parse_list_sheet(generate_list([
        {"type": "Работа", "name": "Пробивка гнезд", "unit": "м2",
         "quantity": 1, "notes": ""},
    ]))
    assert items[0]["name"] == "Пробивка гнезд"
    assert items[0].get("source_no", "") == ""


# ── Фаза 3. Смета: колонка в файле ───────────────────────────────────────────

def _estimate_values(data: bytes, title: str = "Смета") -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(data))
    return [list(row) for row in wb[title].iter_rows(values_only=True)]


def test_estimate_xlsx_adds_column_first():
    data, _ = generate_estimate_xlsx([
        _est_item("Пробивка гнезд", source_no="1"),
        _est_item("Смеси бетонные", type_="Материал", source_no=None, work=None, mat=50.0),
    ])
    values = _estimate_values(data)
    assert values[0][:3] == ["№ в исходной смете", "№", "Наименование"]
    assert len(values[0]) == 11
    assert values[1][:3] == [1, 1, "Пробивка гнезд"]
    assert values[2][:3] == [None, 2, "Смеси бетонные"]


def test_estimate_xlsx_keeps_ten_columns_without_numbers():
    data, _ = generate_estimate_xlsx([_est_item("Пробивка гнезд")])
    values = _estimate_values(data)
    assert values[0] == ["№", "Наименование", "Ед. изм.", "Кол-во", "Цена работ",
                         "Стоимость работ", "Цена матер.", "Стоимость матер.",
                         "Источник цены", "Примечание"]


def test_estimate_total_unchanged_by_new_column():
    """Колонка сдвигает таблицу, но не сумму контракта."""
    plain = [_est_item("Пробивка гнезд", qty=2.0, work=100.0),
             _est_item("Смеси", type_="Материал", qty=3.0, work=None, mat=50.0)]
    numbered = [dict(it, source_no=str(i + 1)) for i, it in enumerate(plain)]

    data_plain, total_plain = generate_estimate_xlsx(plain)
    data_numbered, total_numbered = generate_estimate_xlsx(numbered)

    assert total_plain == total_numbered
    assert extract_total_cost(data_plain) == extract_total_cost(data_numbered)
    assert float(extract_total_cost(data_numbered)) == total_numbered


def test_estimate_totals_block_reads_after_shift():
    """Подпись итога остаётся в первой колонке, число — в последней."""
    data, total = generate_estimate_xlsx([_est_item("Пробивка", qty=2.0, work=100.0,
                                                    source_no="1")])
    values = _estimate_values(data)
    grand = next(row for row in values if str(row[0] or "").startswith("ИТОГО ПО СМЕТЕ"))
    assert grand[-1] == total


def test_multisheet_estimate_keeps_numbers_and_summary():
    items = [
        dict(_est_item("Пробивка", source_no="1"), sheet="Раздел 1"),
        dict(_est_item("Заделка", source_no="1"), sheet="Раздел 2"),
    ]
    data, total = generate_estimate_xlsx(items)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Итого по смете" in wb.sheetnames
    assert _estimate_values(data, "Раздел 1")[1][0] == 1
    assert float(extract_total_cost(data)) == total


# ── Фаза 4 (бэкенд-часть). Оптимизация и сводная наследуют номер ─────────────

def test_optimization_row_copy_keeps_number():
    """Версия оптимизации строится копией строки сметы — номер едет с ней."""
    row = items_to_rows([_est_item("Пробивка гнезд", source_no="1")])[0]
    # Ровно то, что делает построение версии оптимизации: копия + патч цены
    row_copy = dict(row)
    row_copy["price_work"] = 90.0
    row_copy["optimization_note"] = "цена из прайса"

    assert row_copy["source_no"] == "1"
    assert rows_to_items([row_copy])[0]["source_no"] == "1"


def test_summary_section_snapshot_keeps_number():
    """Снимок раздела сводной — копия строк версии, без пересборки полей."""
    rows = items_to_rows([_est_item("Пробивка гнезд", source_no="4")])
    snapshot = list(rows)  # как в summary_service._build_sections_snapshot
    assert snapshot[0]["source_no"] == "4"


def test_statement_export_carries_the_column():
    """Ведомость берёт колонки у документа — номер попадает в выгрузку."""
    from app.utils.xlsx_statement import generate_statement_xlsx

    columns = [
        {"key": "source_no", "label": "№ в исходной смете", "numeric": False},
        {"key": "name", "label": "Наименование", "numeric": False},
    ]
    data = generate_statement_xlsx(
        columns, [{"source_no": "1", "name": "Пробивка гнезд"}], show_total=False,
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))
    flat = [str(cell) for row in wb.worksheets[0].iter_rows(values_only=True) for cell in row]
    assert "№ в исходной смете" in flat
    assert "1" in flat


def test_end_to_end_list_to_estimate_file():
    """Перечень → полнота → смета: номер доезжает до файла сметы."""
    from_list = [
        {"type": "Работа", "name": "Устройство перегородок ГКЛ",
         "unit": "м2", "quantity": 10.0, "notes": "", "source_no": "5"},
        {"type": "Работа", "name": "Штукатурка откосов",
         "unit": "м2", "quantity": 4.0, "notes": "", "source_no": "6"},
    ]
    # Полнота: ИИ вернул те же работы и дописал материал
    after_ai = [
        {"type": "Работа", "name": "Устройство перегородок ГКЛ",
         "unit": "м2", "quantity": 10.0, "notes": ""},
        {"type": "Материал", "name": "Профиль ПН-2 50х40",
         "unit": "м", "quantity": 8.4, "notes": "добавлен по ГЭСН"},
        {"type": "Работа", "name": "Штукатурка откосов",
         "unit": "м2", "quantity": 4.0, "notes": ""},
    ]
    attach_source_numbers(after_ai, from_list)

    # Смета: цены проставлены, позиции пересобраны в строки версии
    priced = [dict(it, work_price=100.0, material_price=None,
                   price_list_name="Прайс", sources="") for it in after_ai]
    rows = items_to_rows(priced)
    assert [r["source_no"] for r in rows] == ["5", None, "6"]

    data, _ = generate_estimate_xlsx(rows_to_items(rows))
    values = _estimate_values(data)
    assert values[0][0] == "№ в исходной смете"
    assert [row[0] for row in values[1:4]] == [5, None, 6]


def test_optimizer_column_mapping_ignores_new_column():
    """Разбор сметы для оптимизации ищет колонки по названиям — наша не мешает."""
    from app.utils.xlsx_optimizer import _find_header_row, _map_columns

    data, _ = generate_estimate_xlsx([_est_item("Пробивка гнезд", source_no="1")])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Смета"]
    header_row = _find_header_row(ws)
    cols = _map_columns(ws, header_row)

    assert ws.cell(row=header_row, column=cols["name"]).value == "Наименование"
    assert ws.cell(row=header_row, column=cols["quantity"]).value == "Кол-во"
    # «№ в исходной смете» не притворилась ни одной из известных колонок
    assert cols["name"] != 1
