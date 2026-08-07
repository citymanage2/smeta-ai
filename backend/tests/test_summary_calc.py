"""Расчёт бланка «Сводная» на сервере — эталон совпадает с экраном.

Фаза 9 плана `plans/2026-08-02-edinyy-redaktor-tablic.md`.

Найдено в Фазе 7: скачиваемый xlsx сводной считал по формулам прежней версии
бланка — без коэффициента, без налогов разделов, без восьми из восемнадцати
строк расходов, с другой формулой прибыли и старым НДС. На одном и том же
наборе данных экран показывал 222 647,82 ₽, а файл — 143 713,91 ₽.

Эталон — `calcSummary` (`frontend/src/stores/summaryEditorStore.ts`). Набор
данных здесь тот же, что во фронтовом регресс-тесте
(`frontend/src/__tests__/summaryRegressFixture.ts`), и ожидаемые числа взяты
оттуда же — до копейки. Если формулы разъедутся, падёт этот тест.
"""
import pytest

from app.utils.summary_calc import calc_summary


def _row(kind, qty, work=None, material=None) -> dict:
    return {
        "id": f"r{qty}-{work}-{material}", "type": kind, "name": "Строка",
        "unit": "м2", "qty": qty, "price_work": work, "price_material": material,
    }


SECTIONS = [
    {
        "card_id": "card-1", "card_name": "АР", "tax_pct": 5,
        "rows": [
            _row("section", None),
            _row("work", 12.5, work=1340.75),
            _row("material", 8, material=2210.4),
            # Вычет: объём < 0 стоимости не даёт.
            _row("work", -3, work=1000),
        ],
    },
    {
        "card_id": "card-2", "card_name": "ОВ", "tax_pct": 0,
        "rows": [
            _row("work", 4, work=999.99),
            _row("material", 2.5, material=15000),
        ],
    },
]

OVERRIDES = {
    "coefficient": 1.07,
    "transport_pct": 3, "cleanup_pct": 2.5, "overhead_pct": 4,
    "daily_workers_cost": 2, "bank_guarantee_cost": 12000, "cleaning_cost": 5000,
    "ppr_cost": 7000, "commissioning_cost": 0, "construction_control_cost": 3000,
    "author_supervision_cost": 0, "passes_cost": 1500, "site_office_cost": 0,
    "travel_cost": 0, "rp_cost": 0, "housing_rent_cost": 0,
    "workers_transport_cost": 800,
    "contingency_pct": 2, "profit_pct": 20, "vat_full_cost_pct": 22, "tax_pct": 2,
    "hidden_fixed_rows": ["cleanup", "ppr"],
    "custom_rows_before": [
        {"id": "c1", "label": "Аренда лесов", "qty_pct": "1 шт", "without_vat": 25000},
    ],
    "custom_rows_after": [],
}


def test_total_for_customer_matches_screen():
    calc = calc_summary(SECTIONS, OVERRIDES)
    assert calc["total_for_customer"] == pytest.approx(222647.81997720266, rel=1e-9)


def test_subtotal_contingency_profit_vat_match_screen():
    calc = calc_summary(SECTIONS, OVERRIDES)
    assert calc["subtotal_with_vat"] == pytest.approx(171809.196946355, rel=1e-9)
    assert calc["subtotal_without_vat"] == pytest.approx(140827.21061176638, rel=1e-9)
    assert calc["contingency_with_vat"] == pytest.approx(3436.1839389271, rel=1e-9)
    assert calc["profit"] == pytest.approx(35910.93870600043, rel=1e-9)
    assert calc["full_cost_without_vat"] == pytest.approx(179554.69353000214, rel=1e-9)
    assert calc["vat"] == pytest.approx(39502.03257660047, rel=1e-9)
    assert calc["other_tax"] == pytest.approx(3591.093870600043, rel=1e-9)


def test_rows_match_screen():
    calc = calc_summary(SECTIONS, OVERRIDES)
    assert calc["works_with_vat"] == pytest.approx(26202.6093465, rel=1e-9)
    assert calc["materials_with_vat"] == pytest.approx(71090.09808, rel=1e-9)
    assert calc["transport_with_vat"] == pytest.approx(2918.781222795, rel=1e-9)
    assert calc["overhead_with_vat"] == pytest.approx(3891.70829706, rel=1e-9)
    assert calc["daily_workers_with_vat"] == pytest.approx(10000, rel=1e-9)
    assert calc["bank_guarantee_with_vat"] == pytest.approx(14640, rel=1e-9)


def test_sections_use_their_own_tax():
    calc = calc_summary(SECTIONS, OVERRIDES)
    ar, ov = calc["section_totals"]
    assert ar["works_raw"] == pytest.approx(17932.53125, rel=1e-9)
    assert ar["works_with_vat"] == pytest.approx(20981.0615625, rel=1e-9)
    assert ar["materials_with_vat"] == pytest.approx(22137.59808, rel=1e-9)
    assert ov["works_with_vat"] == pytest.approx(5221.547784000001, rel=1e-9)
    assert ov["materials_with_vat"] == pytest.approx(48952.5, rel=1e-9)


_TAX_SECTION_ROWS = [_row("work", 10, work=100), _row("material", 2, material=500)]
_TAX_OVERRIDES = {"coefficient": 1}


def test_section_taxes_works_and_materials_are_independent():
    """Работы и материалы раздела облагаются каждый по своей ставке."""
    calc = calc_summary(
        [{
            "card_id": "c", "card_name": "Раздел",
            "tax_pct_works": 0, "tax_pct_materials": 22,
            "rows": _TAX_SECTION_ROWS,
        }],
        _TAX_OVERRIDES,
    )
    section = calc["section_totals"][0]
    assert section["works_raw"] == pytest.approx(1000, rel=1e-9)
    assert section["materials_raw"] == pytest.approx(1000, rel=1e-9)
    # Работы: подрядчик с НДС — добавляем 22%. Материалы: самозанятый — НДС в цене.
    assert section["works_with_vat"] == pytest.approx(1220, rel=1e-9)
    assert section["materials_with_vat"] == pytest.approx(1000, rel=1e-9)


def test_section_with_legacy_single_tax_is_unchanged():
    """Сводная, сохранённая с одной ставкой, считается как считалась."""
    calc = calc_summary(
        [{"card_id": "c", "card_name": "Раздел", "tax_pct": 5, "rows": _TAX_SECTION_ROWS}],
        _TAX_OVERRIDES,
    )
    section = calc["section_totals"][0]
    assert section["tax_pct_works"] == 5
    assert section["tax_pct_materials"] == 5
    assert section["works_with_vat"] == pytest.approx(1170, rel=1e-9)
    assert section["materials_with_vat"] == pytest.approx(1170, rel=1e-9)


def test_hidden_rows_are_counted_but_not_in_subtotal():
    calc = calc_summary(SECTIONS, OVERRIDES)
    # Строки скрыты пользователем: значения считаются и показываются, но в
    # «ИТОГО себестоимость» не входят.
    assert calc["cleanup_with_vat"] > 0
    assert calc["ppr_with_vat"] > 0

    visible = calc_summary(SECTIONS, {**OVERRIDES, "hidden_fixed_rows": []})
    assert visible["subtotal_with_vat"] == pytest.approx(
        calc["subtotal_with_vat"] + calc["cleanup_with_vat"] + calc["ppr_with_vat"],
        rel=1e-9,
    )


def test_defaults_do_not_crash_on_empty_summary():
    calc = calc_summary([], {})
    assert calc["total_for_customer"] == pytest.approx(0.0)
    assert calc["section_totals"] == []


def test_exported_summary_matches_screen(tmp_path):
    """Файл сводной содержит то же «ИТОГО для Заказчика», что и экран."""
    import io

    import openpyxl

    from app.utils.xlsx_summary import generate_summary_xlsx

    class _Summary:
        sections = SECTIONS
        overrides = OVERRIDES

    wb = openpyxl.load_workbook(io.BytesIO(generate_summary_xlsx(_Summary())))
    ws = wb["Сводная"]
    totals = {}
    for row in ws.iter_rows(min_col=1, max_col=4):
        if row[0].value:
            totals[str(row[0].value)] = row[3].value

    grand = next(v for k, v in totals.items() if "Заказчик" in k)
    assert grand == pytest.approx(222647.82, abs=0.01)


def test_exported_sections_show_both_taxes():
    """В файле у раздела две колонки налога — работы и материалы отдельно."""
    import io

    import openpyxl

    from app.utils.xlsx_summary import generate_summary_xlsx

    class _Summary:
        sections = [{
            "card_id": "c", "card_name": "Раздел",
            "tax_pct_works": 0, "tax_pct_materials": 22,
            "rows": _TAX_SECTION_ROWS,
        }]
        overrides = _TAX_OVERRIDES

    wb = openpyxl.load_workbook(io.BytesIO(generate_summary_xlsx(_Summary())))
    ws = wb["Сводная"]
    right = lambda row: [ws.cell(row=row, column=col).value for col in range(6, 13)]  # noqa: E731

    # Колонки F-K — на своих прежних местах, второй налог добавлен в конец:
    # формулы поверх прежних выгрузок не должны съезжать.
    assert right(1) == [
        "Раздел", "Налог работ, %", "Работы (с/с)", "Работы с НДС",
        "Материалы (с/с)", "Материалы с НДС", "Налог матер., %",
    ]
    assert right(2) == ["Раздел", 0, 1000, 1220, 1000, 1000, 22]
    # ИТОГО: суммируются только деньги, налоговые колонки пустые.
    assert right(3) == ["ИТОГО", None, 1000, 1220, 1000, 1000, None]
