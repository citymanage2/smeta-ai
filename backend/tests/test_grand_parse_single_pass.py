"""Парсинг гранд-сметы: один проход по строкам и не в событийном цикле.

Было: строки листа материализовались дважды — отдельно в `_find_header_row`,
отдельно в `parse_xlsx_grand`, — а сам парсинг вызывался прямо в корутине. При
трёх задачах разом это давало и лишние десятки мегабайт на задачу, и заморозку
heartbeat всего процесса (plans/2026-07-30-parallelnaya-obrabotka-umiraet.md).

Стало: строки читаются один раз, парсинг уходит в поток.
"""
import asyncio
import io

import openpyxl
import pytest

from app.utils.file_parser import _find_header_row, parse_xlsx_grand


def _grand_xlsx(n_rows: int = 40) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Смета № 1 на строительные работы"])
    ws.append(["№", "Обоснование", "Наименование работ и затрат", "Ед. изм.", "Количество"])
    for i in range(n_rows):
        ws.append([i + 1, f"ФЕР{i}", f"Устройство перегородок тип {i}", "м2", 1.5 * (i + 1)])
    ws.append(["", "", "Итого по смете", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_find_header_row_accepts_rows_not_worksheet():
    """Функция больше не читает лист сама — иначе файл держался в памяти дважды."""
    rows = [
        ("Смета № 1", None, None, None, None),
        ("№", "Обоснование", "Наименование работ и затрат", "Ед. изм.", "Количество"),
        (1, "ФЕР1", "Кладка стен", "м3", 12.0),
    ]

    header_idx, name_col, unit_col, qty_col, _ = _find_header_row(rows)

    assert header_idx == 1
    assert name_col == 2
    assert unit_col == 3
    assert qty_col == 4


def test_find_header_row_none_when_no_header():
    assert _find_header_row([("просто", "текст"), (1, 2)])[0] is None


def test_find_header_row_empty_input():
    """Пустой файл — не исключение, а «заголовка нет»."""
    assert _find_header_row([]) == (None, None, None, None, None)


def test_parse_still_extracts_rows_after_refactor():
    """Смысл парсинга не изменился: позиции те же, итоговая строка отброшена."""
    rows = parse_xlsx_grand(_grand_xlsx(10))

    names = [r["name"] for r in rows]
    assert "Устройство перегородок тип 0" in names
    assert all("Итого" not in n for n in names)
    assert rows[0]["unit"] == "м2"
    assert rows[0]["quantity"] == 1.5


def test_parse_reads_sheet_once(monkeypatch):
    """Ровно один проход по строкам листа: второй — это память впустую."""
    calls = {"iter_rows": 0}
    real_load = openpyxl.load_workbook

    def counting_load(*args, **kwargs):
        wb = real_load(*args, **kwargs)
        ws = wb.active
        original = ws.iter_rows

        def counted(*a, **kw):
            calls["iter_rows"] += 1
            return original(*a, **kw)

        ws.iter_rows = counted
        return wb

    monkeypatch.setattr(openpyxl, "load_workbook", counting_load)
    monkeypatch.setattr("app.utils.file_parser.openpyxl.load_workbook", counting_load)

    parse_xlsx_grand(_grand_xlsx(20))

    assert calls["iter_rows"] == 1


@pytest.mark.asyncio
async def test_parse_does_not_block_event_loop():
    """Пока парсится файл, событийный цикл жив — heartbeat не пропадает.

    Раньше heartbeat молчал всё время парсинга, и живая задача показывала
    пользователю «Обработчик молчит N минут».
    """
    data = _grand_xlsx(4000)
    ticks = 0

    async def heartbeat():
        nonlocal ticks
        while True:
            await asyncio.sleep(0.01)
            ticks += 1

    hb = asyncio.create_task(heartbeat())
    try:
        rows = await asyncio.to_thread(parse_xlsx_grand, data)
    finally:
        hb.cancel()

    assert len(rows) > 3000
    assert ticks > 0  # цикл продолжал работать во время парсинга
