"""Служебная строка нумерации и числа в пересобранном перечне.

Две вещи, из-за которых перечень и полнота выглядели небрежно:

1. Первой строкой данных шла нумерация колонок «1 2 3 4 5 6». Разбор файла её
   отсекает с 2 августа 2026, но документы, созданные раньше, хранят её в версии
   как обычную строку — редактор показывал мусор.
2. После правки в редакторе число приходит строкой («1234,56»), и в скачанном
   файле ячейка оказывалась текстовой: колонку в Excel не просуммировать.
"""
import io
from types import SimpleNamespace

import openpyxl

from app.services.document_service import strip_service_rows
from app.utils.xlsx_generic import (
    is_numbering_cells,
    parse_xlsx_to_generic_rows,
    rows_to_xlsx,
)


def _sheet(data: bytes):
    return openpyxl.load_workbook(io.BytesIO(data)).active


def _doc(row_format: str) -> SimpleNamespace:
    return SimpleNamespace(row_format=row_format)


# ---------------------------------------------------------------------------
# Опознание строки нумерации
# ---------------------------------------------------------------------------

class TestIsNumberingCells:
    def test_column_numbers_recognised(self):
        assert is_numbering_cells({
            "№ п/п": 1, "Тип": 2, "Наименование": 3,
            "Ед. изм": 4, "Кол-во": 5, "Примечание": 6,
        })

    def test_real_row_is_not_numbering(self):
        assert not is_numbering_cells({
            "№ п/п": 1, "Тип": "Работа", "Наименование": "Демонтаж стен",
            "Ед. изм": "м2", "Кол-во": 12.5, "Примечание": "",
        })

    def test_numeric_data_row_is_not_numbering(self):
        # Позиции с одинаковыми объёмами — числа, но повторяются и уходят далеко
        # за число колонок: данные, а не нумерация.
        assert not is_numbering_cells({"a": 100, "b": 100, "c": 250})

    def test_empty_and_broken_input(self):
        assert not is_numbering_cells({})
        assert not is_numbering_cells(None)


# ---------------------------------------------------------------------------
# Сервис не отдаёт служебную строку редактору
# ---------------------------------------------------------------------------

class TestStripServiceRows:
    def test_numbering_row_removed_from_generic_document(self):
        rows = [
            {"row_id": "a", "cells": {"№ п/п": 1, "Тип": 2, "Наименование": 3}},
            {"row_id": "b", "cells": {"№ п/п": 1, "Тип": "Работа", "Наименование": "Демонтаж"}},
        ]
        left = strip_service_rows(_doc("generic"), rows)

        assert [r["row_id"] for r in left] == ["b"]

    def test_estimate_rows_untouched(self):
        # У сметы строки типизированные: чисел-«нумерации» там быть не может, и
        # трогать их нельзя ни при каких условиях.
        rows = [{"id": "a", "type": "work", "name": "Демонтаж", "qty": 1}]
        assert strip_service_rows(_doc("estimate"), rows) == rows

    def test_only_the_head_of_the_document_is_checked(self):
        # Признак «разные небольшие числа подряд» может совпасть с настоящей
        # позицией. Строка нумерации бывает только первой — ниже не трогаем.
        rows = [
            {"row_id": "a", "cells": {"№": 1, "Тип": "Работа", "Наименование": "Демонтаж"}},
            {"row_id": "b", "cells": {"Кол-во": 1, "Цена": 2, "Стоимость": 3}},
        ]
        assert strip_service_rows(_doc("generic"), rows) == rows

    def test_none_stays_none(self):
        # Черновика может не быть — это не пустой список, а «черновика нет».
        assert strip_service_rows(_doc("generic"), None) is None


# ---------------------------------------------------------------------------
# Генератор перечня больше не пишет строку нумерации
# ---------------------------------------------------------------------------

class TestPerechenSheet:
    def test_data_starts_right_under_header(self):
        from app.services.excel_service import generate_list

        ws = _sheet(generate_list([
            {"type": "Работа", "name": "Демонтаж стен", "unit": "м2", "quantity": 12.5},
            {"type": "Материал", "name": "Кирпич", "unit": "шт", "quantity": 400},
        ]))

        assert ws.cell(row=1, column=1).value == "№ п/п"
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=3).value == "Демонтаж стен"
        assert ws.cell(row=3, column=3).value == "Кирпич"

    def test_sections_sheet_has_no_numbering_row(self):
        from app.services.excel_service import generate_list_project

        ws = _sheet(generate_list_project([
            {"type": "Работа", "name": "Демонтаж стен", "unit": "м2",
             "quantity": 12.5, "section": "Демонтаж"},
        ]))

        # Вторая строка — заголовок раздела, а не «1 2 3 4 5 6».
        assert ws.cell(row=2, column=3).value == "Раздел 1. Демонтаж"
        assert ws.cell(row=3, column=3).value == "Демонтаж стен"

    def test_parsed_back_without_junk_row(self):
        from app.services.excel_service import generate_list

        rows = parse_xlsx_to_generic_rows(generate_list([
            {"type": "Работа", "name": "Демонтаж стен", "unit": "м2", "quantity": 12.5},
        ]))

        assert len(rows) == 1
        assert rows[0]["cells"]["Наименование"] == "Демонтаж стен"

    def test_list_sheet_parser_sees_only_real_items(self):
        # ESTIMATE_FROM_LIST читает наш же перечень: строка нумерации давала
        # позицию с наименованием «3».
        from app.services.excel_service import generate_list
        from app.utils.xlsx_cost_parser import parse_list_sheet

        items = parse_list_sheet(generate_list([
            {"type": "Работа", "name": "Демонтаж стен", "unit": "м2", "quantity": 12.5},
        ]))

        assert [it["name"] for it in items] == ["Демонтаж стен"]


# ---------------------------------------------------------------------------
# Числа в пересобранном файле остаются числами
# ---------------------------------------------------------------------------

class TestRowsToXlsxNumbers:
    def test_edited_value_written_as_number(self):
        rows = [{"row_id": "a", "cells": {
            "Наименование": "Демонтаж стен", "Кол-во": "12,5",
            "Стоимость работ": "1 234,56",
        }}]
        ws = _sheet(rows_to_xlsx(rows))

        assert ws.cell(row=2, column=2).value == 12.5
        assert ws.cell(row=2, column=3).value == 1234.56
        assert ws.cell(row=2, column=3).number_format == "#,##0.00"

    def test_text_stays_text(self):
        rows = [{"row_id": "a", "cells": {
            "Наименование": "Демонтаж стен", "Кол-во": "по проекту", "Артикул": "007",
        }}]
        ws = _sheet(rows_to_xlsx(rows))

        assert ws.cell(row=2, column=2).value == "по проекту"
        # Ведущий ноль — признак кода: превратить «007» в 7 значило бы испортить
        # артикул.
        assert ws.cell(row=2, column=3).value == "007"

    def test_roundtrip_is_stable(self):
        rows = [{"row_id": "a", "cells": {"Наименование": "Демонтаж", "Кол-во": 12.5}}]
        again = parse_xlsx_to_generic_rows(rows_to_xlsx(rows))

        assert again[0]["cells"] == rows[0]["cells"]
