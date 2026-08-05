"""Номер позиции из исходной Гранд-сметы доезжает до перечня.

План: plans/2026-08-06-nomer-pozicii-iz-ishodnoj-smety.md
"""
import io

import openpyxl
import pytest

from app.services.excel_service import generate_list
from app.utils.file_parser import chunk_rows, parse_xlsx_grand, rows_to_text
from app.utils.generic_items import generic_rows_to_items
from app.utils.source_numbers import attach_source_numbers
from app.utils.xlsx_generic import parse_xlsx_to_generic_rows

_HEADER = ["№ п/п", "Обоснование", "Наименование работ и затрат", "Единица измерения", "Количество"]


def _grand_xlsx(rows: list[tuple]) -> bytes:
    """Лист Гранд-сметы: шапка + переданные строки (№, обоснование, наименование, ед., кол-во)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ЛС02-01-03"
    ws.cell(row=1, column=1, value="ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ № 02-01-03")
    for col, head in enumerate(_HEADER, 1):
        ws.cell(row=2, column=col, value=head)
    for index, row in enumerate(rows, 3):
        for col, value in enumerate(row, 1):
            ws.cell(row=index, column=col, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Разбор: номер остаётся при строке ────────────────────────────────────────

def test_parse_keeps_position_number():
    data = _grand_xlsx([
        ("1", "ФЕР46-03-009-03", "Пробивка в кирпичных стенах гнезд", "100 шт", 0.02),
        ("2", "ФЕР46-03-017-05", "Заделка отверстий, гнезд и борозд", "м3", 0.1),
    ])
    rows = parse_xlsx_grand(data)
    assert [r["source_no"] for r in rows] == ["1", "2"]


def test_parse_number_with_dot_and_letter():
    data = _grand_xlsx([
        ("1.1", "ФЕР46-03-009-03", "Пробивка гнезд", "100 шт", 0.02),
        ("2а", "ФЕР46-03-017-05", "Заделка борозд", "м3", 0.1),
        ("3", "ФЕР46-03-017-06", "Штукатурка откосов", "м2", 5),
    ])
    rows = parse_xlsx_grand(data)
    assert [r["source_no"] for r in rows] == ["1.1", "2а", "3"]


def test_parse_row_without_number_gets_empty():
    """Ресурс расценки со своей стоимостью остаётся в перечне, но номера у него нет."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, head in enumerate(_HEADER + ["Сметная стоимость"], 1):
        ws.cell(row=1, column=col, value=head)
    ws.cell(row=2, column=1, value="1")
    ws.cell(row=2, column=3, value="Пробивка гнезд")
    ws.cell(row=2, column=4, value="100 шт")
    ws.cell(row=2, column=5, value=0.02)
    ws.cell(row=2, column=6, value=1632.38)
    ws.cell(row=3, column=1, value="2")
    ws.cell(row=3, column=3, value="Заделка борозд")
    ws.cell(row=3, column=4, value="м3")
    ws.cell(row=3, column=5, value=0.1)
    ws.cell(row=3, column=6, value=2319.52)
    # Неучтённый ресурс: номера нет, но своя стоимость есть — строка остаётся
    ws.cell(row=4, column=3, value="Смеси бетонные тяжелого бетона")
    ws.cell(row=4, column=4, value="м3")
    ws.cell(row=4, column=5, value=0.104)
    ws.cell(row=4, column=6, value=380.78)
    buf = io.BytesIO()
    wb.save(buf)

    rows = parse_xlsx_grand(buf.getvalue())
    by_name = {r["name"]: r["source_no"] for r in rows}
    assert by_name["Пробивка гнезд"] == "1"
    assert by_name["Смеси бетонные тяжелого бетона"] == ""


# ── Сопоставление позиций ИИ со строками чанка ───────────────────────────────

def _rows(*pairs) -> list[dict]:
    return [
        {"name": name, "unit": "м2", "quantity": 1.0, "is_section": False, "source_no": no}
        for name, no in pairs
    ]


def test_attach_exact_match():
    items = [
        {"type": "Работа", "name": "Заделка борозд"},
        {"type": "Работа", "name": "Пробивка гнезд"},
    ]
    attach_source_numbers(items, _rows(("Пробивка гнезд", "1"), ("Заделка борозд", "2")))
    assert [it["source_no"] for it in items] == ["2", "1"]


def test_attach_ignores_case_and_spaces():
    items = [{"type": "Работа", "name": "  пробивка   ГНЕЗД "}]
    attach_source_numbers(items, _rows(("Пробивка гнезд", "7")))
    assert items[0]["source_no"] == "7"


def test_attach_repeated_name_takes_numbers_in_order():
    items = [
        {"type": "Работа", "name": "Штукатурка откосов"},
        {"type": "Работа", "name": "Штукатурка откосов"},
    ]
    attach_source_numbers(items, _rows(("Штукатурка откосов", "4"), ("Штукатурка откосов", "9")))
    assert [it["source_no"] for it in items] == ["4", "9"]


def test_attach_truncated_name_matches_single_candidate():
    items = [{"type": "Работа", "name": "Пробивка в кирпичных стенах гнезд"}]
    attach_source_numbers(
        items,
        _rows(("Пробивка в кирпичных стенах гнезд размером: до 380x380 мм", "3")),
    )
    assert items[0]["source_no"] == "3"


def test_attach_ambiguous_prefix_takes_nothing():
    """Два кандидата с общим началом — чужой номер хуже пустой ячейки."""
    items = [{"type": "Работа", "name": "Пробивка гнезд"}]
    attach_source_numbers(
        items,
        _rows(("Пробивка гнезд размером до 380x380 мм", "3"),
              ("Пробивка гнезд размером до 200x200 мм", "4")),
    )
    assert items[0].get("source_no", "") == ""


def test_attach_unknown_name_gets_nothing():
    items = [{"type": "Материал", "name": "Профиль ПН-2 50х40"}]
    attach_source_numbers(items, _rows(("Пробивка гнезд", "1")))
    assert "source_no" not in items[0]


def test_attach_skips_rows_without_number():
    items = [{"type": "Материал", "name": "Смеси бетонные"}]
    attach_source_numbers(items, _rows(("Смеси бетонные", "")))
    assert "source_no" not in items[0]


def test_attach_ignores_section_rows():
    items = [{"type": "Работа", "name": "Раздел 1. Лестница"}]
    rows = [{"name": "Раздел 1. Лестница", "unit": "", "quantity": None,
             "is_section": True, "source_no": "1"}]
    attach_source_numbers(items, rows)
    assert "source_no" not in items[0]


# ── Колонка в файле ──────────────────────────────────────────────────────────

def _sheet_values(data: bytes, title: str) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(data))
    return [list(row) for row in wb[title].iter_rows(values_only=True)]


def test_generate_list_adds_source_column_first():
    items = [
        {"type": "Работа", "name": "Пробивка гнезд", "unit": "100 шт",
         "quantity": 0.02, "notes": "", "source_no": "1"},
        {"type": "Материал", "name": "Смеси бетонные", "unit": "м3",
         "quantity": 0.104, "notes": "", "source_no": ""},
    ]
    values = _sheet_values(generate_list(items), "Перечень")
    assert values[0] == ["№ в исходной смете", "№ п/п", "Тип", "Наименование",
                         "Ед. изм", "Кол-во", "Примечание"]
    assert values[1][:4] == [1, 1, "Работа", "Пробивка гнезд"]
    assert values[2][:4] == [None, 2, "Материал", "Смеси бетонные"]


def test_generate_list_keeps_six_columns_without_numbers():
    items = [{"type": "Работа", "name": "Пробивка гнезд", "unit": "100 шт",
              "quantity": 0.02, "notes": ""}]
    values = _sheet_values(generate_list(items), "Перечень")
    assert values[0] == ["№ п/п", "Тип", "Наименование", "Ед. изм", "Кол-во", "Примечание"]


def test_generate_list_writes_non_numeric_number_as_text():
    items = [{"type": "Работа", "name": "Пробивка гнезд", "unit": "м2",
              "quantity": 1, "notes": "", "source_no": "1.1"}]
    values = _sheet_values(generate_list(items), "Перечень")
    assert values[1][0] == "1.1"


def test_source_column_in_summary_sheets():
    """Сводки «Работы» и «Материалы» — тот же набор колонок, иначе файл разъедется."""
    items = [
        {"type": "Работа", "name": "Пробивка гнезд", "unit": "м2",
         "quantity": 1, "notes": "", "source_no": "1"},
        {"type": "Материал", "name": "Смеси бетонные", "unit": "м3",
         "quantity": 2, "notes": "", "source_no": "2"},
    ]
    data = generate_list(items)
    assert _sheet_values(data, "Работы")[0][0] == "№ в исходной смете"
    assert _sheet_values(data, "Материалы")[1][0] == 2


# ── Круговой рейс через редактор ─────────────────────────────────────────────

def test_number_survives_editor_round_trip():
    items = [
        {"type": "Работа", "name": "Пробивка гнезд", "unit": "100 шт",
         "quantity": 0.02, "notes": "", "source_no": "1"},
        {"type": "Материал", "name": "Смеси бетонные", "unit": "м3",
         "quantity": 0.104, "notes": "", "source_no": "2"},
    ]
    rows = parse_xlsx_to_generic_rows(generate_list(items), sheets=["Перечень"])
    assert [r["cells"]["№ в исходной смете"] for r in rows] == [1, 2]

    back = generic_rows_to_items(rows)
    assert [it["source_no"] for it in back] == ["1", "2"]
    assert [it["name"] for it in back] == ["Пробивка гнезд", "Смеси бетонные"]


def test_round_trip_without_numbers_has_no_field():
    items = [{"type": "Работа", "name": "Пробивка гнезд", "unit": "м2",
              "quantity": 1, "notes": ""}]
    rows = parse_xlsx_to_generic_rows(generate_list(items), sheets=["Перечень"])
    assert generic_rows_to_items(rows)[0].get("source_no", "") == ""


# ── Полнота номер за собой не тянет ──────────────────────────────────────────

def test_completeness_payload_drops_source_no():
    from app.services.task_processor import _without_source_no

    items = [{"type": "Работа", "name": "Пробивка гнезд", "unit": "м2",
              "quantity": 1, "notes": "", "source_no": "1"}]
    clean = _without_source_no(items)
    assert clean == [{"type": "Работа", "name": "Пробивка гнезд", "unit": "м2",
                      "quantity": 1, "notes": ""}]
    # исходные позиции не тронуты
    assert items[0]["source_no"] == "1"


# ── Сквозной путь: файл → строки → позиции ───────────────────────────────────

def test_end_to_end_numbers_from_grand_file():
    data = _grand_xlsx([
        ("1", "ФЕР46-03-009-03", "Пробивка в кирпичных стенах гнезд", "100 шт", 0.02),
        ("2", "ФЕР46-03-017-05", "Заделка отверстий, гнезд и борозд", "м3", 0.1),
        ("3", "ФССЦ-04.1.02.05", "Смеси бетонные тяжелого бетона", "м3", 0.104),
    ])
    rows = parse_xlsx_grand(data)

    # То, что вернул бы ИИ: работы и материалы в своём порядке, наименования дословно
    items = [
        {"type": "Работа", "name": "Пробивка в кирпичных стенах гнезд",
         "unit": "100 шт", "quantity": 0.02, "notes": ""},
        {"type": "Работа", "name": "Заделка отверстий, гнезд и борозд",
         "unit": "м3", "quantity": 0.1, "notes": ""},
        {"type": "Материал", "name": "Смеси бетонные тяжелого бетона",
         "unit": "м3", "quantity": 0.104, "notes": ""},
    ]
    attach_source_numbers(items, rows)

    values = _sheet_values(generate_list(items), "Перечень")
    assert [row[0] for row in values[1:]] == [1, 2, 3]


def test_number_does_not_go_into_prompt():
    """Номер в промпт не уходит: он стоит токенов на каждом чанке и не нужен ИИ."""
    rows = parse_xlsx_grand(_grand_xlsx([
        ("17", "ФЕР46-03-009-03", "Пробивка гнезд", "100 шт", 0.02),
    ]))
    assert rows[0]["source_no"] == "17"
    assert "17" not in rows_to_text(rows)


def test_numbers_do_not_cross_sheets():
    """У каждого листа своя нумерация: чанк границу листа не пересекает."""
    wb = openpyxl.Workbook()
    for title, names in (("Раздел 1", ["Пробивка гнезд", "Заделка борозд"]),
                         ("Раздел 2", ["Пробивка гнезд", "Штукатурка откосов"])):
        ws = wb.create_sheet(title)
        for col, head in enumerate(_HEADER, 1):
            ws.cell(row=1, column=col, value=head)
        for index, name in enumerate(names, 2):
            ws.cell(row=index, column=1, value=str(index - 1))
            ws.cell(row=index, column=3, value=name)
            ws.cell(row=index, column=4, value="м2")
            ws.cell(row=index, column=5, value=1)
    del wb[wb.sheetnames[0]]
    buf = io.BytesIO()
    wb.save(buf)

    rows = parse_xlsx_grand(buf.getvalue())
    chunks = chunk_rows(rows)
    assert len(chunks) == 2

    # ИИ второго листа вернул «Пробивка гнезд» — номер должен быть листа 2 (1),
    # а не листа 1, где то же наименование стоит под тем же номером.
    second = [{"type": "Работа", "name": "Штукатурка откосов"},
              {"type": "Работа", "name": "Пробивка гнезд"}]
    attach_source_numbers(second, chunks[1])
    assert [it["source_no"] for it in second] == ["2", "1"]
    assert {r["sheet"] for r in chunks[1]} == {"Раздел 2"}
