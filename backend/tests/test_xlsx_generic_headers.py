"""Разбор шапок в generic-парсере xlsx (перечень / полнота).

Задача фазы 4: многострочные шапки Гранд-сметы и строка нумерации колонок
(«1 2 3 4 5 6») не должны попадать в данные. Однострочные шапки, на которых
разобраны все существующие перечни, обязаны работать ровно как раньше.
"""
import io
from typing import Optional

import openpyxl

from app.utils.xlsx_generic import parse_xlsx_to_generic_rows, rows_to_xlsx


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _build(matrix: list, merges: Optional[list] = None) -> bytes:
    """Собрать xlsx из матрицы значений (None = пустая ячейка)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for r, row in enumerate(matrix, start=1):
        for c, value in enumerate(row, start=1):
            if value is not None:
                ws.cell(row=r, column=c, value=value)
    for rng in merges or []:
        ws.merge_cells(rng)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cells(rows: list[dict]) -> list[dict]:
    return [r["cells"] for r in rows]


# ---------------------------------------------------------------------------
# Регресс: однострочные шапки — поведение не меняется
# ---------------------------------------------------------------------------

class TestSingleRowHeaderRegression:
    def test_plain_table_parsed_as_before(self):
        data = _build([
            ["Наименование", "Ед. изм", "Кол-во"],
            ["Демонтаж стен", "м2", 12.5],
            ["Кладка кирпича", "м3", 4],
        ])
        rows = parse_xlsx_to_generic_rows(data)

        assert len(rows) == 2
        assert _cells(rows)[0] == {"Наименование": "Демонтаж стен", "Ед. изм": "м2", "Кол-во": 12.5}
        assert _cells(rows)[1] == {"Наименование": "Кладка кирпича", "Ед. изм": "м3", "Кол-во": 4}

    def test_empty_header_cells_get_placeholder_names(self):
        data = _build([
            ["Наименование", None, "Кол-во"],
            ["Демонтаж стен", "м2", 12.5],
        ])
        rows = parse_xlsx_to_generic_rows(data)

        assert list(_cells(rows)[0].keys()) == ["Наименование", "Col2", "Кол-во"]

    def test_completely_empty_rows_skipped(self):
        data = _build([
            ["Наименование", "Ед. изм"],
            ["Демонтаж стен", "м2"],
            [None, None],
            ["Кладка кирпича", "м3"],
        ])
        rows = parse_xlsx_to_generic_rows(data)

        assert len(rows) == 2

    def test_empty_file_returns_empty_list(self):
        assert parse_xlsx_to_generic_rows(_build([])) == []

    def test_first_data_row_of_unrelated_numbers_is_not_a_header(self):
        """Числа, не похожие на нумерацию колонок, остаются данными."""
        data = _build([
            ["Позиция", "Код", "Цена"],
            [5, 12, 40],
            ["Кладка кирпича", 7, 100],
        ])
        rows = parse_xlsx_to_generic_rows(data)

        assert len(rows) == 2
        assert _cells(rows)[0] == {"Позиция": 5, "Код": 12, "Цена": 40}

    def test_numbering_like_row_deep_in_data_is_kept(self):
        """Строка из мелких чисел далеко от шапки — это данные, а не нумерация."""
        matrix = [["Наименование", "Ед. изм", "Кол-во"]]
        matrix += [[f"Работа {i}", "м2", i] for i in range(1, 20)]
        matrix.append([1, 2, 3])
        rows = parse_xlsx_to_generic_rows(_build(matrix))

        assert len(rows) == 20
        assert _cells(rows)[-1] == {"Наименование": 1, "Ед. изм": 2, "Кол-во": 3}


# ---------------------------------------------------------------------------
# Строка нумерации колонок
# ---------------------------------------------------------------------------

class TestNumberingRow:
    def test_numbering_row_after_header_is_dropped(self):
        data = _build([
            ["№ п/п", "Тип", "Наименование", "Ед. изм", "Кол-во", "Примечание"],
            [1, 2, 3, 4, 5, 6],
            [1, "Работа", "Демонтаж стен", "м2", 12.5, ""],
        ])
        rows = parse_xlsx_to_generic_rows(data)

        assert len(rows) == 1
        assert _cells(rows)[0]["Наименование"] == "Демонтаж стен"

    def test_numbering_row_with_gaps_is_dropped(self):
        """В Гранд-смете часть колонок скрыта — нумерация идёт с пропусками."""
        data = _build([
            ["№", "Наименование", "Ед. изм", "Кол-во"],
            [1, 2, 4, 6],
            [1, "Демонтаж стен", "м2", 12.5],
        ])
        rows = parse_xlsx_to_generic_rows(data)

        assert len(rows) == 1
        assert _cells(rows)[0]["Наименование"] == "Демонтаж стен"

    def test_generated_perechen_has_no_junk_first_row(self):
        """Файл нашего же генератора: строка «1 2 3 4 5 6» в данные не попадает."""
        from app.services.excel_service import generate_list

        xlsx = generate_list([
            {"type": "Работа", "name": "Демонтаж стен", "unit": "м2", "quantity": 12.5},
            {"type": "Материал", "name": "Кирпич", "unit": "шт", "quantity": 400},
        ])
        rows = parse_xlsx_to_generic_rows(xlsx)

        assert len(rows) == 2
        first = _cells(rows)[0]
        assert first["Наименование"] == "Демонтаж стен"
        assert not all(isinstance(v, int) for v in first.values())


# ---------------------------------------------------------------------------
# Многострочная шапка Гранд-сметы
# ---------------------------------------------------------------------------

GRAND_MATRIX = [
    ["ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ № 02-01-01", None, None, None, None, None, None],
    ["Ремонт кровли административного здания", None, None, None, None, None, None],
    [None, None, None, None, None, None, None],
    ["№ п/п", "Обоснование", "Наименование работ и затрат", "Ед. изм.", "Количество",
     "Стоимость, руб.", None],
    [None, None, None, None, None, "на единицу", "всего"],
    [1, 2, 3, 4, 5, 6, 7],
    [1, "ФЕР08-02-001-01", "Кладка стен из кирпича", "м3", 4.5, 1200, 5400],
    [2, "ФЕР09-03-002-05", "Демонтаж кровли", "м2", 120, 85, 10200],
]
GRAND_MERGES = ["A4:A5", "B4:B5", "C4:C5", "D4:D5", "E4:E5", "F4:G4"]


class TestGrandMultiRowHeader:
    def test_first_data_row_is_real(self):
        rows = parse_xlsx_to_generic_rows(_build(GRAND_MATRIX, GRAND_MERGES))

        assert len(rows) == 2
        assert _cells(rows)[0]["Наименование работ и затрат"] == "Кладка стен из кирпича"
        assert _cells(rows)[1]["Наименование работ и затрат"] == "Демонтаж кровли"

    def test_title_block_is_not_data(self):
        rows = parse_xlsx_to_generic_rows(_build(GRAND_MATRIX, GRAND_MERGES))

        joined = " ".join(str(v) for cells in _cells(rows) for v in cells.values())
        assert "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ" not in joined
        assert "Ремонт кровли административного здания" not in joined

    def test_header_rows_joined_into_column_names(self):
        rows = parse_xlsx_to_generic_rows(_build(GRAND_MATRIX, GRAND_MERGES))

        assert list(_cells(rows)[0].keys()) == [
            "№ п/п",
            "Обоснование",
            "Наименование работ и затрат",
            "Ед. изм.",
            "Количество",
            "Стоимость, руб. на единицу",
            "Стоимость, руб. всего",
        ]

    def test_vertically_merged_header_not_duplicated(self):
        rows = parse_xlsx_to_generic_rows(_build(GRAND_MATRIX, GRAND_MERGES))

        assert "Обоснование Обоснование" not in _cells(rows)[0]

    def test_values_land_in_right_columns(self):
        rows = parse_xlsx_to_generic_rows(_build(GRAND_MATRIX, GRAND_MERGES))

        assert _cells(rows)[0]["Стоимость, руб. на единицу"] == 1200
        assert _cells(rows)[0]["Стоимость, руб. всего"] == 5400


# ---------------------------------------------------------------------------
# Имена колонок
# ---------------------------------------------------------------------------

class TestColumnNames:
    def test_duplicate_header_names_do_not_lose_data(self):
        data = _build([
            ["Наименование", "Кол-во", "Кол-во"],
            ["Демонтаж стен", 10, 20],
        ])
        rows = parse_xlsx_to_generic_rows(data)

        cells = _cells(rows)[0]
        assert len(cells) == 3
        assert list(cells.values()) == ["Демонтаж стен", 10, 20]


# ---------------------------------------------------------------------------
# Идемпотентность
# ---------------------------------------------------------------------------

class TestIdempotency:
    def test_reparse_of_saved_rows_gives_same_cells(self):
        rows = parse_xlsx_to_generic_rows(_build(GRAND_MATRIX, GRAND_MERGES))
        again = parse_xlsx_to_generic_rows(rows_to_xlsx(rows))

        assert _cells(again) == _cells(rows)

    def test_generated_perechen_roundtrip_is_stable(self):
        from app.services.excel_service import generate_list

        xlsx = generate_list([
            {"type": "Работа", "name": "Демонтаж стен", "unit": "м2", "quantity": 12.5},
        ])
        first = parse_xlsx_to_generic_rows(xlsx)
        second = parse_xlsx_to_generic_rows(rows_to_xlsx(first))
        third = parse_xlsx_to_generic_rows(rows_to_xlsx(second))

        assert _cells(second) == _cells(first)
        assert _cells(third) == _cells(second)
