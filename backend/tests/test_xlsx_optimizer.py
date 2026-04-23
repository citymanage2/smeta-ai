"""Tests for xlsx_optimizer utility."""
import io
import openpyxl

from app.utils.xlsx_optimizer import (
    parse_estimate_xlsx,
    get_top_items,
    generate_optimized_xlsx,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_smeta_xlsx(items: list[dict]) -> bytes:
    """Build a minimal smeta xlsx in the format excel_service.generate_smeta() produces."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"
    headers = [
        "№", "Тип", "Наименование", "Ед. изм.", "Кол-во",
        "Цена работы (за ед.)", "Цена материала (за ед.)",
        "Стоимость работ", "Стоимость материалов",
        "Итого без НДС", "НДС (22%)", "Итого с НДС",
        "Наименование в прайсе", "Источники", "Примечание",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    VAT = 0.22
    for i, item in enumerate(items, start=1):
        row = i + 1
        qty = item.get("quantity", 1)
        wp = item.get("work_price", 0)
        mp = item.get("material_price", 0)
        subtotal = qty * (wp + mp)
        vat = subtotal * VAT
        total = subtotal + vat
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get("type", "Материал"))
        ws.cell(row=row, column=3, value=item["name"])
        ws.cell(row=row, column=4, value=item.get("unit", "шт"))
        ws.cell(row=row, column=5, value=qty)
        ws.cell(row=row, column=6, value=wp or None)
        ws.cell(row=row, column=7, value=mp or None)
        ws.cell(row=row, column=8, value=qty * wp or None)
        ws.cell(row=row, column=9, value=qty * mp or None)
        ws.cell(row=row, column=10, value=subtotal or None)
        ws.cell(row=row, column=11, value=vat or None)
        ws.cell(row=row, column=12, value=total or None)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_smeta_items():
    """Return 5 test items with varying total costs."""
    return [
        {"name": "Кирпич М150", "type": "Материал", "unit": "шт", "quantity": 1000, "material_price": 10.0},   # total=12000
        {"name": "Монтаж трубы", "type": "Работа", "unit": "м", "quantity": 50, "work_price": 500.0},           # total=30000
        {"name": "Цемент М400", "type": "Материал", "unit": "кг", "quantity": 200, "material_price": 20.0},    # total=4800
        {"name": "Устройство стяжки", "type": "Работа", "unit": "м2", "quantity": 100, "work_price": 300.0},   # total=36000
        {"name": "Гвозди 100мм", "type": "Материал", "unit": "кг", "quantity": 10, "material_price": 50.0},   # total=600
    ]
    # Sorted by total desc: Устройство стяжки(36k), Монтаж трубы(30k), Кирпич М150(12k), Цемент М400(4.8k), Гвозди(600)
    # Grand total = 83400
    # 70% threshold = 58380 → need Устройство(36k) + Монтаж(30k) = 66k > 58380 → first 2 items


# ---------------------------------------------------------------------------
# parse_estimate_xlsx tests
# ---------------------------------------------------------------------------

def test_parse_estimate_xlsx_returns_items():
    """Parsing a well-formed smeta xlsx returns list of items with correct fields."""
    items = _make_smeta_items()[:3]
    file_bytes = _make_smeta_xlsx(items)
    result = parse_estimate_xlsx(file_bytes)
    assert len(result) == 3
    first = result[0]
    assert first["name"] == "Кирпич М150"
    assert first["type"] == "material"
    assert first["unit"] == "шт"
    assert first["quantity"] == 1000
    assert isinstance(first["row_index"], int)
    assert isinstance(first["price_excl_vat"], float)
    assert isinstance(first["price_incl_vat"], float)
    assert isinstance(first["total"], float)
    assert first["price_incl_vat"] > first["price_excl_vat"]


def test_parse_estimate_xlsx_skips_empty_rows():
    """Rows with no name value are skipped."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["№", "Тип", "Наименование", "Ед. изм.", "Кол-во", "Цена работы (за ед.)", "Цена материала (за ед.)",
               "Стоимость работ", "Стоимость материалов", "Итого без НДС", "НДС (20%)", "Итого с НДС",
               "Наименование в прайсе", "Источники", "Примечание"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    # Row 2: valid item
    ws.cell(row=2, column=2, value="Материал")
    ws.cell(row=2, column=3, value="Кирпич М150")
    ws.cell(row=2, column=5, value=100)
    ws.cell(row=2, column=7, value=10.0)
    ws.cell(row=2, column=12, value=1200.0)
    # Row 3: empty name (ИТОГО row)
    ws.cell(row=3, column=1, value="ИТОГО")
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_estimate_xlsx(buf.getvalue())
    assert len(result) == 1
    assert result[0]["name"] == "Кирпич М150"


# ---------------------------------------------------------------------------
# get_top_items tests
# ---------------------------------------------------------------------------

def test_get_top_items_covers_threshold():
    """Selected items' cumulative total covers at least 70% of grand total."""
    items = _make_smeta_items()
    file_bytes = _make_smeta_xlsx(items)
    parsed = parse_estimate_xlsx(file_bytes)
    selected = get_top_items(parsed, categories=["work", "material"], threshold=0.7)
    grand_total = sum(it["total"] for it in parsed)
    selected_total = sum(it["total"] for it in selected)
    assert selected_total >= 0.7 * grand_total


def test_get_top_items_filters_categories():
    """Only items matching the requested categories are returned."""
    items = _make_smeta_items()
    file_bytes = _make_smeta_xlsx(items)
    parsed = parse_estimate_xlsx(file_bytes)
    # Ask only for materials
    selected = get_top_items(parsed, categories=["material"], threshold=0.99)
    assert all(it["type"] == "material" for it in selected)


# ---------------------------------------------------------------------------
# generate_optimized_xlsx tests
# ---------------------------------------------------------------------------

def _make_optimization_results(items_parsed: list[dict]) -> list[dict]:
    """Build mock optimization_results: first item found, rest not found."""
    results = []
    for i, it in enumerate(items_parsed):
        if i == 0:
            results.append({
                "row_index": it["row_index"],
                "name": it["name"],
                "original_price": it["price_incl_vat"],
                "new_price": it["price_incl_vat"] * 0.8,
                "source": "https://example.com/price",
                "savings_abs": it["price_incl_vat"] * 0.2,
                "savings_pct": 20.0,
                "has_vat": True,
            })
        else:
            results.append({
                "row_index": it["row_index"],
                "name": it["name"],
                "original_price": it["price_incl_vat"],
                "new_price": None,
                "source": "Не найдено",
                "savings_abs": None,
                "savings_pct": None,
                "has_vat": True,
            })
    return results


def test_generate_optimized_xlsx_adds_columns():
    """Optimized xlsx has 4 extra columns after original columns."""
    items = _make_smeta_items()[:3]
    original_bytes = _make_smeta_xlsx(items)
    parsed = parse_estimate_xlsx(original_bytes)
    opt_results = _make_optimization_results(parsed)
    result_bytes = generate_optimized_xlsx(original_bytes, opt_results)
    wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
    ws = wb.active
    # Original has 15 columns; optimized adds 4 more
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Цена сниженная" in header_row
    assert "Стоимость сниженная" in header_row
    assert "Источник" in header_row
    assert "Примечание" in header_row


def test_generate_optimized_xlsx_has_comparison_sheet():
    """Result workbook contains a second sheet named 'Сравнение'."""
    items = _make_smeta_items()[:2]
    original_bytes = _make_smeta_xlsx(items)
    parsed = parse_estimate_xlsx(original_bytes)
    opt_results = _make_optimization_results(parsed)
    result_bytes = generate_optimized_xlsx(original_bytes, opt_results)
    wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
    assert "Сравнение" in wb.sheetnames


def test_generate_optimized_xlsx_green_fill_for_found():
    """Rows with a found analogue get green fill (#E2EFDA) in the 'Цена сниженная' column."""
    items = _make_smeta_items()[:2]
    original_bytes = _make_smeta_xlsx(items)
    parsed = parse_estimate_xlsx(original_bytes)
    # First item found
    opt_results = _make_optimization_results(parsed)
    result_bytes = generate_optimized_xlsx(original_bytes, opt_results)
    wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
    ws = wb.active
    # Find the "Цена сниженная" column index
    price_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "Цена сниженная":
            price_col = c
            break
    assert price_col is not None
    # Data row for first item = row 2; first opt_result has new_price (found)
    found_row = parsed[0]["row_index"]
    cell = ws.cell(row=found_row, column=price_col)
    assert cell.fill.fgColor.rgb.upper().endswith("E2EFDA")


def test_generate_optimized_xlsx_yellow_fill_for_not_found():
    """Rows with no analogue get yellow fill (#FFEB9C) in the 'Цена сниженная' column."""
    items = _make_smeta_items()[:2]
    original_bytes = _make_smeta_xlsx(items)
    parsed = parse_estimate_xlsx(original_bytes)
    opt_results = _make_optimization_results(parsed)  # item[1] is not found
    result_bytes = generate_optimized_xlsx(original_bytes, opt_results)
    wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
    ws = wb.active
    price_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "Цена сниженная":
            price_col = c
            break
    # Second item = not found
    not_found_row = parsed[1]["row_index"]
    cell = ws.cell(row=not_found_row, column=price_col)
    assert cell.fill.fgColor.rgb.upper().endswith("FFEB9C")
