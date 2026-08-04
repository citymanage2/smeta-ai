"""Имена листов и группировка строк по листам (план 2026-08-04, Фаза 1)."""
import io

import openpyxl
import pytest

from app.utils.sheet_names import (
    DEFAULT_SHEET_TITLE,
    SHEET_TITLE_LIMIT,
    group_by_sheet,
    safe_sheet_title,
    sheet_of,
    sheet_titles,
)


class TestSafeSheetTitle:
    def test_plain_name_unchanged(self):
        assert safe_sheet_title("Раздел 1") == "Раздел 1"

    @pytest.mark.parametrize("bad", [":", "\\", "/", "?", "*", "[", "]"])
    def test_forbidden_characters_removed(self, bad):
        title = safe_sheet_title(f"Раздел{bad}1")
        assert bad not in title
        # Имя должно остаться осмысленным, а не превратиться в «Лист»
        assert "Раздел" in title

    def test_long_name_truncated_to_excel_limit(self):
        title = safe_sheet_title("Локальный сметный расчёт на общестроительные работы")
        assert len(title) <= SHEET_TITLE_LIMIT

    def test_empty_name_gets_default(self):
        assert safe_sheet_title("") == DEFAULT_SHEET_TITLE
        assert safe_sheet_title(None) == DEFAULT_SHEET_TITLE
        assert safe_sheet_title("///") == DEFAULT_SHEET_TITLE

    def test_duplicates_get_suffix(self):
        used: set = set()
        assert safe_sheet_title("Раздел", used) == "Раздел"
        assert safe_sheet_title("Раздел", used) == "Раздел (2)"
        assert safe_sheet_title("Раздел", used) == "Раздел (3)"

    def test_duplicate_of_long_name_still_fits_limit(self):
        used: set = set()
        long_name = "Локальный сметный расчёт по корпусу номер один"
        first = safe_sheet_title(long_name, used)
        second = safe_sheet_title(long_name, used)
        assert first != second
        assert len(second) <= SHEET_TITLE_LIMIT

    def test_duplicate_check_ignores_case(self):
        # openpyxl сравнивает имена листов без учёта регистра — иначе второй
        # лист молча заменил бы первый.
        used: set = set()
        assert safe_sheet_title("Раздел", used) == "Раздел"
        assert safe_sheet_title("РАЗДЕЛ", used) != "РАЗДЕЛ"

    def test_result_is_accepted_by_openpyxl(self):
        wb = openpyxl.Workbook()
        used: set = set()
        for name in ["Смета: часть 1", "Смета: часть 1", "?", "x" * 60]:
            wb.create_sheet(safe_sheet_title(name, used))
        buf = io.BytesIO()
        wb.save(buf)  # не должно упасть
        assert len(wb.sheetnames) == len(set(n.lower() for n in wb.sheetnames))


class TestGrouping:
    def test_groups_in_first_seen_order(self):
        rows = [
            {"sheet": "Б", "n": 1},
            {"sheet": "А", "n": 2},
            {"sheet": "Б", "n": 3},
        ]
        assert group_by_sheet(rows) == [
            ("Б", [{"sheet": "Б", "n": 1}, {"sheet": "Б", "n": 3}]),
            ("А", [{"sheet": "А", "n": 2}]),
        ]

    def test_rows_without_sheet_form_single_unnamed_group(self):
        rows = [{"n": 1}, {"n": 2}]
        groups = group_by_sheet(rows)
        assert len(groups) == 1
        assert groups[0][0] is None

    def test_blank_sheet_counts_as_absent(self):
        assert sheet_of({"sheet": "   "}) is None
        assert sheet_of({"sheet": None}) is None
        assert sheet_of({}) is None

    def test_sheet_titles_skips_rows_without_sheet(self):
        rows = [{"sheet": "А"}, {}, {"sheet": "Б"}]
        assert sheet_titles(rows) == ["А", "Б"]

    def test_empty_input(self):
        assert group_by_sheet([]) == []
        assert sheet_titles([]) == []
