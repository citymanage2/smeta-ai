"""Дата у позиции прайса — это дата ЦЕНЫ, а не дата правки записи.

План: `plans/2026-09-01-data-tseny-v-prayse.md`.

Менеджер смотрит на колонку «Цена от», чтобы понять, насколько цена свежая.
Пока дату двигало любое касание записи, ответ был неверный: переименование
позиции, заполнение единицы и перезаливка прайса файлом делали дату
сегодняшней, хотя цена не менялась. После перезаливки объединённого файла
колонка обнулялась целиком — сегодняшняя дата у всех тысяч позиций.

Здесь закреплены оба края правила: что дату двигает (число цены, смена единицы
между двумя непустыми написаниями) и что не двигает (имя, появление единицы,
повторная запись того же числа).
"""
import io
from datetime import datetime, timedelta, timezone
from typing import Optional

import openpyxl
import pytest
import pytest_asyncio
from sqlalchemy import delete, select

from app.models.price import PriceMaterial, PriceWork
from app.services import price_bulk
from app.utils.auth import create_access_token
from app.utils.price_change import (
    MONEY_EPS,
    price_changed,
    prices_changed,
    unit_changed,
)
from app.utils.price_min import ESTIMATE_CONTRACTOR

# asyncio_mode=auto (pytest.ini) — async-тесты подхватываются сами, а глобальный
# маркер вешал бы asyncio и на синхронные проверки правила.

OLD = datetime(2026, 3, 1, 10, 0, tzinfo=timezone.utc)


def _admin() -> dict:
    return {"Authorization": f"Bearer {create_access_token(1, 'admin')}"}


def _naive(value: datetime) -> datetime:
    """SQLite отдаёт дату без таймзоны — сравниваем в одном представлении."""
    return value.replace(tzinfo=None) if value.tzinfo else value


def _is_old(value: datetime) -> bool:
    return _naive(value) == _naive(OLD)


async def _fresh(db, model, **where):
    """Перечитать строку из БД, а не из кэша сессии.

    Тестовая сессия создана с `expire_on_commit=False`, поэтому обычный `select`
    вернул бы тот же объект с атрибутами в памяти — и тест зеленел бы, даже если
    дата до базы не доехала.
    """
    db.expire_all()
    q = select(model)
    for field, value in where.items():
        q = q.where(getattr(model, field) == value)
    return (await db.execute(q)).scalar_one()


@pytest_asyncio.fixture(autouse=True)
async def _clean_prices(db_session):
    """Прайс общий на весь прогон — чистим до и после, чтобы тесты не влияли друг на друга."""
    await db_session.execute(delete(PriceWork))
    await db_session.execute(delete(PriceMaterial))
    await db_session.commit()
    yield
    await db_session.execute(delete(PriceWork))
    await db_session.execute(delete(PriceMaterial))
    await db_session.commit()


@pytest.fixture(autouse=True)
def _no_embeddings(monkeypatch):
    """Векторы здесь не проверяются, а модель тяжёлая и лезет в сеть."""
    async def _none(_name: str) -> Optional[list]:
        return None

    def _batch(names, input_type="search_document"):
        return [None] * len(names)

    monkeypatch.setattr("app.routers.prices_catalog._generate_embedding_safe", _none)
    monkeypatch.setattr("app.services.price_bulk._generate_embedding_safe", _none)
    monkeypatch.setattr("app.routers.admin.generate_embeddings_batch", _batch)


@pytest_asyncio.fixture
async def work(db_session) -> PriceWork:
    row = PriceWork(
        name="Кладка стен", unit="м2",
        prices={"Подрядчик А": 500.0}, min_price=500.0, updated_at=OLD,
    )
    db_session.add(row)
    await db_session.commit()
    await db_session.refresh(row)
    return row


@pytest_asyncio.fixture
async def material(db_session) -> PriceMaterial:
    row = PriceMaterial(name="Кирпич", unit="шт", price=25.0, updated_at=OLD)
    db_session.add(row)
    await db_session.commit()
    await db_session.refresh(row)
    return row


# ---------------------------------------------------------------------------
# Правило: что вообще считается изменением цены
# ---------------------------------------------------------------------------

def test_money_eps_is_half_kopeck():
    """Порог закреплён числом: под датой «примерно» превращается в бегущую цифру."""
    assert MONEY_EPS == 0.005


def test_price_changed_below_eps_is_not_a_change():
    """1234.56 из xlsx приезжает как 1234.5600000000001 — это не переоценка."""
    assert price_changed(1234.56, 1234.5640) is False
    assert price_changed(1234.56, 1234.5660) is True


def test_price_appearing_and_disappearing_is_a_change():
    assert price_changed(None, 100.0) is True
    assert price_changed(100.0, None) is True


def test_absence_of_price_has_one_meaning():
    """None, 0, пустая строка и мусор — одно и то же «цены нет»."""
    assert price_changed(None, 0) is False
    assert price_changed(0, "") is False
    assert price_changed("", "—") is False


def test_same_unit_written_differently_is_not_a_change():
    """«м²», «кв. м» и «М2» — одна единица, а не переоценка."""
    assert unit_changed("м2", "м²") is False
    assert unit_changed("м2", "кв. м") is False
    assert unit_changed("м2", " М2 ") is False


def test_different_unit_is_a_change_of_price():
    """«500 ₽ за м2» и «500 ₽ за м3» — разные цены (правило №8 CLAUDE.md)."""
    assert unit_changed("м2", "м3") is True
    assert price_changed(500.0, 500.0, "м2", "м3") is True


def test_filling_in_an_empty_unit_is_not_a_change():
    """Пустая единица ничего не утверждает о цене: её появление — уточнение описания."""
    assert unit_changed("", "м2") is False
    assert unit_changed("м2", "") is False
    assert price_changed(500.0, 500.0, None, "м2") is False


def test_work_prices_compared_by_whole_set():
    """У работы цен несколько: расчёт берёт из набора одно число, и завтра — другое."""
    assert prices_changed({"А": 500.0}, {"А": 500.0}) is False
    assert prices_changed({"А": 500.0}, {"А": 550.0}) is True
    assert prices_changed({"А": 500.0}, {"А": 500.0, "Б": 700.0}) is True
    assert prices_changed({"А": 500.0, "Б": 700.0}, {"А": 500.0}) is True
    assert prices_changed({"А": 500.0}, {}) is True
    assert prices_changed({"А": 500.0, "Б": None}, {"А": 500.0}) is False


# ---------------------------------------------------------------------------
# Ручная правка каталога (критерии 1-4)
# ---------------------------------------------------------------------------

async def test_renaming_work_keeps_price_date(async_client, db_session, work):
    work_id = work.id
    r = await async_client.put(
        f"/prices/catalog/works/{work_id}",
        json={"name": "Кладка наружных стен"}, headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceWork, id=work_id)
    assert row.name == "Кладка наружных стен"
    assert _is_old(row.updated_at)


async def test_repricing_work_moves_date(async_client, db_session, work):
    work_id = work.id
    r = await async_client.put(
        f"/prices/catalog/works/{work_id}",
        json={"prices": {"Подрядчик А": 550.0}}, headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceWork, id=work_id)
    assert not _is_old(row.updated_at)


async def test_changing_work_unit_moves_date(async_client, db_session, work):
    work_id = work.id
    r = await async_client.put(
        f"/prices/catalog/works/{work_id}", json={"unit": "м3"}, headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceWork, id=work_id)
    assert not _is_old(row.updated_at)


async def test_rewriting_work_with_same_price_keeps_date(async_client, db_session, work):
    work_id = work.id
    r = await async_client.put(
        f"/prices/catalog/works/{work_id}",
        json={"prices": {"Подрядчик А": 500.0}, "unit": "кв. м"}, headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceWork, id=work_id)
    assert _is_old(row.updated_at)


async def test_renaming_material_keeps_price_date(async_client, db_session, material):
    material_id = material.id
    r = await async_client.put(
        f"/prices/catalog/materials/{material_id}",
        json={"name": "Кирпич керамический"}, headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceMaterial, id=material_id)
    assert row.name == "Кирпич керамический"
    assert _is_old(row.updated_at)


async def test_repricing_material_moves_date(async_client, db_session, material):
    material_id = material.id
    r = await async_client.put(
        f"/prices/catalog/materials/{material_id}",
        json={"price": 27.5}, headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceMaterial, id=material_id)
    assert not _is_old(row.updated_at)


async def test_changing_material_unit_moves_date(async_client, db_session, material):
    material_id = material.id
    r = await async_client.put(
        f"/prices/catalog/materials/{material_id}",
        json={"unit": "т"}, headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceMaterial, id=material_id)
    assert not _is_old(row.updated_at)


async def test_new_catalog_item_gets_today(async_client, db_session):
    r = await async_client.post(
        "/prices/catalog/materials",
        json={"name": "Песок", "unit": "м3", "price": 900.0}, headers=_admin(),
    )
    assert r.status_code == 201
    row = await _fresh(db_session, PriceMaterial, name="Песок")
    assert not _is_old(row.updated_at)


# ---------------------------------------------------------------------------
# Пакетное добавление из редактора (критерий 5)
# ---------------------------------------------------------------------------

async def test_bulk_add_same_price_keeps_date(db_session):
    row = PriceWork(
        name="Штукатурка", unit="м2",
        prices={ESTIMATE_CONTRACTOR: 300.0}, min_price=300.0, updated_at=OLD,
    )
    db_session.add(row)
    await db_session.commit()

    summary = await price_bulk.add_items(db_session, [{
        "kind": "work", "name": "Штукатурка", "unit": "м2", "price": 300.0,
    }])
    assert summary["updated"] == 1, summary

    fresh = await _fresh(db_session, PriceWork, name="Штукатурка")
    assert _is_old(fresh.updated_at)


async def test_bulk_add_new_price_moves_date(db_session):
    row = PriceWork(
        name="Штукатурка", unit="м2",
        prices={ESTIMATE_CONTRACTOR: 300.0}, min_price=300.0, updated_at=OLD,
    )
    db_session.add(row)
    await db_session.commit()

    summary = await price_bulk.add_items(db_session, [{
        "kind": "work", "name": "Штукатурка", "unit": "м2", "price": 360.0,
    }])
    assert summary["updated"] == 1, summary

    fresh = await _fresh(db_session, PriceWork, name="Штукатурка")
    assert not _is_old(fresh.updated_at)
    assert fresh.min_price == 360.0


async def test_bulk_add_same_material_price_keeps_date(db_session):
    row = PriceMaterial(name="Цемент", unit="кг", price=12.0, updated_at=OLD)
    db_session.add(row)
    await db_session.commit()

    summary = await price_bulk.add_items(db_session, [{
        "kind": "material", "name": "Цемент", "unit": "кг", "price": 12.0,
    }])
    assert summary["updated"] == 1, summary

    fresh = await _fresh(db_session, PriceMaterial, name="Цемент")
    assert _is_old(fresh.updated_at)


# ---------------------------------------------------------------------------
# Загрузка прайса файлом (критерии 6-8)
# ---------------------------------------------------------------------------

def _works_xlsx(rows: list, sheet: str = "Работы") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Наименование", "Ед. изм.", "Подрядчик А"])
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _combined_xlsx(works: list, materials: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Работы"
    ws.append(["Наименование", "Ед. изм.", "Подрядчик А"])
    for r in works:
        ws.append(list(r))
    ws2 = wb.create_sheet("Материалы")
    ws2.append(["Наименование", "Ед. изм.", "Цена"])
    for r in materials:
        ws2.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_price_list_upload_keeps_date_for_unchanged_price(async_client, db_session, work):
    data = _works_xlsx([("Кладка стен", "м2", 500.0)])
    r = await async_client.post(
        "/admin/price-lists/works",
        files={"file": ("price.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceWork, name="Кладка стен")
    assert _is_old(row.updated_at)


async def test_price_list_upload_moves_date_for_changed_price(async_client, db_session, work):
    data = _works_xlsx([("Кладка стен", "м2", 620.0)])
    r = await async_client.post(
        "/admin/price-lists/works",
        files={"file": ("price.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=_admin(),
    )
    assert r.status_code == 200
    row = await _fresh(db_session, PriceWork, name="Кладка стен")
    assert not _is_old(row.updated_at)


async def test_combined_upload_keeps_dates_of_untouched_prices(
    async_client, db_session, work, material,
):
    """Главный случай: в перезалитом файле изменилась одна цена из трёх."""
    db_session.add(PriceWork(
        name="Демонтаж", unit="м2", prices={"Подрядчик А": 100.0},
        min_price=100.0, updated_at=OLD,
    ))
    await db_session.commit()

    data = _combined_xlsx(
        works=[("Кладка стен", "м2", 500.0), ("Демонтаж", "м2", 130.0),
               ("Новая работа", "м2", 90.0)],
        materials=[("Кирпич", "шт", 25.0)],
    )
    r = await async_client.post(
        "/admin/prices/upload",
        files={"file": ("all.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=_admin(),
    )
    assert r.status_code == 200

    db_session.expire_all()
    rows = {
        w.name: w for w in (await db_session.execute(select(PriceWork))).scalars().all()
    }
    assert _is_old(rows["Кладка стен"].updated_at), "цена не менялась — дата прежняя"
    assert not _is_old(rows["Демонтаж"].updated_at), "цена изменилась — дата сегодняшняя"
    assert not _is_old(rows["Новая работа"].updated_at), "новая позиция — дата сегодняшняя"

    kirpich = await _fresh(db_session, PriceMaterial, name="Кирпич")
    assert _is_old(kirpich.updated_at)


# ---------------------------------------------------------------------------
# Запрет на onupdate: иначе всё вышеописанное молча перестанет работать
# ---------------------------------------------------------------------------

def test_updated_at_has_no_onupdate():
    """`onupdate` двигал бы дату на любой UPDATE — ровно та проблема, что чинится."""
    for model in (PriceWork, PriceMaterial):
        column = model.__table__.c.updated_at
        assert column.onupdate is None, (
            f"{model.__name__}.updated_at получил onupdate: дата снова стала бы "
            "датой правки записи, а не датой цены"
        )
