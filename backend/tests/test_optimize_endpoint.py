"""Tests for POST /tasks/{task_id}/optimize/analyze and /optimize/run endpoints.

Порядок поиска цен в _run_optimization_background (важно при мокировании):
  1. Точное совпадение (_exact_match_work / _exact_match_material)
  2. Embedding-поиск (_embedding_match_work / _embedding_match_material) — вызывается ПЕРВЫМ
  3. Веб-поиск через Claude (_web_search_work_price / _web_search_material_price)

При создании тестов, которые мокируют PriceService.find_work_price / find_material_price,
учитывайте что embedding-поиск уже был выполнен внутри этих методов до веб-поиска.
Для изоляции веб-поиска — мокируйте на уровне PriceService, не на уровне отдельных функций.
"""
import io
import pytest
import pytest_asyncio
import openpyxl
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.models.task import Task
from app.models.result import TaskResult
from tests.conftest import store_result_row


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_smeta_xlsx_bytes() -> bytes:
    """Build a minimal valid smeta xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"
    headers = [
        "№", "Тип", "Наименование", "Ед. изм.", "Кол-во",
        "Цена работы (за ед.)", "Цена материала (за ед.)",
        "Стоимость работ", "Стоимость материалов",
        "Итого без НДС", "НДС (20%)", "Итого с НДС",
        "Наименование в прайсе", "Источники", "Примечание",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=2, value="Материал")
    ws.cell(row=2, column=3, value="Кирпич М150")
    ws.cell(row=2, column=4, value="шт")
    ws.cell(row=2, column=5, value=1000)
    ws.cell(row=2, column=7, value=10.0)
    ws.cell(row=2, column=10, value=10000.0)
    ws.cell(row=2, column=11, value=2000.0)
    ws.cell(row=2, column=12, value=12000.0)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


TASK_ID = "b1000000-0000-0000-0000-000000000001"
# SQLite stores UUIDs without dashes
_TASK_ID_HEX = TASK_ID.replace("-", "")


@pytest_asyncio.fixture
async def seed_optimize_task(db_session: AsyncSession):
    """Seed a task with an estimate slot result."""
    task = Task(owner_id=1, 
        id=TASK_ID,
        user_role="user",
        task_type="LIST_FROM_GRAND",
        status="completed",
        estimation_status="estimated",
        input_files=[{"name": "s.pdf", "mime_type": "application/pdf", "size_bytes": 10}],
        input_file_data=[{"name": "s.pdf", "mime_type": "application/pdf", "size_bytes": 10, "content_b64": "dGVzdA=="}],
        chat_history=[],
    )
    db_session.add(task)
    await db_session.flush()

    await store_result_row(
        db_session, TASK_ID, "estimate", "estimate.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        _make_smeta_xlsx_bytes(),
    )
    await db_session.commit()
    yield
    await db_session.execute(text("DELETE FROM task_results WHERE task_id = :tid"), {"tid": _TASK_ID_HEX})
    await db_session.execute(text("DELETE FROM tasks WHERE id = :tid"), {"tid": _TASK_ID_HEX})
    await db_session.commit()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_analyze_returns_items(async_client: AsyncClient, user_token: str, seed_optimize_task):
    """POST /optimize/analyze returns 200 with items list."""
    resp = await async_client.post(
        f"/tasks/{TASK_ID}/optimize/analyze",
        json={"categories": ["work", "material"]},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert "items" in data
    assert "total_analyzed" in data
    assert "total_selected" in data
    assert "coverage_pct" in data
    assert isinstance(data["items"], list)


@pytest.mark.asyncio
async def test_analyze_empty_slot_returns_404(async_client: AsyncClient, user_token: str, db_session: AsyncSession):
    """POST /optimize/analyze returns 404 when no estimate slot exists."""
    no_slot_id = "c1000000-0000-0000-0000-000000000002"
    task = Task(owner_id=1, 
        id=no_slot_id,
        user_role="user",
        task_type="LIST_FROM_GRAND",
        status="completed",
        estimation_status="estimated",
        input_files=[],
        input_file_data=[],
        chat_history=[],
    )
    db_session.add(task)
    await db_session.commit()

    resp = await async_client.post(
        f"/tasks/{no_slot_id}/optimize/analyze",
        json={"categories": ["material"]},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 404

    await db_session.execute(text("DELETE FROM tasks WHERE id = :tid"), {"tid": no_slot_id.replace("-", "")})
    await db_session.commit()


@pytest.mark.asyncio
async def test_run_starts_background(async_client: AsyncClient, user_token: str, seed_optimize_task):
    """POST /optimize/run returns 200 with optimization_started status."""
    items = [
        {
            "row_index": 2,
            "name": "Кирпич М150",
            "type": "material",
            "quantity": 1000,
            "unit": "шт",
            "price_excl_vat": 10.0,
            "price_incl_vat": 12.0,
            "total": 12000.0,
        }
    ]
    resp = await async_client.post(
        f"/tasks/{TASK_ID}/optimize/run",
        json={"items": items, "prompt": "Тест", "categories": ["material"]},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "optimization_started"
    assert data["task_id"] == TASK_ID


@pytest.mark.asyncio
async def test_run_task_not_found(async_client: AsyncClient, user_token: str):
    """POST /optimize/run returns 404 for nonexistent task."""
    resp = await async_client.post(
        f"/tasks/nonexistent-id/optimize/run",
        json={"items": [], "prompt": "", "categories": ["material"]},
        headers={"Authorization": user_token},
    )
    assert resp.status_code == 404
