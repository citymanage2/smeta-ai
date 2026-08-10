"""Лист «Сводная» в скачиваемом файле — те же колонки и числа, что на экране.

Блок «Себестоимость и цена для заказчика» раньше уезжал в файл урезанным: у
верхних строк были обе стоимости, а непредвиденные, прибыль, налоги и итог
писались одной суммой в отдельную колонку. Ввод человека (проценты, количество
людей, коэффициент к ценам) не выгружался вовсе.

Числа берутся из того же набора данных, что и `test_summary_calc.py`, — он
совпадает с фронтовым регресс-тестом до копейки.
"""
import io
from types import SimpleNamespace

import openpyxl
import pytest

from app.utils.summary_calc import calc_summary
from app.utils.xlsx_summary import generate_summary_xlsx

from tests.test_summary_calc import (
    OVERRIDES,
    SECTIONS,
    _TAX_OVERRIDES,
    _TAX_SECTION_ROWS,
)


@pytest.fixture(scope="module")
def sheet():
    summary = SimpleNamespace(sections=SECTIONS, overrides=OVERRIDES)
    wb = openpyxl.load_workbook(io.BytesIO(generate_summary_xlsx(summary)))
    return wb["Сводная"]


@pytest.fixture(scope="module")
def calc():
    return calc_summary(SECTIONS, OVERRIDES)


def _find(ws, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == label:
            return row
    raise AssertionError(f"строка «{label}» не найдена в листе")


def test_header_matches_screen(sheet):
    assert [sheet.cell(row=1, column=c).value for c in range(1, 6)] == [
        "№", "Наименование", "% / Кол-во", "Стоимость с НДС", "Стоимость без НДС",
    ]


def test_fixed_row_has_both_costs(sheet, calc):
    row = _find(sheet, "Работы")
    assert sheet.cell(row=row, column=4).value == round(calc["works_with_vat"], 2)
    assert sheet.cell(row=row, column=5).value == round(calc["works_without_vat"], 2)


def test_manual_row_has_both_costs(sheet, calc):
    row = _find(sheet, "Банковская гарантия")
    assert sheet.cell(row=row, column=4).value == round(calc["bank_guarantee_with_vat"], 2)
    assert sheet.cell(row=row, column=5).value == round(calc["bank_guarantee_without_vat"], 2)


def test_human_inputs_are_exported(sheet):
    """Проценты, количество людей и коэффициент — отдельной колонкой, числом."""
    assert sheet.cell(row=_find(sheet, "Коэффициент к ценам (×все цены)"), column=3).value == 1.07
    assert sheet.cell(row=_find(sheet, "Транспортные расходы"), column=3).value == 3
    assert sheet.cell(row=_find(sheet, "Накладные"), column=3).value == 4
    assert sheet.cell(row=_find(sheet, "Разнорабочие ежедневно"), column=3).value == 2
    assert sheet.cell(row=_find(sheet, "Непредвиденные расходы"), column=3).value == 2
    assert sheet.cell(row=_find(sheet, "Плановая прибыль (без НДС)"), column=3).value == 20
    assert sheet.cell(row=_find(sheet, "НДС от полной себестоимости"), column=3).value == 22


def test_custom_row_has_both_costs(sheet):
    row = _find(sheet, "Аренда лесов")
    assert sheet.cell(row=row, column=3).value == "1 шт"
    assert sheet.cell(row=row, column=4).value == 30500.0
    assert sheet.cell(row=row, column=5).value == 25000.0


def test_hidden_rows_are_not_exported(sheet):
    labels = {sheet.cell(row=r, column=2).value for r in range(1, sheet.max_row + 1)}
    assert "Уборка и вывоз мусора" not in labels
    assert "Рабочая документация (ППР)" not in labels


def test_rows_are_numbered_like_on_screen(sheet):
    """Нумерация — по видимым строкам подряд, ручные строки её продолжают."""
    assert sheet.cell(row=_find(sheet, "Работы"), column=1).value == 1
    # «Уборка и вывоз мусора» скрыта, поэтому её номер занимают «Накладные».
    assert sheet.cell(row=_find(sheet, "Транспортные расходы"), column=1).value == 3
    assert sheet.cell(row=_find(sheet, "Накладные"), column=1).value == 4
    assert sheet.cell(row=_find(sheet, "Аренда лесов"), column=1).value == 17


def test_contingency_has_both_costs(sheet, calc):
    row = _find(sheet, "Непредвиденные расходы")
    assert sheet.cell(row=row, column=4).value == round(calc["contingency_with_vat"], 2)
    assert sheet.cell(row=row, column=5).value == round(calc["contingency_without_vat"], 2)


def test_subtotal_has_both_costs(sheet, calc):
    row = _find(sheet, "ИТОГО себестоимость объекта")
    assert sheet.cell(row=row, column=4).value == round(calc["subtotal_with_vat"], 2)
    assert sheet.cell(row=row, column=5).value == round(calc["subtotal_without_vat"], 2)


def test_single_value_rows_are_merged(sheet, calc):
    """У строк с одной суммой обе денежные колонки объединены, как на экране."""
    merged = {str(rng) for rng in sheet.merged_cells.ranges}
    for label, value in (
        ("Плановая прибыль (без НДС)", calc["profit"]),
        ("Полная себестоимость с учётом прибыли и непредвиденных (без НДС)",
         calc["full_cost_without_vat"]),
        ("НДС от полной себестоимости", calc["vat"]),
        ("Др. налоги от полной себестоимости", calc["other_tax"]),
        ("ИТОГО по смете для Заказчика с учётом налогов", calc["total_for_customer"]),
    ):
        row = _find(sheet, label)
        assert f"D{row}:E{row}" in merged, label
        assert sheet.cell(row=row, column=4).value == round(value, 2), label


def test_total_for_customer_matches_screen(sheet):
    row = _find(sheet, "ИТОГО по смете для Заказчика с учётом налогов")
    assert sheet.cell(row=row, column=4).value == 222647.82


def test_exported_sections_show_both_taxes():
    """В файле у раздела две колонки налога — работы и материалы отдельно.

    Таблица разделов начинается с колонки G: в бланке слева стало пять колонок
    вместо четырёх. Порядок столбцов самого раздела прежний — второй налог
    добавлен в конец, чтобы формулы поверх выгрузок не съезжали.
    """
    summary = SimpleNamespace(
        sections=[{
            "card_id": "c", "card_name": "Раздел",
            "tax_pct_works": 0, "tax_pct_materials": 22,
            "rows": _TAX_SECTION_ROWS,
        }],
        overrides=_TAX_OVERRIDES,
    )
    ws = openpyxl.load_workbook(io.BytesIO(generate_summary_xlsx(summary)))["Сводная"]
    right = lambda row: [ws.cell(row=row, column=col).value for col in range(7, 14)]  # noqa: E731

    assert right(1) == [
        "Раздел", "Налог работ, %", "Работы (с/с)", "Работы с НДС",
        "Материалы (с/с)", "Материалы с НДС", "Налог матер., %",
    ]
    assert right(2) == ["Раздел", 0, 1000, 1220, 1000, 1000, 22]
    # ИТОГО: суммируются только деньги, налоговые колонки пустые.
    assert right(3) == ["ИТОГО", None, 1000, 1220, 1000, 1000, None]
