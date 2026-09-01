"""Цели оптимизации в скачиваемом файле сводной.

План: `plans/2026-09-01-celi-optimizacii.md`, Фаза 5.

Экран и файл считают по одной формуле (правило проекта №4), поэтому числа здесь
сверяются с `calc_summary` — тем самым расчётом, который считает бланк на
экране. Проценты отклонения в файл не выгружаются: и цель, и отклонение в нём
есть отдельными колонками, процент из них получается формулой.

Отдельно проверяется тишина: если целей не задано ни одной, файл выглядит ровно
так, как до появления функции.
"""
import io
from types import SimpleNamespace

import openpyxl
import pytest

from app.utils.summary_calc import calc_summary
from app.utils.xlsx_summary import generate_summary_xlsx

from tests.test_summary_calc import OVERRIDES as BASE_OVERRIDES, SECTIONS as BASE_SECTIONS

SECTIONS = [
    {**BASE_SECTIONS[0], "target_works": 15000, "target_materials": 20000},
    {**BASE_SECTIONS[1], "target_materials": 40000},
]
OVERRIDES = {**BASE_OVERRIDES, "target_total_for_customer": 200000}


def _sheet(sections, overrides):
    summary = SimpleNamespace(sections=sections, overrides=overrides)
    return openpyxl.load_workbook(io.BytesIO(generate_summary_xlsx(summary)))["Сводная"]


@pytest.fixture(scope="module")
def sheet():
    return _sheet(SECTIONS, OVERRIDES)


@pytest.fixture(scope="module")
def calc():
    return calc_summary(SECTIONS, OVERRIDES)


def _find(ws, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == label:
            return row
    raise AssertionError(f"строка «{label}» не найдена в листе")


def _right(ws, row: int) -> list:
    return [ws.cell(row=row, column=col).value for col in range(7, 18)]


def test_target_columns_are_added_after_the_old_ones(sheet):
    """Прежние колонки раздела остались на своих местах — формулы не съезжают."""
    assert _right(sheet, 1) == [
        "Раздел", "Налог работ, %", "Работы (с/с)", "Работы с НДС",
        "Материалы (с/с)", "Материалы с НДС", "Налог матер., %",
        "Цель работ", "Откл. работ", "Цель матер.", "Откл. матер.",
    ]


def test_section_target_and_deviation_match_screen(sheet, calc):
    ar = calc["section_totals"][0]
    row = _right(sheet, 2)
    assert row[0] == "АР"
    assert row[7] == round(ar["target_works"], 2)
    assert row[8] == round(ar["works_deviation"], 2)
    assert row[9] == round(ar["target_materials"], 2)
    assert row[10] == round(ar["materials_deviation"], 2)


def test_section_without_target_leaves_cells_empty(sheet):
    """У «ОВ» цели по работам нет — в файле пусто, а не ноль."""
    row = _right(sheet, 3)
    assert row[0] == "ОВ"
    assert row[7] is None
    assert row[8] is None
    assert row[9] == 40000


def test_totals_row_sums_only_sections_with_target(sheet, calc):
    row = _right(sheet, 4)
    assert row[0] == "ИТОГО"
    assert row[7] == round(calc["targets_total_works"], 2)
    assert row[8] == round(calc["targets_deviation_works"], 2)
    assert row[9] == round(calc["targets_total_materials"], 2)
    assert row[10] == round(calc["targets_deviation_materials"], 2)


def test_object_target_rows_match_screen(sheet, calc):
    target_row = _find(sheet, "Цель по объекту")
    deviation_row = _find(sheet, "Отклонение от цели по объекту")
    assert sheet.cell(row=target_row, column=4).value == 200000
    assert sheet.cell(row=deviation_row, column=4).value == round(calc["total_deviation"], 2)
    # Итог для заказчика стоит выше цели — сначала факт, потом сверка с целью.
    assert target_row > _find(sheet, "ИТОГО по смете для Заказчика с учётом налогов")


def test_file_without_targets_looks_as_before():
    """Целей нет — нет ни колонок, ни строк цели: файл как до этой функции."""
    ws = _sheet(BASE_SECTIONS, BASE_OVERRIDES)
    assert _right(ws, 1)[7:] == [None, None, None, None]
    for label in ("Цель по объекту", "Отклонение от цели по объекту"):
        with pytest.raises(AssertionError):
            _find(ws, label)


def test_object_target_alone_does_not_add_section_columns():
    """Цель только по объекту — колонки разделов не появляются."""
    ws = _sheet(BASE_SECTIONS, {**BASE_OVERRIDES, "target_total_for_customer": 200000})
    assert _right(ws, 1)[7:] == [None, None, None, None]
    assert _find(ws, "Цель по объекту")
