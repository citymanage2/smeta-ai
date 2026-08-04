"""Перечень и полнота по всем листам файла (план 2026-08-04, Фаза 2).

Заказчик присылает гранд-смету, разбитую по листам — по листу на раздел или
корпус. Раньше брался только первый лист, задача завершалась успешно, и
пропажа двух третей работ обнаруживалась уже на аукционе.
"""
import io

import openpyxl

from app.services.excel_service import data_sheet_titles, generate_list
from app.services.task_processor import _chunk_by_work_boundaries, _chunk_sheet, _tag_sheet
from app.utils.file_parser import chunk_rows, parse_xlsx_grand
from app.utils.xlsx_generic import parse_xlsx_to_generic_rows

_HEADER = ["№", "Наименование", "Ед. изм.", "Кол-во"]


def _grand(sheets: dict) -> bytes:
    """Гранд-смета из нескольких листов: {имя листа: [(наименование, ед, кол-во)]}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title)
        if rows is None:  # лист без таблицы — титульный
            ws.cell(row=1, column=1, value="ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ № 02-01-01")
            ws.cell(row=2, column=1, value="Составил: Иванов И.И.")
            continue
        for col, head in enumerate(_HEADER, 1):
            ws.cell(row=1, column=col, value=head)
        for index, (name, unit, qty) in enumerate(rows, start=2):
            ws.cell(row=index, column=1, value=index - 1)
            ws.cell(row=index, column=2, value=name)
            ws.cell(row=index, column=3, value=unit)
            ws.cell(row=index, column=4, value=qty)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseGrandAllSheets:
    def test_rows_of_every_sheet_are_taken(self):
        data = _grand({
            "Раздел 1": [("Демонтаж стен", "м2", 12.5)],
            "Раздел 2": [("Кладка кирпича", "м3", 4), ("Штукатурка", "м2", 30)],
        })
        rows = parse_xlsx_grand(data)

        assert [r["name"] for r in rows] == ["Демонтаж стен", "Кладка кирпича", "Штукатурка"]
        assert [r["sheet"] for r in rows] == ["Раздел 1", "Раздел 2", "Раздел 2"]

    def test_sheet_without_table_does_not_create_a_tab(self):
        data = _grand({
            "Титульный": None,
            "Раздел 1": [("Демонтаж стен", "м2", 12.5)],
            "Раздел 2": [("Кладка кирпича", "м3", 4)],
        })
        rows = parse_xlsx_grand(data)

        assert [r["sheet"] for r in rows] == ["Раздел 1", "Раздел 2"]
        assert all("Иванов" not in r["name"] for r in rows)

    def test_title_sheet_alone_with_one_table_leaves_document_without_tabs(self):
        # Лист данных остался один — вкладки не нужны, поведение прежнее.
        data = _grand({
            "Титульный": None,
            "Раздел 1": [("Демонтаж стен", "м2", 12.5)],
        })
        rows = parse_xlsx_grand(data)

        assert [r["name"] for r in rows] == ["Демонтаж стен"]
        assert all("sheet" not in r for r in rows)

    def test_single_sheet_file_is_not_marked(self):
        # Один лист — документ остаётся без вкладок и ведёт себя как раньше.
        data = _grand({"Смета": [("Демонтаж стен", "м2", 12.5)]})
        rows = parse_xlsx_grand(data)

        assert [r["name"] for r in rows] == ["Демонтаж стен"]
        assert all("sheet" not in r for r in rows)

    def test_fallback_still_works_when_no_sheet_has_a_header(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Демонтаж стен")
        ws.cell(row=2, column=1, value="Кладка кирпича")
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_xlsx_grand(buf.getvalue())

        assert [r["name"] for r in rows] == ["Демонтаж стен", "Кладка кирпича"]


class TestChunksStayWithinSheet:
    def test_chunk_rows_does_not_mix_sheets(self):
        rows = (
            [{"name": f"А{i}", "unit": "м2", "quantity": 1, "sheet": "Раздел 1"} for i in range(3)]
            + [{"name": f"Б{i}", "unit": "м2", "quantity": 1, "sheet": "Раздел 2"} for i in range(3)]
        )
        chunks = chunk_rows(rows, chunk_size=250)

        assert len(chunks) == 2
        assert {_chunk_sheet(c) for c in chunks} == {"Раздел 1", "Раздел 2"}
        for chunk in chunks:
            assert len({r["sheet"] for r in chunk}) == 1

    def test_work_boundary_chunking_does_not_mix_sheets(self):
        items = (
            [{"type": "Работа", "name": "А", "sheet": "Раздел 1"}]
            + [{"type": "Материал", "name": "а", "sheet": "Раздел 1"}]
            + [{"type": "Работа", "name": "Б", "sheet": "Раздел 2"}]
        )
        chunks = _chunk_by_work_boundaries(items, max_chunk_size=25)

        assert len(chunks) == 2
        for chunk in chunks:
            assert len({it["sheet"] for it in chunk}) == 1

    def test_items_returned_by_ai_get_the_chunk_sheet(self):
        chunk = [{"name": "А", "sheet": "Раздел 2"}]
        produced = [{"type": "Работа", "name": "Демонтаж"}, {"type": "Материал", "name": "Кирпич"}]
        _tag_sheet(produced, _chunk_sheet(chunk))

        assert [it["sheet"] for it in produced] == ["Раздел 2", "Раздел 2"]

    def test_existing_sheet_is_not_overwritten(self):
        produced = [{"name": "А", "sheet": "Раздел 1"}]
        _tag_sheet(produced, "Раздел 9")

        assert produced[0]["sheet"] == "Раздел 1"

    def test_single_sheet_file_leaves_items_unmarked(self):
        produced = [{"name": "А"}]
        _tag_sheet(produced, _chunk_sheet([{"name": "исходная"}]))

        assert "sheet" not in produced[0]


class TestGeneratedListFile:
    def test_sheet_per_source_sheet_plus_note(self):
        items = [
            {"type": "Работа", "name": "Демонтаж", "unit": "м2", "quantity": 1, "sheet": "Раздел 1"},
            {"type": "Работа", "name": "Кладка", "unit": "м3", "quantity": 2, "sheet": "Раздел 2"},
        ]
        wb = openpyxl.load_workbook(io.BytesIO(generate_list(items)))

        assert wb.sheetnames == ["Раздел 1", "Раздел 2", "Пояснительная записка"]

    def test_single_sheet_file_keeps_the_old_four_sheets(self):
        items = [{"type": "Работа", "name": "Демонтаж", "unit": "м2", "quantity": 1}]
        wb = openpyxl.load_workbook(io.BytesIO(generate_list(items)))

        assert wb.sheetnames == ["Перечень", "Работы", "Материалы", "Пояснительная записка"]

    def test_items_without_sheet_are_not_lost(self):
        # Позиция без листа не должна молча подмешаться к чужому разделу.
        items = [
            {"type": "Работа", "name": "Демонтаж", "unit": "м2", "quantity": 1, "sheet": "Раздел 1"},
            {"type": "Работа", "name": "Кладка", "unit": "м3", "quantity": 2, "sheet": "Раздел 2"},
            {"type": "Работа", "name": "Ничей", "unit": "шт", "quantity": 3},
        ]
        wb = openpyxl.load_workbook(io.BytesIO(generate_list(items)))

        assert "Прочее" in wb.sheetnames
        assert wb["Прочее"].cell(row=2, column=3).value == "Ничей"

    def test_document_gets_a_tab_per_source_sheet(self):
        items = [
            {"type": "Работа", "name": "Демонтаж", "unit": "м2", "quantity": 1, "sheet": "Раздел 1"},
            {"type": "Работа", "name": "Кладка", "unit": "м3", "quantity": 2, "sheet": "Раздел 2"},
        ]
        rows = parse_xlsx_to_generic_rows(
            generate_list(items), sheets=data_sheet_titles(items),
        )

        assert [r["sheet"] for r in rows] == ["Раздел 1", "Раздел 2"]
        assert [r["cells"]["Наименование"] for r in rows] == ["Демонтаж", "Кладка"]

    def test_long_sheet_name_is_trimmed_to_excel_limit(self):
        long_name = "Локальный сметный расчёт на общестроительные работы"
        items = [
            {"type": "Работа", "name": "Демонтаж", "unit": "м2", "quantity": 1, "sheet": long_name},
            {"type": "Работа", "name": "Кладка", "unit": "м3", "quantity": 2, "sheet": "Раздел 2"},
        ]
        titles = data_sheet_titles(items)
        wb = openpyxl.load_workbook(io.BytesIO(generate_list(items)))

        assert all(len(t) <= 31 for t in titles)
        # Имя листа в файле и имя вкладки в документе — одно и то же.
        assert wb.sheetnames[: len(titles)] == titles


class TestPipelineKeepsSheets:
    def test_list_to_completeness_to_estimate_keeps_the_sheet(self):
        """Лист едет по циклу вместе с позицией — без связей между задачами."""
        from app.utils.estimate_rows import items_to_rows, rows_to_items

        items = [
            {"type": "Работа", "name": "Демонтаж", "unit": "м2", "quantity": 1,
             "work_price": 100, "sheet": "Раздел 1"},
            {"type": "Материал", "name": "Кирпич", "unit": "шт", "quantity": 2,
             "material_price": 50, "sheet": "Раздел 2"},
        ]
        rows = items_to_rows(items)
        assert [r["sheet"] for r in rows] == ["Раздел 1", "Раздел 2"]

        back = rows_to_items(rows)
        assert [it["sheet"] for it in back] == ["Раздел 1", "Раздел 2"]
