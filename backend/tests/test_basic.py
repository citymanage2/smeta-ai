"""Basic smoke tests for Smeta AI backend."""
import pytest


def test_imports():
    """Verify core modules can be imported."""
    from app.config import settings
    assert settings.JWT_ALGORITHM == "HS256"
    assert settings.MAX_FILE_SIZE_MB == 20


def test_hash_password():
    """Test password hashing utility."""
    from app.utils.auth import hash_password, verify_password
    hashed = hash_password("test123")
    assert hashed != "test123"
    assert verify_password("test123", hashed)
    assert not verify_password("wrong", hashed)


def test_create_access_token():
    """Test JWT token creation and verification."""
    from app.utils.auth import create_access_token, verify_token
    token = create_access_token(1, "user", "ivan")
    payload = verify_token(token)
    assert payload["role"] == "user"
    assert payload["sub"] == "1"
    assert payload["username"] == "ivan"


def test_normalize_text():
    """Test text normalization for price lookup."""
    from app.services.price_service import normalize_text
    assert normalize_text("  Штукатурка стен  ") == "штукатурка стен"
    assert normalize_text("Ёлка") == "елка"


def test_parse_xlsx_empty():
    """Test xlsx parser with minimal data."""
    import io
    import openpyxl
    from app.utils.file_parser import parse_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    ws["A1"] = "Наименование"
    ws["B1"] = "Количество"
    ws["A2"] = "Бетон М200"
    ws["B2"] = 10

    buf = io.BytesIO()
    wb.save(buf)
    result = parse_xlsx(buf.getvalue())
    assert "Наименование" in result
    assert "Бетон М200" in result


def test_generate_list():
    """Test Excel list generation."""
    from app.services.excel_service import generate_list

    items = [
        {"type": "Работа", "name": "Штукатурка стен", "unit": "м2", "quantity": 100},
        {"type": "Материал", "name": "Цемент М400", "unit": "кг", "quantity": 500},
    ]
    result = generate_list(items)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_smeta():
    """Test Excel smeta generation."""
    from app.services.excel_service import generate_smeta

    items = [
        {
            "type": "Работа",
            "name": "Кладка кирпича",
            "unit": "м3",
            "quantity": 10,
            "work_price": 5000,
            "material_price": None,
        },
        {
            "type": "Материал",
            "name": "Кирпич рядовой",
            "unit": "шт",
            "quantity": 1000,
            "work_price": None,
            "material_price": 15,
        },
    ]
    result = generate_smeta(items)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_parse_xml():
    """Test XML parser."""
    from app.utils.file_parser import parse_xml

    xml_data = '<?xml version="1.0" encoding="UTF-8"?><smeta><item name="test" unit="m2" qty="100" price="500"/></smeta>'.encode("utf-8")
    result = parse_xml(xml_data)
    assert "smeta" in result or "item" in result
