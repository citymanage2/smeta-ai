"""
CRUD-роутер для Каталога расценок.

Позволяет просматривать, добавлять, редактировать и удалять отдельные позиции
из price_works и price_materials. При каждом изменении генерирует эмбеддинг
и перезагружает in-memory кэш price_service.
"""
import asyncio
import io
from datetime import datetime, timezone
from typing import Optional, Literal
from urllib.parse import quote

import openpyxl
import structlog
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.price import PriceWork, PriceMaterial
from app.services import price_service
from app.services.embedding_service import (
    normalize_name,
    generate_embedding,
    EmbeddingUnavailableError,
)
from app.utils.auth import get_current_user
from app.utils.permissions import get_manager_user
from app.utils.price_change import price_changed, prices_changed
from app.utils.price_min import compute_min_price

logger = structlog.get_logger()

router = APIRouter(prefix="/prices", tags=["catalog"])


# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------

class CatalogItem(BaseModel):
    id: int
    kind: Literal["work", "material"]
    name: str
    unit: Optional[str]
    price: Optional[float]          # min_price для работ, price для материалов
    prices: Optional[dict]          # только для работ: {contractor: price}
    updated_at: datetime

    model_config = {"from_attributes": True}


class CatalogResponse(BaseModel):
    items: list[CatalogItem]
    total: int


class CreateWorkBody(BaseModel):
    name: str
    unit: Optional[str] = None
    prices: Optional[dict] = None   # {contractor_name: price}


class CreateMaterialBody(BaseModel):
    name: str
    unit: Optional[str] = None
    price: Optional[float] = None


class UpdateWorkBody(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    prices: Optional[dict] = None


class UpdateMaterialBody(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _work_to_item(w: PriceWork) -> CatalogItem:
    return CatalogItem(
        id=w.id,
        kind="work",
        name=w.name,
        unit=w.unit,
        price=w.min_price,
        prices=w.prices,
        updated_at=w.updated_at,
    )


def _material_to_item(m: PriceMaterial) -> CatalogItem:
    return CatalogItem(
        id=m.id,
        kind="material",
        name=m.name,
        unit=m.unit,
        price=m.price,
        prices=None,
        updated_at=m.updated_at,
    )


async def _generate_embedding_safe(name: str) -> Optional[list]:
    """Generate embedding in thread; return None on any error."""
    try:
        return await asyncio.to_thread(generate_embedding, normalize_name(name), "search_document")
    except EmbeddingUnavailableError:
        return None


async def _reload_cache(db: AsyncSession) -> None:
    await price_service.load_cache(db)


def _escape_like(term: str) -> str:
    """Экранировать спецсимволы LIKE, чтобы поиск по '%' / '_' был буквальным."""
    return term.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def _apply_sort(items: list[CatalogItem], sort: str) -> list[CatalogItem]:
    if sort == "name_asc":
        return sorted(items, key=lambda x: x.name.lower())
    if sort == "name_desc":
        return sorted(items, key=lambda x: x.name.lower(), reverse=True)
    if sort == "price_asc":
        return sorted(items, key=lambda x: (x.price is None, x.price or 0))
    if sort == "price_desc":
        return sorted(items, key=lambda x: (x.price is None, -(x.price or 0)))
    if sort == "date_asc":
        return sorted(items, key=lambda x: x.updated_at)
    if sort == "date_desc":
        return sorted(items, key=lambda x: x.updated_at, reverse=True)
    return sorted(items, key=lambda x: x.name.lower())


# ---------------------------------------------------------------------------
# GET /prices/catalog
# ---------------------------------------------------------------------------

@router.get("/catalog", response_model=CatalogResponse)
async def list_catalog(
    tab: str = Query("all", pattern="^(all|works|materials)$"),
    search: Optional[str] = Query(None),
    sort: str = Query("name_asc", pattern="^(name_asc|name_desc|price_asc|price_desc|date_asc|date_desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    items: list[CatalogItem] = []
    search_term = search.strip() if search else None
    # Экранированный шаблон для ILIKE — фильтр выполняется в БД, а не в Python.
    pattern = f"%{_escape_like(search_term)}%" if search_term else None

    # Явный список колонок: тяжёлая колонка `embedding` (≈8-10 КБ JSON/строку)
    # НЕ грузится — она не нужна в списке. Это снимает мегабайты лишнего трафика
    # и памяти на каждый показ каталога.
    if tab in ("all", "works"):
        q = select(
            PriceWork.id, PriceWork.name, PriceWork.unit,
            PriceWork.min_price, PriceWork.prices, PriceWork.updated_at,
        )
        if pattern:
            q = q.where(PriceWork.name.ilike(pattern, escape="\\"))
        rows = (await db.execute(q)).all()
        items.extend(
            CatalogItem(
                id=r.id, kind="work", name=r.name, unit=r.unit,
                price=r.min_price, prices=r.prices, updated_at=r.updated_at,
            )
            for r in rows
        )

    if tab in ("all", "materials"):
        q = select(
            PriceMaterial.id, PriceMaterial.name, PriceMaterial.unit,
            PriceMaterial.price, PriceMaterial.updated_at,
        )
        if pattern:
            q = q.where(PriceMaterial.name.ilike(pattern, escape="\\"))
        rows = (await db.execute(q)).all()
        items.extend(
            CatalogItem(
                id=r.id, kind="material", name=r.name, unit=r.unit,
                price=r.price, prices=None, updated_at=r.updated_at,
            )
            for r in rows
        )

    # Сортировка/пагинация — в Python, но уже на «лёгких» строках без embedding.
    # Единый путь для all/works/materials исключает расхождение порядка между
    # вкладками (SQL lower() vs Python .lower() дают разный collation для кириллицы).
    items = _apply_sort(items, sort)
    total = len(items)
    start = (page - 1) * page_size
    page_items = items[start: start + page_size]

    return CatalogResponse(items=page_items, total=total)


# ---------------------------------------------------------------------------
# POST /prices/match-preview — «почему позиция не нашлась в прайсе»
# ---------------------------------------------------------------------------


class MatchPreviewBody(BaseModel):
    name: str
    kind: Literal["work", "material"]


class MatchCandidate(BaseModel):
    name: str
    score: float
    unit: Optional[str] = None
    price: Optional[float] = None
    would_match: bool


class MatchPreviewResponse(BaseModel):
    threshold: float
    catalog_size: int
    vectors_ready: bool
    matched: bool
    candidates: list[MatchCandidate]
    hint: str


@router.post("/match-preview", response_model=MatchPreviewResponse)
async def match_preview(
    body: MatchPreviewBody,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_manager_user),
):
    """Показать, что прайс отвечает на конкретное название и с какой похожестью.

    Нужен, чтобы отличать три разные причины «прайс: найдено 16 из 1220», которые
    снаружи выглядят одинаково:
    - в каталоге нет такой позиции (лучший кандидат — про другое);
    - позиция есть, но похожесть чуть ниже порога (порог надо снизить);
    - у прайса нет векторов, и поиск по смыслу вообще не работает.
    """
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Пустое название")

    if body.kind == "work":
        catalog_size = len(price_service._works_cache)
        vectors_ready = price_service._works_embeddings is not None
        raw = await price_service.find_top_n_works(name, n=5)
    else:
        catalog_size = len(price_service._materials_cache)
        vectors_ready = price_service._materials_embeddings is not None
        raw = await price_service.find_top_n_materials(name, n=5)

    threshold = price_service.SIMILARITY_THRESHOLD
    candidates = [
        MatchCandidate(
            name=c.get("text", ""),
            score=round(float(c.get("score") or 0.0), 3),
            unit=c.get("unit"),
            price=c.get("min_price") if body.kind == "work" else c.get("price"),
            would_match=float(c.get("score") or 0.0) >= threshold,
        )
        for c in raw
    ]
    matched = any(c.would_match for c in candidates)
    best = candidates[0].score if candidates else None

    if not vectors_ready:
        hint = (
            "У прайса нет векторов — поиск по смыслу отключён, работает только "
            "точное совпадение названия. Сгенерируйте векторы в админке."
        )
    elif catalog_size == 0:
        hint = "Каталог пуст — сопоставлять не с чем."
    elif matched:
        hint = "Позиция нашлась бы в прайсе."
    elif best is not None and best >= threshold - 0.08:
        hint = (
            f"Лучший кандидат чуть ниже порога ({best} против {threshold}). "
            "Если он по смыслу тот же — стоит снизить порог PRICE_SIMILARITY_THRESHOLD."
        )
    else:
        hint = (
            "Похожих позиций в прайсе нет — лучший кандидат про другое. "
            "Такую позицию цену придётся искать через ИИ, пока её не добавят в прайс."
        )

    return MatchPreviewResponse(
        threshold=threshold,
        catalog_size=catalog_size,
        vectors_ready=vectors_ready,
        matched=matched,
        candidates=candidates,
        hint=hint,
    )


# ---------------------------------------------------------------------------
# POST /prices/catalog/works
# ---------------------------------------------------------------------------

@router.post("/catalog/works", response_model=CatalogItem, status_code=status.HTTP_201_CREATED)
async def create_work(
    body: CreateWorkBody,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_manager_user),
):
    prices = body.prices or {}
    min_price = compute_min_price(prices)

    embedding = await _generate_embedding_safe(body.name)

    work = PriceWork(
        name=body.name,
        unit=body.unit,
        prices=prices,
        min_price=min_price,
        embedding=embedding,
        updated_at=datetime.now(timezone.utc),
    )
    db.add(work)
    await db.commit()
    await db.refresh(work)
    await _reload_cache(db)
    logger.info("Created work", id=work.id, name=work.name)
    return _work_to_item(work)


# ---------------------------------------------------------------------------
# POST /prices/catalog/materials
# ---------------------------------------------------------------------------

@router.post("/catalog/materials", response_model=CatalogItem, status_code=status.HTTP_201_CREATED)
async def create_material(
    body: CreateMaterialBody,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_manager_user),
):
    embedding = await _generate_embedding_safe(body.name)

    material = PriceMaterial(
        name=body.name,
        unit=body.unit,
        price=body.price,
        embedding=embedding,
        updated_at=datetime.now(timezone.utc),
    )
    db.add(material)
    await db.commit()
    await db.refresh(material)
    await _reload_cache(db)
    logger.info("Created material", id=material.id, name=material.name)
    return _material_to_item(material)


# ---------------------------------------------------------------------------
# PUT /prices/catalog/works/{id}
# ---------------------------------------------------------------------------

@router.put("/catalog/works/{work_id}", response_model=CatalogItem)
async def update_work(
    work_id: int,
    body: UpdateWorkBody,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_manager_user),
):
    result = await db.execute(select(PriceWork).where(PriceWork.id == work_id))
    work = result.scalar_one_or_none()
    if not work:
        raise HTTPException(status_code=404, detail="Работа не найдена")

    name_changed = body.name is not None and body.name != work.name

    # Снимаем «до» перед присваиванием: дату двигает только переоценка, а
    # переименование позиции — нет (см. `utils/price_change.py`).
    prev_prices = dict(work.prices or {})
    prev_unit = work.unit

    if body.name is not None:
        work.name = body.name
    if body.unit is not None:
        work.unit = body.unit
    if body.prices is not None:
        work.prices = body.prices
        work.min_price = compute_min_price(body.prices)

    if name_changed:
        work.embedding = await _generate_embedding_safe(work.name)

    if prices_changed(prev_prices, work.prices, prev_unit, work.unit):
        work.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(work)
    await _reload_cache(db)
    return _work_to_item(work)


# ---------------------------------------------------------------------------
# PUT /prices/catalog/materials/{id}
# ---------------------------------------------------------------------------

@router.put("/catalog/materials/{material_id}", response_model=CatalogItem)
async def update_material(
    material_id: int,
    body: UpdateMaterialBody,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_manager_user),
):
    result = await db.execute(select(PriceMaterial).where(PriceMaterial.id == material_id))
    material = result.scalar_one_or_none()
    if not material:
        raise HTTPException(status_code=404, detail="Материал не найден")

    name_changed = body.name is not None and body.name != material.name

    # Как и у работы: дата означает дату цены, а не дату правки записи.
    prev_price = material.price
    prev_unit = material.unit

    if body.name is not None:
        material.name = body.name
    if body.unit is not None:
        material.unit = body.unit
    if body.price is not None:
        material.price = body.price

    if name_changed:
        material.embedding = await _generate_embedding_safe(material.name)

    if price_changed(prev_price, material.price, prev_unit, material.unit):
        material.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(material)
    await _reload_cache(db)
    return _material_to_item(material)


# ---------------------------------------------------------------------------
# DELETE /prices/catalog/works/{id}
# ---------------------------------------------------------------------------

@router.delete("/catalog/works/{work_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_work(
    work_id: int,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_manager_user),
):
    result = await db.execute(select(PriceWork).where(PriceWork.id == work_id))
    work = result.scalar_one_or_none()
    if not work:
        raise HTTPException(status_code=404, detail="Работа не найдена")
    await db.delete(work)
    await db.commit()
    await _reload_cache(db)


# ---------------------------------------------------------------------------
# DELETE /prices/catalog/materials/{id}
# ---------------------------------------------------------------------------

@router.delete("/catalog/materials/{material_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_material(
    material_id: int,
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_manager_user),
):
    result = await db.execute(select(PriceMaterial).where(PriceMaterial.id == material_id))
    material = result.scalar_one_or_none()
    if not material:
        raise HTTPException(status_code=404, detail="Материал не найден")
    await db.delete(material)
    await db.commit()
    await _reload_cache(db)


# ---------------------------------------------------------------------------
# GET /prices/catalog/template
# ---------------------------------------------------------------------------

@router.get("/catalog/template")
async def download_template(
    type: str = Query(..., pattern="^(works|materials)$"),
    _user=Depends(get_current_user),
):
    wb = openpyxl.Workbook()
    ws = wb.active

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if type == "works":
        ws.title = "Работы"
        headers = ["Наименование", "Ед. изм.", "Подрядчик 1", "Подрядчик 2", "Подрядчик 3"]
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        # Example row
        ws.append(["Кладка кирпича", "м²", 1500, 1400, ""])
    else:
        ws.title = "Материалы"
        headers = ["Наименование", "Ед. изм.", "Цена"]
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18
        # Example row
        ws.append(["Кирпич керамический М100", "шт", 18.5])

    # Write headers in row 1 (push example to row 2)
    ws.insert_rows(1)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"template_{type}.xlsx"
    encoded = quote(filename)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ---------------------------------------------------------------------------
# GET /prices/catalog/export
# ---------------------------------------------------------------------------

@router.get("/catalog/export")
async def export_catalog(
    tab: str = Query("all", pattern="^(all|works|materials)$"),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    search_term = search.strip() if search else None
    pattern = f"%{_escape_like(search_term)}%" if search_term else None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог расценок"

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Тип", "Наименование", "Ед. изм.", "Цена (мин.)", "Подрядчики / детали", "Цена от"]
    col_widths = [12, 50, 12, 15, 40, 22]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    row = 2

    if tab in ("all", "works"):
        q = select(
            PriceWork.name, PriceWork.unit, PriceWork.min_price,
            PriceWork.prices, PriceWork.updated_at,
        )
        if pattern:
            q = q.where(PriceWork.name.ilike(pattern, escape="\\"))
        works = (await db.execute(q)).all()
        for w in works:
            contractors = "; ".join(f"{k}: {v}" for k, v in (w.prices or {}).items())
            ws.cell(row=row, column=1, value="Работа")
            ws.cell(row=row, column=2, value=w.name)
            ws.cell(row=row, column=3, value=w.unit or "")
            ws.cell(row=row, column=4, value=w.min_price)
            ws.cell(row=row, column=5, value=contractors)
            ws.cell(row=row, column=6, value=w.updated_at.strftime("%d.%m.%Y %H:%M") if w.updated_at else "")
            row += 1

    if tab in ("all", "materials"):
        q = select(
            PriceMaterial.name, PriceMaterial.unit,
            PriceMaterial.price, PriceMaterial.updated_at,
        )
        if pattern:
            q = q.where(PriceMaterial.name.ilike(pattern, escape="\\"))
        materials = (await db.execute(q)).all()
        for m in materials:
            ws.cell(row=row, column=1, value="Материал")
            ws.cell(row=row, column=2, value=m.name)
            ws.cell(row=row, column=3, value=m.unit or "")
            ws.cell(row=row, column=4, value=m.price)
            ws.cell(row=row, column=5, value="")
            ws.cell(row=row, column=6, value=m.updated_at.strftime("%d.%m.%Y %H:%M") if m.updated_at else "")
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "catalog_export.xlsx"
    encoded = quote(filename)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
