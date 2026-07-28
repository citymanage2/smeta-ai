import csv
import io
from datetime import datetime, timezone, timedelta
from typing import Optional
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, update, func
import openpyxl
import structlog

from app.config import settings
from app.database import get_db
from app.models.job import Job
from app.models.task import Task
from app.models.task_input_file import TaskInputFile
from app.services import storage_service
from app.models.price import PriceWork, PriceMaterial
from app.models.price_list import PriceList
from app.models.price_cache import PriceCacheWork, PriceCacheMaterial
from app.utils.auth import get_admin_user
from app.services import price_service
from app.services.embedding_service import (
    normalize_name,
    generate_embeddings_batch,
    EmbeddingUnavailableError,
)

logger = structlog.get_logger()

router = APIRouter(prefix="/admin", tags=["admin"])

PRICE_LIST_ALLOWED_EXTS = {".xlsx", ".xls", ".csv", ".txt", ".pdf", ".docx"}
MAX_PRICE_FILE_BYTES = 20 * 1024 * 1024  # 20 MB


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class TaskListItem(BaseModel):
    id: str
    user_role: str
    task_type: str
    status: str
    progress_message: Optional[str]
    error_message: Optional[str]
    created_at: datetime
    updated_at: datetime
    deleted_at: Optional[datetime]
    files_count: int


class TaskDetail(BaseModel):
    id: str
    user_role: str
    task_type: str
    status: str
    input_files: list
    user_prompt: Optional[str]
    chat_history: list
    progress_message: Optional[str]
    error_message: Optional[str]
    created_at: datetime
    updated_at: datetime


class TaskListResponse(BaseModel):
    items: list[TaskListItem]
    total: int


class PriceUploadResponse(BaseModel):
    works_loaded: int
    materials_loaded: int


class PriceListInfo(BaseModel):
    type: str
    filename: Optional[str]
    updated_at: Optional[str]
    embedding_status: str = "pending"


class PriceListsInfoResponse(BaseModel):
    works: PriceListInfo
    materials: PriceListInfo


class SinglePriceUploadResponse(BaseModel):
    loaded: int
    message: str
    added: Optional[int] = None
    updated: Optional[int] = None


class GenerateEmbeddingsResponse(BaseModel):
    status: str
    updated: int = 0
    error: Optional[str] = None


class PriceCacheItem(BaseModel):
    id: str
    name: str
    unit: Optional[str]
    price: float
    sources: Optional[str]
    updated_at: datetime
    expires_in_days: int


class PriceCacheListResponse(BaseModel):
    items: list[PriceCacheItem]
    total: int
    page: int
    page_size: int


class PriceCacheCreateRequest(BaseModel):
    name: str
    unit: Optional[str] = None
    price: float
    sources: Optional[str] = None


class PriceCachePatchRequest(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    sources: Optional[str] = None


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _parse_price_rows(rows: list, is_works: bool) -> list[dict]:
    """
    Parse price data from a list of rows (list of cell values).
    Auto-detects header row by looking for 'Наименование' or falling back to row 0.
    """
    if not rows:
        return []

    # Find header row
    header_row_idx = 0
    headers: list[str] = []
    for i, row in enumerate(rows):
        for cell in row:
            if cell and isinstance(cell, str) and "наименование" in cell.lower():
                header_row_idx = i
                headers = [str(h).strip() if h is not None else "" for h in row]
                break
        if headers:
            break

    if not headers:
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        header_row_idx = 0

    # Map column indices
    name_col: Optional[int] = None
    unit_col: Optional[int] = None
    price_col: Optional[int] = None
    contractor_cols: list[tuple[int, str]] = []

    for i, h in enumerate(headers):
        lower = h.lower()
        if "наименование" in lower or "название" in lower or "name" in lower:
            name_col = i
        elif ("ед" in lower and ("изм" in lower or "." in lower)) or lower in ("ед", "unit"):
            unit_col = i
        elif "цена" in lower or "стоимость" in lower or "price" in lower:
            price_col = i

    if name_col is None:
        return []

    # For works: all non-name/unit columns with headers are contractor price cols
    if is_works:
        for i, h in enumerate(headers):
            if i != name_col and i != unit_col and h:
                contractor_cols.append((i, h))

    result = []
    for row in rows[header_row_idx + 1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        name = row[name_col] if name_col < len(row) else None
        if not name:
            continue
        name = str(name).strip()
        if not name or name.lower() in ("итого", "всего", "total"):
            continue

        unit = ""
        if unit_col is not None and unit_col < len(row) and row[unit_col] is not None:
            unit = str(row[unit_col]).strip()

        if is_works:
            prices: dict[str, float] = {}
            for col_idx, contractor_name in contractor_cols:
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        prices[contractor_name] = float(row[col_idx])
                    except (ValueError, TypeError):
                        pass
            min_price = min(prices.values()) if prices else None
            result.append({"name": name, "unit": unit, "prices": prices, "min_price": min_price})
        else:
            price: Optional[float] = None
            col = price_col if price_col is not None else (
                # Fall back to last non-name/unit numeric column
                next(
                    (i for i in range(len(row) - 1, -1, -1)
                     if i != name_col and i != unit_col and row[i] is not None),
                    None,
                )
            )
            if col is not None and col < len(row) and row[col] is not None:
                try:
                    price = float(row[col])
                except (ValueError, TypeError):
                    pass
            result.append({"name": name, "unit": unit, "price": price})

    return result


def _parse_xlsx_single(data: bytes, is_works: bool) -> list[dict]:
    """Parse a single-type xlsx/xls file (all sheets tried)."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        items = _parse_price_rows(rows, is_works)
        if items:
            return items
    return []


def _parse_xlsx_combined(data: bytes) -> tuple[list[dict], list[dict]]:
    """
    Parse combined xlsx with 'Работы' and 'Материалы' sheets.
    Used by the legacy /prices/upload endpoint.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    works: list[dict] = []
    materials: list[dict] = []
    for name in wb.sheetnames:
        lower = name.lower()
        rows = list(wb[name].iter_rows(values_only=True))
        if "работ" in lower or "work" in lower:
            works = _parse_price_rows(rows, is_works=True)
        elif "материал" in lower or "material" in lower:
            materials = _parse_price_rows(rows, is_works=False)
    return works, materials


def _parse_csv_single(data: bytes, is_works: bool) -> list[dict]:
    """Parse a CSV or TXT price list file."""
    text = data.decode("utf-8-sig", errors="replace")
    # Detect delimiter
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [list(row) for row in reader]
    return _parse_price_rows(rows, is_works)


def _parse_file_for_type(data: bytes, filename: str, mime_type: str, is_works: bool) -> list[dict]:
    """Route file to the correct parser based on extension."""
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return _parse_xlsx_single(data, is_works)
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return _parse_csv_single(data, is_works)
    # PDF / DOCX — cannot parse automatically
    return []


def _ext_from_filename(filename: str) -> str:
    return "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


# ---------------------------------------------------------------------------
# Task endpoints
# ---------------------------------------------------------------------------

@router.get("/tasks", response_model=TaskListResponse)
async def list_tasks(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    date_from: Optional[datetime] = Query(None),
    date_to: Optional[datetime] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    task_type: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Get paginated list of all tasks with optional filters."""
    # Build WHERE conditions separately so they apply to both count and data queries
    conditions = [Task.deleted_at.is_(None)]
    if date_from:
        conditions.append(Task.created_at >= date_from)
    if date_to:
        conditions.append(Task.created_at <= date_to)
    if status_filter:
        conditions.append(Task.status == status_filter)
    if task_type:
        conditions.append(Task.task_type == task_type)

    # COUNT: select only id — avoids loading input_file_data / chat_history
    count_query = select(func.count(Task.id)).where(*conditions)
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # DATA: select only the columns needed for TaskListItem
    data_query = select(
        Task.id, Task.user_role, Task.task_type, Task.status,
        Task.progress_message, Task.error_message,
        Task.created_at, Task.updated_at, Task.deleted_at, Task.input_files,
    ).where(*conditions).order_by(Task.created_at.desc(), Task.id.desc()).offset((page - 1) * limit).limit(limit)

    result = await db.execute(data_query)
    rows = result.all()

    items = [
        TaskListItem(
            id=str(row.id),
            user_role=row.user_role,
            task_type=row.task_type,
            status=row.status,
            progress_message=row.progress_message,
            error_message=row.error_message,
            created_at=row.created_at,
            updated_at=row.updated_at,
            deleted_at=row.deleted_at,
            files_count=len(row.input_files or []),
        )
        for row in rows
    ]
    return TaskListResponse(items=items, total=total)


@router.get("/tasks/trash", response_model=TaskListResponse)
async def list_trash(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Список задач в корзине (deleted_at IS NOT NULL)."""
    conditions = [Task.deleted_at.is_not(None)]

    count_query = select(func.count(Task.id)).where(*conditions)
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    data_query = select(
        Task.id, Task.user_role, Task.task_type, Task.status,
        Task.progress_message, Task.error_message,
        Task.created_at, Task.updated_at, Task.deleted_at, Task.input_files,
    ).where(*conditions).order_by(Task.deleted_at.desc(), Task.id.desc()).offset((page - 1) * limit).limit(limit)

    result = await db.execute(data_query)
    rows = result.all()

    items = [
        TaskListItem(
            id=str(row.id),
            user_role=row.user_role,
            task_type=row.task_type,
            status=row.status,
            progress_message=row.progress_message,
            error_message=row.error_message,
            created_at=row.created_at,
            updated_at=row.updated_at,
            deleted_at=row.deleted_at,
            files_count=len(row.input_files or []),
        )
        for row in rows
    ]
    return TaskListResponse(items=items, total=total)


@router.get("/tasks/{task_id}", response_model=TaskDetail)
async def get_task(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    # Trashed tasks (deleted_at set) are not visible via the normal detail
    # endpoint — consistent with list_tasks and the projects soft-delete
    # contract; they remain accessible only through the /tasks/trash list.
    result = await db.execute(
        select(Task).where(Task.id == task_id, Task.deleted_at.is_(None))
    )
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Задача не найдена")
    return TaskDetail(
        id=str(task.id),
        user_role=task.user_role,
        task_type=task.task_type,
        status=task.status,
        input_files=task.input_files or [],
        user_prompt=task.user_prompt,
        chat_history=task.chat_history or [],
        progress_message=task.progress_message,
        error_message=task.error_message,
        created_at=task.created_at,
        updated_at=task.updated_at,
    )


@router.get("/tasks/{task_id}/download-input/{file_index}")
async def download_input_file(
    task_id: str,
    file_index: int,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Download an original input file for a task by index."""
    result = await db.execute(select(Task).where(Task.id == task_id))
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Задача не найдена")

    # Файлы задачи хранятся в task_input_files (байты в S3 по storage_key).
    tif_res = await db.execute(
        select(TaskInputFile).where(
            TaskInputFile.task_id == task_id,
            TaskInputFile.file_index == file_index,
        )
    )
    tif = tif_res.scalar_one_or_none()
    if tif is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл не найден")

    data = await storage_service.load_bytes(tif.storage_key)
    return Response(
        content=data,
        media_type=tif.mime_type or "application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(tif.file_name)}"},
    )


@router.delete("/tasks/trash", status_code=status.HTTP_204_NO_CONTENT)
async def clear_trash(
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Очистить корзину — окончательно удалить все задачи с deleted_at IS NOT NULL."""
    await db.execute(delete(Task).where(Task.deleted_at.is_not(None)))
    await db.commit()
    logger.info("Trash cleared by admin")


@router.delete("/tasks/{task_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_task(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Мягкое удаление — перемещает задачу в корзину."""
    result = await db.execute(select(Task).where(Task.id == task_id))
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Задача не найдена")
    task.deleted_at = datetime.now(timezone.utc)
    await db.commit()
    logger.info("Task moved to trash by admin", task_id=task_id)


@router.post("/tasks/{task_id}/restore", status_code=status.HTTP_204_NO_CONTENT)
async def restore_task(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Восстановить задачу из корзины."""
    result = await db.execute(select(Task).where(Task.id == task_id))
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Задача не найдена")
    task.deleted_at = None
    await db.commit()
    logger.info("Task restored from trash by admin", task_id=task_id)


@router.delete("/tasks/{task_id}/permanent", status_code=status.HTTP_204_NO_CONTENT)
async def permanent_delete_task(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Окончательное удаление задачи и всех связанных данных."""
    result = await db.execute(select(Task).where(Task.id == task_id))
    task = result.scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Задача не найдена")
    await db.delete(task)
    await db.commit()
    logger.info("Task permanently deleted by admin", task_id=task_id)


# ---------------------------------------------------------------------------
# Price list info endpoint
# ---------------------------------------------------------------------------

@router.get("/price-lists/info", response_model=PriceListsInfoResponse)
async def get_price_lists_info(
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Return metadata for currently stored works and materials price lists."""
    async def _get_info(pl_type: str) -> PriceListInfo:
        res = await db.execute(
            select(PriceList)
            .where(PriceList.type == pl_type)
            .order_by(PriceList.updated_at.desc())
            .limit(1)
        )
        row = res.scalar_one_or_none()
        if row:
            return PriceListInfo(
                type=pl_type,
                filename=row.filename,
                updated_at=row.updated_at.isoformat(),
                embedding_status=row.embedding_status,
            )
        return PriceListInfo(type=pl_type, filename=None, updated_at=None, embedding_status="pending")

    return PriceListsInfoResponse(
        works=await _get_info("works"),
        materials=await _get_info("materials"),
    )


# ---------------------------------------------------------------------------
# Separate upload endpoints
# ---------------------------------------------------------------------------

_INSERT_BATCH_SIZE = 200


async def _handle_price_upload(
    file: UploadFile,
    pl_type: str,
    is_works: bool,
    db: AsyncSession,
) -> SinglePriceUploadResponse:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Файл не выбран")

    ext = _ext_from_filename(file.filename)
    if ext not in PRICE_LIST_ALLOWED_EXTS:
        raise HTTPException(
            status_code=415,
            detail=f"Недопустимый формат. Разрешены: {', '.join(PRICE_LIST_ALLOWED_EXTS)}",
        )

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой")
    if len(data) > MAX_PRICE_FILE_BYTES:
        raise HTTPException(status_code=413, detail="Файл превышает 20 МБ")

    mime_type = file.content_type or "application/octet-stream"
    now = datetime.now(timezone.utc)

    # Try to parse data into price rows
    try:
        items = _parse_file_for_type(data, file.filename, mime_type, is_works)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Ошибка разбора файла: {e}")

    # Generate embeddings before touching the DB (one batch API call)
    embeddings: list = []
    embedding_status = "pending"
    if items:
        try:
            normalized = [normalize_name(item["name"]) for item in items]
            embeddings = generate_embeddings_batch(normalized, input_type="search_document")
            embedding_status = "ready"
            logger.info("Embeddings generated", type=pl_type, count=len(items))
        except EmbeddingUnavailableError as e:
            embeddings = [None] * len(items)
            embedding_status = "failed"
            logger.warning("Embeddings generation failed, price list saved without vectors",
                           type=pl_type, error=str(e))

    # Load existing price rows into memory by normalised name for merge
    if is_works:
        res = await db.execute(select(PriceWork.id, PriceWork.name, PriceWork.unit, PriceWork.prices))
        existing_by_norm: dict = {
            normalize_name(row.name): {"id": row.id, "unit": row.unit, "prices": row.prices or {}}
            for row in res
        }
    else:
        res = await db.execute(select(PriceMaterial.id, PriceMaterial.name, PriceMaterial.unit))
        existing_by_norm = {
            normalize_name(row.name): {"id": row.id, "unit": row.unit}
            for row in res
        }

    # Upsert PriceList metadata (keep one record per type, update on re-upload)
    pl_res = await db.execute(
        select(PriceList).where(PriceList.type == pl_type).limit(1)
    )
    pl_obj = pl_res.scalar_one_or_none()
    if pl_obj:
        pl_obj.filename = file.filename
        pl_obj.mime_type = mime_type
        pl_obj.content = data
        pl_obj.updated_at = now
        pl_obj.embedding_status = embedding_status
    else:
        db.add(PriceList(
            type=pl_type,
            filename=file.filename,
            mime_type=mime_type,
            content=data,
            updated_at=now,
            embedding_status=embedding_status,
        ))
    await db.commit()

    # Separate new items into updates (existing name match) and inserts (new)
    model_cls = PriceWork if is_works else PriceMaterial
    update_ops: list = []   # (id, values_dict)
    insert_objs: list = []  # new ORM objects

    for item, emb in zip(items, embeddings):
        key = normalize_name(item["name"])
        if key in existing_by_norm:
            existing = existing_by_norm[key]
            if is_works:
                merged = {**existing["prices"], **item.get("prices", {})}
                positive = [v for v in merged.values() if v and v > 0]
                values: dict = {
                    "prices": merged,
                    "min_price": min(positive) if positive else None,
                    "unit": item.get("unit") or existing["unit"],
                    "embedding": emb,
                    "updated_at": now,
                }
            else:
                values = {
                    "price": item.get("price"),
                    "unit": item.get("unit") or existing["unit"],
                    "embedding": emb,
                    "updated_at": now,
                }
            update_ops.append((existing["id"], values))
        else:
            if is_works:
                obj = PriceWork(
                    name=item["name"],
                    unit=item.get("unit", ""),
                    prices=item.get("prices", {}),
                    min_price=item.get("min_price"),
                    embedding=emb,
                    updated_at=now,
                )
            else:
                obj = PriceMaterial(
                    name=item["name"],
                    unit=item.get("unit", ""),
                    price=item.get("price"),
                    embedding=emb,
                    updated_at=now,
                )
            insert_objs.append(obj)

    added = len(insert_objs)
    updated_count = len(update_ops)

    # Process updates and inserts in small batches to avoid dropping the DB
    # connection on a single huge transaction (embeddings add ~13 KB per row)
    all_ops = [("update", op) for op in update_ops] + [("insert", obj) for obj in insert_objs]
    for batch_start in range(0, len(all_ops), _INSERT_BATCH_SIZE):
        batch = all_ops[batch_start: batch_start + _INSERT_BATCH_SIZE]
        for kind, payload in batch:
            if kind == "update":
                row_id, values = payload
                await db.execute(
                    update(model_cls).where(model_cls.id == row_id).values(**values)
                )
            else:
                db.add(payload)
        await db.commit()

    await price_service.load_cache(db)

    loaded = len(items)
    if loaded == 0 and ext in (".pdf", ".docx"):
        msg = "Файл сохранён. Автоматический разбор PDF/DOCX не поддерживается — данные в базе не обновлены."
    elif loaded == 0:
        msg = "Файл сохранён, но позиции не найдены. Проверьте структуру файла (нужны колонки Наименование, Цена)."
    else:
        kind = "работ" if is_works else "материалов"
        msg = f"Прайс {kind} обновлён: добавлено {added}, обновлено {updated_count} позиций."

    logger.info("Price list merged", type=pl_type, filename=file.filename, added=added, updated=updated_count)
    return SinglePriceUploadResponse(loaded=loaded, message=msg, added=added, updated=updated_count)


@router.post("/price-lists/works", response_model=SinglePriceUploadResponse)
async def upload_works_price_list(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Upload price list for works. Merges with existing works prices (new names added, matching names updated)."""
    return await _handle_price_upload(file, "works", is_works=True, db=db)


@router.post("/price-lists/materials", response_model=SinglePriceUploadResponse)
async def upload_materials_price_list(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Upload price list for materials. Merges with existing materials prices (new names added, matching names updated)."""
    return await _handle_price_upload(file, "materials", is_works=False, db=db)


@router.post("/price-lists/{pl_type}/generate-embeddings", response_model=GenerateEmbeddingsResponse)
async def generate_price_embeddings(
    pl_type: str,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """
    Generate (or regenerate) embedding vectors for all rows of a price list.
    Returns HTTP 200 in both success and failure cases — check 'status' field.
    """
    if pl_type not in ("works", "materials"):
        raise HTTPException(status_code=400, detail="Тип должен быть 'works' или 'materials'")

    # Find the price list record
    res = await db.execute(
        select(PriceList)
        .where(PriceList.type == pl_type)
        .order_by(PriceList.updated_at.desc())
        .limit(1)
    )
    price_list = res.scalar_one_or_none()
    if not price_list:
        raise HTTPException(status_code=404, detail="Прайс-лист не найден")

    # Load all rows
    if pl_type == "works":
        rows_res = await db.execute(select(PriceWork))
    else:
        rows_res = await db.execute(select(PriceMaterial))
    rows = rows_res.scalars().all()

    if not rows:
        price_list.embedding_status = "ready"
        await db.commit()
        return GenerateEmbeddingsResponse(status="ready", updated=0)

    try:
        normalized = [normalize_name(row.name) for row in rows]
        embeddings = generate_embeddings_batch(normalized, input_type="search_document")
        for row, emb in zip(rows, embeddings):
            row.embedding = emb
        price_list.embedding_status = "ready"
        await db.commit()
        await price_service.load_cache(db)
        logger.info("Embeddings regenerated via endpoint", type=pl_type, count=len(rows))
        return GenerateEmbeddingsResponse(status="ready", updated=len(rows))
    except EmbeddingUnavailableError as e:
        price_list.embedding_status = "failed"
        await db.commit()
        logger.warning("Embeddings generation failed via endpoint", type=pl_type, error=str(e))
        return GenerateEmbeddingsResponse(status="failed", error=str(e))


# ---------------------------------------------------------------------------
# Legacy combined upload (kept for backward compatibility)
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Price cache endpoints
# ---------------------------------------------------------------------------

def _expires_in_days(updated_at: datetime) -> int:
    now = datetime.now(timezone.utc)
    updated = updated_at if updated_at.tzinfo else updated_at.replace(tzinfo=timezone.utc)
    return 30 - (now - updated).days


def _cache_item(row) -> PriceCacheItem:
    return PriceCacheItem(
        id=str(row.id),
        name=row.name,
        unit=row.unit,
        price=float(row.price),
        sources=row.sources,
        updated_at=row.updated_at,
        expires_in_days=_expires_in_days(row.updated_at),
    )


async def _list_cache(
    model_cls,
    db: AsyncSession,
    page: int,
    page_size: int,
    search: Optional[str],
) -> PriceCacheListResponse:
    conditions = []
    if search:
        conditions.append(model_cls.name.ilike(f"%{search}%"))

    count_q = select(func.count(model_cls.id))
    data_q = select(model_cls)
    if conditions:
        count_q = count_q.where(*conditions)
        data_q = data_q.where(*conditions)

    total = (await db.execute(count_q)).scalar() or 0
    rows = (
        await db.execute(
            data_q.order_by(model_cls.updated_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    ).scalars().all()

    return PriceCacheListResponse(
        items=[_cache_item(r) for r in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/price-cache/works", response_model=PriceCacheListResponse)
async def list_cache_works(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    return await _list_cache(PriceCacheWork, db, page, page_size, search)


@router.get("/price-cache/materials", response_model=PriceCacheListResponse)
async def list_cache_materials(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    return await _list_cache(PriceCacheMaterial, db, page, page_size, search)


@router.post("/price-cache/works", response_model=PriceCacheItem, status_code=status.HTTP_201_CREATED)
async def create_cache_work(
    body: PriceCacheCreateRequest,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    now = datetime.now(timezone.utc)
    obj = PriceCacheWork(
        name=body.name, name_norm=price_service.normalize_text(body.name),
        unit=body.unit, price=body.price, sources=body.sources, updated_at=now,
    )
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    await price_service.load_cache(db)
    logger.info("Cache work created by admin", name=body.name)
    return _cache_item(obj)


@router.post("/price-cache/materials", response_model=PriceCacheItem, status_code=status.HTTP_201_CREATED)
async def create_cache_material(
    body: PriceCacheCreateRequest,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    now = datetime.now(timezone.utc)
    obj = PriceCacheMaterial(
        name=body.name, name_norm=price_service.normalize_text(body.name),
        unit=body.unit, price=body.price, sources=body.sources, updated_at=now,
    )
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    await price_service.load_cache(db)
    logger.info("Cache material created by admin", name=body.name)
    return _cache_item(obj)


@router.patch("/price-cache/works/{item_id}", response_model=PriceCacheItem)
async def update_cache_work(
    item_id: str,
    body: PriceCachePatchRequest,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    res = await db.execute(select(PriceCacheWork).where(PriceCacheWork.id == item_id))
    obj = res.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Запись не найдена")

    now = datetime.now(timezone.utc)
    if body.name is not None:
        obj.name = body.name
        obj.name_norm = price_service.normalize_text(body.name)
    if body.unit is not None:
        obj.unit = body.unit
    if body.price is not None:
        obj.price = body.price
    if body.sources is not None:
        obj.sources = body.sources
    obj.updated_at = now

    await db.commit()
    await db.refresh(obj)
    await price_service.load_cache(db)
    logger.info("Cache work updated by admin", id=item_id)
    return _cache_item(obj)


@router.patch("/price-cache/materials/{item_id}", response_model=PriceCacheItem)
async def update_cache_material(
    item_id: str,
    body: PriceCachePatchRequest,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    res = await db.execute(select(PriceCacheMaterial).where(PriceCacheMaterial.id == item_id))
    obj = res.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Запись не найдена")

    now = datetime.now(timezone.utc)
    if body.name is not None:
        obj.name = body.name
        obj.name_norm = price_service.normalize_text(body.name)
    if body.unit is not None:
        obj.unit = body.unit
    if body.price is not None:
        obj.price = body.price
    if body.sources is not None:
        obj.sources = body.sources
    obj.updated_at = now

    await db.commit()
    await db.refresh(obj)
    await price_service.load_cache(db)
    logger.info("Cache material updated by admin", id=item_id)
    return _cache_item(obj)


@router.delete("/price-cache/works/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_cache_work(
    item_id: str,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    res = await db.execute(select(PriceCacheWork).where(PriceCacheWork.id == item_id))
    obj = res.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Запись не найдена")
    await db.delete(obj)
    await db.commit()
    await price_service.load_cache(db)
    logger.info("Cache work deleted by admin", id=item_id)


@router.delete("/price-cache/materials/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_cache_material(
    item_id: str,
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    res = await db.execute(select(PriceCacheMaterial).where(PriceCacheMaterial.id == item_id))
    obj = res.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Запись не найдена")
    await db.delete(obj)
    await db.commit()
    await price_service.load_cache(db)
    logger.info("Cache material deleted by admin", id=item_id)


# ---------------------------------------------------------------------------
# Legacy combined upload (kept for backward compatibility)
# ---------------------------------------------------------------------------

@router.post("/prices/upload", response_model=PriceUploadResponse)
async def upload_prices(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """
    Upload combined price list Excel file with 'Работы' and 'Материалы' sheets.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Файл не выбран")

    allowed_exts = {".xlsx", ".xls"}
    if _ext_from_filename(file.filename) not in allowed_exts:
        raise HTTPException(status_code=415, detail="Принимаются только файлы Excel (.xlsx, .xls)")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Файл пустой")

    try:
        works_data, materials_data = _parse_xlsx_combined(data)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Ошибка разбора файла: {e}")

    now = datetime.now(timezone.utc)
    await db.execute(delete(PriceWork))
    await db.execute(delete(PriceMaterial))
    await db.commit()

    for item in works_data:
        db.add(PriceWork(
            name=item["name"], unit=item.get("unit", ""),
            prices=item.get("prices", {}), min_price=item.get("min_price"), updated_at=now,
        ))
    for item in materials_data:
        db.add(PriceMaterial(
            name=item["name"], unit=item.get("unit", ""),
            price=item.get("price"), updated_at=now,
        ))

    await db.commit()
    await price_service.load_cache(db)

    logger.info("Combined prices uploaded", works=len(works_data), materials=len(materials_data))
    return PriceUploadResponse(works_loaded=len(works_data), materials_loaded=len(materials_data))


# ---------------------------------------------------------------------------
# Queue health — диагностика durable-очереди jobs
# ---------------------------------------------------------------------------

# queued-job не должна ждать дольше этого при живом свободном worker'е; дольше —
# признак, что очередь не разбирается (verdict "stalled").
QUEUE_STALL_THRESHOLD_S = 120


class QueueCounts(BaseModel):
    queued: int = 0
    running: int = 0
    done: int = 0
    failed: int = 0


class QueuedInfo(BaseModel):
    count: int
    oldest_age_s: Optional[float]


class RunningInfo(BaseModel):
    count: int
    oldest_claimed_age_s: Optional[float]
    stale_count: int


class QueueHealthResponse(BaseModel):
    checked_at: datetime
    counts: QueueCounts
    queued: QueuedInfo
    running: RunningInfo
    visibility_timeout_s: int
    verdict: str  # idle | ok | busy | stalled
    hint: str


def _age_s(dt, now: datetime) -> Optional[float]:
    """Возраст в секундах; naive/строковые даты (SQLite) трактуем как UTC. None → None."""
    if dt is None:
        return None
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return max(0.0, (now - dt).total_seconds())


@router.get("/queue-health", response_model=QueueHealthResponse)
async def queue_health(
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Состояние durable-очереди `jobs` для быстрой диагностики «задача виснет в
    Ожидании» без SQL по проду. verdict:
    - idle    — очередь пуста, ничего не считается;
    - ok      — очередь движется штатно (свежие queued либо идут running);
    - busy    — есть бэклог, но worker жив (свежие running) — легитимно;
    - stalled — есть старые queued и НЕТ живых running → worker не разбирает очередь.
    """
    now = datetime.now(timezone.utc)
    vis = settings.JOB_VISIBILITY_TIMEOUT_S

    # 1) Счётчики по статусам + возраст самой старой в каждом статусе.
    rows = (
        await db.execute(
            select(Job.status, func.count(Job.id), func.min(Job.created_at)).group_by(Job.status)
        )
    ).all()
    counts = QueueCounts()
    oldest_created: dict[str, object] = {}
    for status_val, cnt, oldest in rows:
        if hasattr(counts, status_val):
            setattr(counts, status_val, cnt)
        oldest_created[status_val] = oldest

    # 2) running: самый старый claimed_at + число протухших (кандидаты на reclaim).
    stale_cutoff = now - timedelta(seconds=vis)
    running_oldest_claim = (
        await db.execute(select(func.min(Job.claimed_at)).where(Job.status == "running"))
    ).scalar_one_or_none()
    stale_count = (
        await db.execute(
            select(func.count(Job.id)).where(Job.status == "running", Job.claimed_at < stale_cutoff)
        )
    ).scalar_one() or 0

    oldest_queued_age = _age_s(oldest_created.get("queued"), now)
    running_info = RunningInfo(
        count=counts.running,
        oldest_claimed_age_s=_age_s(running_oldest_claim, now),
        stale_count=stale_count,
    )
    live_running = counts.running - stale_count

    if counts.queued == 0 and counts.running == 0:
        verdict, hint = "idle", "Очередь пуста — задач в работе нет."
    elif (
        counts.queued > 0
        and oldest_queued_age is not None
        and oldest_queued_age > QUEUE_STALL_THRESHOLD_S
        and live_running == 0
    ):
        verdict = "stalled"
        hint = (
            "Очередь не разбирается: есть queued-job, но нет живых running. "
            "Проверьте контейнер worker и его логи (подключение к БД / SSL)."
        )
    elif counts.queued > 0 and live_running > 0:
        verdict, hint = "busy", "Есть бэклог, но worker жив — задачи считаются."
    else:
        verdict, hint = "ok", "Очередь движется штатно."

    return QueueHealthResponse(
        checked_at=now,
        counts=counts,
        queued=QueuedInfo(count=counts.queued, oldest_age_s=oldest_queued_age),
        running=running_info,
        visibility_timeout_s=vis,
        verdict=verdict,
        hint=hint,
    )


# ---------------------------------------------------------------------------
# Диагностика: доступен ли AI-API прямо сейчас (баланс/ключ/прокси)
# ---------------------------------------------------------------------------


class ApiHealthResponse(BaseModel):
    checked_at: datetime
    ok: bool
    status_code: Optional[int] = None
    error: Optional[str] = None
    error_code: Optional[str] = None
    is_balance_error: bool = False
    base_url: str
    via_proxy: bool
    api_key_set: bool
    proxy_secret_set: bool
    model: str
    paused_tasks: int
    verdict: str  # ok | no_balance | auth | unavailable | misconfigured
    hint: str


@router.get("/api-health", response_model=ApiHealthResponse)
async def api_health(
    db: AsyncSession = Depends(get_db),
    admin: dict = Depends(get_admin_user),
):
    """Пробный минимальный вызов AI-API — «пополнение дошло или нет».

    У Anthropic нет API проверки баланса, поэтому проверяем запросом (max_tokens=1,
    стоимость пренебрежимо мала). Отвечает на вопрос, ради которого иначе надо
    ждать 10 минут до следующего тика resume_poller: деньги есть или нет, и если
    нет — что именно ответил API. Важно: на сервере запросы могут идти через
    агрегатор (`base_url`), и «баланс исчерпан» тогда относится к его счёту, а не
    к Anthropic.
    """
    from app.services.claude_service import api_ping

    ping = await api_ping()
    paused_count = (
        await db.execute(select(func.count(Task.id)).where(Task.status == "paused"))
    ).scalar_one() or 0

    if ping["ok"]:
        verdict = "ok"
        hint = (
            f"API отвечает. Задач на паузе: {paused_count} — "
            "они возобновятся автоматически в течение 10 минут (resume_poller) "
            "или сразу по кнопке «Продолжить сейчас»."
        )
    elif ping["is_balance_error"]:
        verdict = "no_balance"
        where = "агрегатора (base_url задан)" if ping["via_proxy"] else "Anthropic"
        hint = (
            f"Баланс {where} всё ещё исчерпан — пополнение не дошло или пополнен "
            f"не тот счёт. Ответ API: {ping.get('status_code')} {ping.get('error') or ''}"
        )
    elif not ping["api_key_set"]:
        verdict = "misconfigured"
        hint = "ANTHROPIC_API_KEY не задан в env приложения."
    elif ping["status_code"] in (401, 403):
        verdict = "auth"
        hint = (
            "API отклонил ключ (401/403): неверный/отозванный ключ либо не принят "
            f"X-Proxy-Secret. Ответ: {ping.get('error') or ''}"
        )
    else:
        verdict = "unavailable"
        hint = f"API недоступен: {ping.get('status_code')} {ping.get('error') or ''}"

    return ApiHealthResponse(
        checked_at=datetime.now(timezone.utc),
        paused_tasks=paused_count,
        verdict=verdict,
        hint=hint,
        **{k: ping.get(k) for k in (
            "ok", "status_code", "error", "error_code", "is_balance_error",
            "base_url", "via_proxy", "api_key_set", "proxy_secret_set", "model",
        )},
    )
