import io
from decimal import Decimal
from typing import Optional

import openpyxl


def extract_total_cost(file_bytes: bytes) -> Optional[Decimal]:
    """
    Ищет строку где первая ячейка содержит 'итого' или 'всего'
    (регистронезависимо, после strip).
    Возвращает последнее числовое значение в этой строке.
    Если найдено несколько таких строк — берёт последнюю.
    Если ничего не найдено — возвращает None.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception:
        return None

    last_cost: Optional[Decimal] = None

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            if not row:
                continue
            first_val = row[0].value
            if first_val is None:
                continue
            normalized = str(first_val).strip().lower()
            if "итого" not in normalized and "всего" not in normalized:
                continue
            # Find last numeric value in this row
            for cell in reversed(row):
                val = cell.value
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    last_cost = Decimal(str(val))
                    break

    return last_cost


# Сводки, которые наш собственный генератор перечня кладёт рядом с данными.
# Их позиции — те же самые, и второй раз их брать нельзя.
_LEGACY_SUMMARY_SHEETS = ("работы", "материалы")
_NOTE_SHEET = "пояснительная записка"


def _is_legacy_list_workbook(wb) -> bool:
    """Файл нашего генератора до появления вкладок: «Перечень» + две сводки.

    В нём лист «Перечень» — единственный с данными, а «Работы» и «Материалы»
    повторяют его позиции. Признак строгий (нужны все три листа), чтобы файл
    заказчика с разделом, честно названным «Работы», не потерял этот раздел.
    """
    titles = {ws.title.strip().lower() for ws in wb.worksheets}
    has_list = any("перечень" in title for title in titles)
    return has_list and all(name in titles for name in _LEGACY_SUMMARY_SHEETS)


def _parse_list_worksheet(ws) -> list[dict]:
    """Позиции одного листа перечня. Пустой лист даёт пустой список."""
    items: list[dict] = []
    # Detect column positions from header row (row 1)
    header_map: dict[str, int] = {}  # canonical_key -> 0-based col index
    header_aliases = {
        "тип": "type",
        "наименование": "name",
        "ед": "unit",
        "ед.": "unit",
        "кол": "quantity",
        "кол-во": "quantity",
        "примечание": "notes",
        # Перечень могли скачать и загрузить обратно задачей сметы. Номер
        # позиции исходной сметы нужно прочитать, иначе он потеряется ровно
        # на том пути, где человек работает файлами, а не кнопкой «Далее».
        "№ в исходной смете": "source_no",
    }

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return []

    for col_idx, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        normalized = str(cell_val).strip().lower().rstrip(".")
        for alias, key in header_aliases.items():
            if normalized.startswith(alias):
                if key not in header_map:
                    header_map[key] = col_idx
                break

    # Fallback: assume fixed column layout (№, Тип, Наименование, Ед. изм., Кол-во, Примечание)
    if "name" not in header_map:
        header_map = {"type": 1, "name": 2, "unit": 3, "quantity": 4, "notes": 5}

    for row in rows_iter:
        if not row or all(v is None for v in row):
            continue

        name_val = row[header_map["name"]] if header_map.get("name") is not None and header_map["name"] < len(row) else None
        if name_val is None or str(name_val).strip() == "":
            continue

        item_type = ""
        if "type" in header_map and header_map["type"] < len(row):
            item_type = str(row[header_map["type"]] or "").strip()

        unit = ""
        if "unit" in header_map and header_map["unit"] < len(row):
            unit = str(row[header_map["unit"]] or "").strip()

        quantity = None
        if "quantity" in header_map and header_map["quantity"] < len(row):
            qty_raw = row[header_map["quantity"]]
            if isinstance(qty_raw, (int, float)) and not isinstance(qty_raw, bool):
                quantity = float(qty_raw)
            elif qty_raw is not None:
                try:
                    quantity = float(str(qty_raw).replace(",", "."))
                except (ValueError, TypeError):
                    quantity = None

        notes = ""
        if "notes" in header_map and header_map["notes"] < len(row):
            notes = str(row[header_map["notes"]] or "").strip()

        source_no = ""
        if "source_no" in header_map and header_map["source_no"] < len(row):
            source_no = str(row[header_map["source_no"]] or "").strip()

        item = {
            "type": item_type,
            "name": str(name_val).strip(),
            "unit": unit,
            "quantity": quantity,
            "notes": notes,
        }
        if source_no:
            item["source_no"] = source_no
        items.append(item)

    return items


def parse_list_sheet(file_bytes: bytes) -> list[dict]:
    """
    Parse Excel file with a list of works/materials.
      Columns: №, Тип, Наименование, Ед. изм., Кол-во, Примечание
    Returns list of dicts {type, name, unit, quantity, notes, sheet}.
    Raises ValueError if nothing recognisable was found.

    Разбираются все листы файла: смету на основе загруженного Excel считают по
    файлу заказчика, а он бывает разбит по листам — по листу на раздел. Взятый
    один лист терял остальные, и сумма контракта выходила заниженной.

    Исключение — наш собственный перечень старого вида: рядом с «Перечнем» в
    нём лежат сводки «Работы» и «Материалы» с теми же позициями, и брать их
    второй раз нельзя.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Не удалось открыть xlsx-файл: {exc}") from exc

    if _is_legacy_list_workbook(wb):
        targets = [ws for ws in wb.worksheets if "перечень" in ws.title.strip().lower()]
    else:
        targets = [
            ws for ws in wb.worksheets
            if ws.title.strip().lower() != _NOTE_SHEET
        ]

    by_sheet = [(ws, _parse_list_worksheet(ws)) for ws in targets]
    multi_sheet = sum(1 for _, sheet_items in by_sheet if sheet_items) > 1

    items: list[dict] = []
    for ws, sheet_items in by_sheet:
        for item in sheet_items:
            # Один лист — признак не ставим: документ остаётся без вкладок и
            # ведёт себя ровно как до появления многолистовых файлов.
            if multi_sheet:
                item["sheet"] = ws.title
            items.append(item)

    if not items:
        available = ", ".join(f'"{ws.title}"' for ws in wb.worksheets)
        raise ValueError(
            "В файле не найдено ни одной позиции. "
            f"Доступные листы: {available}. "
            "Проверьте, что в файле есть таблица с колонкой «Наименование»."
        )

    return items
