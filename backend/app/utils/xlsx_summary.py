import io

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.summary_estimate import SummaryEstimate
from app.utils.price_coercion import coerce_qty, coerce_qty_signed
from app.utils.summary_calc import FIXED_ROW_KEYS, FIXED_ROW_LABELS, calc_summary

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_GRAND_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
_WORK_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_COEF_FILL = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
_BOLD_WHITE = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_NUM_FMT = "#,##0.00"
# Ввод человека из колонки «% / Кол-во»: число остаётся числом, подпись — в формате.
_PCT_FMT = '0.##"%"'
_QTY_FMT = '0.##" чел"'
_COEF_FMT = "0.####"

# Левый блок бланка: № | Наименование | % / Кол-во | Стоимость с НДС | без НДС.
_COL_NUM, _COL_NAME, _COL_INPUT, _COL_WITH_VAT, _COL_WITHOUT_VAT = 1, 2, 3, 4, 5
_LEFT_COLS = 5


def _to_f(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _pct(v) -> float:
    return _to_f(v) / 100.0


def _billable_qty(row: dict) -> float:
    """Объём для умножения на цену: вычет (< 0) даёт 0.

    Строка с отрицательным объёмом корректирует объём соседней позиции, а не
    является работой: `qty × цена` по ней дала бы отрицательную сумму и занизила
    сводную. Правило то же, что в `price_coercion.coerce_qty`.
    """
    return coerce_qty(row.get("quantity") or row.get("qty"))


def generate_summary_xlsx(summary: SummaryEstimate) -> bytes:
    """Многолистовой xlsx: лист на каждый раздел + лист «Сводная»."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sections = summary.sections or []
    overrides = summary.overrides or {}

    for sec in sections:
        raw_name = sec.get("card_name") or "Раздел"
        sheet_name = raw_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_section_sheet(ws, sec)

    ws_sum = wb.create_sheet(title="Сводная")
    _write_summary_sheet(ws_sum, sections, overrides)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_section_sheet(ws, section: dict) -> None:
    headers = [
        "№", "Наименование", "Ед. изм.", "Кол-во",
        "Цена работ", "Стоимость работ", "Цена матер.", "Стоимость матер.",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _BOLD_WHITE
        cell.fill = _HEADER_FILL

    for i, w in enumerate([5, 50, 10, 10, 14, 18, 14, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, row in enumerate(section.get("rows") or [], 1):
        qty = _billable_qty(row)
        # В колонке «Кол-во» — исходное число со знаком: вычет должен остаться
        # видимым, стоимости по нему просто нет.
        qty_shown = coerce_qty_signed(row.get("quantity") or row.get("qty"))
        wp = _to_f(row.get("work_price") or row.get("price_work"))
        mp = _to_f(row.get("material_price") or row.get("price_material"))
        work_cost = round(qty * wp, 2)
        mat_cost = round(qty * mp, 2)
        is_work = str(row.get("type", "")).strip() == "Работа"
        fill = _WORK_FILL if is_work else None
        r = idx + 1

        values = [idx, row.get("name", ""), row.get("unit", ""), qty_shown or None,
                  wp or None, work_cost or None, mp or None, mat_cost or None]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            if fill:
                cell.fill = fill
            if col in (5, 6, 7, 8) and isinstance(val, (int, float)):
                cell.number_format = _NUM_FMT


def _write_summary_sheet(ws, sections: list, overrides: dict) -> None:
    """Лист «Сводная» — ровно то, что человек видит в бланке на экране.

    Числа считает `utils/summary_calc` — построчный повтор клиентской
    `calcSummary`. Раньше здесь была своя, более старая формула: без
    коэффициента, без налогов разделов, без восьми строк расходов, с другой
    прибылью и НДС. Файл расходился с экраном на десятки процентов.

    Колонки бланка те же, что на экране (`components/summary/SummarySheet.tsx`):
    № | Наименование | % / Кол-во | Стоимость с НДС | Стоимость без НДС.
    Раскладка закреплена тестом `backend/tests/test_xlsx_summary.py`.
    """
    calc = calc_summary(sections, overrides)
    hidden = calc["hidden_fixed_rows"]

    # ── Левая таблица: бланк «Себестоимость и цена для заказчика» (A-E) ────
    # Колонки те же, что на экране, и у каждой строки те же числа: и с НДС, и
    # без НДС. Раньше половина бланка (непредвиденные, прибыль, налоги, итог)
    # уезжала одной суммой в отдельную колонку «Сумма», а ввод человека (%,
    # количество, коэффициент) в файл не попадал вовсе.
    headers = ["№", "Наименование", "% / Кол-во", "Стоимость с НДС", "Стоимость без НДС"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _BOLD_WHITE
        cell.fill = _HEADER_FILL

    for i, w in enumerate([5, 52, 14, 20, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_idx = 2

    ws.cell(row=row_idx, column=_COL_NAME, value="Коэффициент к ценам (×все цены)").font = _BOLD
    _input_cell(ws, row_idx, calc["coefficient"], _COEF_FMT)
    for col in range(1, _LEFT_COLS + 1):
        ws.cell(row=row_idx, column=col).fill = _COEF_FILL
    row_idx += 1

    # Строки, где в колонке «% / Кол-во» стоит ввод человека.
    row_inputs = {
        "transport": (calc["transport_pct"], _PCT_FMT),
        "cleanup": (calc["cleanup_pct"], _PCT_FMT),
        "overhead": (calc["overhead_pct"], _PCT_FMT),
        "daily_workers": (calc["daily_workers_qty"], _QTY_FMT),
    }

    num = 0
    for key in FIXED_ROW_KEYS:
        if key in hidden:
            continue
        num += 1
        ws.cell(row=row_idx, column=_COL_NUM, value=num)
        ws.cell(row=row_idx, column=_COL_NAME, value=FIXED_ROW_LABELS[key])
        if key in row_inputs:
            _input_cell(ws, row_idx, *row_inputs[key])
        _money(ws, row_idx, _COL_WITH_VAT, calc[f"{key}_with_vat"])
        _money(ws, row_idx, _COL_WITHOUT_VAT, calc[f"{key}_without_vat"])
        row_idx += 1

    for custom in calc["custom_rows_before"]:
        num += 1
        ws.cell(row=row_idx, column=_COL_NUM, value=num)
        row_idx = _custom_row(ws, row_idx, custom)

    _total_row(ws, row_idx, "ИТОГО себестоимость объекта", _TOTAL_FILL,
               calc["subtotal_with_vat"], calc["subtotal_without_vat"])
    row_idx += 1

    # Строки без номера под итогом себестоимости — справочные.
    for custom in calc["custom_rows_after"]:
        row_idx = _custom_row(ws, row_idx, custom)

    ws.cell(row=row_idx, column=_COL_NAME, value="Непредвиденные расходы")
    _input_cell(ws, row_idx, calc["contingency_pct"], _PCT_FMT)
    _money(ws, row_idx, _COL_WITH_VAT, calc["contingency_with_vat"])
    _money(ws, row_idx, _COL_WITHOUT_VAT, calc["contingency_without_vat"])
    row_idx += 1

    # Ниже — строки, у которых сумма одна на обе колонки (как на экране, где
    # ячейка растянута): прибыль и полная себестоимость считаются без НДС, НДС и
    # налоги — сами налоги, итог для заказчика уже с ними.
    ws.cell(row=row_idx, column=_COL_NAME, value="Плановая прибыль (без НДС)")
    _input_cell(ws, row_idx, calc["profit_pct"], _PCT_FMT)
    _merged_money(ws, row_idx, calc["profit"])
    row_idx += 1

    _total_row(ws, row_idx,
               "Полная себестоимость с учётом прибыли и непредвиденных (без НДС)",
               _TOTAL_FILL, calc["full_cost_without_vat"], None)
    row_idx += 1

    for label, pct, value in (
        ("НДС от полной себестоимости", calc["vat_pct"], calc["vat"]),
        ("Др. налоги от полной себестоимости", calc["other_tax_pct"], calc["other_tax"]),
    ):
        ws.cell(row=row_idx, column=_COL_NAME, value=label)
        _input_cell(ws, row_idx, pct, _PCT_FMT)
        _merged_money(ws, row_idx, value)
        row_idx += 1

    _total_row(ws, row_idx, "ИТОГО по смете для Заказчика с учётом налогов",
               _GRAND_FILL, calc["total_for_customer"], None)
    row_idx += 1

    # Цель по объекту — только если она задана: без целей файл обязан выглядеть
    # так же, как до появления этой функции.
    if calc["target_total_for_customer"] is not None:
        ws.cell(row=row_idx, column=_COL_NAME, value="Цель по объекту").font = _BOLD
        _merged_money(ws, row_idx, calc["target_total_for_customer"], bold=True)
        for col in range(1, _LEFT_COLS + 1):
            ws.cell(row=row_idx, column=col).fill = _COEF_FILL
        row_idx += 1

        ws.cell(row=row_idx, column=_COL_NAME, value="Отклонение от цели по объекту")
        _merged_money(ws, row_idx, calc["total_deviation"])
        row_idx += 1

    # ── Правая таблица: разделы со своими налогами (с колонки G) ──────────
    # Налогов у раздела два: работы и материалы облагаются независимо
    # (работы — самозанятый, материалы — подрядчик с НДС, и наоборот).
    #
    # Второй налог добавлен колонкой в конец, а не рядом с материалами: порядок
    # столбцов раздела не меняется, и формулы, написанные людьми поверх прежних
    # выгрузок, не съезжают на соседний столбец. Сама таблица сдвинута на один
    # столбец вправо — в бланке слева стало пять колонок вместо четырёх.
    OFF = 7
    right_headers = [
        "Раздел", "Налог работ, %", "Работы (с/с)", "Работы с НДС",
        "Материалы (с/с)", "Материалы с НДС", "Налог матер., %",
    ]
    # Цели оптимизации — снова в конец таблицы, по той же причине, что и второй
    # налог: порядок прежних столбцов не должен меняться. Колонок нет вовсе,
    # пока не задана хоть одна цель раздела. Процента отклонения в файле нет:
    # цель и отклонение стоят рядом, процент считается формулой.
    with_targets = calc["has_section_targets"]
    if with_targets:
        right_headers += ["Цель работ", "Откл. работ", "Цель матер.", "Откл. матер."]
    for col, h in enumerate(right_headers, OFF):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _BOLD_WHITE
        cell.fill = _HEADER_FILL

    widths = [30, 14, 18, 18, 18, 18, 14] + ([18, 18, 18, 18] if with_targets else [])
    for i, w in enumerate(widths, OFF):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Колонки денег (смещение от OFF) — их суммируем и форматируем как рубли.
    MONEY_COLS = (2, 3, 4, 5)
    totals_right = dict.fromkeys(MONEY_COLS, 0.0)
    # Колонки целей суммируются не построчно: в ИТОГО идут только разделы с
    # заданной целью, а это уже посчитал `calc_summary`.
    TARGET_COLS = (7, 8, 9, 10)
    for r_idx, section in enumerate(calc["section_totals"], 2):
        values = [
            section["card_name"], section["tax_pct_works"],
            section["works_raw"], section["works_with_vat"],
            section["materials_raw"], section["materials_with_vat"],
            section["tax_pct_materials"],
        ]
        if with_targets:
            values += [
                section["target_works"], section["works_deviation"],
                section["target_materials"], section["materials_deviation"],
            ]
        for offset, value in enumerate(values):
            cell = ws.cell(
                row=r_idx, column=OFF + offset,
                value=round(value, 2) if isinstance(value, float) else value,
            )
            if offset in MONEY_COLS:
                totals_right[offset] += value
                cell.number_format = _NUM_FMT
            elif offset in TARGET_COLS and value is not None:
                cell.number_format = _NUM_FMT

    target_totals = {
        7: calc["targets_total_works"], 8: calc["targets_deviation_works"],
        9: calc["targets_total_materials"], 10: calc["targets_deviation_materials"],
    }
    total_row = len(calc["section_totals"]) + 2
    for offset in range(len(right_headers)):
        if offset == 0:
            value = "ИТОГО"
        elif offset in MONEY_COLS:
            value = round(totals_right[offset], 2)
        elif offset in TARGET_COLS:
            total = target_totals[offset]
            value = round(total, 2) if total is not None else None
        else:
            value = None
        cell = ws.cell(row=total_row, column=OFF + offset, value=value)
        cell.font = _BOLD
        cell.fill = _TOTAL_FILL
        if offset in MONEY_COLS or (offset in TARGET_COLS and value is not None):
            cell.number_format = _NUM_FMT


def _money(ws, row: int, col: int, value, *, bold: bool = False):
    """Денежная ячейка бланка."""
    cell = ws.cell(row=row, column=col, value=round(_to_f(value), 2))
    cell.number_format = _NUM_FMT
    if bold:
        cell.font = _BOLD
    return cell


def _merged_money(ws, row: int, value, *, bold: bool = False):
    """Одна сумма на обе денежные колонки — как растянутая ячейка на экране."""
    cell = _money(ws, row, _COL_WITH_VAT, value, bold=bold)
    ws.merge_cells(start_row=row, start_column=_COL_WITH_VAT,
                   end_row=row, end_column=_COL_WITHOUT_VAT)
    return cell


def _input_cell(ws, row: int, value, fmt: str):
    """Колонка «% / Кол-во»: число человека остаётся числом, подпись — формат."""
    cell = ws.cell(row=row, column=_COL_INPUT, value=round(_to_f(value), 4))
    cell.number_format = fmt
    return cell


def _custom_row(ws, row: int, custom: dict) -> int:
    """Строка, добавленная человеком: хранится без НДС, показывается в обеих колонках."""
    ws.cell(row=row, column=_COL_NAME, value=str(custom.get("label") or ""))
    qty_pct = str(custom.get("qty_pct") or "")
    if qty_pct:
        ws.cell(row=row, column=_COL_INPUT, value=qty_pct)
    without_vat = _to_f(custom.get("without_vat"))
    _money(ws, row, _COL_WITH_VAT, without_vat * 1.22)
    _money(ws, row, _COL_WITHOUT_VAT, without_vat)
    return row + 1


def _total_row(ws, row: int, label: str, fill: PatternFill,
               with_vat: float, without_vat) -> None:
    """Итоговая строка. `without_vat=None` — сумма одна на обе колонки."""
    for col in range(1, _LEFT_COLS + 1):
        ws.cell(row=row, column=col).fill = fill
    lc = ws.cell(row=row, column=_COL_NAME, value=label)
    lc.font = _BOLD
    if without_vat is None:
        _merged_money(ws, row, with_vat, bold=True)
    else:
        _money(ws, row, _COL_WITH_VAT, with_vat, bold=True)
        _money(ws, row, _COL_WITHOUT_VAT, without_vat, bold=True)
