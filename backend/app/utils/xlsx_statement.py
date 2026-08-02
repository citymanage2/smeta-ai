"""Выгрузка-ведомость: любой документ → xlsx в корпоративном оформлении.

Один генератор на все пять типов документов. Он ничего не знает про смету:
колонки приходят от документа (у перечня свои, у сметы свои), строки — из
предпросмотра, где человек мог их поправить или удалить.

Цены приходят уже с коэффициентом: его применяет редактор при показе, и в
выгрузку попадает ровно то, что человек видел на экране.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Корпоративное оформление: синяя шапка, светлая заливка итога.
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=14, color="1F3864")
_META_FONT = Font(size=10, color="595959")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NUM_FMT = "#,##0.00"

# Числовые колонки, которые складывать нельзя: количество и цена за единицу —
# сумма по ним смысла не имеет и только вводит в заблуждение.
_NOT_SUMMABLE = {"num", "qty", "quantity", "price_work", "price_material", "price"}
_NOT_SUMMABLE_WORDS = ("кол-во", "количество", "цена", "№", "ед. изм")


def _is_summable(key: str, label: str) -> bool:
    if key in _NOT_SUMMABLE:
        return False
    lowered = str(label or "").strip().lower()
    return not any(word in lowered for word in _NOT_SUMMABLE_WORDS)


def _to_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _column_width(label: str, key: str, rows: list) -> int:
    longest = len(str(label))
    for row in rows:
        value = row.get(key)
        if value is not None:
            longest = max(longest, len(str(value)))
    return max(10, min(55, longest + 4))


def generate_statement_xlsx(
    columns: list,
    rows: list,
    *,
    title: str = "",
    object_name: str = "",
    project_name: str = "",
    show_date: bool = True,
    show_total: bool = True,
    sheet_name: str = "Выгрузка",
    generated_at: Optional[datetime] = None,
) -> bytes:
    """Собрать ведомость. `columns` — [{key, label, numeric}], `rows` — словари."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Выгрузка")[:31]

    keys = [str(c.get("key")) for c in columns]
    labels = [str(c.get("label") or c.get("key") or "") for c in columns]
    numeric = [bool(c.get("numeric")) for c in columns]

    row_idx = 1
    if title:
        cell = ws.cell(row=row_idx, column=1, value=title)
        cell.font = _TITLE_FONT
        if len(keys) > 1:
            ws.merge_cells(start_row=row_idx, start_column=1,
                           end_row=row_idx, end_column=len(keys))
        row_idx += 1

    for line in (object_name, project_name):
        if line:
            cell = ws.cell(row=row_idx, column=1, value=line)
            cell.font = _META_FONT
            row_idx += 1

    if show_date:
        stamp = (generated_at or datetime.now(timezone.utc)).strftime("%d.%m.%Y")
        cell = ws.cell(row=row_idx, column=1, value=f"Дата формирования: {stamp}")
        cell.font = _META_FONT
        row_idx += 1

    if row_idx > 1:
        row_idx += 1  # пустая строка между шапкой и таблицей

    header_row = row_idx
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = _column_width(
            label, keys[col - 1], rows,
        )
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    row_idx += 1

    totals = [0.0] * len(keys)
    has_total = [False] * len(keys)

    for row in rows:
        for col, key in enumerate(keys, 1):
            raw = row.get(key)
            value: Any = raw
            if numeric[col - 1]:
                value = _to_number(raw)
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = _BORDER
            if isinstance(value, float):
                cell.number_format = _NUM_FMT
                if _is_summable(key, labels[col - 1]):
                    totals[col - 1] += value
                    has_total[col - 1] = True
        row_idx += 1

    if show_total and rows:
        label_cell = ws.cell(row=row_idx, column=1, value="ИТОГО")
        label_cell.font = _BOLD
        label_cell.fill = _TOTAL_FILL
        label_cell.border = _BORDER
        for col in range(1, len(keys) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = _TOTAL_FILL
            cell.border = _BORDER
            if col == 1:
                continue
            if has_total[col - 1]:
                cell.value = round(totals[col - 1], 2)
                cell.font = _BOLD
                cell.number_format = _NUM_FMT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
