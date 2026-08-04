"""Generic xlsx parser/generator for LIST and COMPLETENESS tasks.

Does NOT map columns to estimate schema — preserves xlsx structure as-is.
Row format: {"row_id": str(uuid), "sheet": "Раздел 1", "cells": {"ColName": value, ...}}

Разбираются все листы файла: заказчики разбивают смету по листам — по листу на
раздел или корпус, — и взятый один первый лист терял остальные молча. Лист
остаётся при строке полем `sheet` и становится вкладкой в редакторе.
"""
import io
import uuid
from datetime import datetime, date
from typing import Optional

import openpyxl

from app.utils.sheet_names import group_by_sheet, safe_sheet_title

# Строку нумерации колонок («1 2 3 4 5 6») ищем только рядом с началом листа:
# ниже такая строка — уже данные, а не часть шапки.
_NUMBERING_SCAN_LIMIT = 15
# Сколько строк максимум склеиваем в одно название колонки.
_MAX_HEADER_ROWS = 4


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _non_empty_count(row) -> int:
    return sum(1 for v in row if _text(v))


def _is_numbering_row(row) -> bool:
    """Строка нумерации колонок Гранд-сметы и нашего же перечня: «1 2 3 4 5 6».

    Признаки: минимум три ячейки, все целые числа, все разные, начинается с 1 или 2,
    и значения не убегают далеко за число колонок (допускаем скрытые колонки —
    нумерация тогда идёт с пропусками).
    """
    values = [v for v in row if _text(v)]
    if len(values) < 3:
        return False

    numbers = []
    for v in values:
        if isinstance(v, bool):
            return False
        if isinstance(v, int):
            number = v
        elif isinstance(v, float) and v == int(v):
            number = int(v)
        elif _text(v).isdigit():
            number = int(_text(v))
        else:
            return False
        if not 1 <= number <= 99:
            return False
        numbers.append(number)

    if len(set(numbers)) != len(numbers):
        return False
    return min(numbers) <= 2 and max(numbers) <= len(numbers) + 3


def is_numbering_cells(cells) -> bool:
    """Строка нумерации колонок в уже разобранном документе.

    Разбор отсекает её с 2 августа 2026, но документы, созданные раньше, хранят
    её как обычную строку данных. Сервис документов не отдаёт такие строки
    редактору — по тем же признакам, что и парсер.
    """
    if not isinstance(cells, dict):
        return False
    return _is_numbering_row(list(cells.values()))


def _detect_header(all_rows) -> tuple:
    """Вернуть (индексы строк шапки, индекс первой строки данных).

    Опорный признак — строка нумерации колонок. Она есть и в выгрузках
    Гранд-сметы, и в перечнях нашего генератора, и всегда стоит сразу под шапкой.
    Если её нет, поведение прежнее: шапка — первая строка листа.
    """
    limit = min(_NUMBERING_SCAN_LIMIT, len(all_rows))
    for idx in range(1, limit):
        if not _is_numbering_row(all_rows[idx]):
            continue
        # Шапка — строки над нумерацией, пока в них есть хотя бы две заполненные
        # ячейки. Титульный блок сметы («ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЁТ…») так отсекается:
        # его строки объединённые, заполненная ячейка в них одна.
        start = idx
        while start > 0 and idx - start < _MAX_HEADER_ROWS and _non_empty_count(all_rows[start - 1]) >= 2:
            start -= 1
        if start == idx:
            start = idx - 1
        return list(range(start, idx)), idx + 1

    return [0], 1


def _header_grid(ws, all_rows, header_idx: list, width: int) -> dict:
    """Матрица ячеек шапки с раскрытыми объединёнными ячейками.

    В xlsx у объединённой ячейки значение хранится только в левом верхнем углу,
    остальные приходят пустыми. Без раскрытия колонка «Стоимость, руб. / всего»
    получила бы имя «всего».
    """
    grid = {i: (list(all_rows[i]) + [None] * width)[:width] for i in header_idx}
    top, bottom = min(header_idx), max(header_idx)

    for rng in ws.merged_cells.ranges:
        r0, r1 = rng.min_row - 1, rng.max_row - 1
        c0, c1 = rng.min_col - 1, rng.max_col - 1
        if r1 < top or r0 > bottom or c0 >= width:
            continue
        if r0 >= len(all_rows) or c0 >= len(all_rows[r0]):
            continue
        value = all_rows[r0][c0]
        if value is None:
            continue
        for r in range(max(r0, top), min(r1, bottom) + 1):
            for c in range(c0, min(c1, width - 1) + 1):
                if grid[r][c] is None:
                    grid[r][c] = value

    return grid


def _build_headers(ws, all_rows, header_idx: list, width: int) -> list:
    """Склеить строки шапки в названия колонок, без пустых и без дублей."""
    grid = _header_grid(ws, all_rows, header_idx, width)

    headers = []
    for col in range(width):
        parts = []
        for row_idx in header_idx:
            part = _text(grid[row_idx][col])
            # объединённая по вертикали ячейка повторяется — не дублируем её в имени
            if part and (not parts or parts[-1] != part):
                parts.append(part)
        headers.append(" ".join(parts))

    # Одинаковые имена колонок схлопнулись бы в один ключ и потеряли данные
    seen: dict = {}
    unique = []
    for i, name in enumerate(headers):
        name = name or f"Col{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        unique.append(name)

    return unique


def _parse_sheet(ws) -> list[dict]:
    """Строки одного листа. Шапка ищется в нём же — у листов она своя."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    header_idx, data_start = _detect_header(all_rows)
    width = max(len(all_rows[i]) for i in header_idx)
    headers = _build_headers(ws, all_rows, header_idx, width)

    result = []
    for raw_row in all_rows[data_start:]:
        # Skip completely empty rows
        if all(v is None for v in raw_row):
            continue

        cells: dict = {}
        for header, val in zip(headers, raw_row):
            if val is None:
                cells[header] = ""
            elif isinstance(val, (datetime, date)):
                cells[header] = val.isoformat()
            elif isinstance(val, float) and val == int(val):
                cells[header] = int(val)
            else:
                cells[header] = val

        result.append({"row_id": str(uuid.uuid4()), "sheet": ws.title, "cells": cells})

    return result


def parse_xlsx_to_generic_rows(
    file_bytes: bytes, sheets: Optional[list] = None
) -> list[dict]:
    """Строки всех листов файла подряд, каждая помечена своим листом.

    `sheets` ограничивает разбор перечисленными листами и задаёт их порядок.
    Нужен там, где часть листов файла служебная: наш собственный перечень несёт
    рядом с данными сводки «Работы» и «Материалы», и без ограничения документ
    получил бы вкладки-двойники с теми же позициями.

    Если ни одно из перечисленных имён в файле не нашлось — берём первый лист.
    Так открываются перечни, пересобранные до появления вкладок: у их файла
    лист назван по умолчанию, и по списку имён документ оказался бы пустым.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if sheets is None:
        targets = list(wb.worksheets)
    else:
        by_title = {ws.title: ws for ws in wb.worksheets}
        targets = [by_title[title] for title in sheets if title in by_title]
        if not targets and wb.worksheets:
            targets = [wb.worksheets[0]]

    result: list[dict] = []
    for ws in targets:
        # Лист без единой строки данных вкладки не создаёт: титульные листы и
        # листы подписей есть почти в каждой выгрузке Гранд-сметы.
        result.extend(_parse_sheet(ws))
    return result


_MONEY_FMT = "#,##0.00"
_MONEY_WORDS = ("цена", "стоимость", "сумма")


def _as_number(value):
    """Числовая строка → число. Иначе None.

    После правки в редакторе значение приходит строкой: человек набрал «1234,56»
    или вставил «1 234,56» из другой таблицы. Записанное строкой, оно попадёт в
    Excel текстом, и колонку в скачанном файле не просуммировать.

    Ведущий ноль («007», «0123») — признак кода, а не числа: такие значения
    оставляем как есть, иначе артикул превратится в 7.
    """
    if isinstance(value, bool) or isinstance(value, (int, float)):
        return None
    text = str(value or "").strip().replace(" ", "").replace(" ", "")
    if not text:
        return None
    normalized = text.replace(",", ".")
    lead = normalized[1:] if normalized[:1] in "+-" else normalized
    if len(lead) > 1 and lead[0] == "0" and not lead.startswith("0."):
        return None
    try:
        number = float(normalized)
    except ValueError:
        return None
    return int(number) if number == int(number) and "." not in normalized else number


def _write_sheet(ws, rows: list[dict]) -> None:
    """Один лист: колонки берутся у его же строк.

    Набор колонок считается по строкам этого листа, а не всего документа: у
    разных листов исходного файла шапки разные, и общий набор дал бы каждому
    листу пустые колонки соседа.
    """
    all_keys: list[str] = list(dict.fromkeys(k for row in rows for k in row.get("cells", {}).keys()))
    money_cols = {
        i for i, key in enumerate(all_keys, 1)
        if any(word in str(key).lower() for word in _MONEY_WORDS)
    }

    ws.append(all_keys)
    for row_idx, row in enumerate(rows, 2):
        cells = row.get("cells", {})
        for col_idx, key in enumerate(all_keys, 1):
            value = cells.get(key, "")
            number = _as_number(value)
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=value if number is None else number)
            if col_idx in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = _MONEY_FMT


def rows_to_xlsx(rows: list[dict]) -> bytes:
    """Строки документа → xlsx: по листу на каждую вкладку, в её порядке."""
    wb = openpyxl.Workbook()

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    used: set = set()
    for index, (title, sheet_rows) in enumerate(group_by_sheet(rows)):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = safe_sheet_title(title, used)
        _write_sheet(ws, sheet_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
