import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.constants import TASK_TYPE_LABELS, ESTIMATION_STATUS_LABELS
from app.utils.price_coercion import coerce_price, coerce_qty, coerce_qty_signed
from app.utils.sheet_names import group_by_sheet, safe_sheet_title
from app.utils.source_numbers import SOURCE_NO_HEADER, has_source_numbers, source_no_value

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_BOLD = Font(bold=True)


def generate_project_xlsx(project, tasks: list, slot_results: dict, base_url: str) -> bytes:
    # project is accepted for API consistency with pdf_exporter (name, description used there)
    """
    Generate xlsx bytes for a project export.

    Args:
        project: Project ORM instance (id, name, description)
        tasks: list of Task ORM instances
        slot_results: dict with keys 'source'/'estimate'/'optimized',
                      each value is list of (Task, TaskResult) tuples
        base_url: backend base URL for hyperlinks, e.g. "https://host.com"

    Returns: bytes of the xlsx file
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Задачи ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Задачи"

    headers = ["Тип задачи", "Статус сметы", "Стоимость (₽)", "Дата создания"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20

    total_cost = 0.0
    has_cost = False
    unestimated = estimated = optimized = 0

    for row_idx, task in enumerate(tasks, 2):
        type_label = TASK_TYPE_LABELS.get(task.task_type, task.task_type)
        status_label = ESTIMATION_STATUS_LABELS.get(task.estimation_status, task.estimation_status)
        cost = float(task.cost) if task.cost is not None else None
        if isinstance(task.created_at, datetime):
            created = task.created_at.strftime("%d.%m.%Y")
        else:
            created = str(task.created_at)

        ws.cell(row=row_idx, column=1, value=type_label)
        ws.cell(row=row_idx, column=2, value=status_label)
        cost_cell = ws.cell(row=row_idx, column=3, value=cost)
        cost_cell.number_format = "#,##0.00"
        ws.cell(row=row_idx, column=4, value=created)

        if cost is not None:
            total_cost += cost
            has_cost = True
        if task.estimation_status == "unestimated":
            unestimated += 1
        elif task.estimation_status == "estimated":
            estimated += 1
        elif task.estimation_status == "optimized":
            optimized += 1

    # Total row
    total_row = len(tasks) + 2
    summary = f"не рассчитано: {unestimated} / рассчитано: {estimated} / оптимизировано: {optimized}"
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).font = _BOLD
        ws.cell(row=total_row, column=col).fill = _TOTAL_FILL
    ws.cell(row=total_row, column=1, value="ИТОГО")
    ws.cell(row=total_row, column=2, value=summary)
    total_cell = ws.cell(row=total_row, column=3, value=total_cost if has_cost else None)
    total_cell.number_format = "#,##0.00"

    # ── Sheets 2-4: Slots ───────────────────────────────────────────────
    slot_config = [
        ("source", "Исходные файлы"),
        ("estimate", "Расчёты"),
        ("optimized", "Оптимизированные"),
    ]
    for slot, sheet_name in slot_config:
        ws_s = wb.create_sheet(title=sheet_name)
        for col, h in enumerate(["Тип задачи", "Имя файла", "Ссылка"], 1):
            cell = ws_s.cell(row=1, column=col, value=h)
            cell.font = _BOLD
            cell.fill = _HEADER_FILL
        ws_s.column_dimensions["A"].width = 35
        ws_s.column_dimensions["B"].width = 30
        ws_s.column_dimensions["C"].width = 50

        pairs = slot_results.get(slot, [])
        if not pairs:
            ws_s.cell(row=2, column=1, value="Файлы отсутствуют")
        else:
            for row_idx, (task, task_result) in enumerate(pairs, 2):
                url = f"{base_url}/tasks/{task.id}/files/{slot}/download"
                ws_s.cell(row=row_idx, column=1, value=TASK_TYPE_LABELS.get(task.task_type, task.task_type))
                ws_s.cell(row=row_idx, column=2, value=task_result.file_name)
                ws_s.cell(row=row_idx, column=3, value=f'=HYPERLINK("{url}")')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


DEFAULT_OVERHEAD_PCT = 3.0
DEFAULT_TRANSPORT_PCT = 3.0


def coefficient_for(coefficient: Optional[dict], row_id) -> tuple:
    """Множители (работы, материалы) для конкретной строки.

    Коэффициент — настройка документа: `{"work": 1.05, "material": 1.0,
    "scope": "all" | ["row_id", ...]}`. Исходные цены он не меняет никогда,
    поэтому его можно снять и получить ровно прежние числа.
    """
    if not isinstance(coefficient, dict):
        return 1.0, 1.0

    scope = coefficient.get("scope", "all")
    if isinstance(scope, (list, tuple, set)):
        if row_id is None or str(row_id) not in {str(x) for x in scope}:
            return 1.0, 1.0

    def _k(value) -> float:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return 1.0
        # Ноль и минус коэффициентом не бывают: они молча обнулили бы смету.
        return number if number > 0 else 1.0

    return _k(coefficient.get("work", 1.0)), _k(coefficient.get("material", 1.0))



# Лист «Смета» — единственный, когда исходный файл был из одного листа. Строки
# без листа собираются в «Прочее»: подмешать их к чужому разделу нельзя.
ESTIMATE_SHEET_TITLE = "Смета"
UNSORTED_SHEET_TITLE = "Прочее"
SUMMARY_SHEET_TITLE = "Итого по смете"

_EST_HEADERS = [
    "№",
    "Наименование",
    "Ед. изм.",
    "Кол-во",
    "Цена работ",
    "Стоимость работ",
    "Цена матер.",
    "Стоимость матер.",
    "Источник цены",
    "Примечание",
]
_EST_COL_WIDTHS = [5, 50, 10, 10, 14, 18, 14, 18, 12, 60]

# Номер позиции в смете заказчика — им сверяют строку с исходником. Колонка
# необязательная: смета, посчитанная по файлу без нумерации, выглядит как
# раньше. Заголовок и правило «только если номер есть» — общие с перечнем.
_EST_SOURCE_NO_WIDTH = 12

_HEADER_FILL_EST = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_WORK_ROW_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_TOTAL_FILL_EST = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_GRAND_TOTAL_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")


def _pct_label(value: float) -> str:
    return f"{value:g}"


def _estimate_sheet_groups(items: list) -> list:
    """Листы сметы: `[(имя листа, позиции), ...]` в порядке появления.

    Один лист (или ни одного) — прежний единственный «Смета»: у всех
    существующих смет имя листа менять незачем.
    """
    groups = group_by_sheet(items)
    named = [title for title, _ in groups if title is not None]
    if len(named) < 2:
        return [(ESTIMATE_SHEET_TITLE, list(items or []))]

    used: set = set()
    return [
        (safe_sheet_title(title if title is not None else UNSORTED_SHEET_TITLE, used), group)
        for title, group in groups
    ]


def _write_estimate_sheet(ws, items: list, coefficient: Optional[dict],
                          with_source_no: bool = False) -> tuple:
    """Позиции одного листа. Возвращает (сумма работ, сумма материалов, след. строка).

    `with_source_no` добавляет первой колонку «№ в исходной смете». Решение
    принимается один раз на весь файл: у листов одной сметы набор колонок обязан
    совпадать, иначе блок итогов встанет в разные колонки на разных листах.
    """
    from openpyxl.utils import get_column_letter

    headers = ([SOURCE_NO_HEADER] if with_source_no else []) + _EST_HEADERS
    widths = ([_EST_SOURCE_NO_WIDTH] if with_source_no else []) + _EST_COL_WIDTHS

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL_EST

    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    total_works = 0.0
    total_materials = 0.0
    row_num = 2
    idx = 0
    for item in items:
        idx += 1
        # Цены приходят из ответа ИИ и могут быть строкой, отрицательными или
        # мусором. Раньше строка роняла round() — и задача становилась
        # неизлечимой, потому что возобновление читало тот же чекпоинт и падало
        # снова. Здесь последний рубеж: непригодная цена = «цены нет».
        # coerce_qty обнуляет минус — стоимость по вычету не считается и в итог
        # не попадает. Но показать в колонке «Кол-во» нужно исходное число со
        # знаком: иначе строка выглядит как позиция с потерянным объёмом.
        qty = coerce_qty(item.get("quantity"))
        qty_shown = coerce_qty_signed(item.get("quantity"))
        work_price = coerce_price(item.get("work_price"))
        mat_price = coerce_price(item.get("material_price"))

        k_work, k_material = coefficient_for(coefficient, item.get("row_id"))
        if work_price is not None and k_work != 1.0:
            work_price = round(work_price * k_work, 2)
        if mat_price is not None and k_material != 1.0:
            mat_price = round(mat_price * k_material, 2)

        work_cost = round(qty * work_price, 2) if work_price is not None and qty else None
        mat_cost = round(qty * mat_price, 2) if mat_price is not None and qty else None

        if work_cost is not None:
            total_works += work_cost
        if mat_cost is not None:
            total_materials += mat_cost

        is_work = str(item.get("type", "")).strip() == "Работа"

        row_fill = _WORK_ROW_FILL if is_work else None

        values = ([source_no_value(item)] if with_source_no else []) + [
            idx,
            item.get("name", ""),
            item.get("unit", ""),
            # Приведённый объём, а не сырой: строка «3» попала бы в числовую
            # колонку текстом, и Excel не сложил бы её в сумме.
            qty_shown if qty_shown else None,
            work_price,
            work_cost,
            mat_price,
            mat_cost,
            item.get("price_list_name", "") or "",
            item.get("notes", "") or "",
        ]
        off = 1 if with_source_no else 0
        money_cols = {5 + off, 6 + off, 7 + off, 8 + off}
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = Font(bold=is_work)
            if row_fill:
                cell.fill = row_fill
            if col in money_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
            if col == 4 + off and isinstance(val, (int, float)):
                cell.number_format = "#,##0.##"
        row_num += 1

    return total_works, total_materials, row_num


def _write_totals_block(ws, row_num: int, rows: list, col_count: int = len(_EST_HEADERS)) -> int:
    """Блок итогов: подпись слева, число в последней колонке.

    `col_count` — ширина таблицы этого файла. Подпись всегда начинается с первой
    колонки, поэтому `extract_total_cost` находит строку «ИТОГО» независимо от
    того, добавлена ли колонка «№ в исходной смете».
    """
    for label, value in rows:
        is_grand = label.startswith("ИТОГО") or label.startswith("ВСЕГО")
        fill = _GRAND_TOTAL_FILL if is_grand else _TOTAL_FILL_EST
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.fill = fill
        # Подпись занимает все колонки, кроме последней — в ней число
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=col_count - 1)
        val_cell = ws.cell(row=row_num, column=col_count, value=value)
        val_cell.font = Font(bold=True)
        val_cell.fill = fill
        val_cell.number_format = "#,##0.00"
        row_num += 1
    return row_num


def generate_estimate_xlsx(
    items: list[dict],
    *,
    overhead_pct: float = DEFAULT_OVERHEAD_PCT,
    transport_pct: float = DEFAULT_TRANSPORT_PCT,
    coefficient: Optional[dict] = None,
) -> tuple:
    """
    Generate Excel estimate file for ESTIMATE_FROM_LIST task.

    Each item dict must have:
      type, name, unit, quantity,
      work_price, material_price,   (float | None)
      price_list_name,              (str | None — "Прайс" / "Кеш" / "Интернет")
      notes,                        (str | None — примечание: источники / дата кеша / наименование в прайсе)
      sheet                         (str | None — лист исходного файла)

    Проценты доп. расходов приходят снаружи (настройка проекта или версии) —
    раньше здесь были зашиты 3%, и файл расходился с экраном у любого проекта
    с другими ставками. Коэффициент применяется к ценам: в файл они попадают
    уже умноженными (решение пользователя 4.5).

    Позиции, размеченные листами исходного файла, раскладываются по листам с
    теми же именами. У каждого листа свой блок итогов — лист гранд-сметы это
    самостоятельная локальная смета, — а последним идёт «Итого по смете» со
    сводкой по листам и общей суммой.

    Возвращает (байты файла, общий итог).
    """
    wb = openpyxl.Workbook()

    overhead_rate = float(overhead_pct or 0)
    transport_rate = float(transport_pct or 0)

    groups = _estimate_sheet_groups(items)
    per_sheet: list = []

    # Колонка «№ в исходной смете» — общая на весь файл: иначе блок итогов встал
    # бы в разные колонки на разных листах, а сумму читают из последней.
    with_source_no = has_source_numbers(items)
    col_count = len(_EST_HEADERS) + (1 if with_source_no else 0)

    for index, (title, group) in enumerate(groups):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = title

        works, materials, row_num = _write_estimate_sheet(
            ws, group, coefficient, with_source_no=with_source_no,
        )
        overhead = round(works * overhead_rate / 100, 2)
        transport = round(materials * transport_rate / 100, 2)
        sheet_total = round(works + overhead + materials + transport, 2)

        row_num += 1  # blank separator
        _write_totals_block(ws, row_num, [
            ("Сумма по работам:", works),
            (f"Накладные расходы {_pct_label(overhead_rate)}%:", overhead),
            ("Сумма по материалам:", materials),
            (f"Транспортные расходы {_pct_label(transport_rate)}%:", transport),
            ("ИТОГО ПО СМЕТЕ:" if len(groups) == 1 else f"ИТОГО ПО ЛИСТУ «{title}»:", sheet_total),
        ], col_count)
        per_sheet.append((title, sheet_total))

    grand_total = round(sum(total for _, total in per_sheet), 2)

    if len(groups) > 1:
        # Сумма контракта — одно число, и искать его по последнему листу нельзя:
        # листы равноправны. Отдельный лист со сводкой отвечает на вопрос «сколько
        # всего» сразу, а `extract_total_cost` берёт из него же строку «ВСЕГО».
        ws_total = wb.create_sheet(SUMMARY_SHEET_TITLE)
        from openpyxl.utils import get_column_letter

        widths = ([_EST_SOURCE_NO_WIDTH] if with_source_no else []) + _EST_COL_WIDTHS
        for i, width in enumerate(widths, 1):
            ws_total.column_dimensions[get_column_letter(i)].width = width
        header = ws_total.cell(row=1, column=1, value="Лист")
        header.font = Font(bold=True, color="FFFFFF")
        header.fill = _HEADER_FILL_EST
        ws_total.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count - 1)
        header_total = ws_total.cell(row=1, column=col_count, value="Итого, руб.")
        header_total.font = Font(bold=True, color="FFFFFF")
        header_total.fill = _HEADER_FILL_EST

        _write_totals_block(
            ws_total, 2,
            [(title, total) for title, total in per_sheet] + [("ВСЕГО ПО СМЕТЕ:", grand_total)],
            col_count,
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), grand_total
