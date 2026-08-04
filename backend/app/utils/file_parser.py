import base64
import io
import re
import xml.etree.ElementTree as ET
from typing import Any, Optional, Tuple
import openpyxl
import structlog

from app.utils.sheet_names import group_by_sheet

logger = structlog.get_logger()


# ---------------------------------------------------------------------------
# Ключевые слова для определения колонок гранд-сметы
# ---------------------------------------------------------------------------

_NAME_KEYWORDS = ["наимен"]
_UNIT_KEYWORDS = ["ед", "изм", "мер"]  # ед. изм. / единица измерения
_QTY_KEYWORDS = ["колич", "объем", "объём", "кол-во", "кол.", "кол "]
# Ключевое слово для колонки "всего с учётом коэффициентов" (итоговое кол-во)
_QTY_TOTAL_KEYWORDS = ["всего с учет", "всего с коэф"]

# Слова в строке-наименовании, указывающие на итоговую/служебную строку (пропускаем)
_TOTAL_KEYWORDS = re.compile(
    r"итого|всего|в том числе|накладные|сметная прибыль|нр\b|сп\b|зп\b|"
    r"поправочн|индекс|лимитир|непредвиден|ндс|налог|фот\b",
    re.IGNORECASE,
)

# Паттерны мусорных строк Гранд-Сметы: нормо-часы труда и машин (3Т, 3ТМ, ЗТ, ЗТМ)
_LABOR_ROW = re.compile(r"^[3з]\s*[тТ][мМ]?\s*$", re.IGNORECASE)
# Одиночные цифры 1-15 (строка с номерами колонок в заголовке Гранд-Сметы)
_COLNUM_ROW = re.compile(r"^\d{1,2}$")
# Составляющие затрат Гранд-Сметы: «1 ОТ», «2 ЭМ», «3 в т.ч. ОТм», «4 М» и аналогичные
# (номер + аббревиатура или номер + «в т.ч.»)
_COMPONENT_ROW = re.compile(r"^\d+\s+(ОТ|ЭМ|М|ОТм|в т\.ч\.)", re.IGNORECASE)
# Нормативные коды расценок (ФЕР, ТЕР, ГЭСН и т.п.) — строки-наименования позиций
_NORM_CODE = re.compile(r"^(фер|тер|гэсн|фсн|фснб|пр/|ерер|гэснр|фсем)", re.IGNORECASE)
# Заголовки разделов/глав сметы.
# Слов «блок», «узел», «часть», «отдел» здесь быть не должно: с них начинаются
# наименования материалов («Блок дверной деревянный…», «Узел учёта»), и такая
# позиция уезжала в перечень заголовком раздела — без единицы, без объёма и с
# прямым указанием промпта «заголовки разделов пропускай». Материал исчезал.
_SECTION_HEADER = re.compile(r"^(раздел|подраздел|глава)\s", re.IGNORECASE)


def _col_score(header: str, keywords: list[str]) -> int:
    """Количество ключевых слов из списка, найденных в заголовке."""
    h = header.lower()
    return sum(1 for kw in keywords if kw in h)


def _find_header_row(all_rows) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int], Optional[int]]:
    """
    Сканируем первые 40 строк листа в поиске строки с заголовками колонок.
    Также сканируем следующие строки после найденного заголовка, чтобы найти
    колонку «всего с учётом коэффициентов» (в Гранд-Смете она в строке подзаголовков).
    Возвращает (header_row_idx, name_col, unit_col, qty_col, qty_total_col) — 0-based.
    qty_total_col — колонка итогового количества (или None если не найдена).

    Принимает УЖЕ прочитанные строки, а не лист: раньше функция материализовала
    их сама, и весь файл оказывался в памяти дважды (второй раз — в
    `parse_xlsx_grand`). На гранд-смете в 20 000 строк это лишние ~50 МБ на каждую
    задачу — при трёх задачах разом контейнер выходил за лимит и получал OOM-kill
    (plans/2026-07-30-parallelnaya-obrabotka-umiraet.md).
    """
    scan_limit = min(80, len(all_rows))

    best_row_idx = None
    best_score = 0
    best_cols = (None, None, None)

    for row_idx in range(scan_limit):
        row = all_rows[row_idx]
        cells = [str(c).strip() if c is not None else "" for c in row]

        name_col = unit_col = qty_col = None
        name_score = unit_score = qty_score = 0

        for col_idx, cell in enumerate(cells):
            if not cell or cell == "None":
                continue

            ns = _col_score(cell, _NAME_KEYWORDS)
            us = _col_score(cell, _UNIT_KEYWORDS)
            qs = _col_score(cell, _QTY_KEYWORDS)

            if ns > name_score:
                name_score = ns
                name_col = col_idx
            if us > unit_score:
                unit_score = us
                unit_col = col_idx
            if qs > qty_score:
                qty_score = qs
                qty_col = col_idx

        # Строка-заголовок: нашли хотя бы колонку с наименованием
        row_score = name_score + unit_score + qty_score
        if name_score > 0 and row_score > best_score:
            best_score = row_score
            best_row_idx = row_idx
            best_cols = (name_col, unit_col, qty_col)

    if best_row_idx is None:
        return (None, None, None, None, None)

    name_col, unit_col, qty_col = best_cols

    # Ищем колонку итогового количества в строках подзаголовков после основного заголовка
    # В Гранд-Смете строка подзаголовков содержит «всего с учётом коэффициентов»
    qty_total_col = None
    for sub_idx in range(best_row_idx + 1, min(best_row_idx + 5, len(all_rows))):
        sub_cells = [str(c).strip().lower() if c is not None else "" for c in all_rows[sub_idx]]
        for col_idx, cell in enumerate(sub_cells):
            if any(kw in cell for kw in _QTY_TOTAL_KEYWORDS):
                qty_total_col = col_idx
                break
        if qty_total_col is not None:
            break

    return (best_row_idx, name_col, unit_col, qty_col, qty_total_col)


def _is_total_row(name: str) -> bool:
    """True если строка — итог/суммарная, её нужно пропустить."""
    return bool(_TOTAL_KEYWORDS.search(name))


def _parse_quantity(value) -> Optional[float]:
    """Преобразуем значение ячейки в число или None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").replace(" ", "").strip())
    except (ValueError, TypeError):
        return None


def _is_trash_row(name: str) -> bool:
    """True для мусорных строк Гранд-Сметы: нормо-часы, номера колонок, составляющие затрат."""
    return bool(_LABOR_ROW.match(name) or _COLNUM_ROW.match(name) or _COMPONENT_ROW.match(name))


def _parse_grand_sheet(ws) -> "list[dict]":
    """Значимые строки одного листа гранд-сметы. Шапка ищется в нём же."""
    all_rows = list(ws.iter_rows(values_only=True))
    header_row_idx, name_col, unit_col, qty_col, qty_total_col = _find_header_row(all_rows)
    if header_row_idx is None or name_col is None:
        return []

    # Используем колонку итогового количества если найдена, иначе базовую
    effective_qty_col = qty_total_col if qty_total_col is not None else qty_col
    logger.info(
        "Grand-смета: заголовок найден",
        sheet=ws.title,
        header_row=header_row_idx,
        name_col=name_col,
        unit_col=unit_col,
        qty_col=qty_col,
        qty_total_col=qty_total_col,
    )

    rows: list[dict] = []
    for row in all_rows[header_row_idx + 1:]:
        cells = [str(c).strip() if c is not None else "" for c in row]

        name = cells[name_col] if name_col < len(cells) else ""
        if not name or name == "None":
            # Проверяем col 0 — там могут быть заголовки разделов (объединённые ячейки)
            col0 = cells[0] if cells else ""
            if col0 and col0 != "None" and _SECTION_HEADER.match(col0):
                rows.append({"name": col0, "unit": "", "quantity": None, "is_section": True})
            continue

        if _is_total_row(name):
            continue
        if _is_trash_row(name):
            continue

        # Заголовки разделов: нет единицы и количества, текст похож на раздел
        unit = cells[unit_col] if unit_col is not None and unit_col < len(cells) else ""
        unit = unit if unit != "None" else ""
        qty_raw = row[effective_qty_col] if effective_qty_col is not None and effective_qty_col < len(row) else None
        qty = _parse_quantity(qty_raw)

        # Пропускаем строки с процентными единицами (НР/СП не отфильтрованные по имени)
        if unit.strip() == "%":
            continue

        # Пропускаем строки с единицами нормо-часов (остатки труда/машин)
        unit_lower = unit.lower().replace(".", "").replace("-", "").strip()
        if unit_lower in ("челч", "чч", "мчч", "машч"):
            continue

        # Заголовок раздела не имеет ни единицы измерения, ни объёма.
        # Строка с ними — позиция сметы, как бы она ни называлась.
        is_section = bool(_SECTION_HEADER.match(name)) and not unit and qty is None
        rows.append({"name": name, "unit": unit, "quantity": qty, "is_section": is_section})

    return rows


def _parse_grand_sheet_fallback(ws) -> "list[dict]":
    """Шапку не нашли — берём все непустые строки, первая ячейка = наименование."""
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        non_empty = [c for c in cells if c and c != "None"]
        if not non_empty:
            continue
        name = non_empty[0]
        if _is_total_row(name) or _is_trash_row(name):
            continue
        rows.append({"name": name, "unit": "", "quantity": None, "is_section": False})
    return rows


def parse_xlsx_grand(data: bytes) -> "list[dict]":
    """
    Умный парсинг Excel-файла гранд-сметы.

    Возвращает список словарей {name, unit, quantity, is_section, sheet} — только
    значимые строки с работами/материалами и заголовки разделов.
    Без итоговых строк, строк НР/СП/ФОТ, нормо-часов и пустых строк.

    Разбираются все листы файла: гранд-смету присылают разбитой по листам — по
    листу на раздел или корпус, — и взятый один первый лист терял остальные, а
    задача при этом завершалась успешно. Лист остаётся при строке полем `sheet`.

    Если колонки определить не удалось ни на одном листе — возвращает весь текст
    построчно в поле name (fallback-режим).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

        # Лист без шапки строк не даёт: титульный лист и лист подписей есть почти
        # в каждой выгрузке, и их содержимое в перечне — мусор.
        by_sheet = [(ws, _parse_grand_sheet(ws)) for ws in wb.worksheets]

        if not any(sheet_rows for _, sheet_rows in by_sheet):
            # Шапку не нашли нигде — прежний fallback, теперь по всем листам.
            logger.warning("Grand-смета: заголовок не найден, fallback-режим")
            by_sheet = [(ws, _parse_grand_sheet_fallback(ws)) for ws in wb.worksheets]

        multi_sheet = sum(1 for _, sheet_rows in by_sheet if sheet_rows) > 1
        rows: list[dict] = []
        for ws, sheet_rows in by_sheet:
            for row in sheet_rows:
                # Один лист — признак не ставим: документ остаётся без вкладок и
                # ведёт себя ровно как до появления многолистовых файлов.
                if multi_sheet:
                    row["sheet"] = ws.title
                rows.append(row)

        work_rows = sum(1 for r in rows if not r.get("is_section") and r.get("unit"))
        logger.info(
            "Grand-смета: строк извлечено",
            count=len(rows),
            work_rows=work_rows,
            sheets=[ws.title for ws, sheet_rows in by_sheet if sheet_rows],
        )
        if rows and work_rows == 0:
            logger.warning(
                "Grand-смета: не найдено ни одной строки с работами/материалами — "
                "возможно, неверно определена колонка наименования",
            )
        return rows

    except Exception as e:
        logger.error("Ошибка парсинга гранд-сметы", error=str(e))
        return []


def _chunk_one_sheet(rows: "list[dict]", chunk_size: int) -> "list[list[dict]]":
    chunks = []
    start = 0
    total = len(rows)

    while start < total:
        end = min(start + chunk_size, total)

        # Пытаемся не резать на полуслове: откатываемся до строки без quantity
        # (скорее всего это начало новой работы), но не более чем на 20 строк
        if end < total:
            for lookback in range(end - 1, max(start, end - 20) - 1, -1):
                if rows[lookback].get("quantity") is None and rows[lookback].get("unit") == "":
                    end = lookback
                    break

        chunks.append(rows[start:end])
        start = end

    return chunks


def chunk_rows(rows: "list[dict]", chunk_size: int = 250) -> "list[list[dict]]":
    """
    Делит список строк на чанки по chunk_size.
    По возможности не разрывает группу: если последняя строка чанка —
    не «Работа» (нет unit/quantity) — сдвигаем границу назад до ближайшей
    строки без unit (предполагаем работу).

    Чанк не пересекает границу листа: позиции чанка получают лист этого чанка,
    и смешанный чанк отправил бы половину строк на чужую вкладку.
    """
    if not rows:
        return []

    chunks: list = []
    for _, sheet_rows in group_by_sheet(rows):
        chunks.extend(_chunk_one_sheet(sheet_rows, chunk_size))
    return chunks


def rows_to_text(rows: list[dict]) -> str:
    """Компактное текстовое представление строк для отправки в Claude."""
    lines = []
    for r in rows:
        if r.get("is_section"):
            lines.append(f"\n=== {r['name']} ===")
        else:
            qty = str(r["quantity"]) if r["quantity"] is not None else ""
            unit = r.get("unit", "")
            lines.append(f"{r['name']}\t{unit}\t{qty}")
    return "\n".join(lines)


def parse_xlsx(data: bytes) -> str:
    """Convert xlsx bytes to readable text representation."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        result_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result_parts.append(f"=== Лист: {sheet_name} ===")

            rows_data = []
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if all(cell is None for cell in row):
                    continue
                row_str = "\t".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                rows_data.append(row_str)

            result_parts.extend(rows_data)
            result_parts.append("")

        return "\n".join(result_parts)
    except Exception as e:
        logger.error("Failed to parse xlsx", error=str(e))
        return f"[Ошибка при разборе Excel файла: {e}]"


def parse_xml(data: bytes) -> str:
    """Parse XML bytes and return formatted text."""
    try:
        # Try to decode as text first
        text = data.decode("utf-8", errors="replace")
        # Try to parse and re-format if it's valid XML
        try:
            root = ET.fromstring(data)
            return _xml_to_text(root)
        except ET.ParseError:
            return text
    except Exception as e:
        logger.error("Failed to parse XML", error=str(e))
        return f"[Ошибка при разборе XML файла: {e}]"


def _xml_to_text(element: ET.Element, indent: int = 0) -> str:
    """Recursively convert XML element to readable text."""
    prefix = "  " * indent
    parts = []

    tag = element.tag
    # Strip namespace
    if "}" in tag:
        tag = tag.split("}")[1]

    attrs = ""
    if element.attrib:
        attr_parts = [f'{k}="{v}"' for k, v in element.attrib.items()]
        attrs = " " + " ".join(attr_parts)

    text = (element.text or "").strip()

    if text:
        parts.append(f"{prefix}<{tag}{attrs}> {text}")
    else:
        parts.append(f"{prefix}<{tag}{attrs}>")

    for child in element:
        parts.append(_xml_to_text(child, indent + 1))

    return "\n".join(parts)


# Claude Vision API отклоняет изображения, у которых любая сторона > 8000 px.
# Берём с запасом для крупных сканов.
_MAX_IMAGE_DIMENSION = 7500


def _downscale_if_needed(data: bytes, mime_type: str) -> tuple[bytes, str]:
    """Уменьшает изображение, если любая сторона превышает _MAX_IMAGE_DIMENSION px.

    Возвращает (bytes, mime_type). При любой ошибке/неподдерживаемом формате —
    отдаёт исходные данные без изменений.
    """
    if mime_type not in ("image/jpeg", "image/png"):
        return data, mime_type
    try:
        from io import BytesIO

        from PIL import Image

        with Image.open(BytesIO(data)) as img:
            if max(img.size) <= _MAX_IMAGE_DIMENSION:
                return data, mime_type
            scale = _MAX_IMAGE_DIMENSION / max(img.size)
            new_size = (max(1, round(img.width * scale)), max(1, round(img.height * scale)))
            resized = img.resize(new_size, Image.LANCZOS)
            buf = BytesIO()
            if mime_type == "image/png":
                resized.save(buf, format="PNG")
            else:
                resized.convert("RGB").save(buf, format="JPEG", quality=90)
            return buf.getvalue(), mime_type
    except Exception:
        return data, mime_type


def file_to_base64(data: bytes, mime_type: str) -> dict:
    """Return a base64-encoded content block suitable for Claude Vision."""
    data, mime_type = _downscale_if_needed(data, mime_type)
    encoded = base64.b64encode(data).decode("utf-8")
    return {
        "type": "image",
        "source": {
            "type": "base64",
            "media_type": mime_type,
            "data": encoded,
        },
    }


def extract_text_from_image_for_claude(data: bytes, mime_type: str) -> dict:
    """Return an image content block for Claude to extract text from."""
    return file_to_base64(data, mime_type)


_PDF_MAX_SIZE = 32 * 1024 * 1024  # 32 MB — Claude API hard limit


def pdf_to_content_block(data: bytes) -> dict:
    """Return a PDF document content block for Claude.

    Raises ValueError with a human-readable Russian message if the file
    exceeds the 32 MB API limit or does not start with the %PDF magic bytes.
    """
    if len(data) > _PDF_MAX_SIZE:
        raise ValueError("Файл слишком большой. Максимальный размер PDF: 32MB")
    if not data.startswith(b"%PDF"):
        raise ValueError("Файл не является валидным PDF")
    encoded = base64.b64encode(data).decode("utf-8")
    return {
        "type": "document",
        "source": {
            "type": "base64",
            "media_type": "application/pdf",
            "data": encoded,
        },
    }


def parse_file(name: str, mime_type: str, content_b64: str) -> Any:
    """
    Parse a file based on its MIME type.
    Returns either a text string or a content block dict for Claude.
    """
    data = base64.b64decode(content_b64)

    if mime_type in ("image/jpeg", "image/png", "image/gif", "image/webp"):
        return file_to_base64(data, mime_type)

    if mime_type == "application/pdf":
        return pdf_to_content_block(data)

    if mime_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return parse_xlsx(data)

    if mime_type == "text/xml" or mime_type == "application/xml" or name.endswith(".xml"):
        return parse_xml(data)

    # Fallback: try to decode as text
    try:
        return data.decode("utf-8", errors="replace")
    except Exception:
        return f"[Не удалось прочитать файл: {name}]"
