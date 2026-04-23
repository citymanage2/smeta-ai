"""Utilities for parsing and optimizing estimate xlsx files."""
import io
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill
from app.config import settings

# Heuristic keywords for "work" type detection
_WORK_KEYWORDS = ("монтаж", "устройство", "разборка", "прокладка", "установка",
                  "укладка", "демонтаж", "сборка", "покраска", "штукатурка",
                  "кладка", "бурение", "сварка", "резка")

# "Extra" row keywords — накладные расходы, итого, НДС
_EXTRA_KEYWORDS = ("накладные", "прибыль", "ндс", "итого", "сметная прибыль", "всего")

_GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")

def _detect_type(type_cell_value: Optional[str], name: str) -> str:
    """Determine item type from explicit column or heuristic."""
    if type_cell_value:
        v = type_cell_value.strip().lower()
        if v in ("работа", "работы", "work"):
            return "work"
        if v in ("материал", "материалы", "material"):
            return "material"
        if any(k in v for k in ("накладн", "ндс", "прибыль", "итого")):
            return "extra"

    name_lower = name.lower()
    if any(k in name_lower for k in _EXTRA_KEYWORDS):
        return "extra"
    if any(k in name_lower for k in _WORK_KEYWORDS):
        return "work"
    return "material"


def _find_header_row(ws) -> Optional[int]:
    """Find the header row index (1-based) in the first 10 rows."""
    for row_idx in range(1, min(11, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col).value
            if val and isinstance(val, str):
                lower = val.lower()
                if any(k in lower for k in ("наименование", "назван")):
                    return row_idx
    return None


def _map_columns(ws, header_row: int) -> dict:
    """Return mapping of logical field -> column index (1-based)."""
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if not val or not isinstance(val, str):
            continue
        lower = val.strip().lower()
        if "наименование" in lower or "назван" in lower:
            mapping.setdefault("name", col)
        elif "тип" in lower:
            mapping.setdefault("type", col)
        elif "ед" in lower and ("изм" in lower or "." in lower):
            mapping.setdefault("unit", col)
        elif "кол" in lower:
            mapping.setdefault("quantity", col)
        elif "итого с" in lower or ("итого" in lower and "ндс" in lower and "без" not in lower):
            mapping.setdefault("total_incl_vat", col)
        elif "итого без" in lower or ("итого" in lower and "без" in lower):
            mapping.setdefault("total_excl_vat", col)
        elif "ндс" in lower and "%" in lower:
            mapping.setdefault("vat_col", col)
        elif ("цена" in lower or "стоим" in lower) and ("работ" in lower or "труд" in lower):
            mapping.setdefault("work_price", col)
        elif ("цена" in lower or "стоим" in lower) and "матер" in lower:
            mapping.setdefault("material_price", col)
        elif "цена" in lower and "ед" in lower:
            mapping.setdefault("unit_price", col)
    return mapping


def _to_float(val) -> float:
    """Safely convert a cell value to float, returning 0.0 on failure."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def parse_estimate_xlsx(file_bytes: bytes) -> list[dict]:
    """
    Parse estimate xlsx. Finds header row by 'наименование' keyword in first 10 rows.
    Returns list of items: {row_index, name, type, quantity, unit,
                             price_excl_vat, price_incl_vat, total}
    Skips rows without a name value and rows matching 'extra' keywords (totals/НДС rows).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Не удалось открыть xlsx файл: {e}") from e
    ws = wb.active

    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError("Не удалось найти строку заголовков в xlsx (поиск по 'наименование')")

    cols = _map_columns(ws, header_row)
    if "name" not in cols:
        raise ValueError("Колонка 'Наименование' не найдена в xlsx")

    items = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=cols["name"])
        name = name_cell.value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()

        # Skip extra / total rows
        type_raw = ""
        if "type" in cols:
            type_val = ws.cell(row=row_idx, column=cols["type"]).value
            type_raw = str(type_val).strip() if type_val else ""

        item_type = _detect_type(type_raw, name)
        if item_type == "extra":
            continue

        quantity = 0.0
        if "quantity" in cols:
            q = ws.cell(row=row_idx, column=cols["quantity"]).value
            quantity = _to_float(q)

        unit = ""
        if "unit" in cols:
            u = ws.cell(row=row_idx, column=cols["unit"]).value
            unit = str(u).strip() if u else ""

        # Calculate price_excl_vat (per unit)
        price_excl_vat = 0.0
        if "work_price" in cols:
            wp = ws.cell(row=row_idx, column=cols["work_price"]).value
            price_excl_vat += _to_float(wp)
        if "material_price" in cols:
            mp = ws.cell(row=row_idx, column=cols["material_price"]).value
            price_excl_vat += _to_float(mp)

        # Fallback: if no work/material price cols, try unit_price col
        if price_excl_vat == 0.0 and "unit_price" in cols:
            up = ws.cell(row=row_idx, column=cols["unit_price"]).value
            price_excl_vat = _to_float(up)

        # Fallback: derive from total_excl_vat / quantity
        if price_excl_vat == 0.0 and "total_excl_vat" in cols and quantity:
            tv = ws.cell(row=row_idx, column=cols["total_excl_vat"]).value
            if tv:
                price_excl_vat = _to_float(tv) / quantity

        price_incl_vat = price_excl_vat * (1 + settings.VAT_RATE)

        # Total (incl VAT)
        total = 0.0
        if "total_incl_vat" in cols:
            t = ws.cell(row=row_idx, column=cols["total_incl_vat"]).value
            total = _to_float(t)
        if total == 0.0:
            total = quantity * price_incl_vat

        items.append({
            "row_index": row_idx,
            "name": name,
            "type": item_type,
            "quantity": quantity,
            "unit": unit,
            "price_excl_vat": round(price_excl_vat, 4),
            "price_incl_vat": round(price_incl_vat, 4),
            "total": round(total, 2),
        })

    return items


def get_top_items(items: list[dict], categories: list[str], threshold: float = 0.7) -> list[dict]:
    """
    Filter by categories, sort by total descending, return items until
    cumulative sum crosses threshold * grand_total.
    All returned items have selected=True added.
    """
    filtered = [it for it in items if it["type"] in categories]
    if not filtered:
        return []

    grand_total = sum(it["total"] for it in filtered)
    if grand_total == 0:
        return []

    sorted_items = sorted(filtered, key=lambda x: x["total"], reverse=True)
    target = threshold * grand_total
    cumulative = 0.0
    result = []
    for it in sorted_items:
        cumulative += it["total"]
        result.append({**it, "selected": True})
        if cumulative >= target:
            break

    return result


def generate_optimized_xlsx(original_bytes: bytes, optimization_results: list[dict]) -> bytes:
    """
    Open original xlsx, add 4 columns to first sheet:
      'Цена сниженная', 'Стоимость сниженная', 'Источник', 'Примечание'
    Apply green fill (#E2EFDA) for found analogues, yellow (#FFEB9C) for 'Не найдено'.
    Add sheet 'Сравнение' with before/after comparison table.
    Returns modified xlsx as bytes.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    except Exception as e:
        raise ValueError(f"Не удалось открыть xlsx файл для оптимизации: {e}") from e
    ws = wb.active

    # Find actual header row (may not be row 1)
    header_row = _find_header_row(ws) or 1

    # Find quantity column in original data (before appending new columns)
    qty_col = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if h and isinstance(h, str) and "кол" in h.lower():
            qty_col = c
            break

    # Find how many columns exist and append new ones
    last_col = ws.max_column
    new_price_col = last_col + 1
    new_total_col = last_col + 2
    source_col = last_col + 3
    note_col = last_col + 4

    ws.cell(row=header_row, column=new_price_col, value="Цена сниженная")
    ws.cell(row=header_row, column=new_total_col, value="Стоимость сниженная")
    ws.cell(row=header_row, column=source_col, value="Источник")
    ws.cell(row=header_row, column=note_col, value="Примечание")

    for opt in optimization_results:
        row_idx = opt["row_index"]
        new_price = opt.get("new_price")
        source = opt.get("source", "Не найдено")

        fill = _GREEN_FILL if new_price is not None else _YELLOW_FILL

        # Determine quantity from the row using pre-scanned qty_col
        qty = None
        if qty_col is not None:
            qty = ws.cell(row=row_idx, column=qty_col).value

        new_total = None
        if new_price is not None and qty:
            try:
                new_total = round(float(new_price) * float(qty), 2)
            except (TypeError, ValueError):
                pass

        price_cell = ws.cell(row=row_idx, column=new_price_col)
        price_cell.value = round(new_price, 4) if new_price is not None else "Не найдено"
        price_cell.fill = fill

        total_cell = ws.cell(row=row_idx, column=new_total_col)
        total_cell.value = round(new_total, 2) if new_total is not None else None
        total_cell.fill = fill

        source_cell = ws.cell(row=row_idx, column=source_col)
        source_cell.value = source
        source_cell.fill = fill

        note_cell = ws.cell(row=row_idx, column=note_col)
        savings_pct = opt.get("savings_pct")
        if savings_pct is not None:
            note_cell.value = f"Экономия {savings_pct:.1f}%"
        note_cell.fill = fill

    # Add comparison sheet
    ws_cmp = wb.create_sheet("Сравнение")
    cmp_headers = ["Наименование", "Цена было", "Цена стало",
                   "Экономия на ед.", "Экономия %", "Источник"]
    for col, h in enumerate(cmp_headers, start=1):
        ws_cmp.cell(row=1, column=col, value=h)

    total_savings = 0.0
    total_original = 0.0

    for i, opt in enumerate(optimization_results, start=2):
        ws_cmp.cell(row=i, column=1, value=opt.get("name", ""))
        ws_cmp.cell(row=i, column=2, value=round(opt.get("original_price", 0), 4))
        new_price = opt.get("new_price")
        ws_cmp.cell(row=i, column=3, value=round(new_price, 4) if new_price is not None else "Не найдено")
        savings_abs = opt.get("savings_abs")
        ws_cmp.cell(row=i, column=4, value=round(savings_abs, 4) if savings_abs is not None else None)
        savings_pct = opt.get("savings_pct")
        ws_cmp.cell(row=i, column=5, value=round(savings_pct, 2) if savings_pct is not None else None)
        ws_cmp.cell(row=i, column=6, value=opt.get("source", ""))
        if savings_abs is not None:
            total_savings += savings_abs
        total_original += opt.get("original_price", 0)

    # Summary row
    summary_row = len(optimization_results) + 2
    ws_cmp.cell(row=summary_row, column=1, value="ИТОГО экономия")
    ws_cmp.cell(row=summary_row, column=4, value=round(total_savings, 2))
    if total_original > 0:
        ws_cmp.cell(row=summary_row, column=5, value=round(total_savings / total_original * 100, 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
