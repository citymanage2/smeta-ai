import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import structlog
from app.config import settings
from app.utils.price_coercion import coerce_price, coerce_qty
from app.utils.sheet_names import group_by_sheet, safe_sheet_title

logger = structlog.get_logger()

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Row highlight fills
FILL_ADDED      = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # новая позиция
FILL_ADJUSTED   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # скорректированный объём
FILL_UNKNOWN    = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # объём неизвестен
FILL_CALCULATED = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")  # объём рассчитан по чертежам

_CALCULATED_NOTES_MARKERS = (
    "определён по чертежу",
    "определён по документу",
    "подсчитано по",
    "рассчитано по чертежам",
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Перечень sheet style constants ───────────────────────────────────────────
_P_FONT_BLACK = Font(name="Arial", size=9, color="000000")
_P_FONT_BLACK_BOLD = Font(name="Arial", size=9, bold=True, color="000000")
_P_FONT_BLUE_ITALIC = Font(name="Arial", size=9, italic=True, color="4284F3")
_P_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
_P_COL_WIDTHS = [6.5, 10.33, 67.33, 10.33, 10.33, 57.5]
_P_HEADERS = ["№ п/п", "Тип", "Наименование", "Ед. изм", "Кол-во", "Примечание"]
_P_QTY_FMT = "0.00"

# Номер позиции исходной сметы — им менеджер сверяет перечень со сметой
# заказчика построчно. Колонка необязательная: у перечня из проекта и у
# PDF-скана исходного номера нет, и пустой столбец им ни к чему.
SOURCE_NO_HEADER = "№ в исходной смете"
_P_SOURCE_NO_WIDTH = 12.0


def has_source_numbers(items: list) -> bool:
    """True, если хоть одна позиция знает свой номер в исходной смете."""
    return any(str((it or {}).get("source_no") or "").strip() for it in items or [])


def _source_no_value(item: dict):
    """Номер для ячейки: «12» числом, «1.1» и «2а» — текстом, как в смете."""
    number = str(item.get("source_no") or "").strip()
    if not number:
        return None
    return int(number) if number.isdigit() else number


def _write_perechen_sheet(ws, items: list, with_sections: bool = False,
                          with_source_no: bool = False) -> None:
    """Fill worksheet with the standard Перечень format (6 columns).

    `with_source_no` добавляет первой колонку «№ в исходной смете». Решение
    принимается один раз на весь файл (`generate_list`), а не по листу: сводки
    «Работы» и «Материалы» обязаны иметь тот же набор колонок, что и данные.
    """
    headers = ([SOURCE_NO_HEADER] if with_source_no else []) + _P_HEADERS
    widths = ([_P_SOURCE_NO_WIDTH] if with_source_no else []) + _P_COL_WIDTHS
    # Сдвиг колонок из-за необязательного первого столбца.
    off = 1 if with_source_no else 0
    wrapped = {3 + off, 6 + off}

    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Row 1: headers — bold, center, Arial 9
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _P_FONT_BLACK_BOLD
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=(col in wrapped or col == 1 and with_source_no))
        c.border = _P_BORDER

    # Строки нумерации колонок («1 2 3 4 5 6») здесь нет намеренно. В Гранд-смете
    # она часть шапки, а у нас лист уезжает в редактор: там она садилась первой
    # строкой данных и выглядела мусором. Разбор без неё работает — шапкой
    # считается первая строка листа.
    data_row = 2
    item_num = 0
    current_section = object()  # sentinel — no section seen yet
    section_num = 0

    for item in items:
        item_type = str(item.get("type", "")).strip()
        is_material = item_type.lower() in ("материал", "material", "материалы")

        # Section header row (only when with_sections and section changes)
        if with_sections:
            sec = item.get("section", "") or ""
            if sec != current_section:
                current_section = sec
                section_num += 1
                label = f"Раздел {section_num}. {sec}" if sec else f"Раздел {section_num}"
                for col in range(1, 7 + off):
                    c = ws.cell(row=data_row, column=col,
                                value=label if col == 3 + off else None)
                    c.font = _P_FONT_BLACK_BOLD if col == 3 + off else _P_FONT_BLACK
                    c.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=(col == 3 + off))
                    c.border = _P_BORDER
                data_row += 1

        item_num += 1

        # Col 0 (опционально): номер позиции в исходной смете — center
        if with_source_no:
            c0 = ws.cell(row=data_row, column=1, value=_source_no_value(item))
            c0.font = _P_FONT_BLACK
            c0.alignment = Alignment(horizontal="center", vertical="center")
            c0.border = _P_BORDER

        # Col 1: порядковый номер — integer, center
        c1 = ws.cell(row=data_row, column=1 + off, value=item_num)
        c1.font = _P_FONT_BLACK
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.number_format = "0"
        c1.border = _P_BORDER

        # Col 2: тип — center; materials: blue italic
        c2 = ws.cell(row=data_row, column=2 + off, value=item_type)
        c2.font = _P_FONT_BLUE_ITALIC if is_material else _P_FONT_BLACK
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = _P_BORDER

        # Col 3: наименование — left (work) / right (material), wrap
        name = str(item.get("name", "") or "")
        c3 = ws.cell(row=data_row, column=3 + off, value=name)
        c3.font = _P_FONT_BLUE_ITALIC if is_material else _P_FONT_BLACK
        c3.alignment = Alignment(
            horizontal="right" if is_material else "left",
            vertical="center", wrap_text=True,
        )
        c3.border = _P_BORDER

        # Col 4: ед. изм. — center (work) / right (material)
        unit = str(item.get("unit", "") or "")
        c4 = ws.cell(row=data_row, column=4 + off, value=unit)
        c4.font = _P_FONT_BLUE_ITALIC if is_material else _P_FONT_BLACK
        c4.alignment = Alignment(
            horizontal="right" if is_material else "center",
            vertical="center",
        )
        c4.border = _P_BORDER

        # Col 5: кол-во — center (work) / right (material), 2 decimals
        qty = item.get("quantity")
        c5 = ws.cell(row=data_row, column=5 + off, value=qty)
        c5.font = _P_FONT_BLUE_ITALIC if is_material else _P_FONT_BLACK
        c5.alignment = Alignment(
            horizontal="right" if is_material else "center",
            vertical="center",
        )
        if qty is not None:
            c5.number_format = _P_QTY_FMT
        c5.border = _P_BORDER

        # Col 6: примечание — left, black, wrap
        notes_text = str(item.get("notes", "") or "")
        if _is_calculated_from_drawing(item) and qty is not None:
            notes_text = (f"рассчитано по чертежам | {notes_text}"
                          if notes_text else "рассчитано по чертежам")
        c6 = ws.cell(row=data_row, column=6 + off, value=notes_text)
        c6.font = _P_FONT_BLACK
        c6.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c6.border = _P_BORDER

        data_row += 1


def _is_calculated_from_drawing(item: dict) -> bool:
    """True если объём рассчитан по чертежам/документу (не взят из спецификации напрямую)."""
    if item.get("_calculated"):
        return True
    notes = str(item.get("notes") or "").lower()
    return any(marker in notes for marker in _CALCULATED_NOTES_MARKERS)


def _row_fill(item: dict) -> Optional[PatternFill]:
    """Return fill for a data row based on item status."""
    notes = str(item.get("notes") or "").lower()
    qty = item.get("quantity")
    if qty is None:
        return FILL_UNKNOWN
    if "добавлен" in notes:
        return FILL_ADDED
    if "скорректирован" in notes:
        return FILL_ADJUSTED
    if _is_calculated_from_drawing(item):
        return FILL_CALCULATED
    return None


def _apply_row_fill(ws, row: int, col_count: int, fill: Optional[PatternFill]) -> None:
    if fill is None:
        return
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).fill = fill


def _style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, col_count: int, bold: bool = False) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BOLD_FONT if bold else NORMAL_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# Листы перечня. «Перечень» — единственный лист данных, когда исходный файл был
# из одного листа; «Прочее» собирает позиции, у которых лист не проставлен, — их
# нельзя молча подмешать к чужому разделу.
LIST_SHEET_TITLE = "Перечень"
NOTE_SHEET_TITLE = "Пояснительная записка"
UNSORTED_SHEET_TITLE = "Прочее"


def _list_sheet_groups(items: list) -> list:
    """Листы данных перечня: `[(имя листа, позиции), ...]` в порядке появления.

    Исходный файл из одного листа даёт прежний единственный «Перечень» — иначе
    у всех существующих смет разъехались бы имена листов. Несколько листов —
    лист на исходный лист, имя то же самое.
    """
    groups = group_by_sheet(items)
    named = [title for title, _ in groups if title is not None]
    if len(named) < 2:
        return [(LIST_SHEET_TITLE, list(items or []))]

    used: set = set()
    return [
        (safe_sheet_title(title if title is not None else UNSORTED_SHEET_TITLE, used), group)
        for title, group in groups
    ]


def data_sheet_titles(items: list) -> list:
    """Имена листов данных, которые создаст `generate_list` для этих позиций.

    Нужны разбору результата: в файле рядом с данными лежат сводки «Работы» и
    «Материалы» и пояснительная записка, и без этого списка документ получил бы
    вкладки-двойники.
    """
    return [title for title, _ in _list_sheet_groups(items)]


def generate_list(items: list, changes_summary: Optional[str] = None) -> bytes:
    """
    Generate Excel file with list of works/materials.
    items: list of dicts with keys: type, name, unit, quantity, sheet
    changes_summary: optional explanatory text about deviations from TZ

    Позиции, размеченные листами исходного файла, раскладываются по листам с
    теми же именами. Сводки «Работы» и «Материалы» при этом не создаются: из
    пяти разделов получилось бы шестнадцать листов, в которых не разобраться.

    Колонка «№ в исходной смете» появляется первой, если номер известен хоть у
    одной позиции. Решение общее на весь файл: у листов одного файла набор
    колонок обязан совпадать.
    """
    wb = openpyxl.Workbook()

    groups = _list_sheet_groups(items)
    single_sheet = len(groups) == 1
    with_source_no = has_source_numbers(items)

    for index, (title, group) in enumerate(groups):
        ws = wb.active if index == 0 else wb.create_sheet()
        ws.title = title
        _write_perechen_sheet(ws, group, with_sections=False, with_source_no=with_source_no)
        ws.freeze_panes = "A2"

    if single_sheet:
        # Sheet 2: Работы
        works = [it for it in items if it.get("type", "").lower() in ("работа", "work", "работы")]
        ws_works = wb.create_sheet("Работы")
        _write_perechen_sheet(ws_works, works, with_sections=False, with_source_no=with_source_no)
        ws_works.freeze_panes = "A2"

        # Sheet 3: Материалы
        materials = [it for it in items if it.get("type", "").lower() in ("материал", "material", "материалы")]
        ws_mats = wb.create_sheet("Материалы")
        _write_perechen_sheet(ws_mats, materials, with_sections=False, with_source_no=with_source_no)
        ws_mats.freeze_panes = "A2"

    # Последний лист: Пояснительная записка
    ws_note = wb.create_sheet(NOTE_SHEET_TITLE)
    note_text = (
        changes_summary
        if changes_summary
        else "Перечень соответствует ТЗ, дополнений не требуется"
    )
    ws_note.cell(row=1, column=1, value="Пояснительная записка").font = BOLD_FONT
    ws_note.cell(row=1, column=1).fill = HEADER_FILL
    ws_note.cell(row=1, column=1).font = HEADER_FONT
    ws_note.row_dimensions[1].height = 30

    # Write text wrapped across rows (split by newline for readability)
    for row_offset, line in enumerate(note_text.splitlines(), start=2):
        cell = ws_note.cell(row=row_offset, column=1, value=line)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_note.column_dimensions["A"].width = 120
    ws_note.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_list_project(items: list, changes_summary: Optional[str] = None) -> bytes:
    """
    Generate Excel for LIST_FROM_TZ_PROJECT tasks.
    Same 3 item sheets as generate_list, plus a 2-section "Пояснительная записка" sheet
    that splits changes_summary at "Раздел 2".
    """
    wb = openpyxl.Workbook()

    # Sheet 1: Перечень (All items) — sections shown as header rows
    ws_all = wb.active
    ws_all.title = "Перечень"
    _write_perechen_sheet(ws_all, items, with_sections=True)
    ws_all.freeze_panes = "A2"

    # Sheet 2: Работы
    works = [it for it in items if it.get("type", "").lower() in ("работа", "work", "работы")]
    ws_works = wb.create_sheet("Работы")
    _write_perechen_sheet(ws_works, works, with_sections=False)
    ws_works.freeze_panes = "A2"

    # Sheet 3: Материалы
    materials = [it for it in items if it.get("type", "").lower() in ("материал", "material", "материалы")]
    ws_mats = wb.create_sheet("Материалы")
    _write_perechen_sheet(ws_mats, materials, with_sections=False)
    ws_mats.freeze_panes = "A2"

    # Sheet 4: Пояснительная записка (two sections)
    ws_note = wb.create_sheet("Пояснительная записка")
    ws_note.column_dimensions["A"].width = 120

    FALLBACK = "Документация соответствует друг другу, дополнений не требуется"

    if not changes_summary:
        # No summary — write fallback under a single header
        _write_note_section(ws_note, 1, "Пояснительная записка", FALLBACK)
    else:
        # Split on "Раздел 2" (case-insensitive, strip surrounding whitespace)
        import re
        split_match = re.search(r"(?i)раздел\s*2", changes_summary)
        if split_match:
            part1 = changes_summary[: split_match.start()].strip()
            part2 = changes_summary[split_match.start() :].strip()
        else:
            part1 = changes_summary.strip()
            part2 = ""

        current_row = _write_note_section(
            ws_note, 1,
            "Раздел 1 — Сравнение ТЗ и проекта",
            part1 or FALLBACK,
        )
        if part2:
            _write_note_section(
                ws_note, current_row + 1,
                "Раздел 2 — Изменения по нормативной базе",
                part2,
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_note_section(ws, start_row: int, title: str, body: str) -> int:
    """
    Write a bold section header at start_row, then body lines below it.
    Returns the last row written.
    """
    # Section header
    header_cell = ws.cell(row=start_row, column=1, value=title)
    header_cell.font = BOLD_FONT
    header_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[start_row].height = 22

    current_row = start_row + 1
    for line in body.splitlines():
        cell = ws.cell(row=current_row, column=1, value=line)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        current_row += 1

    return current_row


def generate_smeta_from_project(
    smeta_items: list,
    list_items: list,
    research_result: Optional[str] = None,
    changes_summary: Optional[str] = None,
) -> bytes:
    """
    Generate 5-sheet Excel for SMETA_FROM_PROJECT (3-stage) tasks.
    Sheet 1 "Смета"                      — priced estimate (Stage 3 items)
    Sheet 2 "Перечень работ и материалов" — all items from Stage 2
    Sheet 3 "Перечень работ"              — works from Stage 2
    Sheet 4 "Перечень материалов"         — materials from Stage 2
    Sheet 5 "Пояснительная записка"       — research_result + changes_summary
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Смета ─────────────────────────────────────────────────────
    ws_smeta = wb.active
    ws_smeta.title = "Смета"

    smeta_headers = [
        "№", "Тип", "Наименование", "Ед. изм.", "Кол-во",
        "Цена работы (за ед.)", "Цена материала (за ед.)",
        "Стоимость работ", "Стоимость материалов",
        "Итого без НДС", f"НДС ({int(settings.VAT_RATE * 100)}%)", "Итого с НДС",
        "Источник цены", "Примечание",
    ]
    for col, h in enumerate(smeta_headers, start=1):
        ws_smeta.cell(row=1, column=col, value=h)
    _style_header_row(ws_smeta, 1, len(smeta_headers))
    ws_smeta.row_dimensions[1].height = 40

    for i, item in enumerate(smeta_items, start=1):
        row = i + 1
        qty = coerce_qty(item.get("quantity"))
        wp = coerce_price(item.get("work_price")) or 0
        mp = coerce_price(item.get("material_price")) or 0
        work_total = qty * wp
        mat_total = qty * mp
        subtotal = work_total + mat_total
        vat = subtotal * settings.VAT_RATE
        total = subtotal + vat

        ws_smeta.cell(row=row, column=1, value=i)
        ws_smeta.cell(row=row, column=2, value=item.get("type", ""))
        ws_smeta.cell(row=row, column=3, value=item.get("name", ""))
        ws_smeta.cell(row=row, column=4, value=item.get("unit", ""))
        ws_smeta.cell(row=row, column=5, value=qty)
        ws_smeta.cell(row=row, column=6, value=wp if wp else None)
        ws_smeta.cell(row=row, column=7, value=mp if mp else None)
        ws_smeta.cell(row=row, column=8, value=work_total if work_total else None)
        ws_smeta.cell(row=row, column=9, value=mat_total if mat_total else None)
        ws_smeta.cell(row=row, column=10, value=subtotal if subtotal else None)
        ws_smeta.cell(row=row, column=11, value=vat if vat else None)
        ws_smeta.cell(row=row, column=12, value=total if total else None)
        ws_smeta.cell(row=row, column=13, value=item.get("price_list_name", "") or "")
        ws_smeta.cell(row=row, column=14, value=item.get("notes", ""))
        _style_data_row(ws_smeta, row, len(smeta_headers))

    if smeta_items:
        total_row = len(smeta_items) + 2
        ws_smeta.cell(row=total_row, column=1, value="ИТОГО")
        ws_smeta.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
        data_end = len(smeta_items) + 1
        for col in [8, 9, 10, 11, 12]:
            cl = get_column_letter(col)
            ws_smeta.cell(row=total_row, column=col,
                          value=f"=SUM({cl}2:{cl}{data_end})")
            ws_smeta.cell(row=total_row, column=col).number_format = '#,##0.00'
        for col in range(1, len(smeta_headers) + 1):
            c = ws_smeta.cell(row=total_row, column=col)
            c.fill = TOTAL_FILL
            c.font = BOLD_FONT
            c.border = THIN_BORDER

    for col in [6, 7, 8, 9, 10, 11, 12]:
        for row in range(2, len(smeta_items) + 2):
            ws_smeta.cell(row=row, column=col).number_format = '#,##0.00'

    _auto_fit_columns(ws_smeta)
    ws_smeta.freeze_panes = "A2"

    # ── Sheets 2-4: Перечень from Stage 2 ──────────────────────────────────
    list_headers_all = ["№", "Тип", "Раздел", "Наименование", "Ед. изм.", "Кол-во", "Примечание"]
    list_headers_sub = ["№", "Раздел", "Наименование", "Ед. изм.", "Кол-во", "Примечание"]

    def _write_list_sheet(ws, rows, headers, include_type=False):
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, 1, len(headers))
        ws.row_dimensions[1].height = 30
        for i, item in enumerate(rows, start=1):
            r = i + 1
            if include_type:
                vals = [
                    i, item.get("type", ""), item.get("section", ""),
                    item.get("name", ""), item.get("unit", ""),
                    item.get("quantity"), item.get("notes", ""),
                ]
            else:
                vals = [
                    i, item.get("section", ""), item.get("name", ""),
                    item.get("unit", ""), item.get("quantity"), item.get("notes", ""),
                ]
            for col, val in enumerate(vals, start=1):
                ws.cell(row=r, column=col, value=val)
            _style_data_row(ws, r, len(headers))
        _auto_fit_columns(ws)
        ws.freeze_panes = "A2"

    ws_all = wb.create_sheet("Перечень работ и материалов")
    _write_list_sheet(ws_all, list_items, list_headers_all, include_type=True)

    works = [it for it in list_items if it.get("type", "").lower() in ("работа", "work", "работы")]
    ws_works = wb.create_sheet("Перечень работ")
    _write_list_sheet(ws_works, works, list_headers_sub)

    materials = [it for it in list_items if it.get("type", "").lower() in ("материал", "material", "материалы")]
    ws_mats = wb.create_sheet("Перечень материалов")
    _write_list_sheet(ws_mats, materials, list_headers_sub)

    # ── Sheet 5: Пояснительная записка ────────────────────────────────────
    ws_note = wb.create_sheet("Пояснительная записка")
    ws_note.column_dimensions["A"].width = 120

    current_row = _write_note_section(
        ws_note, 1,
        "Проверка проекта",
        research_result or "Результаты проверки проекта отсутствуют.",
    )
    _write_note_section(
        ws_note, current_row + 1,
        "Замечания к перечню",
        changes_summary or "Перечень соответствует документации, дополнений не требуется.",
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_smeta_from_tz_project(
    smeta_items: list,
    list_items: list,
    changes_summary: Optional[str] = None,
) -> bytes:
    """
    Generate 5-sheet Excel for SMETA_FROM_TZ_PROJECT tasks.
    Sheet 1 "Смета"                      — priced estimate (smeta_items from Stage 2)
    Sheet 2 "Перечень работ и материалов" — all items from Stage 1
    Sheet 3 "Перечень работ"              — works from Stage 1
    Sheet 4 "Перечень материалов"         — materials from Stage 1
    Sheet 5 "Пояснительная записка"       — changes_summary (two sections)
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Смета (priced estimate, Stage 2 items) ────────────────────
    ws_smeta = wb.active
    ws_smeta.title = "Смета"

    smeta_headers = [
        "№", "Тип", "Наименование", "Ед. изм.", "Кол-во",
        "Цена работы (за ед.)", "Цена материала (за ед.)",
        "Стоимость работ", "Стоимость материалов",
        "Итого без НДС", f"НДС ({int(settings.VAT_RATE * 100)}%)", "Итого с НДС",
        "Источник цены", "Примечание",
    ]
    for col, h in enumerate(smeta_headers, start=1):
        ws_smeta.cell(row=1, column=col, value=h)
    _style_header_row(ws_smeta, 1, len(smeta_headers))
    ws_smeta.row_dimensions[1].height = 40

    for i, item in enumerate(smeta_items, start=1):
        row = i + 1
        qty = coerce_qty(item.get("quantity"))
        wp = coerce_price(item.get("work_price")) or 0
        mp = coerce_price(item.get("material_price")) or 0
        work_total = qty * wp
        mat_total = qty * mp
        subtotal = work_total + mat_total
        vat = subtotal * settings.VAT_RATE
        total = subtotal + vat

        ws_smeta.cell(row=row, column=1, value=i)
        ws_smeta.cell(row=row, column=2, value=item.get("type", ""))
        ws_smeta.cell(row=row, column=3, value=item.get("name", ""))
        ws_smeta.cell(row=row, column=4, value=item.get("unit", ""))
        ws_smeta.cell(row=row, column=5, value=qty)
        ws_smeta.cell(row=row, column=6, value=wp if wp else None)
        ws_smeta.cell(row=row, column=7, value=mp if mp else None)
        ws_smeta.cell(row=row, column=8, value=work_total if work_total else None)
        ws_smeta.cell(row=row, column=9, value=mat_total if mat_total else None)
        ws_smeta.cell(row=row, column=10, value=subtotal if subtotal else None)
        ws_smeta.cell(row=row, column=11, value=vat if vat else None)
        ws_smeta.cell(row=row, column=12, value=total if total else None)
        ws_smeta.cell(row=row, column=13, value=item.get("price_list_name", "") or "")
        ws_smeta.cell(row=row, column=14, value=item.get("notes", ""))
        _style_data_row(ws_smeta, row, len(smeta_headers))

    # Totals row
    if smeta_items:
        total_row = len(smeta_items) + 2
        ws_smeta.cell(row=total_row, column=1, value="ИТОГО")
        ws_smeta.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
        data_end = len(smeta_items) + 1
        for col in [8, 9, 10, 11, 12]:
            cl = get_column_letter(col)
            ws_smeta.cell(row=total_row, column=col,
                          value=f"=SUM({cl}2:{cl}{data_end})")
            ws_smeta.cell(row=total_row, column=col).number_format = '#,##0.00'
        for col in range(1, len(smeta_headers) + 1):
            c = ws_smeta.cell(row=total_row, column=col)
            c.fill = TOTAL_FILL
            c.font = BOLD_FONT
            c.border = THIN_BORDER

    for col in [6, 7, 8, 9, 10, 11, 12]:
        for row in range(2, len(smeta_items) + 2):
            ws_smeta.cell(row=row, column=col).number_format = '#,##0.00'

    _auto_fit_columns(ws_smeta)
    ws_smeta.freeze_panes = "A2"

    # ── Sheets 2-4: Перечень from Stage 1 ──────────────────────────────────
    list_headers_all = ["№", "Тип", "Раздел", "Наименование", "Ед. изм.", "Кол-во", "Примечание"]
    list_headers_sub = ["№", "Раздел", "Наименование", "Ед. изм.", "Кол-во", "Примечание"]

    def _write_list_sheet(ws, rows, headers, include_type=False):
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        _style_header_row(ws, 1, len(headers))
        ws.row_dimensions[1].height = 30
        for i, item in enumerate(rows, start=1):
            r = i + 1
            if include_type:
                vals = [
                    i,
                    item.get("type", ""),
                    item.get("section", ""),
                    item.get("name", ""),
                    item.get("unit", ""),
                    item.get("quantity"),
                    item.get("notes", ""),
                ]
            else:
                vals = [
                    i,
                    item.get("section", ""),
                    item.get("name", ""),
                    item.get("unit", ""),
                    item.get("quantity"),
                    item.get("notes", ""),
                ]
            for col, val in enumerate(vals, start=1):
                ws.cell(row=r, column=col, value=val)
            _style_data_row(ws, r, len(headers))
        _auto_fit_columns(ws)
        ws.freeze_panes = "A2"

    ws_all = wb.create_sheet("Перечень работ и материалов")
    _write_list_sheet(ws_all, list_items, list_headers_all, include_type=True)

    works = [it for it in list_items if it.get("type", "").lower() in ("работа", "work", "работы")]
    ws_works = wb.create_sheet("Перечень работ")
    _write_list_sheet(ws_works, works, list_headers_sub)

    materials = [it for it in list_items if it.get("type", "").lower() in ("материал", "material", "материалы")]
    ws_mats = wb.create_sheet("Перечень материалов")
    _write_list_sheet(ws_mats, materials, list_headers_sub)

    # ── Sheet 5: Пояснительная записка (two sections) ──────────────────────
    ws_note = wb.create_sheet("Пояснительная записка")
    ws_note.column_dimensions["A"].width = 120

    FALLBACK = "Документация соответствует друг другу, дополнений не требуется"

    if not changes_summary:
        _write_note_section(ws_note, 1, "Пояснительная записка", FALLBACK)
    else:
        import re
        split_match = re.search(r"(?i)раздел\s*2", changes_summary)
        if split_match:
            part1 = changes_summary[: split_match.start()].strip()
            part2 = changes_summary[split_match.start():].strip()
        else:
            part1 = changes_summary.strip()
            part2 = ""

        current_row = _write_note_section(
            ws_note, 1,
            "Раздел 1 — Сравнение ТЗ и проекта",
            part1 or FALLBACK,
        )
        if part2:
            _write_note_section(
                ws_note, current_row + 1,
                "Раздел 2 — Изменения по нормативной базе",
                part2,
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_smeta(items: list) -> bytes:
    """
    Generate full smeta Excel.
    items: list of dicts with: type, name, unit, quantity,
           work_price (per unit), material_price (per unit), notes
    """
    wb = openpyxl.Workbook()

    # Sheet 1: Смета
    ws = wb.active
    ws.title = "Смета"

    headers = [
        "№",
        "Тип",
        "Наименование",
        "Ед. изм.",
        "Кол-во",
        "Цена работы (за ед.)",
        "Цена материала (за ед.)",
        "Стоимость работ",
        "Стоимость материалов",
        "Итого без НДС",
        f"НДС ({int(settings.VAT_RATE * 100)}%)",
        "Итого с НДС",
        "Наименование в прайсе",
        "Источники",
        "Примечание",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 40

    for i, item in enumerate(items, start=1):
        row = i + 1
        qty = coerce_qty(item.get("quantity"))
        wp = coerce_price(item.get("work_price")) or 0
        mp = coerce_price(item.get("material_price")) or 0

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item.get("type", ""))
        ws.cell(row=row, column=3, value=item.get("name", ""))
        ws.cell(row=row, column=4, value=item.get("unit", ""))
        ws.cell(row=row, column=5, value=qty)
        ws.cell(row=row, column=6, value=wp if wp else None)
        ws.cell(row=row, column=7, value=mp if mp else None)

        work_total = qty * wp
        mat_total = qty * mp
        subtotal = work_total + mat_total
        vat = subtotal * settings.VAT_RATE
        total = subtotal + vat

        ws.cell(row=row, column=8, value=work_total if work_total else None)
        ws.cell(row=row, column=9, value=mat_total if mat_total else None)
        ws.cell(row=row, column=10, value=subtotal if subtotal else None)
        ws.cell(row=row, column=11, value=vat if vat else None)
        ws.cell(row=row, column=12, value=total if total else None)
        ws.cell(row=row, column=13, value=item.get("price_list_name", "") or "")
        ws.cell(row=row, column=14, value=item.get("sources", "") or "")
        ws.cell(row=row, column=15, value=item.get("notes", ""))
        _style_data_row(ws, row, len(headers))

    # Totals row
    total_row = len(items) + 2
    ws.cell(row=total_row, column=1, value="ИТОГО")
    ws.merge_cells(
        start_row=total_row, start_column=1, end_row=total_row, end_column=7
    )
    data_start = 2
    data_end = len(items) + 1

    for col in [8, 9, 10, 11, 12]:
        col_letter = get_column_letter(col)
        ws.cell(
            row=total_row,
            column=col,
            value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
        )
        ws.cell(row=total_row, column=col).number_format = '#,##0.00'

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER

    # Number formats for currency columns
    for col in [6, 7, 8, 9, 10, 11, 12]:
        for row in range(2, len(items) + 2):
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    _auto_fit_columns(ws)
    ws.freeze_panes = "A2"

    # Sheet 2: Работы
    works = [it for it in items if it.get("type", "").lower() in ("работа", "work", "работы")]
    ws_w = wb.create_sheet("Работы")
    w_headers = ["№", "Наименование", "Ед. изм.", "Кол-во", "Цена за ед.", "Стоимость", "Наименование в прайсе", "Источники"]
    for col, h in enumerate(w_headers, start=1):
        ws_w.cell(row=1, column=col, value=h)
    _style_header_row(ws_w, 1, len(w_headers))
    ws_w.row_dimensions[1].height = 30

    for i, item in enumerate(works, start=1):
        row = i + 1
        qty = coerce_qty(item.get("quantity"))
        price = coerce_price(item.get("work_price")) or 0
        ws_w.cell(row=row, column=1, value=i)
        ws_w.cell(row=row, column=2, value=item.get("name", ""))
        ws_w.cell(row=row, column=3, value=item.get("unit", ""))
        ws_w.cell(row=row, column=4, value=qty)
        ws_w.cell(row=row, column=5, value=price if price else None)
        ws_w.cell(row=row, column=6, value=qty * price if price else None)
        ws_w.cell(row=row, column=7, value=item.get("price_list_name", "") or "")
        ws_w.cell(row=row, column=8, value=item.get("sources", "") or "")
        _style_data_row(ws_w, row, len(w_headers))

    if works:
        sum_row = len(works) + 2
        ws_w.cell(row=sum_row, column=1, value="ИТОГО")
        ws_w.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=5)
        ws_w.cell(
            row=sum_row,
            column=6,
            value=f"=SUM(F2:F{len(works)+1})",
        )
        for col in range(1, len(w_headers) + 1):
            cell = ws_w.cell(row=sum_row, column=col)
            cell.fill = TOTAL_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER

    _auto_fit_columns(ws_w)
    ws_w.freeze_panes = "A2"

    # Sheet 3: Материалы
    materials = [it for it in items if it.get("type", "").lower() in ("материал", "material", "материалы")]
    ws_m = wb.create_sheet("Материалы")
    m_headers = ["№", "Наименование", "Ед. изм.", "Кол-во", "Цена за ед.", "Стоимость", "Наименование в прайсе", "Источники"]
    for col, h in enumerate(m_headers, start=1):
        ws_m.cell(row=1, column=col, value=h)
    _style_header_row(ws_m, 1, len(m_headers))
    ws_m.row_dimensions[1].height = 30

    for i, item in enumerate(materials, start=1):
        row = i + 1
        qty = coerce_qty(item.get("quantity"))
        price = coerce_price(item.get("material_price")) or 0
        ws_m.cell(row=row, column=1, value=i)
        ws_m.cell(row=row, column=2, value=item.get("name", ""))
        ws_m.cell(row=row, column=3, value=item.get("unit", ""))
        ws_m.cell(row=row, column=4, value=qty)
        ws_m.cell(row=row, column=5, value=price if price else None)
        ws_m.cell(row=row, column=6, value=qty * price if price else None)
        ws_m.cell(row=row, column=7, value=item.get("price_list_name", "") or "")
        ws_m.cell(row=row, column=8, value=item.get("sources", "") or "")
        _style_data_row(ws_m, row, len(m_headers))

    if materials:
        sum_row = len(materials) + 2
        ws_m.cell(row=sum_row, column=1, value="ИТОГО")
        ws_m.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=5)
        ws_m.cell(
            row=sum_row,
            column=6,
            value=f"=SUM(F2:F{len(materials)+1})",
        )
        for col in range(1, len(m_headers) + 1):
            cell = ws_m.cell(row=sum_row, column=col)
            cell.fill = TOTAL_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER

    _auto_fit_columns(ws_m)
    ws_m.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_smeta_detailed(items: list) -> bytes:
    """
    Generate detailed smeta Excel with 3 sheets and VAT sub-columns.
    items: list of dicts with:
        type, name, unit, quantity,
        work_price (ex. VAT), mat_price (ex. VAT),
        usn (bool: if True, VAT=0), price_list_name, notes
    """
    wb = openpyxl.Workbook()

    def _vat_triple(price, usn: bool):
        """Return (ex_vat, vat, inc_vat) for a given price."""
        if not price:
            return None, None, None
        vat = 0.0 if usn else price * settings.VAT_RATE
        return price, vat, price + vat

    def _write_sheet1(ws, rows):
        """Sheet 1: Перечень работ и материалов (all items)."""
        # Row 1: group headers with merges
        groups = [
            (1, 1, "№ п/п"),
            (2, 2, "Работа/\nМатериал"),
            (3, 3, "Наименование"),
            (4, 4, "Ед. изм."),
            (5, 5, "Кол-во"),
            (6, 8, "Цена за ед. изм.\nРаботы"),
            (9, 11, "Стоимость работ\nруб."),
            (12, 14, "Цена за ед. изм.\nМатериала"),
            (15, 17, "Стоимость\nМатериала руб."),
            (18, 18, "Наименование\nв прайсе"),
            (19, 19, "Примечание"),
        ]
        for c_start, c_end, title in groups:
            ws.cell(row=1, column=c_start, value=title)
            if c_start == c_end:
                ws.merge_cells(start_row=1, start_column=c_start, end_row=2, end_column=c_start)
            else:
                ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
        # Row 2: sub-headers for VAT columns
        for col, label in [
            (6, "без НДС"), (7, "НДС"), (8, "с НДС"),
            (9, "без НДС"), (10, "НДС"), (11, "с НДС"),
            (12, "без НДС"), (13, "НДС"), (14, "с НДС"),
            (15, "без НДС"), (16, "НДС"), (17, "с НДС"),
        ]:
            ws.cell(row=2, column=col, value=label)
        total_cols = 19
        _style_header_row(ws, 1, total_cols)
        _style_header_row(ws, 2, total_cols)
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 30

        for i, item in enumerate(rows, start=1):
            r = i + 2
            qty = coerce_qty(item.get("quantity"))
            wp = coerce_price(item.get("work_price")) or 0
            mp = coerce_price(item.get("mat_price")) or 0
            usn = item.get("usn", False)

            wp_ex, wp_vat, wp_inc = _vat_triple(wp, usn)
            mp_ex, mp_vat, mp_inc = _vat_triple(mp, usn)
            wt_ex = (wp_ex or 0) * qty or None
            wt_vat = (wp_vat or 0) * qty or None
            wt_inc = (wp_inc or 0) * qty or None
            mt_ex = (mp_ex or 0) * qty or None
            mt_vat = (mp_vat or 0) * qty or None
            mt_inc = (mp_inc or 0) * qty or None

            vals = [
                i, item.get("type", ""), item.get("name", ""),
                item.get("unit", ""), qty if qty else None,
                wp_ex, wp_vat, wp_inc,
                wt_ex, wt_vat, wt_inc,
                mp_ex, mp_vat, mp_inc,
                mt_ex, mt_vat, mt_inc,
                item.get("price_list_name", ""),
                item.get("notes", ""),
            ]
            for col, val in enumerate(vals, start=1):
                ws.cell(row=r, column=col, value=val)
            _style_data_row(ws, r, total_cols)

        # Currency format
        for col in range(6, 18):
            for r in range(3, len(rows) + 3):
                ws.cell(row=r, column=col).number_format = '#,##0.00'

        # Totals row
        if rows:
            tr = len(rows) + 3
            ws.cell(row=tr, column=1, value="ИТОГО")
            ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)
            data_start, data_end = 3, len(rows) + 2
            for col in range(9, 18):
                cl = get_column_letter(col)
                ws.cell(row=tr, column=col,
                        value=f"=SUM({cl}{data_start}:{cl}{data_end})")
                ws.cell(row=tr, column=col).number_format = '#,##0.00'
            for col in range(1, total_cols + 1):
                c = ws.cell(row=tr, column=col)
                c.fill = TOTAL_FILL
                c.font = BOLD_FONT
                c.border = THIN_BORDER

        _auto_fit_columns(ws)
        ws.freeze_panes = "A3"

    def _write_sheet_works(ws, works):
        """Sheet 2: Перечень работ."""
        total_cols = 12
        groups = [
            (1, 1, "№ п/п"),
            (2, 2, "Наименование"),
            (3, 3, "Ед. изм."),
            (4, 4, "Кол-во"),
            (5, 7, "Цена за ед. изм.\nРаботы"),
            (8, 10, "Стоимость работ\nруб."),
            (11, 11, "Наименование\nв прайсе"),
            (12, 12, "Примечание"),
        ]
        for c_start, c_end, title in groups:
            ws.cell(row=1, column=c_start, value=title)
            if c_start == c_end:
                ws.merge_cells(start_row=1, start_column=c_start, end_row=2, end_column=c_start)
            else:
                ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
        for col, label in [(5, "без НДС"), (6, "НДС"), (7, "с НДС"),
                           (8, "без НДС"), (9, "НДС"), (10, "с НДС")]:
            ws.cell(row=2, column=col, value=label)
        _style_header_row(ws, 1, total_cols)
        _style_header_row(ws, 2, total_cols)
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 30

        for i, item in enumerate(works, start=1):
            r = i + 2
            qty = coerce_qty(item.get("quantity"))
            wp = coerce_price(item.get("work_price")) or 0
            usn = item.get("usn", False)
            wp_ex, wp_vat, wp_inc = _vat_triple(wp, usn)
            wt_ex = (wp_ex or 0) * qty or None
            wt_vat = (wp_vat or 0) * qty or None
            wt_inc = (wp_inc or 0) * qty or None
            vals = [
                i, item.get("name", ""), item.get("unit", ""),
                qty if qty else None,
                wp_ex, wp_vat, wp_inc,
                wt_ex, wt_vat, wt_inc,
                item.get("price_list_name", ""), item.get("notes", ""),
            ]
            for col, val in enumerate(vals, start=1):
                ws.cell(row=r, column=col, value=val)
            _style_data_row(ws, r, total_cols)

        for col in range(5, 11):
            for r in range(3, len(works) + 3):
                ws.cell(row=r, column=col).number_format = '#,##0.00'

        if works:
            tr = len(works) + 3
            ws.cell(row=tr, column=1, value="ИТОГО")
            ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
            for col in range(8, 11):
                cl = get_column_letter(col)
                ws.cell(row=tr, column=col,
                        value=f"=SUM({cl}3:{cl}{len(works)+2})")
                ws.cell(row=tr, column=col).number_format = '#,##0.00'
            for col in range(1, total_cols + 1):
                c = ws.cell(row=tr, column=col)
                c.fill = TOTAL_FILL
                c.font = BOLD_FONT
                c.border = THIN_BORDER

        _auto_fit_columns(ws)
        ws.freeze_panes = "A3"

    def _write_sheet_materials(ws, materials):
        """Sheet 3: Перечень материалов."""
        total_cols = 12
        groups = [
            (1, 1, "№ п/п"),
            (2, 2, "Наименование"),
            (3, 3, "Ед. изм."),
            (4, 4, "Кол-во"),
            (5, 7, "Цена за ед. изм.\nМатериала"),
            (8, 10, "Стоимость\nМатериала руб."),
            (11, 11, "Наименование\nв прайсе"),
            (12, 12, "Примечание"),
        ]
        for c_start, c_end, title in groups:
            ws.cell(row=1, column=c_start, value=title)
            if c_start == c_end:
                ws.merge_cells(start_row=1, start_column=c_start, end_row=2, end_column=c_start)
            else:
                ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
        for col, label in [(5, "без НДС"), (6, "НДС"), (7, "с НДС"),
                           (8, "без НДС"), (9, "НДС"), (10, "с НДС")]:
            ws.cell(row=2, column=col, value=label)
        _style_header_row(ws, 1, total_cols)
        _style_header_row(ws, 2, total_cols)
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 30

        for i, item in enumerate(materials, start=1):
            r = i + 2
            qty = coerce_qty(item.get("quantity"))
            mp = coerce_price(item.get("mat_price")) or 0
            usn = item.get("usn", False)
            mp_ex, mp_vat, mp_inc = _vat_triple(mp, usn)
            mt_ex = (mp_ex or 0) * qty or None
            mt_vat = (mp_vat or 0) * qty or None
            mt_inc = (mp_inc or 0) * qty or None
            vals = [
                i, item.get("name", ""), item.get("unit", ""),
                qty if qty else None,
                mp_ex, mp_vat, mp_inc,
                mt_ex, mt_vat, mt_inc,
                item.get("price_list_name", ""), item.get("notes", ""),
            ]
            for col, val in enumerate(vals, start=1):
                ws.cell(row=r, column=col, value=val)
            _style_data_row(ws, r, total_cols)

        for col in range(5, 11):
            for r in range(3, len(materials) + 3):
                ws.cell(row=r, column=col).number_format = '#,##0.00'

        if materials:
            tr = len(materials) + 3
            ws.cell(row=tr, column=1, value="ИТОГО")
            ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)
            for col in range(8, 11):
                cl = get_column_letter(col)
                ws.cell(row=tr, column=col,
                        value=f"=SUM({cl}3:{cl}{len(materials)+2})")
                ws.cell(row=tr, column=col).number_format = '#,##0.00'
            for col in range(1, total_cols + 1):
                c = ws.cell(row=tr, column=col)
                c.fill = TOTAL_FILL
                c.font = BOLD_FONT
                c.border = THIN_BORDER

        _auto_fit_columns(ws)
        ws.freeze_panes = "A3"

    ws1 = wb.active
    ws1.title = "Перечень работ и материалов"
    _write_sheet1(ws1, items)

    works = [it for it in items if it.get("type", "").lower() in ("работа", "work", "работы")]
    ws2 = wb.create_sheet("Перечень работ")
    _write_sheet_works(ws2, works)

    materials = [it for it in items if it.get("type", "").lower() in ("материал", "material", "материалы")]
    ws3 = wb.create_sheet("Перечень материалов")
    _write_sheet_materials(ws3, materials)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_scan_result(data: dict) -> bytes:
    """
    Generate Excel from scan/OCR result.
    data: {
      header: {title, date, contractor, object},
      sections: [{title, items: [{name, unit, qty, price, total, notes}]}],
      summary: {total_works, total_materials, total_vat, grand_total}
    }
    """
    wb = openpyxl.Workbook()

    # Sheet 1: Информация
    ws_info = wb.active
    ws_info.title = "Информация"

    header = data.get("header", {})
    info_rows = [
        ("Документ", header.get("title", "")),
        ("Дата", header.get("date", "")),
        ("Подрядчик", header.get("contractor", "")),
        ("Объект", header.get("object", "")),
    ]

    for row_idx, (key, val) in enumerate(info_rows, start=1):
        ws_info.cell(row=row_idx, column=1, value=key).font = BOLD_FONT
        ws_info.cell(row=row_idx, column=2, value=val)

    _auto_fit_columns(ws_info)

    # Sheet 2: Данные
    ws_data = wb.create_sheet("Данные")
    headers = ["№", "Раздел", "Наименование", "Ед. изм.", "Кол-во", "Цена", "Сумма", "Примечание"]
    for col, h in enumerate(headers, start=1):
        ws_data.cell(row=1, column=col, value=h)
    _style_header_row(ws_data, 1, len(headers))
    ws_data.row_dimensions[1].height = 30

    row_idx = 2
    item_num = 0
    sections = data.get("sections", [])

    for section in sections:
        section_title = section.get("title", "")
        # Section header row
        ws_data.cell(row=row_idx, column=1, value="")
        ws_data.cell(row=row_idx, column=2, value=section_title)
        ws_data.merge_cells(
            start_row=row_idx, start_column=2, end_row=row_idx, end_column=len(headers)
        )
        for col in range(1, len(headers) + 1):
            cell = ws_data.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
        row_idx += 1

        for item in section.get("items", []):
            item_num += 1
            ws_data.cell(row=row_idx, column=1, value=item_num)
            ws_data.cell(row=row_idx, column=2, value=section_title)
            ws_data.cell(row=row_idx, column=3, value=item.get("name", ""))
            ws_data.cell(row=row_idx, column=4, value=item.get("unit", ""))
            ws_data.cell(row=row_idx, column=5, value=item.get("qty"))
            ws_data.cell(row=row_idx, column=6, value=item.get("price"))
            ws_data.cell(row=row_idx, column=7, value=item.get("total"))
            ws_data.cell(row=row_idx, column=8, value=item.get("notes", ""))
            _style_data_row(ws_data, row_idx, len(headers))
            row_idx += 1

    _auto_fit_columns(ws_data)
    ws_data.freeze_panes = "A2"

    # Sheet 3: Итоги
    ws_summary = wb.create_sheet("Итоги")
    summary = data.get("summary", {})
    summary_rows = [
        ("Итого работ", summary.get("total_works", "")),
        ("Итого материалов", summary.get("total_materials", "")),
        (f"НДС ({int(settings.VAT_RATE * 100)}%)", summary.get("total_vat", "")),
        ("Итого с НДС", summary.get("grand_total", "")),
    ]

    ws_summary.cell(row=1, column=1, value="Показатель").font = HEADER_FONT
    ws_summary.cell(row=1, column=1).fill = HEADER_FILL
    ws_summary.cell(row=1, column=2, value="Значение").font = HEADER_FONT
    ws_summary.cell(row=1, column=2).fill = HEADER_FILL

    for row_idx, (key, val) in enumerate(summary_rows, start=2):
        ws_summary.cell(row=row_idx, column=1, value=key).font = BOLD_FONT
        cell_val = ws_summary.cell(row=row_idx, column=2, value=val)
        if isinstance(val, (int, float)):
            cell_val.number_format = '#,##0.00'

    _auto_fit_columns(ws_summary)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Estimate Optimization export helpers
# ---------------------------------------------------------------------------

_SECTION_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_SUBTOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_GRAND_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_CHEAPER_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_DEARER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_NUMBER_FMT = '#,##0.00'
_VAT_RATE = 0.22


def _row_cost_dict(row: dict) -> float:
    """Стоимость строки, округлённая до копеек.

    Округление именно здесь, а не только при выводе: Excel показывает 2 знака, но
    складывает полные значения, поэтому «сумма показанных строк» расходилась с
    «показанным итогом». Проверено: три строки по 33.4665 → пользователь видит
    33.47 трижды (100.41), а итог печатался 100.40. На тендере такая копейка —
    повод для придирки.
    """
    qty = coerce_qty(row.get("qty"))
    pw = coerce_price(row.get("price_work")) or 0
    pm = coerce_price(row.get("price_material")) or 0
    return round(qty * (pw + pm), 2)


def _safe_cell(value):
    """Strip illegal XML control characters that crash openpyxl."""
    if isinstance(value, str):
        import re as _re
        return _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
    return value


def generate_estimate_export(
    rows: list,
    overhead_pct: float,
    transport_pct: float,
    contingency_pct: float,
    version_display_name: str,
    coefficient: Optional[dict] = None,
) -> bytes:
    """Export a single EstimateVersion to xlsx.

    Цены выходят с коэффициентом документа (решение пользователя 4.5), а
    накладные и транспортные считаются той же формулой, что везде в проекте:
    накладные — от работ, транспортные — от материалов. Раньше здесь обе
    брались от общего базиса, и выгруженная версия не сходилась с экраном.
    """
    from app.utils.xlsx_exporter import coefficient_for
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"

    COLS = ["№", "Тип", "Наименование", "Ед. изм.", "Кол-во",
            "Цена работы, руб", "Цена материала, руб", "Стоимость, руб"]

    for ci, h in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    col_widths = [6, 10, 50, 10, 10, 20, 22, 20]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    TYPE_LABELS = {"work": "Работа", "material": "Материал", "section": "—"}

    total_works = 0.0
    total_materials = 0.0
    row_idx = 2

    for r in rows:
        if r.get("is_excluded"):
            continue
        rtype = r.get("type", "")
        is_section = rtype == "section"

        k_work, k_material = coefficient_for(coefficient, r.get("id"))
        if k_work != 1.0 or k_material != 1.0:
            r = dict(r)
            if r.get("price_work") is not None:
                r["price_work"] = round(float(r["price_work"]) * k_work, 2)
            if r.get("price_material") is not None:
                r["price_material"] = round(float(r["price_material"]) * k_material, 2)
        cost = _row_cost_dict(r)

        work_cost = _row_cost_dict({**r, "price_material": None})
        material_cost = _row_cost_dict({**r, "price_work": None})
        if not is_section:
            total_works += work_cost
            total_materials += material_cost

        ws.cell(row=row_idx, column=1, value=r.get("num") if not is_section else None)
        ws.cell(row=row_idx, column=2, value=_safe_cell(TYPE_LABELS.get(rtype, rtype)))
        ws.cell(row=row_idx, column=3, value=_safe_cell(r.get("name", "")))
        ws.cell(row=row_idx, column=4, value=_safe_cell(r.get("unit", "")) if not is_section else None)
        ws.cell(row=row_idx, column=5, value=r.get("qty") if not is_section else None)
        ws.cell(row=row_idx, column=6, value=r.get("price_work") if not is_section else None)
        ws.cell(row=row_idx, column=7, value=r.get("price_material") if not is_section else None)
        ws.cell(row=row_idx, column=8, value=cost if not is_section else None)

        for ci in range(1, len(COLS) + 1):
            cell = ws.cell(row=row_idx, column=ci)
            cell.border = THIN_BORDER
            if is_section:
                cell.font = BOLD_FONT
                cell.fill = _SECTION_FILL
            else:
                cell.font = NORMAL_FONT
        for ci in (6, 7, 8):
            ws.cell(row=row_idx, column=ci).number_format = _NUMBER_FMT
        row_idx += 1

    row_idx += 1

    base = total_works + total_materials
    # Единое правило проекта: накладные — от работ, транспортные — от материалов.
    overhead = total_works * overhead_pct / 100
    transport = total_materials * transport_pct / 100
    contingency = base * contingency_pct / 100
    total = base + overhead + transport + contingency
    vat = total * _VAT_RATE
    grand_total = total + vat

    def _write_total(label: str, value: float, fill: PatternFill, bold: bool = False):
        nonlocal row_idx
        ws.cell(row=row_idx, column=3, value=label)
        c_value = ws.cell(row=row_idx, column=8, value=value)
        c_value.number_format = _NUMBER_FMT
        for ci in range(1, len(COLS) + 1):
            cell = ws.cell(row=row_idx, column=ci)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.font = BOLD_FONT if bold else NORMAL_FONT
        row_idx += 1

    _grand_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    _write_total("Работы (итого)", total_works, _SUBTOTAL_FILL)
    _write_total("Материалы (итого)", total_materials, _SUBTOTAL_FILL)
    _write_total("Итого (базис)", base, _SUBTOTAL_FILL, bold=True)
    _write_total(f"Накладные расходы ({overhead_pct}%)", overhead, _GRAND_FILL)
    _write_total(f"Транспортные расходы ({transport_pct}%)", transport, _GRAND_FILL)
    _write_total(f"Непредвиденные расходы ({contingency_pct}%)", contingency, _GRAND_FILL)
    _write_total("Итого", total, _GRAND_FILL, bold=True)
    _write_total(f"НДС {int(_VAT_RATE * 100)}%", vat, _GRAND_FILL)
    _write_total("ИТОГО с НДС", grand_total, _grand_blue, bold=True)

    # White font for last (blue) row
    for ci in range(1, len(COLS) + 1):
        ws.cell(row=row_idx - 1, column=ci).font = Font(bold=True, color="FFFFFF", size=11)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_svodная(wb: "openpyxl.Workbook", versions: list, customer_estimate: Optional[dict]) -> None:
    """Add 'Сводная' sheet as the first sheet with summary totals table."""
    ws = wb.create_sheet("Сводная", 0)

    VAT_RATE = _VAT_RATE
    AMBER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    AMBER_HEADER = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    GRAND_BLUE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    PCT_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    VERSION_FILLS = [
        PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid"),
        PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid"),
    ]

    def _calc_totals(v: dict) -> dict:
        rows = v["rows"]
        works = sum(_row_cost_dict(r) for r in rows if r.get("type") == "work" and not r.get("is_excluded"))
        materials = sum(_row_cost_dict(r) for r in rows if r.get("type") == "material" and not r.get("is_excluded"))
        base = works + materials
        ovh = base * v["overhead_pct"] / 100
        trp = base * v["transport_pct"] / 100
        cng = base * v["contingency_pct"] / 100
        total = base + ovh + trp + cng
        vat = total * VAT_RATE
        return {"works": works, "materials": materials, "base": base, "total": total, "vat": vat, "grand_total": total + vat}

    all_totals = [_calc_totals(v) for v in versions]

    has_customer = customer_estimate is not None and customer_estimate.get("grand_total", 0) > 0

    # Column layout: col1=Показатель, col2=Смета от заказчика (optional), col3..N=versions
    col_offset = 2 if has_customer else 1

    # Header row
    h = ws.cell(row=1, column=1, value="Показатель")
    h.font = Font(bold=True, color="FFFFFF", size=11)
    h.fill = HEADER_FILL
    h.border = THIN_BORDER
    h.alignment = Alignment(horizontal="left")
    ws.column_dimensions["A"].width = 30

    if has_customer:
        hc = ws.cell(row=1, column=2, value="Смета от заказчика")
        hc.font = Font(bold=True, color="FFFFFF", size=11)
        hc.fill = AMBER_HEADER
        hc.border = THIN_BORDER
        hc.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(2)].width = 20

    for vi, v in enumerate(versions):
        ci = col_offset + 1 + vi
        hv = ws.cell(row=1, column=ci, value=v["version_display_name"])
        hv.font = Font(bold=True, color="FFFFFF", size=11)
        hv.fill = VERSION_FILLS[vi % len(VERSION_FILLS)]
        hv.border = THIN_BORDER
        hv.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = 20

    ws.row_dimensions[1].height = 28

    # Row definitions: (key, label, customer_field_or_none)
    ROWS = [
        ("works",       "Работы, руб",              "works"),
        ("materials",   "Материалы, руб",            "materials"),
        ("base",        "Итого (базис), руб",         None),
        ("total",       "Итого с расходами, руб",     None),
        ("vat",         f"НДС {int(VAT_RATE*100)}%, руб", "vat"),
        ("grand_total", "ИТОГО с НДС, руб",           "grand_total"),
    ]

    orig_totals = all_totals[0] if all_totals else {}

    for ri, (key, label, cust_field) in enumerate(ROWS):
        row_num = ri + 2
        is_grand = key == "grand_total"
        row_fill = GRAND_BLUE if is_grand else None

        lc = ws.cell(row=row_num, column=1, value=label)
        lc.font = Font(bold=is_grand, color="FFFFFF" if is_grand else "000000", size=11)
        lc.fill = row_fill or PatternFill(fill_type=None)
        lc.border = THIN_BORDER

        if has_customer:
            cval = customer_estimate.get(cust_field) if cust_field else None
            cc = ws.cell(row=row_num, column=2, value=cval if cval is not None else "—")
            cc.fill = AMBER_FILL
            cc.border = THIN_BORDER
            cc.alignment = Alignment(horizontal="right")
            if isinstance(cval, (int, float)):
                cc.number_format = _NUMBER_FMT
                cc.font = Font(bold=is_grand, size=11)

        for vi, vt in enumerate(all_totals):
            ci = col_offset + 1 + vi
            val = vt.get(key, 0)
            orig_val = orig_totals.get(key, 0)
            vc = ws.cell(row=row_num, column=ci, value=val)
            vc.number_format = _NUMBER_FMT
            vc.border = THIN_BORDER
            vc.alignment = Alignment(horizontal="right")
            if is_grand:
                vc.fill = GRAND_BLUE
                vc.font = Font(bold=True, color="FFFFFF", size=11)
            else:
                if vi > 0 and val < orig_val - 0.01:
                    vc.fill = _CHEAPER_FILL
                    vc.font = BOLD_FONT
                elif vi > 0 and val > orig_val + 0.01:
                    vc.fill = _DEARER_FILL
                    vc.font = BOLD_FONT
                else:
                    vc.font = Font(bold=is_grand, size=11)

    # "% к смете заказчика" row — only if customer grand_total > 0
    if has_customer and customer_estimate.get("grand_total", 0) > 0:
        pct_row = len(ROWS) + 2
        base_row = ws.cell(row=pct_row, column=1, value="% к смете заказчика")
        base_row.font = Font(italic=True, size=10, color="374151")
        base_row.fill = PCT_FILL
        base_row.border = THIN_BORDER

        if has_customer:
            bc = ws.cell(row=pct_row, column=2, value="базис")
            bc.fill = AMBER_FILL
            bc.border = THIN_BORDER
            bc.alignment = Alignment(horizontal="center")
            bc.font = Font(bold=True, color="92400E", size=10)

        cgt = customer_estimate["grand_total"]
        for vi, vt in enumerate(all_totals):
            ci = col_offset + 1 + vi
            pct = ((vt["grand_total"] - cgt) / cgt) * 100
            pct_str = f"{'+' if pct > 0 else ''}{pct:.1f}%"
            pc = ws.cell(row=pct_row, column=ci, value=pct_str)
            pc.fill = PCT_FILL
            pc.border = THIN_BORDER
            pc.alignment = Alignment(horizontal="right")
            color = "166534" if pct < -0.05 else ("dc2626" if pct > 0.05 else "475569")
            pc.font = Font(bold=True, color=color, size=11)

    ws.freeze_panes = "B2"


def generate_comparison_export(versions: list, customer_estimate: Optional[dict] = None) -> bytes:
    """
    Export multi-version comparison to xlsx.

    versions: list of dicts:
        { id, version_display_name, rows: list[dict], overhead_pct, transport_pct, contingency_pct }
    customer_estimate: optional dict { works, materials, vat, grand_total } — manual user values.
    Rows aligned by lineage_id. Cells cheaper than original are green, dearer — red.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сравнение"

    if not versions:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    _build_svodная(wb, versions, customer_estimate)

    # Align rows by lineage_id
    original_rows = versions[0]["rows"]
    lineage_order: list[str] = []
    seen: set[str] = set()
    for r in original_rows:
        lid = r.get("lineage_id") or r.get("id", "")
        if lid and lid not in seen:
            lineage_order.append(lid)
            seen.add(lid)
    for v in versions[1:]:
        for r in v["rows"]:
            lid = r.get("lineage_id") or r.get("id", "")
            if lid and lid not in seen:
                lineage_order.append(lid)
                seen.add(lid)

    lookup: dict[str, dict[str, dict]] = {}
    for v in versions:
        for r in v["rows"]:
            lid = r.get("lineage_id") or r.get("id", "")
            if lid:
                lookup.setdefault(lid, {})[v["id"]] = r

    FIXED = 4
    vcols = 4  # qty, price_work, price_material, cost per version

    VERSION_FILLS = [
        PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid"),
        PatternFill(start_color="F4B942", end_color="F4B942", fill_type="solid"),
    ]

    # Row 1: merged version name headers
    for ci, label in enumerate(["№", "Тип", "Наименование", "Ед. изм."], start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

    for vi, v in enumerate(versions):
        start_col = FIXED + 1 + vi * vcols
        end_col = start_col + vcols - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        hc = ws.cell(row=1, column=start_col, value=v["version_display_name"])
        vfill = VERSION_FILLS[vi % len(VERSION_FILLS)]
        hc.fill = vfill
        hc.font = Font(bold=True, color="FFFFFF", size=11)
        hc.alignment = Alignment(horizontal="center")
        hc.border = THIN_BORDER

    # Row 2: sub-headers
    for ci, label in enumerate(["№", "Тип", "Наименование", "Ед. изм."], start=1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    sub_headers = ["Кол-во", "Цена работы", "Цена матер.", "Стоимость"]
    for vi in range(len(versions)):
        vfill = VERSION_FILLS[vi % len(VERSION_FILLS)]
        for si, sh in enumerate(sub_headers):
            ci = FIXED + 1 + vi * vcols + si
            cell = ws.cell(row=2, column=ci, value=sh)
            cell.fill = PatternFill(
                start_color=vfill.start_color.rgb,
                end_color=vfill.end_color.rgb,
                fill_type="solid",
            )
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 35
    ws.freeze_panes = "A3"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 10
    for vi in range(len(versions)):
        for si in range(vcols):
            ws.column_dimensions[get_column_letter(FIXED + 1 + vi * vcols + si)].width = 15

    TYPE_LABELS = {"work": "Работа", "material": "Материал", "section": "—"}

    orig_cost_by_lineage: dict[str, float] = {
        (r.get("lineage_id") or r.get("id", "")): _row_cost_dict(r) for r in original_rows
    }

    # Data rows
    row_idx = 3
    _absent_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    for lid in lineage_order:
        any_row = lookup.get(lid, {})
        rep = next(iter(any_row.values()), None)
        if rep is None:
            continue

        rtype = rep.get("type", "")
        is_section = rtype == "section"
        is_added = lid not in {(r.get("lineage_id") or r.get("id", "")) for r in original_rows}

        ws.cell(row=row_idx, column=1, value=rep.get("num") if not is_section else None)
        ws.cell(row=row_idx, column=2, value=TYPE_LABELS.get(rtype, rtype))
        ws.cell(row=row_idx, column=3, value=rep.get("name", ""))
        ws.cell(row=row_idx, column=4, value=rep.get("unit", "") if not is_section else None)

        _added_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for ci in range(1, FIXED + 1):
            cell = ws.cell(row=row_idx, column=ci)
            cell.border = THIN_BORDER
            if is_section:
                cell.font = BOLD_FONT
                cell.fill = _SECTION_FILL
            elif is_added:
                cell.fill = _added_fill

        orig_cost = orig_cost_by_lineage.get(lid, 0.0)

        for vi, v in enumerate(versions):
            r = lookup.get(lid, {}).get(v["id"])
            start_ci = FIXED + 1 + vi * vcols

            if r is None:
                for si in range(vcols):
                    cell = ws.cell(row=row_idx, column=start_ci + si, value="—")
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = _absent_fill
            else:
                cost = _row_cost_dict(r)
                values = [r.get("qty"), r.get("price_work"), r.get("price_material"),
                          cost if not is_section else None]
                for si, val in enumerate(values):
                    cell = ws.cell(row=row_idx, column=start_ci + si, value=val)
                    cell.border = THIN_BORDER
                    if val is not None and isinstance(val, (int, float)):
                        cell.number_format = _NUMBER_FMT
                    if is_section:
                        cell.font = BOLD_FONT
                        cell.fill = _SECTION_FILL
                    elif vi > 0 and si == vcols - 1 and not is_section:
                        if cost < orig_cost - 0.01:
                            cell.fill = _CHEAPER_FILL
                        elif cost > orig_cost + 0.01:
                            cell.fill = _DEARER_FILL

        row_idx += 1

    # Totals block
    row_idx += 1

    def _calc_totals(v: dict) -> dict:
        v_rows = v["rows"]
        works = sum(_row_cost_dict(r) for r in v_rows if r.get("type") == "work")
        materials = sum(_row_cost_dict(r) for r in v_rows if r.get("type") == "material")
        base = works + materials
        ovh = base * v["overhead_pct"] / 100
        trp = base * v["transport_pct"] / 100
        cng = base * v["contingency_pct"] / 100
        total = base + ovh + trp + cng
        vat = total * _VAT_RATE
        return {
            "Работы": works, "Материалы": materials,
            "Итого (базис)": base,
            f"Накладные ({v['overhead_pct']}%)": ovh,
            f"Транспорт ({v['transport_pct']}%)": trp,
            f"Непредвиден. ({v['contingency_pct']}%)": cng,
            "Итого": total,
            f"НДС {int(_VAT_RATE * 100)}%": vat,
            "ИТОГО с НДС": total + vat,
        }

    all_totals = [_calc_totals(v) for v in versions]
    total_keys = list(all_totals[0].keys())
    _grand_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for key in total_keys:
        is_grand = key == "ИТОГО с НДС"
        row_fill = _grand_blue if is_grand else _GRAND_FILL

        label_cell = ws.cell(row=row_idx, column=3, value=key)
        label_cell.font = Font(bold=True, color="FFFFFF" if is_grand else "000000", size=11)
        label_cell.fill = row_fill
        label_cell.border = THIN_BORDER

        for ci in [1, 2, 4]:
            cell = ws.cell(row=row_idx, column=ci)
            cell.fill = row_fill
            cell.border = THIN_BORDER

        orig_val = all_totals[0].get(key, 0)
        for vi, vt in enumerate(all_totals):
            val = vt.get(key, 0)
            cost_col = FIXED + vi * vcols + vcols
            cell = ws.cell(row=row_idx, column=cost_col, value=val)
            cell.number_format = _NUMBER_FMT
            cell.border = THIN_BORDER
            if is_grand:
                cell.fill = _grand_blue
                cell.font = Font(bold=True, color="FFFFFF", size=11)
            elif vi > 0 and val < orig_val - 0.01:
                cell.fill = _CHEAPER_FILL
                cell.font = BOLD_FONT
            elif vi > 0 and val > orig_val + 0.01:
                cell.fill = _DEARER_FILL
                cell.font = BOLD_FONT
            else:
                cell.fill = _GRAND_FILL
                cell.font = BOLD_FONT

            for si in range(vcols - 1):
                c = ws.cell(row=row_idx, column=FIXED + 1 + vi * vcols + si)
                c.fill = _grand_blue if is_grand else _GRAND_FILL
                c.border = THIN_BORDER

        row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
