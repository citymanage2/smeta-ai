"""Parse an estimate Excel file into a list of EstimateRow dicts.

Each returned dict matches EstimateRowSchema.
"""
import io
import uuid
from typing import Optional

import openpyxl


# Keywords used to detect column roles (case-insensitive, stripped)
_NUM_KW = ("№", "num", "n", "п/п", "пп")
_TYPE_KW = ("тип", "вид", "type")
_NAME_KW = ("наименование", "наименов", "name", "описание")
_UNIT_KW = ("ед", "unit", "ед.изм", "единица")
_QTY_KW = ("кол", "qty", "количество", "объем", "объём")
_PRICE_WORK_KW = ("цена работ", "стоимость работ", "стоим. работ", "стоим работ", "стоим.работ",
                  "труд", "labor", "работа", "price_work", "цр", "работ")
_PRICE_MAT_KW = ("цена матер", "стоимость матер", "стоим. матер", "стоим матер", "стоим.матер",
                 "матер", "material", "price_material", "цм")
_SKIP_KW = ("сумма", "итого", "total", "cost", "всего")

# Russian prepositions/conjunctions that mark comment rows (start with lowercase)
_COMMENT_STARTS = ("с ", "без ", "при ", "в ", "на ", "из ", "по ", "за ",
                   "до ", "от ", "со ", "не ", "и ", "или ")

# Minimum merged-cell span to treat a row as a section header
_SECTION_MERGE_SPAN = 3


def _header_matches(cell_val: str, keywords: tuple) -> bool:
    v = cell_val.lower().strip()
    return any(v.startswith(k) or k in v for k in keywords)


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        f = float(val)
        return f if f != 0.0 else None
    s = str(val).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        f = float(s)
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None


_BARE_PRICE_KW = ("цена", "price", "стоимость", "цена (руб", "цена руб")


def _detect_columns(header_row: list) -> dict:
    """Return {role: col_index} mapping from a header row (0-based)."""
    cols: dict = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        v = str(cell)
        if "num" not in cols and _header_matches(v, _NUM_KW):
            cols["num"] = idx
        elif "type" not in cols and _header_matches(v, _TYPE_KW):
            cols["type"] = idx
        elif "name" not in cols and _header_matches(v, _NAME_KW):
            cols["name"] = idx
        elif "unit" not in cols and _header_matches(v, _UNIT_KW):
            cols["unit"] = idx
        elif "qty" not in cols and _header_matches(v, _QTY_KW):
            cols["qty"] = idx
        elif "price_work" not in cols and _header_matches(v, _PRICE_WORK_KW):
            cols["price_work"] = idx
        elif "price_material" not in cols and _header_matches(v, _PRICE_MAT_KW):
            cols["price_material"] = idx
        elif "price_material_total" not in cols and _header_matches(v, _PRICE_MAT_KW):
            cols["price_material_total"] = idx

    # Fallback: bare "цена" / "стоимость" column → assign to whichever price slot is missing.
    # In estimates where work and materials sit on separate rows the single unqualified
    # "цена" header is always the unit-price column for the row type (work or material).
    # We map it to price_work first; if price_work is already taken, to price_material.
    if "price_work" not in cols or "price_material" not in cols:
        assigned = set(cols.values())
        for idx, cell in enumerate(header_row):
            if cell is None or idx in assigned:
                continue
            v = str(cell).lower().strip()
            # Match bare price keywords but NOT qualified ones (already handled above)
            is_bare = any(v == k or v.startswith(k + " ") or v.startswith(k + "(")
                          for k in _BARE_PRICE_KW)
            if not is_bare:
                continue
            # Skip if it already looks like a work/material qualified header
            if _header_matches(v, _PRICE_WORK_KW) or _header_matches(v, _PRICE_MAT_KW):
                continue
            if "price_work" not in cols:
                cols["price_work"] = idx
                assigned.add(idx)
            elif "price_material" not in cols:
                cols["price_material"] = idx
                assigned.add(idx)

    return cols


def _infer_type(price_work: Optional[float], price_material: Optional[float]) -> str:
    pw = price_work or 0.0
    pm = price_material or 0.0
    if pw > 0 and pm == 0:
        return "work"
    if pm > 0 and pw == 0:
        return "material"
    if pw > 0 and pm > 0:
        return "work"
    return "section"


def _is_comment_name(name: str) -> bool:
    """Return True if name looks like a comment/continuation clause to skip."""
    if not name:
        return False
    # Starts with lowercase → likely a note/continuation (e.g. "с сохранением")
    if name[0].islower():
        return True
    # Very short text with no uppercase start (after stripping punctuation)
    stripped = name.lstrip("- •·–—")
    if stripped and stripped[0].islower():
        return True
    return False


def link_materials_to_works(rows: list[dict]) -> list[dict]:
    """Set work_row_id on each material row pointing to the preceding work row.

    Section rows do not reset the link — materials after a section header
    still belong to the last seen work. Multiple works in a row: subsequent
    materials link to the last work (known limitation, documented in plan).
    """
    last_work_id = None
    for row in rows:
        if row.get("type") == "work":
            last_work_id = row["id"]
        elif row.get("type") == "material" and last_work_id is not None:
            row["work_row_id"] = last_work_id
    return rows


def _parse_estimate_sheet(ws) -> list[dict]:
    """Строки одного листа сметы. Шапка ищется в нём же — у листов она своя."""
    # Build merged-cell lookup: (1-based row, 1-based col) → span width
    # Section headers in construction estimates are typically merged across all columns.
    merged_spans: dict[tuple[int, int], int] = {}
    for merged_range in ws.merged_cells.ranges:
        span = merged_range.max_col - merged_range.min_col + 1
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                merged_spans[(r, c)] = span

    # Iterate rows as Cell objects (not values_only) so we can inspect formatting
    rows_raw = list(ws.iter_rows(values_only=False))
    if not rows_raw:
        return []

    # Find header row: first row that contains a name-like keyword
    header_idx: Optional[int] = None
    cols: dict = {}
    for i, row in enumerate(rows_raw):
        row_strs = [str(c.value) if c.value is not None else "" for c in row]
        combined = " ".join(row_strs).lower()
        if any(k in combined for k in ("наименование", "наименов", "name")):
            cols = _detect_columns(row_strs)
            if "name" in cols:
                header_idx = i
                break

    if header_idx is None or "name" not in cols:
        return []

    result = []
    num_counter = 0

    for row_offset, row in enumerate(rows_raw[header_idx + 1:]):
        # Excel row number (1-based) for merged-cell lookup
        excel_row = header_idx + 2 + row_offset

        name_col_idx = cols["name"]
        name_cell = row[name_col_idx] if name_col_idx < len(row) else None
        name_val = name_cell.value if name_cell is not None else None
        if name_val is None or str(name_val).strip() == "":
            continue

        name = str(name_val).strip()

        # Skip obvious total/footer rows
        low = name.lower()
        if any(k in low for k in ("итого", "всего", "total", "grand total", "в том числе")):
            continue

        # Skip comment/continuation lines (e.g. "с сохранением")
        if _is_comment_name(name):
            continue

        # Read quantitative fields
        unit_val = row[cols["unit"]].value if "unit" in cols and cols["unit"] < len(row) else None
        unit = str(unit_val).strip() if unit_val is not None else ""

        qty_val = row[cols["qty"]].value if "qty" in cols and cols["qty"] < len(row) else None
        qty = _to_float(qty_val)

        pw_val = row[cols["price_work"]].value if "price_work" in cols and cols["price_work"] < len(row) else None
        pm_val = row[cols["price_material"]].value if "price_material" in cols and cols["price_material"] < len(row) else None
        pm_total_val = row[cols["price_material_total"]].value if "price_material_total" in cols and cols["price_material_total"] < len(row) else None
        price_work = _to_float(pw_val)
        price_material = _to_float(pm_val)
        price_material_total = _to_float(pm_total_val)

        # For type detection use either unit price or total (unit price preferred)
        mat_price_signal = price_material if price_material is not None else price_material_total
        has_prices = price_work is not None or mat_price_signal is not None
        has_qty = qty is not None
        has_unit = bool(unit)

        # Determine section vs work/material
        # Check Excel formatting: merged wide → definitive section header
        name_col_1based = name_col_idx + 1
        merge_span = merged_spans.get((excel_row, name_col_1based), 1)
        is_merged_wide = merge_span >= _SECTION_MERGE_SPAN

        is_bold = bool(
            name_cell is not None
            and hasattr(name_cell, "font")
            and name_cell.font is not None
            and name_cell.font.bold
        )

        if is_merged_wide or (is_bold and not has_prices and not has_qty and not has_unit):
            # Explicit structural section header from Excel formatting
            row_type = "section"
        elif not has_prices and not has_qty and not has_unit:
            # No data → section header (name-only row)
            row_type = "section"
        else:
            # Has some quantitative data → determine work/material
            if "type" in cols and cols["type"] < len(row):
                raw_type = str(row[cols["type"]].value or "").lower().strip()
                if "работ" in raw_type or raw_type in ("work", "w", "р"):
                    row_type = "work"
                elif "матер" in raw_type or raw_type in ("material", "m", "м"):
                    row_type = "material"
                else:
                    row_type = _infer_type(price_work, mat_price_signal)
            else:
                row_type = _infer_type(price_work, mat_price_signal)

            # If infer returned "section" but row has unit/qty, keep it as work
            if row_type == "section" and (has_qty or has_unit):
                row_type = "work"

        # Recover material unit price from total when unit price column is empty
        if row_type == "material" and price_material is None and price_material_total is not None:
            if qty and qty > 0:
                price_material = round(price_material_total / qty, 2)
            else:
                price_material = price_material_total

        # Row number
        num_val = row[cols["num"]].value if "num" in cols and cols["num"] < len(row) else None
        num = None
        if num_val is not None:
            try:
                num = int(float(str(num_val)))
            except (ValueError, TypeError):
                pass
        if num is None:
            num_counter += 1
            num = num_counter

        cost: Optional[float] = None
        if qty is not None:
            pw = price_work or 0.0
            pm = price_material or 0.0
            cost = round(qty * (pw + pm), 2) if (pw + pm) > 0 else None

        row_id = str(uuid.uuid4())
        result.append({
            "id": row_id,
            "lineage_id": row_id,
            "num": num,
            "type": row_type,
            "name": name,
            "unit": unit,
            "qty": qty,
            "price_work": price_work,
            "price_material": price_material,
            "cost": cost,
            "selected": False,
            "abc_group": None,
            "optimization_note": None,
        })

    return link_materials_to_works(result)


def parse_estimate_excel(file_bytes: bytes) -> list[dict]:
    """Parse xlsx bytes into a list of EstimateRow dicts.

    Uses openpyxl with data_only=True so formula cells return computed values.
    Returns an empty list if the file cannot be parsed or no recognizable header found.

    Разбираются все листы файла: смету присылают разбитой по листам — по листу
    на раздел или корпус, — и взятый один первый лист терял остальные молча.
    Лист остаётся при строке полем `sheet` и становится вкладкой в редакторе.
    Лист без опознаваемой шапки (титульный, сводка итогов) строк не даёт.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return []

    by_sheet = [(ws, _parse_estimate_sheet(ws)) for ws in wb.worksheets]
    multi_sheet = sum(1 for _, rows in by_sheet if rows) > 1

    result: list[dict] = []
    for ws, rows in by_sheet:
        for row in rows:
            # Один лист — признак не ставим: документ остаётся без вкладок и
            # ведёт себя ровно как до появления многолистовых файлов.
            if multi_sheet:
                row["sheet"] = ws.title
            result.append(row)
    return result
