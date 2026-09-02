"""Эталонный прайс: цены из файла становятся единственно верными.

План `plans/2026-09-02-etalonnyy-prays-iz-smety.md`.

Обычная загрузка прайса (`admin._handle_price_upload`) — это **слияние**: цены
подрядчиков по той же позиции остаются рядом с новой. Здесь наоборот: файл
объявлен эталоном, и по его позициям в прайсе должна остаться ровно одна цена —
из файла. Операция необратима и трогает общий на всех справочник, поэтому
модуль разделён на два шага, и первый ничего не меняет:

- `parse_reference_file` — что вообще написано в файле (и что из него выброшено
  и почему);
- `build_plan` — что произойдёт с прайсом: где добавим, где сотрём чужую цену и
  чью именно, а где вытеснять отказываемся.

Отказ вытеснять — не осторожность ради осторожности. Единица измерения это
часть цены, а не подпись (правило №8 CLAUDE.md): цена за мешок и цена за кг
назначены за разное, и стереть первую ради второй значит молча исказить прайс
подрядчика. Такие позиции показываются человеку, а не чинятся догадкой.
"""
import asyncio
import csv
import io
from datetime import datetime, timezone
from typing import Optional

import openpyxl
import structlog
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.price import PriceMaterial, PriceWork
from app.models.price_cache import PriceCacheMaterial, PriceCacheWork
from app.services import price_service
from app.services.embedding_service import (
    EmbeddingUnavailableError,
    generate_embeddings_batch,
    normalize_name,
)
from app.services.price_bulk import SKIP_NO_NAME, SKIP_NO_PRICE, SKIP_NOT_PRICEABLE
from app.utils.price_change import MONEY_EPS, price_changed, prices_changed
from app.utils.price_min import ESTIMATE_CONTRACTOR, compute_min_price
from app.utils.unit_compat import STATUS_INCOMPATIBLE, compare_units
from app.utils.unit_normalizer import unit_price_factor

logger = structlog.get_logger()

# Потолок на один файл. Реальная сводная смета на объект — это полторы тысячи
# позиций, разложенных по десятку листов-разделов, и отклонять её незачем: цены
# там ровно те, ради которых функция и делалась. Файл сверх потолка отклоняется
# целиком — молча взять первые N строк хуже, чем не взять ничего.
MAX_REFERENCE_ITEMS = 3000

ACTION_ADD = "add"          # позиции в прайсе нет — стирать нечего
ACTION_REPRICE = "reprice"  # позиция есть — цена из файла вытеснит остальные
ACTION_BLOCKED = "blocked"  # вытеснять нельзя, решает человек

BLOCK_UNIT_MISMATCH = "ед. изм. несводима"

# Предупреждения разбора: не ошибки, но человек обязан их увидеть до записи.
NOTE_BOTH_PRICES = "в строке две цены — взята цена материала"
NOTE_PRICE_CONFLICT = "в файле разные цены у одной позиции — взята последняя"
NOTE_SPLIT_BY_SECTION = "одноимённые позиции разведены по разделам файла"

# Сколько кандидатов-дублей показывать на одну позицию файла.
DUPLICATE_TOP_N = 5

# Насколько порог показа дубля ниже порога, по которому цена подставляется в
# смету. Порог подбора настроен так, чтобы **не взять** чужую цену; здесь цена
# не подставляется, а показывается человеку — и скрытый дубль обходится дороже
# лишней строки в списке, потому что молча продолжит подставляться в расчёт.
DUPLICATE_MARGIN = 0.08

KIND_WORK = "work"
KIND_MATERIAL = "material"

# Что в колонке «Тип» файла сметы означает работу и материал. Всё остальное
# (раздел, пустая ячейка) ценой не является.
_TYPE_WORK = "работа"
_TYPE_MATERIAL = "материал"


class ReferenceFileError(Exception):
    """Файл нельзя принять целиком: не разобрали или он больше потолка."""


# ---------------------------------------------------------------------------
# Разбор файла
# ---------------------------------------------------------------------------

def match_key(name: str) -> str:
    """Ключ «та же самая позиция» — общий для всех путей записи в прайс.

    Это `embedding_service.normalize_name`, тот же, которым ищут существующую
    позицию загрузка прайса файлом и «Добавить в прайс». Иначе эталон заводил
    бы вторую строку там, где загрузка файлом сливает две: «Раствор М-100» и
    «Раствор М 100» — одна позиция, записанная по-разному.

    С 02.09.2026 этой функции можно доверять: до правки она сжимала «буква +
    слова + число» до марки материала и склеивала разные позиции (шесть планок
    кровли в одну), и эталон вынужден был считать ключ сам. Теперь правило
    марок применяется только к отдельно стоящей букве, склеек нет, и второе
    определение «та же позиция» в проекте не нужно (правило №19 CLAUDE.md).
    """
    return normalize_name(name)


def _sheets_from_bytes(data: bytes, filename: str) -> "list[tuple[str, list[list]]]":
    """Листы файла как таблицы строк.

    Из книги берутся **все** листы, а не первый подходящий: сводная смета
    разложена по разделам — «Кровля», «АР», «ЭОМ» и ещё десяток (правило №14
    CLAUDE.md). Какой лист про цены, а какой служебный, решает уже разбор
    колонок: у листа-раздела есть цена за единицу, у «Сводной» — только суммы.
    """
    lower = filename.lower()

    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:  # noqa: BLE001 — наружу уходит человеческий текст
            raise ReferenceFileError(f"Не удалось прочитать книгу Excel: {e}") from e
        return [
            (name, [list(row) for row in wb[name].iter_rows(values_only=True)])
            for name in wb.sheetnames
        ]

    if lower.endswith(".csv") or lower.endswith(".txt"):
        text = data.decode("utf-8-sig", errors="replace")
        try:
            delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
        return [("", [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)])]

    raise ReferenceFileError("Принимаются файлы Excel, CSV и TXT.")


def _header_index(rows: list[list]) -> Optional[int]:
    """Номер строки заголовка — первой, где встречается «наименование».

    В реальной сводной шапка живёт на восьмой строке, над ней — название
    объекта и пустые строки, поэтому фиксированный номер строки не годится.
    """
    for i, row in enumerate(rows):
        for cell in row:
            if isinstance(cell, str) and "наименование" in cell.lower():
                return i
    return None


def _columns(headers: list[str]) -> dict:
    """Заголовки → номера нужных колонок.

    Цены разбираются **до** единицы и до общей «Цены»: настоящий заголовок
    выглядит как «Цена за ед. изм. Работы» — в нём есть и «ед.», и «цена», и
    любая другая ветка забрала бы колонку себе.

    Колонка «Стоимость» ценой не считается вовсе. В смете это сумма за всю
    позицию: взять её ценой за единицу значит завысить прайс во столько раз,
    сколько было объёма.
    """
    columns: dict = {
        "name": None, "unit": None, "type": None,
        "price_work": None, "price_material": None, "price": None,
    }

    for i, raw in enumerate(headers):
        lower = str(raw or "").replace("\n", " ").strip().lower()
        if not lower:
            continue
        if "наименование" in lower or "название" in lower or lower == "name":
            if columns["name"] is None:
                columns["name"] = i
        elif lower == "тип":
            columns["type"] = i
        elif "цена" in lower and "работ" in lower:
            if columns["price_work"] is None:
                columns["price_work"] = i
        elif "цена" in lower and "материал" in lower:
            if columns["price_material"] is None:
                columns["price_material"] = i
        elif "цена" in lower or lower == "price":
            if columns["price"] is None:
                columns["price"] = i
        elif ("ед" in lower and ("изм" in lower or "." in lower)) or lower in ("ед", "unit"):
            if columns["unit"] is None:
                columns["unit"] = i

    return columns


def _cell(row: list, index: Optional[int]):
    if index is None or index >= len(row):
        return None
    return row[index]


def _amount(value: object) -> Optional[float]:
    """Цена из ячейки. Ноль, пустая строка и мусор — это «цены нет»."""
    if value is None:
        return None
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if number > 0 else None


def _kind_from_type(value: object) -> Optional[str]:
    text = str(value or "").strip().lower()
    if text.startswith(_TYPE_WORK):
        return KIND_WORK
    if text.startswith(_TYPE_MATERIAL):
        return KIND_MATERIAL
    return None


def _is_column_numbers_row(row: list, columns: dict) -> bool:
    """Строка «1, 2, 3 …» под шапкой — разметка листа, а не позиция.

    У неё есть и «наименование», и «единица», и «цена» — цифры соседних
    колонок, — поэтому отбрасывать её нужно целиком (то же правило, что в
    `file_parser._is_trash_row`): иначе в прайс уедет позиция с именем «2» и
    ценой 5.
    """
    seen = 0
    for index in (columns["name"], columns["unit"], columns["price_work"],
                  columns["price_material"], columns["price"]):
        value = _cell(row, index)
        if value is None or str(value).strip() == "":
            continue
        try:
            number = float(str(value).strip().replace(",", "."))
        except ValueError:
            return False
        if number != int(number) or not (1 <= number <= 99):
            return False
        seen += 1
    return seen >= 2


def _parse_sheet(
    sheet: str,
    rows: list[list],
    kind: Optional[str],
    prepared: dict,
    skipped: dict,
    notes: dict,
    conflicts: dict,
) -> int:
    """Позиции одного листа — в общий `prepared`. Возвращает число принятых строк.

    Лист без колонки цены за единицу пропускается целиком и молча: в книге
    рядом с разделами лежат «Сводная» и «Целевые показатели», где в шапке тоже
    есть «Наименование», а в строках — «Работы» и «Материалы» с суммами в
    десятки миллионов. Попади они в прайс — там появилась бы позиция «Работы»
    по 54 миллиона за штуку.
    """
    header_idx = _header_index(rows)
    if header_idx is None:
        return 0

    headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
    columns = _columns(headers)
    if columns["name"] is None:
        return 0

    split_prices = columns["price_work"] is not None or columns["price_material"] is not None
    if not split_prices:
        if columns["price"] is None:
            return 0
        if kind not in (KIND_WORK, KIND_MATERIAL):
            raise ReferenceFileError(
                "В файле не написано, где работы, а где материалы — выберите тип."
            )

    def _count(store: dict, reason: str) -> None:
        store[reason] = store.get(reason, 0) + 1

    accepted = 0
    for row in rows[header_idx + 1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        if _is_column_numbers_row(row, columns):
            continue

        name = str(_cell(row, columns["name"]) or "").strip()

        if split_prices:
            row_kind = (
                _kind_from_type(_cell(row, columns["type"]))
                if columns["type"] is not None else None
            )
            work_price = _amount(_cell(row, columns["price_work"]))
            material_price = _amount(_cell(row, columns["price_material"]))
            if row_kind is None:
                if work_price is not None and material_price is not None:
                    # Кабель с ценой монтажа: названия работы в такой строке
                    # нет — в колонке написан кабель. Взять её работой значило
                    # бы завести в прайсе работ позицию «Кабель витая пара».
                    row_kind = KIND_MATERIAL
                    _count(notes, NOTE_BOTH_PRICES)
                elif work_price is not None:
                    row_kind = KIND_WORK
                elif material_price is not None:
                    row_kind = KIND_MATERIAL
            if row_kind is None:
                if not name:
                    continue
                # Тип написан в файле и это не работа и не материал (раздел) —
                # причина точнее, чем «без цены»: у раздела её и не бывает.
                declared = str(_cell(row, columns["type"]) or "").strip() \
                    if columns["type"] is not None else ""
                _count(skipped, SKIP_NOT_PRICEABLE if declared else SKIP_NO_PRICE)
                continue
            price = work_price if row_kind == KIND_WORK else material_price
        else:
            row_kind = kind
            price = _amount(_cell(row, columns["price"]))

        if not name:
            _count(skipped, SKIP_NO_NAME)
            continue
        if name.lower() in ("итого", "всего", "total"):
            continue
        if price is None:
            _count(skipped, SKIP_NO_PRICE)
            continue

        # Цена за «100 м2» — это не цена за м2 (правило №3 плана).
        unit, factor = unit_price_factor(_cell(row, columns["unit"]))
        item = {
            "kind": row_kind,
            "name": name,
            "unit": unit,
            "price": round(price / factor, 2),
            "sheet": sheet,
        }

        # Лист сводной — это система (ОПС, СОТ, ЛВС), и одноимённая работа в
        # разных системах стоит по-разному не по ошибке (решение пользователя
        # 02.09.2026). Поэтому внутри разбора позиции разведены по листам, а
        # схлопывать их или нет — решается потом, по совпадению цены.
        key = (row_kind, match_key(name), sheet)
        previous = prepared.get(key)
        if previous is not None and abs(previous["price"] - item["price"]) > MONEY_EPS:
            # Файл сам себе противоречит. Побеждает последняя цена — но молчать
            # об этом нельзя: эталон объявляет единственно верную цену, и какая
            # из нескольких верна, решает человек, а не порядок строк.
            if key not in conflicts:
                _count(notes, NOTE_PRICE_CONFLICT)
            conflicts.setdefault(key, [previous["price"]]).append(item["price"])
        prepared[key] = item
        accepted += 1

    return accepted




def _merge_across_sheets(prepared: dict, notes: dict) -> list[dict]:
    """Одноимённые позиции из разных разделов — одна или несколько?

    Решает цена. Совпала во всех разделах — это и правда одна позиция, и
    уточнять нечего. Разошлась — разделы говорят о разных вещах («Демонтаж
    кабеля» в пожарке и в видеонаблюдении), и в прайс они идут порознь, с
    названием раздела в скобках. Иначе один кабель молча вытеснил бы другой.
    """
    groups: dict = {}
    for (kind, name_key, sheet), item in prepared.items():
        groups.setdefault((kind, name_key), []).append(item)

    items: list[dict] = []
    for group in groups.values():
        prices = {round(item["price"], 2) for item in group}
        if len(group) == 1 or len(prices) == 1:
            items.append(group[0])
            continue

        notes[NOTE_SPLIT_BY_SECTION] = notes.get(NOTE_SPLIT_BY_SECTION, 0) + len(group)
        for item in group:
            named = dict(item)
            if item["sheet"]:
                named["name"] = f"{item['name']} ({item['sheet']})"
            items.append(named)

    return items


def parse_reference_file(
    data: bytes,
    filename: str,
    kind: Optional[str],
) -> dict:
    """Файл → {items, skipped, notes}.

    Поддерживаются два вида файла:

    - **смета с ценами за единицу** — есть колонки «Цена за ед. изм. Работы» и
      «Цена за ед. изм. Материала» (или «Цена работы (за ед.)» нашего
      экспорта); тип строки читается из колонки «Тип», а при её отсутствии — из
      того, в какой колонке стоит цена. Работы и материалы приезжают одним
      файлом, разложенным хоть по десятку листов-разделов;
    - **простой прайс** — «Наименование / Ед. изм. / Цена»; тип в файле не
      написан, поэтому его называет человек (`kind`). Выдумывать тип по имени
      позиции нельзя: ошибка отправит цену работы в прайс материалов.

    `notes` — счётчики того, что человек обязан увидеть до записи; `conflicts`
    — поимённый список позиций, у которых в самом файле цены разошлись, со
    всеми найденными ценами и той, что будет записана.
    """
    sheets = _sheets_from_bytes(data, filename)

    prepared: dict = {}
    skipped: dict[str, int] = {}
    notes: dict[str, int] = {}
    conflicts: dict = {}
    accepted = 0

    for sheet, rows in sheets:
        accepted += _parse_sheet(sheet, rows, kind, prepared, skipped, notes, conflicts)
        if accepted > MAX_REFERENCE_ITEMS:
            raise ReferenceFileError(
                f"В файле больше {MAX_REFERENCE_ITEMS} позиций с ценой — "
                "столько за раз не принимаем."
            )

    if not prepared and not skipped:
        raise ReferenceFileError(
            "В файле не нашлось ни одной позиции с ценой за единицу. "
            "Нужны колонки «Наименование» и «Цена» (или «Цена за ед. изм. Работы» "
            "и «Цена за ед. изм. Материала»)."
        )

    conflict_list = [
        {
            "kind": key[0],
            "name": prepared[key]["name"],
            "unit": prepared[key]["unit"],
            "prices": sorted(set(prices)),
            "taken": prepared[key]["price"],
        }
        for key, prices in conflicts.items()
        if key in prepared
    ]

    items = _merge_across_sheets(prepared, notes)

    return {
        "items": items,
        "skipped": skipped,
        "notes": notes,
        "conflicts": conflict_list,
    }


# ---------------------------------------------------------------------------
# Дубли: то же самое под другим названием
# ---------------------------------------------------------------------------

def duplicate_threshold() -> float:
    """Порог показа кандидата. Читается на каждый вызов: порог подбора живёт в
    env и меняется без деплоя (`PRICE_SIMILARITY_THRESHOLD`)."""
    return price_service.SIMILARITY_THRESHOLD - DUPLICATE_MARGIN


# ---------------------------------------------------------------------------
# План вытеснения
# ---------------------------------------------------------------------------

def _removed_work_prices(prices: Optional[dict]) -> list[dict]:
    """Чужие цены работы, которые исчезнут. Цена «Из смет» чужой не считается:
    её перезапись — обычное обновление, а не потеря прайса подрядчика."""
    if not isinstance(prices, dict):
        return []
    removed = []
    for contractor, value in prices.items():
        if str(contractor) == ESTIMATE_CONTRACTOR:
            continue
        try:
            amount = float(value)
        except (TypeError, ValueError):
            continue
        if amount > 0:
            removed.append({"contractor": str(contractor), "price": amount})
    return removed


async def find_duplicates(items: list[dict]) -> dict:
    """Позиции прайса и кеша, названные иначе, но, похоже, про то же самое.

    Ничего не удаляет и не отмечает: решение за человеком. Позиция, которая и
    так будет переоценена (точное совпадение названия), дублем не считается —
    предложить её к удалению значило бы предложить стереть ту самую строку,
    куда мы записываем эталонную цену.
    """
    kinds = {item["kind"] for item in items} or {KIND_WORK, KIND_MATERIAL}
    vectors_ready = all(price_service.duplicate_vectors_ready(kind) for kind in kinds)
    threshold = duplicate_threshold()

    # Имена ищутся пачкой на каждый тип: модель прогоняется один раз на файл,
    # а не по разу на позицию (см. `find_duplicate_candidates_batch`).
    by_kind: dict[str, list[int]] = {}
    for index, item in enumerate(items):
        by_kind.setdefault(item["kind"], []).append(index)

    found_per_item: list[list[dict]] = [[] for _ in items]
    for kind, indexes in by_kind.items():
        rows = await price_service.find_duplicate_candidates_batch(
            [items[i]["name"] for i in indexes], kind, n=DUPLICATE_TOP_N,
        )
        for index, row in zip(indexes, rows):
            found_per_item[index] = row

    seen: set = set()
    candidates: list[dict] = []

    for item, found in zip(items, found_per_item):
        own_key = match_key(item["name"])
        for candidate in found:
            score = float(candidate.get("score") or 0.0)
            if score < threshold:
                continue
            if match_key(str(candidate.get("name") or "")) == own_key:
                continue
            key = (candidate.get("source"), item["kind"], candidate.get("id"))
            if key in seen:
                continue
            seen.add(key)
            candidates.append({
                "source": candidate.get("source"),
                "kind": item["kind"],
                "id": candidate.get("id"),
                "name": candidate.get("name"),
                "unit": candidate.get("unit"),
                "price": candidate.get("price"),
                "score": round(score, 3),
                "for_name": item["name"],
            })

    candidates.sort(key=lambda c: c["score"], reverse=True)
    return {"vectors_ready": vectors_ready, "candidates": candidates}


async def build_plan(db: AsyncSession, items: list[dict]) -> list[dict]:
    """Что произойдёт с прайсом. Ничего не пишет — только читает.

    У каждой позиции файла один из трёх исходов: `add` (в прайсе такой нет),
    `reprice` (есть, цена из файла вытеснит остальные) и `blocked` (единица
    несводима с той, за которую назначена цена в прайсе).
    """
    work_rows = (await db.execute(
        select(PriceWork.id, PriceWork.name, PriceWork.unit, PriceWork.prices)
        .order_by(PriceWork.id)
    )).all()
    material_rows = (await db.execute(
        select(PriceMaterial.id, PriceMaterial.name, PriceMaterial.unit, PriceMaterial.price)
        .order_by(PriceMaterial.id)
    )).all()

    # Дубли имён в самом прайсе разрешаем в пользу самой ранней записи
    # (строки читаются по возрастанию id) — как это делает загрузка прайса
    # файлом: иначе результат зависел бы от порядка выдачи БД.
    works: dict = {}
    for row in work_rows:
        works.setdefault(match_key(row.name), row)
    materials: dict = {}
    for row in material_rows:
        materials.setdefault(match_key(row.name), row)

    plan: list[dict] = []
    for item in items:
        kind = item["kind"]
        found = (works if kind == KIND_WORK else materials).get(match_key(item["name"]))

        entry = {
            "kind": kind,
            "name": item["name"],
            "unit": item["unit"],
            "price": item["price"],
            "action": ACTION_ADD,
            "match": None,
            "removed": [],
            "reason": None,
        }

        if found is not None:
            entry["match"] = {
                "id": found.id,
                "name": found.name,
                "unit": found.unit,
            }
            status, _ = compare_units(found.unit, item["unit"])
            if status == STATUS_INCOMPATIBLE:
                entry["action"] = ACTION_BLOCKED
                entry["reason"] = BLOCK_UNIT_MISMATCH
            else:
                entry["action"] = ACTION_REPRICE
                if kind == KIND_WORK:
                    entry["removed"] = _removed_work_prices(found.prices)
                else:
                    old = found.price
                    try:
                        old_amount = float(old) if old is not None else None
                    except (TypeError, ValueError):
                        old_amount = None
                    if old_amount is not None and old_amount > 0 \
                            and abs(old_amount - item["price"]) > MONEY_EPS:
                        entry["removed"] = [{"contractor": None, "price": old_amount}]

        plan.append(entry)

    return plan


# ---------------------------------------------------------------------------
# Применение
# ---------------------------------------------------------------------------

# Что откуда удаляется. Прайс и кеш веб-поиска — разные таблицы, но для расчёта
# сметы это два одинаково живых источника цены.
_REMOVE_MODELS = {
    ("price", KIND_WORK): PriceWork,
    ("price", KIND_MATERIAL): PriceMaterial,
    ("cache", KIND_WORK): PriceCacheWork,
    ("cache", KIND_MATERIAL): PriceCacheMaterial,
}


async def _embeddings_safe(names: list[str]) -> list[Optional[list]]:
    """Векторы для поиска по смыслу — одной прогонкой модели на все имена.

    Позиция без вектора всё равно найдётся точным совпадением названия, поэтому
    недоступная модель не повод отказывать в записи (так же в `price_bulk`).
    По вектору за раз здесь нельзя: эталон из сметы — это сотни новых позиций,
    и поштучная генерация превратила бы запись в минуты ожидания.
    """
    if not names:
        return []
    try:
        return await asyncio.to_thread(
            generate_embeddings_batch, names, "search_document",
        )
    except EmbeddingUnavailableError:
        return [None] * len(names)


def _mark_key(source: object, kind: object, item_id: object) -> tuple:
    """Ключ позиции для сверки «это не та строка, куда мы только что писали».
    Идентификаторы приезжают из JSON то числом, то строкой — сравниваем текстом."""
    return (str(source), str(kind), str(item_id))


async def apply_reference(
    db: AsyncSession,
    items: list[dict],
    remove: list[dict],
) -> dict:
    """Записать эталонные цены и удалить отмеченное человеком.

    План строится заново, а не принимается с фронта: между предпросмотром и
    применением прайс мог измениться, а решение «вытеснять или нет» зависит от
    единицы измерения той позиции, что лежит в базе **сейчас**.

    Удаление идёт последним и никогда не трогает строки, в которые эта же
    операция записала цену: иначе отметка, поставленная по недосмотру, стёрла
    бы собственный результат операции.
    """
    plan = await build_plan(db, items)
    by_name = {(item["kind"], match_key(item["name"])): item for item in items}
    now = datetime.now(timezone.utc)

    added = 0
    updated = 0
    blocked = 0
    protected: set[tuple] = set()

    # Векторы для новых позиций — одной прогонкой до записи. У существующих
    # имя не меняется (эталон объявляет цену, а не переименовывает позицию),
    # поэтому их вектор трогать незачем.
    new_names = [
        entry["name"] for entry in plan
        if entry["action"] == ACTION_ADD
    ]
    embeddings = await _embeddings_safe(new_names)
    embedding_by_name = dict(zip(new_names, embeddings))

    for entry in plan:
        if entry["action"] == ACTION_BLOCKED:
            blocked += 1
            continue

        item = by_name[(entry["kind"], match_key(entry["name"]))]
        is_work = entry["kind"] == KIND_WORK
        model = PriceWork if is_work else PriceMaterial

        if entry["match"] is None:
            values = {
                "name": item["name"],
                "unit": item["unit"],
                "embedding": embedding_by_name.get(entry["name"]),
                "updated_at": now,
            }
            if is_work:
                prices = {ESTIMATE_CONTRACTOR: item["price"]}
                row = PriceWork(prices=prices, min_price=compute_min_price(prices), **values)
            else:
                row = PriceMaterial(price=item["price"], **values)
            db.add(row)
            await db.flush()
            protected.add(_mark_key("price", entry["kind"], row.id))
            added += 1
            continue

        found = await db.get(model, entry["match"]["id"])
        if found is None:  # позицию удалили между предпросмотром и применением
            continue

        # Единица из файла главнее: эталон объявляет, за что назначена цена.
        # Пустую единицу файла за объявление не считаем — она ничего не говорит.
        new_unit = item["unit"] or found.unit

        if is_work:
            prices = {ESTIMATE_CONTRACTOR: item["price"]}
            repriced = prices_changed(found.prices, prices, found.unit, new_unit)
            found.prices = prices
            found.min_price = compute_min_price(prices)
        else:
            repriced = price_changed(found.price, item["price"], found.unit, new_unit)
            found.price = item["price"]

        found.unit = new_unit
        if repriced:
            found.updated_at = now

        protected.add(_mark_key("price", entry["kind"], found.id))
        updated += 1

    removed = 0
    for mark in remove:
        source = str(mark.get("source") or "")
        kind = str(mark.get("kind") or "")
        item_id = mark.get("id")
        model = _REMOVE_MODELS.get((source, kind))
        if model is None or item_id is None:
            continue
        if _mark_key(source, kind, item_id) in protected:
            continue
        # Идентификатор приезжает из JSON как придётся, а типы колонок разные:
        # у прайса целое, у кеша строка UUID. Несовпадение типа в PostgreSQL —
        # ошибка запроса, а не «ничего не нашлось».
        if source == "price":
            try:
                item_id = int(item_id)
            except (TypeError, ValueError):
                continue
        else:
            item_id = str(item_id)
        result = await db.execute(delete(model).where(model.id == item_id))
        removed += result.rowcount or 0

    await db.commit()

    # Расчёт сметы читает прайс из памяти: без перезагрузки эталонная цена
    # заработала бы только после перезапуска сервера.
    await price_service.load_cache(db)

    logger.info(
        "reference_price_applied",
        added=added, updated=updated, blocked=blocked, removed=removed,
    )
    return {
        "added": added,
        "updated": updated,
        "blocked": blocked,
        "removed": removed,
    }
